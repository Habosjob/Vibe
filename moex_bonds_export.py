"""Скрипт выгружает торгуемые облигации MOEX и сохраняет Excel-отчёты."""

from __future__ import annotations

import json
import logging
import shutil
import time
import zipfile
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import moex_bonds_config as cfg


@dataclass
class MoexBlocks:
    descriptions: list[dict[str, Any]]
    coupons: list[dict[str, Any]]
    amortizations: list[dict[str, Any]]
    offers: list[dict[str, Any]]


RUSSIAN_COLUMN_NAMES = {
    "SECID": "Код бумаги",
    "BOARDID": "Код режима торгов",
    "SHORTNAME": "Краткое наименование",
    "PREVWAPRICE": "Средневзвешенная цена предыдущего дня",
    "YIELDATPREVWAPRICE": "Доходность по средневзвешенной цене",
    "NEXTCOUPON": "Дата следующего купона",
    "ACCRUEDINT": "НКД",
    "LATNAME": "Латинское наименование",
    "NAME": "Полное наименование",
    "ISIN": "ISIN",
    "REGNUMBER": "Регистрационный номер",
    "LISTLEVEL": "Уровень листинга",
    "FACEUNIT": "Валюта номинала",
    "PREVPRICE": "Цена предыдущей сделки",
    "LOTSIZE": "Лот",
    "FACEVALUE": "Номинал",
    "BOARDNAME": "Режим торгов",
    "MATDATE": "Дата погашения",
    "DECIMALS": "Знаков после запятой",
    "COUPONPERIOD": "Купонный период, дней",
    "ISSUESIZE": "Объем выпуска",
    "PREVLEGALCLOSEPRICE": "Официальная цена закрытия (пред.)",
    "PREVDATE": "Дата предыдущих торгов",
    "SECNAME": "Наименование ценной бумаги",
    "REMARKS": "Примечание",
    "MARKETCODE": "Код рынка",
    "INSTRID": "Код группы инструмента",
    "SECTORID": "Код сектора",
    "MINSTEP": "Минимальный шаг цены",
    "COUPONFREQUENCY": "Частота купона",
    "COUPONPERCENT": "Ставка купона, %",
    "COUPONVALUE": "Размер купона",
    "CURRENCYID": "Код валюты",
    "ISSUESIZEPLACED": "Объем размещения",
    "SECTYPE": "Тип ценной бумаги",
    "OFFERDATE": "Дата оферты (описание)",
    "SETTLEDATE": "Дата расчетов",
    "LOTVALUE": "Стоимость лота",
    "FACEVALUEONSETTLEDATE": "Номинал на дату расчетов",
    "CALLOPTIONDATE": "Дата call-оферты",
    "PUTOPTIONDATE": "Дата put-оферты",
    "DATEYIELDFROMISSUER": "Дата расчета доходности эмитентом",
    "BONDTYPE": "Тип облигации",
    "BONDSUBTYPE": "Подтип облигации",
    "BUYBACKPRICE": "Цена оферты",
    "BUYBACKDATE": "Дата оферты",
    "EMITENT_ID": "Код эмитента",
    "EMITENT_TITLE": "Эмитент",
    "EMITENT_INN": "ИНН эмитента",
    "EMITENT_OKPO": "ОКПО эмитента",
    "TYPE": "Тип инструмента",
    "GROUP": "Группа инструментов",
    "PRIMARY_BOARDID": "Основной режим торгов",
    "MARKETPRICE_BOARDID": "Режим цены рынка",
    "STATUS": "Статус",
    "IS_TRADED": "Торгуется",
    "NUMTRADES": "Количество сделок",
    "VOLTODAY": "Объем за день",
    "VALTODAY": "Оборот за день",
}

OUTPUT_COLUMNS_TO_DROP = {
    "Код режима торгов",
    "Лот",
    "Режим торгов",
    "Статус",
    "Знаков после запятой",
    "Наименование ценной бумаги",
    "Примечание",
    "Код рынка",
    "Код группы инструмента",
    "Код сектора",
    "Минимальный шаг цены",
    "Латинское наименование",
    "Регистрационный номер",
    "Код валюты",
    "Уровень листинга",
    "Тип ценной бумаги",
    "Дата расчетов",
    "Дата расчета доходности эмитентом",
    "Торгуется",
    "ОКПО эмитента",
    "Группа инструментов",
    "Основной режим торгов",
    "Режим цены рынка",
    "Государственный регистрационный номер программы облигации",
    "Дата государственной регистрации ценной бумаги",
    "Дата начала торгов",
    "Дата начала торгов на Московской Бирже",
    "Дата принятия решения организатором торговли о включении ценной бумаги в Список",
    "Допуск к вечерней дополнительной торговой сессии",
    "Допуск к утренней дополнительной торговой сессии",
    "ИЦБ допущена к орг. торгам по инициативе биржи",
    "Категория квалифицированного инвестора",
    "Код типа инструмента",
    "Наличие проспекта",
    "Номер государственной регистрации",
    "Облигации размещены с целью финансирования соглашений о партнерстве",
    "Полное наименование",
    "Сектор компаний повышенного инвестиционного риска (ПИР)",
    "Тип бумаги",
    "Участник программы создания акционерной стоимости",
    "Эмитент не соответствует требованию на текущий Список",
    "Код эмитента",
    "Тип инструмента",
}

COLUMN_DICTIONARY_DESCRIPTIONS = {
    "Код бумаги": "Уникальный код облигации на бирже.",
    "Краткое наименование": "Короткое название бумаги, чтобы быстро ее найти.",
    "ISIN": "Международный идентификатор ценной бумаги.",
    "Эмитент": "Название компании или организации, выпустившей облигацию.",
    "ИНН эмитента": "Налоговый номер эмитента.",
    "Номинал": "Базовая стоимость одной облигации.",
    "Валюта номинала": "В какой валюте выражен номинал.",
    "Дата погашения": "Дата, когда эмитент должен вернуть номинал.",
    "Дата следующего купона": "Ближайшая дата выплаты купонного дохода.",
    "Ставка купона, %": "Процентная ставка купона.",
    "Размер купона": "Сумма одной купонной выплаты.",
    "Купонный период, дней": "Количество дней между купонными выплатами.",
    "Доходность по средневзвешенной цене": "Оценка доходности к цене прошлых торгов.",
    "Средневзвешенная цена предыдущего дня": "Средняя цена бумаги за прошлый торговый день.",
    "Цена предыдущей сделки": "Цена последней сделки из прошлой торговой сессии.",
    "Оборот за день": "Денежный оборот по бумаге за текущий день.",
    "Объем за день": "Количество бумаг, которое прошло в сделках за день.",
    "Количество сделок": "Сколько сделок по бумаге было заключено за день.",
    "Официальная цена закрытия (пред.)": "Официальная цена закрытия предыдущего торгового дня.",
    "Дата предыдущих торгов": "Дата предыдущей торговой сессии для бумаги.",
    "Объем выпуска": "Полный объем облигаций в выпуске.",
    "Объем размещения": "Какая часть выпуска была фактически размещена.",
    "Дата оферты": "Дата, когда может сработать оферта на выкуп.",
    "Цена оферты": "Цена выкупа бумаги по оферте.",
}

DESCRIPTION_KEY_ALIASES = {
    "ISIN код": "ISIN",
    "Код ценной бумаги": "Код бумаги",
    "Наименование ценной бумаги": "Краткое наименование",
    "Английское наименование": "Латинское наименование",
    "Номинальная стоимость": "Номинал",
    "Сумма купона, в валюте номинала": "Размер купона",
    "Периодичность выплаты купона в год": "Частота купона",
    "Типа инструмента": "Тип инструмента",
}


# Группы колонок для удобного сворачивания/разворачивания в Excel.
# Порядок важен: именно в таком порядке группы будут показаны в листе.
COLUMN_GROUP_DEFINITIONS: list[tuple[str, list[str]]] = [
    (
        "Идентификация",
        [
            "Код бумаги",
            "ISIN",
            "Краткое наименование",
            "Полное наименование",
            "Тип инструмента",
            "Тип облигации",
            "Подтип облигации",
        ],
    ),
    (
        "Эмитент",
        [
            "Эмитент",
            "ИНН эмитента",
            "Код эмитента",
        ],
    ),
    (
        "Параметры выпуска",
        [
            "Номинал",
            "Валюта номинала",
            "Объем выпуска",
            "Объем размещения",
            "Дата погашения",
            "Дата оферты",
            "Цена оферты",
        ],
    ),
    (
        "Купоны",
        [
            "Дата следующего купона",
            "Ставка купона, %",
            "Размер купона",
            "Купонный период, дней",
            "Частота купона",
            "НКД",
        ],
    ),
    (
        "Торговые показатели",
        [
            "Доходность по средневзвешенной цене",
            "Средневзвешенная цена предыдущего дня",
            "Цена предыдущей сделки",
            "Официальная цена закрытия (пред.)",
            "Дата предыдущих торгов",
            "Количество сделок",
            "Объем за день",
            "Оборот за день",
        ],
    ),
]

GROUP_SEPARATOR_STYLES: list[tuple[str, str]] = [
    ("D9E1F2", "1F4E78"),
    ("E2F0D9", "2F6B2F"),
    ("FCE4D6", "9C5700"),
    ("F4DFEC", "6A1B57"),
    ("E4DFEC", "3E2F78"),
]


def normalize_description_key(raw_key: str | None) -> str:
    """Возвращает каноническое имя поля из карточки облигации."""
    if raw_key is None:
        return "unknown"
    key = str(raw_key).strip()
    if not key:
        return "unknown"
    return DESCRIPTION_KEY_ALIASES.get(key, key)


def setup_environment() -> None:
    """Готовит папки проекта и очищает временные данные перед запуском."""
    cfg.LOGS_DIR.mkdir(parents=True, exist_ok=True)
    cfg.RAW_DIR.mkdir(parents=True, exist_ok=True)
    cfg.CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cfg.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    cfg.CHECKPOINT_DIR.mkdir(parents=True, exist_ok=True)
    cfg.RAW_ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)

    if cfg.LOG_FILE.exists():
        cfg.LOG_FILE.unlink()

    for item in cfg.RAW_DIR.iterdir():
        if item.name == ".gitkeep":
            continue
        if item.is_dir():
            shutil.rmtree(item, ignore_errors=True)
        else:
            item.unlink(missing_ok=True)


def validate_required_columns(df: pd.DataFrame, required_columns: list[str], context: str) -> None:
    """Проверяет, что в таблице есть обязательные колонки; иначе прерывает запуск."""
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        missing_list = ", ".join(missing)
        raise RuntimeError(f"{context}: отсутствуют обязательные колонки: {missing_list}")


def enforce_cache_soft_limit() -> None:
    """Ограничивает разрастание кэша, удаляя самые старые файлы при превышении лимита."""
    limit_bytes = cfg.CACHE_SOFT_LIMIT_MB * 1024 * 1024
    if limit_bytes <= 0 or not cfg.CACHE_DIR.exists():
        return

    all_files = [path for path in cfg.CACHE_DIR.rglob("*") if path.is_file()]
    total_size = sum(path.stat().st_size for path in all_files)
    if total_size <= limit_bytes:
        return

    target_size = int(limit_bytes * 0.9)
    removed_files = 0
    for file_path in sorted(all_files, key=lambda p: p.stat().st_mtime):
        file_size = file_path.stat().st_size
        file_path.unlink(missing_ok=True)
        total_size -= file_size
        removed_files += 1
        if total_size <= target_size:
            break

    logging.warning(
        "Кэш превысил лимит %s МБ, удалено %s старых файлов. Текущий размер: %.2f МБ",
        cfg.CACHE_SOFT_LIMIT_MB,
        removed_files,
        total_size / (1024 * 1024),
    )


def setup_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(cfg.LOG_FILE, mode="w", encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )


def request_json(url: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
    """Делает HTTP-запрос к ISS MOEX c повторными попытками при временных сбоях."""
    last_exception: Exception | None = None
    for attempt in range(1, cfg.RETRY_COUNT + 1):
        try:
            response = requests.get(url, params=params, timeout=cfg.REQUEST_TIMEOUT_SECONDS)
            response.raise_for_status()
            return response.json()
        except Exception as exc:  # noqa: BLE001
            last_exception = exc
            logging.warning("Попытка %s/%s для %s завершилась ошибкой: %s", attempt, cfg.RETRY_COUNT, url, exc)
            time.sleep(1.5 * attempt)

    raise RuntimeError(f"Не удалось получить данные: {url}") from last_exception


def fetch_all_pages(url: str, block_name: str, extra_params: dict[str, Any] | None = None) -> pd.DataFrame:
    """Собирает все страницы ISS-блока через параметр start."""
    start = 0
    all_rows: list[list[Any]] = []
    columns: list[str] | None = None
    previous_rows: list[list[Any]] | None = None
    page_counter = 0
    while True:
        page_counter += 1
        params = {"iss.meta": "off", "start": start}
        if extra_params:
            params.update(extra_params)

        payload = request_json(url, params=params)
        block = payload.get(block_name)
        if not block:
            raise RuntimeError(f"В ответе нет блока '{block_name}' для URL {url}")

        if columns is None:
            columns = block["columns"]

        rows = block["data"]
        if not rows:
            break

        if previous_rows is not None and rows == previous_rows:
            logging.warning(
                "Пагинация для %s перестала сдвигаться (start=%s). Останавливаю цикл, чтобы избежать дублей.",
                url,
                start,
            )
            break

        all_rows.extend(rows)
        if page_counter == 1 or page_counter % 10 == 0:
            logging.info(
                "Пагинация %s: страница=%s, start=%s, строк в странице=%s, накоплено=%s",
                block_name,
                page_counter,
                start,
                len(rows),
                len(all_rows),
            )

        if len(rows) < 100:
            break

        previous_rows = rows
        start += len(rows)

    return pd.DataFrame(all_rows, columns=columns or [])


def is_cache_valid(cache_file: Path, ttl_hours: int) -> bool:
    if not cache_file.exists():
        return False
    max_age = timedelta(hours=ttl_hours)
    age = datetime.now() - datetime.fromtimestamp(cache_file.stat().st_mtime)
    return age <= max_age


def fetch_reference_row(secid: str) -> dict[str, Any] | None:
    ref_cache_dir = cfg.CACHE_DIR / "reference_rows"
    ref_cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = ref_cache_dir / f"{secid}.json"

    if is_cache_valid(cache_file, cfg.CACHE_TTL_HOURS["reference_rows"]):
        return json.loads(cache_file.read_text(encoding="utf-8"))

    payload = request_json(
        f"{cfg.MOEX_BASE_URL}/securities.json",
        params={"iss.meta": "off", "engine": "stock", "market": "bonds", "q": secid},
    )
    block = payload.get("securities", {"columns": [], "data": []})
    columns = block.get("columns", [])
    rows = block.get("data", [])

    secid_index = columns.index("secid") if "secid" in columns else None
    target_row = None
    if secid_index is not None:
        for row in rows:
            if str(row[secid_index]) == secid:
                target_row = dict(zip(columns, row))
                break

    if target_row is not None:
        cache_file.write_text(json.dumps(target_row, ensure_ascii=False), encoding="utf-8")

    return target_row


def collect_reference_data_for_secids(secids: list[str]) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    total = len(secids)
    logging.info("Загружаю справочные данные по %s облигациям через точечные запросы", total)

    with ThreadPoolExecutor(max_workers=cfg.MAX_WORKERS) as executor:
        future_map = {executor.submit(fetch_reference_row, secid): secid for secid in secids}
        for index, future in enumerate(as_completed(future_map), start=1):
            secid = future_map[future]
            try:
                row = future.result()
                if row:
                    records.append(row)
            except Exception as exc:  # noqa: BLE001
                logging.error("Ошибка загрузки справочника для %s: %s", secid, exc)

            if index % 200 == 0 or index == total:
                logging.info("Справочник: обработано %s/%s", index, total)

    return pd.DataFrame(records)


def to_records(payload_block: dict[str, Any]) -> list[dict[str, Any]]:
    columns = payload_block.get("columns", [])
    return [dict(zip(columns, row)) for row in payload_block.get("data", [])]


def fetch_security_details(secid: str) -> MoexBlocks:
    cache_file = cfg.CACHE_DIR / f"{secid}.json"

    if is_cache_valid(cache_file, cfg.CACHE_TTL_HOURS["security_details"]):
        cached = json.loads(cache_file.read_text(encoding="utf-8"))
        details_payload = cached["details"]
        bondization_payload = cached["bondization"]
    else:
        details_payload = request_json(f"{cfg.MOEX_BASE_URL}/securities/{secid}.json", params={"iss.meta": "off"})
        bondization_payload = request_json(f"{cfg.MOEX_BASE_URL}/securities/{secid}/bondization.json", params={"iss.meta": "off"})
        cache_file.write_text(
            json.dumps({"details": details_payload, "bondization": bondization_payload}, ensure_ascii=False),
            encoding="utf-8",
        )

    description_rows = to_records(details_payload.get("description", {"columns": [], "data": []}))
    descriptions = [
        {
            "secid": secid,
            "field": row.get("name"),
            "title": row.get("title"),
            "value": row.get("value"),
        }
        for row in description_rows
    ]

    coupons = to_records(bondization_payload.get("coupons", {"columns": [], "data": []}))
    amortizations = to_records(bondization_payload.get("amortizations", {"columns": [], "data": []}))
    offers = to_records(bondization_payload.get("offers", {"columns": [], "data": []}))

    for row in coupons:
        row["secid"] = secid
    for row in amortizations:
        row["secid"] = secid
    for row in offers:
        row["secid"] = secid

    return MoexBlocks(
        descriptions=descriptions,
        coupons=coupons,
        amortizations=amortizations,
        offers=offers,
    )


def chunk_list(items: list[str], chunks: int) -> list[list[str]]:
    if not items:
        return []
    chunk_size = max(1, len(items) // chunks + (1 if len(items) % chunks else 0))
    return [items[i : i + chunk_size] for i in range(0, len(items), chunk_size)]


def load_checkpoint_state() -> dict[str, Any]:
    if not is_cache_valid(cfg.CHECKPOINT_STATE_FILE, cfg.CACHE_TTL_HOURS["checkpoint"]):
        return {"completed_secids": []}
    return json.loads(cfg.CHECKPOINT_STATE_FILE.read_text(encoding="utf-8"))


def _load_checkpoint_rows(path: Path) -> list[dict[str, Any]]:
    if not is_cache_valid(path, cfg.CACHE_TTL_HOURS["checkpoint"]):
        return []
    return json.loads(path.read_text(encoding="utf-8"))


def save_checkpoint(blocks: MoexBlocks, completed_secids: list[str]) -> None:
    cfg.CHECKPOINT_STATE_FILE.write_text(
        json.dumps({"completed_secids": completed_secids, "updated_at": datetime.now().isoformat()}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    cfg.CHECKPOINT_DESCRIPTIONS_FILE.write_text(json.dumps(blocks.descriptions, ensure_ascii=False), encoding="utf-8")
    cfg.CHECKPOINT_COUPONS_FILE.write_text(json.dumps(blocks.coupons, ensure_ascii=False), encoding="utf-8")
    cfg.CHECKPOINT_AMORTIZATIONS_FILE.write_text(json.dumps(blocks.amortizations, ensure_ascii=False), encoding="utf-8")
    cfg.CHECKPOINT_OFFERS_FILE.write_text(json.dumps(blocks.offers, ensure_ascii=False), encoding="utf-8")


def load_checkpoint_blocks() -> MoexBlocks:
    return MoexBlocks(
        descriptions=_load_checkpoint_rows(cfg.CHECKPOINT_DESCRIPTIONS_FILE),
        coupons=_load_checkpoint_rows(cfg.CHECKPOINT_COUPONS_FILE),
        amortizations=_load_checkpoint_rows(cfg.CHECKPOINT_AMORTIZATIONS_FILE),
        offers=_load_checkpoint_rows(cfg.CHECKPOINT_OFFERS_FILE),
    )


def clear_checkpoint() -> None:
    for file_path in [
        cfg.CHECKPOINT_STATE_FILE,
        cfg.CHECKPOINT_DESCRIPTIONS_FILE,
        cfg.CHECKPOINT_COUPONS_FILE,
        cfg.CHECKPOINT_AMORTIZATIONS_FILE,
        cfg.CHECKPOINT_OFFERS_FILE,
    ]:
        file_path.unlink(missing_ok=True)


def collect_extended_data(secids: list[str]) -> MoexBlocks:
    state = load_checkpoint_state()
    completed_secids = set(state.get("completed_secids", []))
    blocks = load_checkpoint_blocks()

    if completed_secids:
        logging.info("Найден checkpoint: пропускаю уже обработанные облигации (%s шт.)", len(completed_secids))

    secids_to_process = [secid for secid in secids if secid not in completed_secids]
    grouped = chunk_list(secids_to_process, cfg.CHUNK_COUNT)

    logging.info(
        "Старт загрузки расширенных данных: всего=%s, осталось после checkpoint=%s, чанков=%s",
        len(secids),
        len(secids_to_process),
        len(grouped),
    )

    failed_secids: set[str] = set()
    for chunk_index, secid_chunk in enumerate(grouped, start=1):
        print(f"Обрабатываю чанк {chunk_index}/{len(grouped)} ({len(secid_chunk)} облигаций)...")
        with ThreadPoolExecutor(max_workers=cfg.MAX_WORKERS) as executor:
            future_map = {executor.submit(fetch_security_details, secid): secid for secid in secid_chunk}
            for index, future in enumerate(as_completed(future_map), start=1):
                secid = future_map[future]
                try:
                    item = future.result()
                    blocks.descriptions.extend(item.descriptions)
                    blocks.coupons.extend(item.coupons)
                    blocks.amortizations.extend(item.amortizations)
                    blocks.offers.extend(item.offers)
                    completed_secids.add(secid)
                except Exception as exc:  # noqa: BLE001
                    failed_secids.add(secid)
                    logging.error("Ошибка при загрузке %s: %s", secid, exc)

                if index % 50 == 0 or index == len(secid_chunk):
                    logging.info(
                        "Чанк %s/%s: обработано %s/%s",
                        chunk_index,
                        len(grouped),
                        index,
                        len(secid_chunk),
                    )

        save_checkpoint(blocks, sorted(completed_secids))
        logging.info("Checkpoint сохранен после чанка %s/%s", chunk_index, len(grouped))

    if failed_secids:
        print(f"Дополнительная попытка для проблемных бумаг: {len(failed_secids)} шт.")
        logging.warning("Запускаю повторную загрузку проблемных SECID: %s шт.", len(failed_secids))
        with ThreadPoolExecutor(max_workers=cfg.MAX_WORKERS) as executor:
            future_map = {executor.submit(fetch_security_details, secid): secid for secid in sorted(failed_secids)}
            for future in as_completed(future_map):
                secid = future_map[future]
                try:
                    item = future.result()
                    blocks.descriptions.extend(item.descriptions)
                    blocks.coupons.extend(item.coupons)
                    blocks.amortizations.extend(item.amortizations)
                    blocks.offers.extend(item.offers)
                    completed_secids.add(secid)
                except Exception as exc:  # noqa: BLE001
                    logging.error("Повторная попытка также завершилась ошибкой для %s: %s", secid, exc)

        save_checkpoint(blocks, sorted(completed_secids))

    return blocks


def save_raw(df: pd.DataFrame, file_name: str) -> None:
    (cfg.RAW_DIR / file_name).write_text(df.to_json(orient="records", force_ascii=False, indent=2), encoding="utf-8")


def save_parquet(df: pd.DataFrame, file_path: Path) -> None:
    """Сохраняет таблицу в Parquet для быстрого машинного чтения."""
    df.to_parquet(file_path, index=False)


def build_emitents_sheet(traded_bonds: pd.DataFrame) -> pd.DataFrame:
    emitent_cols_upper = ["EMITENT_ID", "EMITENT_TITLE", "EMITENT_INN", "EMITENT_OKPO"]
    existing_cols = [col for col in emitent_cols_upper if col in traded_bonds.columns]
    if not existing_cols:
        return pd.DataFrame(columns=emitent_cols_upper + ["bonds_count", "secids"])

    emitent_id_column = "EMITENT_ID" if "EMITENT_ID" in existing_cols else existing_cols[0]
    emitents = traded_bonds[existing_cols + ["SECID"]].dropna(subset=[emitent_id_column], how="all").copy()

    return emitents.groupby(existing_cols, dropna=False, as_index=False).agg(
        bonds_count=("SECID", "count"),
        secids=("SECID", lambda x: ", ".join(sorted(set(map(str, x))))),
    )


def merge_emitents_incremental(new_emitents: pd.DataFrame) -> pd.DataFrame:
    """Добавляет только новых эмитентов к существующему справочнику."""
    if "EMITENT_ID" not in new_emitents.columns:
        return new_emitents

    if not cfg.EMITENTS_OUTPUT_FILE.exists():
        return new_emitents.sort_values("EMITENT_TITLE", na_position="last").reset_index(drop=True)

    existing_emitents = pd.read_parquet(cfg.EMITENTS_OUTPUT_FILE)
    if "EMITENT_ID" not in existing_emitents.columns:
        return new_emitents.sort_values("EMITENT_TITLE", na_position="last").reset_index(drop=True)

    existing_ids = set(existing_emitents["EMITENT_ID"].dropna().astype(str))
    candidates = new_emitents.copy()
    candidates["EMITENT_ID"] = candidates["EMITENT_ID"].astype(str)
    new_only = candidates[~candidates["EMITENT_ID"].isin(existing_ids)]

    if new_only.empty:
        return existing_emitents.sort_values("EMITENT_TITLE", na_position="last").reset_index(drop=True)

    merged_emitents = pd.concat([existing_emitents, new_only], ignore_index=True)
    return merged_emitents.sort_values("EMITENT_TITLE", na_position="last").reset_index(drop=True)


def build_descriptions_wide_sheet(descriptions_df: pd.DataFrame) -> pd.DataFrame:
    """Превращает длинный формат описаний в широкий (1 строка = 1 облигация)."""
    if descriptions_df.empty:
        return pd.DataFrame(columns=["secid"])

    descriptions_df = descriptions_df.copy()
    descriptions_df["key"] = descriptions_df["title"].fillna(descriptions_df["field"]).map(normalize_description_key)
    descriptions_df = descriptions_df.drop_duplicates(subset=["secid", "key"], keep="last")

    wide = descriptions_df.pivot(index="secid", columns="key", values="value").reset_index()
    wide.columns.name = None
    return wide


def merge_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Устраняет дубли названий колонок (с учетом пробелов/регистра) без потери значений."""

    def normalize_column_label(column_name: Any) -> str:
        return " ".join(str(column_name).strip().lower().split())

    groups: dict[str, list[int]] = defaultdict(list)
    first_name_by_group: dict[str, Any] = {}
    for index, column_name in enumerate(df.columns):
        normalized = normalize_column_label(column_name)
        groups[normalized].append(index)
        first_name_by_group.setdefault(normalized, column_name)

    merged_data: dict[Any, pd.Series] = {}
    column_order: list[Any] = []
    for normalized_name, indexes in groups.items():
        first_column_name = first_name_by_group[normalized_name]
        combined = df.iloc[:, indexes[0]].copy()
        for index in indexes[1:]:
            right = df.iloc[:, index]
            combined = combined.where(combined.notna(), right)
        merged_data[first_column_name] = combined
        column_order.append(first_column_name)

    return pd.DataFrame(merged_data, columns=column_order)


def build_column_dictionary_sheet(columns: list[str]) -> pd.DataFrame:
    """Формирует словарь колонок прямо внутри Excel для пользователей без техподготовки."""
    rows = []
    for column in columns:
        rows.append(
            {
                "Колонка": column,
                "Описание": COLUMN_DICTIONARY_DESCRIPTIONS.get(column, "Техническое поле из выгрузки MOEX."),
            }
        )
    return pd.DataFrame(rows)


def remove_unwanted_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Удаляет лишние столбцы из итогового листа по списку бизнес-требований."""
    columns_to_drop = [column for column in OUTPUT_COLUMNS_TO_DROP if column in df.columns]
    if not columns_to_drop:
        return df
    logging.info("Удаляю из итогового листа %s столбцов по бизнес-правилам", len(columns_to_drop))
    return df.drop(columns=columns_to_drop)


def build_merged_bonds_sheet(traded_bonds_df: pd.DataFrame, descriptions_wide_df: pd.DataFrame) -> pd.DataFrame:
    """Объединяет торговые данные и описание облигаций в один лист без дублей."""
    merged = traded_bonds_df.copy()
    merged["secid"] = merged["SECID"]

    merged = merged.merge(descriptions_wide_df, how="left", on="secid")

    suffix_x_columns = [col for col in merged.columns if col.endswith("_x")]
    for column_x in suffix_x_columns:
        base_name = column_x[:-2]
        column_y = f"{base_name}_y"
        if column_y in merged.columns:
            merged[base_name] = merged[column_x].combine_first(merged[column_y])
            merged = merged.drop(columns=[column_x, column_y])

    duplicate_pairs = [
        ("SECID", "secid"),
        ("SECID", "Код ценной бумаги"),
        ("SHORTNAME", "Краткое наименование"),
        ("SHORTNAME", "Наименование ценной бумаги"),
        ("LATNAME", "Латинское наименование"),
        ("LATNAME", "Английское наименование"),
        ("NAME", "Полное наименование"),
        ("ISIN", "ISIN код"),
        ("REGNUMBER", "Регистрационный номер"),
        ("LISTLEVEL", "Уровень листинга"),
        ("FACEUNIT", "Валюта номинала"),
        ("PREVPRICE", "Цена предыдущей сделки"),
        ("LOTSIZE", "Лот"),
        ("FACEVALUE", "Номинал"),
        ("FACEVALUE", "Номинальная стоимость"),
        ("MATDATE", "Дата погашения"),
        ("COUPONFREQUENCY", "Частота купона"),
        ("COUPONFREQUENCY", "Периодичность выплаты купона в год"),
        ("COUPONPERCENT", "Ставка купона, %"),
        ("COUPONVALUE", "Размер купона"),
        ("COUPONVALUE", "Сумма купона, в валюте номинала"),
        ("BUYBACKPRICE", "Цена оферты"),
        ("BUYBACKDATE", "Дата оферты"),
    ]
    drop_columns = [desc_col for src_col, desc_col in duplicate_pairs if src_col in merged.columns and desc_col in merged.columns]
    if drop_columns:
        merged = merged.drop(columns=drop_columns)

    renamed_columns = {col: RUSSIAN_COLUMN_NAMES[col] for col in merged.columns if col in RUSSIAN_COLUMN_NAMES}
    merged = merged.rename(columns=renamed_columns)
    merged = merge_duplicate_columns(merged)
    merged = remove_unwanted_columns(merged)
    return merged


def build_data_quality_sheet(merged_bonds_df: pd.DataFrame) -> pd.DataFrame:
    """Формирует простой отчет по заполненности ключевых полей."""
    important_columns = [
        "Код бумаги",
        "ISIN",
        "Код эмитента",
        "Эмитент",
        "Торгуется",
        "Статус",
    ]
    rows: list[dict[str, Any]] = []
    total = len(merged_bonds_df)
    for column in important_columns:
        if column in merged_bonds_df.columns:
            column_data = merged_bonds_df[column]
            if isinstance(column_data, pd.DataFrame):
                empty_count = int(column_data.isna().all(axis=1).sum())
            else:
                empty_count = int(column_data.isna().sum())
        else:
            empty_count = total
        fill_rate = 0.0 if total == 0 else round((1 - empty_count / total) * 100, 2)
        rows.append({
            "Поле": column,
            "Всего строк": total,
            "Пустых значений": empty_count,
            "Заполнено, %": fill_rate,
        })
    return pd.DataFrame(rows)


def archive_raw_data() -> None:
    """Архивирует raw JSON после успешного запуска и удаляет старые архивы."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_file = cfg.RAW_ARCHIVE_DIR / f"raw_{timestamp}.zip"
    raw_files = sorted(cfg.RAW_DIR.glob("*.json"))
    if not raw_files:
        return

    with zipfile.ZipFile(archive_file, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file_path in raw_files:
            zf.write(file_path, arcname=file_path.name)

    old_archives = sorted(cfg.RAW_ARCHIVE_DIR.glob("raw_*.zip"), key=lambda p: p.stat().st_mtime, reverse=True)
    for stale_archive in old_archives[cfg.RAW_ARCHIVE_KEEP_LAST :]:
        stale_archive.unlink(missing_ok=True)


def beautify_sheet(worksheet: Any) -> None:
    """Улучшает читаемость листа: стиль заголовка, автофильтр, freeze pane и ширина."""
    if worksheet.max_row < 1 or worksheet.max_column < 1:
        return

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        first_cell = column_cells[0]
        col_letter = first_cell.column_letter
        values = [str(c.value) if c.value is not None else "" for c in column_cells[:200]]
        max_len = max((len(v) for v in values), default=10)
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)


def build_grouped_bonds_sheet(df: pd.DataFrame) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    """Добавляет в таблицу визуальные разделители и возвращает метаданные групп для Excel."""
    if df.empty:
        return df.copy(), []

    grouped_columns: list[str] = []
    group_metadata: list[dict[str, Any]] = []
    used_columns: set[str] = set()
    separator_index = 1

    for group_name, candidate_columns in COLUMN_GROUP_DEFINITIONS:
        actual_columns = [column for column in candidate_columns if column in df.columns and column not in used_columns]
        if not actual_columns:
            continue

        separator_column = f"Разделитель {separator_index:02d}"
        grouped_columns.append(separator_column)
        start_position = len(grouped_columns) + 1
        grouped_columns.extend(actual_columns)
        end_position = len(grouped_columns)

        group_metadata.append(
            {
                "separator_column": separator_column,
                "group_name": group_name,
                "start": start_position,
                "end": end_position,
            }
        )
        used_columns.update(actual_columns)
        separator_index += 1

    remaining_columns = [column for column in df.columns if column not in used_columns]
    if remaining_columns:
        separator_column = f"Разделитель {separator_index:02d}"
        grouped_columns.append(separator_column)
        start_position = len(grouped_columns) + 1
        grouped_columns.extend(remaining_columns)
        end_position = len(grouped_columns)
        group_metadata.append(
            {
                "separator_column": separator_column,
                "group_name": "Прочие поля",
                "start": start_position,
                "end": end_position,
            }
        )

    grouped_df = pd.DataFrame(index=df.index)
    for meta in group_metadata:
        grouped_df[meta["separator_column"]] = meta["group_name"]
    for column in df.columns:
        grouped_df[column] = df[column]

    grouped_df = grouped_df[grouped_columns]
    return grouped_df, group_metadata


def apply_column_groups_to_sheet(worksheet: Any, group_metadata: list[dict[str, Any]]) -> None:
    """Применяет сворачиваемые группы и цветные разделители для листа с облигациями."""
    if worksheet.max_column < 1 or not group_metadata:
        return

    for idx, meta in enumerate(group_metadata):
        separator_col_index = meta["start"] - 1
        separator_col_letter = get_column_letter(separator_col_index)
        fill_color, font_color = GROUP_SEPARATOR_STYLES[idx % len(GROUP_SEPARATOR_STYLES)]

        worksheet.column_dimensions[separator_col_letter].width = 20
        for row_idx in range(1, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_idx, column=separator_col_index)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(color=font_color, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx > 1:
                cell.value = None

        start_letter = get_column_letter(meta["start"])
        end_letter = get_column_letter(meta["end"])
        worksheet.column_dimensions.group(start_letter, end_letter, outline_level=1, hidden=False)


def write_excel(file_path: Path, sheet_name: str, df: pd.DataFrame) -> None:
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        beautify_sheet(writer.book[sheet_name])


def write_core_excel(merged_bonds_df: pd.DataFrame, quality_df: pd.DataFrame) -> None:
    dictionary_df = build_column_dictionary_sheet(merged_bonds_df.columns.tolist())
    grouped_bonds_df, group_metadata = build_grouped_bonds_sheet(merged_bonds_df)
    with pd.ExcelWriter(cfg.CORE_OUTPUT_FILE, engine="openpyxl") as writer:
        grouped_bonds_df.to_excel(writer, sheet_name="bonds_traded", index=False)
        quality_df.to_excel(writer, sheet_name="data_quality", index=False)
        dictionary_df.to_excel(writer, sheet_name="column_dictionary", index=False)
        beautify_sheet(writer.book["bonds_traded"])
        apply_column_groups_to_sheet(writer.book["bonds_traded"], group_metadata)
        beautify_sheet(writer.book["data_quality"])
        beautify_sheet(writer.book["column_dictionary"])


def update_emitents_history(new_emitents_df: pd.DataFrame) -> None:
    """Ведет историю появления новых эмитентов (дата первого обнаружения + идентификатор)."""
    if "EMITENT_ID" not in new_emitents_df.columns:
        return

    history_file = cfg.OUTPUT_DIR / "moex_emitents_history.csv"
    current_date = datetime.now().strftime("%Y-%m-%d")
    snapshot = new_emitents_df[["EMITENT_ID", "EMITENT_TITLE"]].dropna(subset=["EMITENT_ID"]).copy()
    snapshot["EMITENT_ID"] = snapshot["EMITENT_ID"].astype(str)
    snapshot = snapshot.drop_duplicates(subset=["EMITENT_ID"], keep="last")

    if history_file.exists():
        history_df = pd.read_csv(history_file)
    else:
        history_df = pd.DataFrame(columns=["first_seen", "EMITENT_ID", "EMITENT_TITLE"])

    known_ids = set(history_df["EMITENT_ID"].dropna().astype(str)) if not history_df.empty else set()
    new_rows = snapshot[~snapshot["EMITENT_ID"].isin(known_ids)].copy()
    if new_rows.empty:
        return

    new_rows.insert(0, "first_seen", current_date)
    updated_history = pd.concat([history_df, new_rows], ignore_index=True)
    updated_history = updated_history.sort_values(["first_seen", "EMITENT_TITLE"], na_position="last")
    updated_history.to_csv(history_file, index=False, encoding="utf-8")


def main() -> None:
    start_time = time.perf_counter()
    setup_environment()
    setup_logging()

    print("[1/7] Загружаю список облигаций по рынку MOEX...")
    bonds_market_df = fetch_all_pages(
        f"{cfg.MOEX_BASE_URL}/engines/stock/markets/bonds/securities.json",
        block_name="securities",
    )
    validate_required_columns(bonds_market_df, ["SECID", "ISIN", "STATUS"], "Блок market/securities")

    print("[2/7] Загружаю справочник эмитентов и торговых атрибутов...")
    source_secids = bonds_market_df["SECID"].dropna().astype(str).unique().tolist()
    if cfg.MAX_BONDS_TO_PROCESS:
        source_secids = source_secids[: cfg.MAX_BONDS_TO_PROCESS]
        bonds_market_df = bonds_market_df[bonds_market_df["SECID"].isin(source_secids)].copy()
        logging.warning("Для отладки ограничен список облигаций до %s штук", cfg.MAX_BONDS_TO_PROCESS)

    all_securities_df = collect_reference_data_for_secids(source_secids)
    all_securities_df = all_securities_df.rename(columns=str.upper)
    validate_required_columns(all_securities_df, ["SECID"], "Справочник /iss/securities?q=<SECID>")

    required_reference_cols = [
        "SECID",
        "IS_TRADED",
        "EMITENT_ID",
        "EMITENT_TITLE",
        "EMITENT_INN",
        "EMITENT_OKPO",
        "TYPE",
        "GROUP",
        "PRIMARY_BOARDID",
        "MARKETPRICE_BOARDID",
    ]
    for column in required_reference_cols:
        if column not in all_securities_df.columns:
            all_securities_df[column] = None

    merged_df = bonds_market_df.merge(
        all_securities_df[required_reference_cols],
        how="left",
        on="SECID",
    )

    traded_bonds_df = merged_df[(merged_df["IS_TRADED"] == 1) | (merged_df["STATUS"] == "A")].copy()
    traded_bonds_df = traded_bonds_df.sort_values(["EMITENT_TITLE", "SECID"], na_position="last")

    secids = traded_bonds_df["SECID"].dropna().astype(str).unique().tolist()
    print(f"Найдено торгуемых облигаций: {len(secids)}")

    print("[3/7] Загружаю расширенные данные (10 чанков + checkpoint)...")
    blocks = collect_extended_data(secids)

    descriptions_df = pd.DataFrame(blocks.descriptions)
    descriptions_wide_df = build_descriptions_wide_sheet(descriptions_df)
    coupons_df = pd.DataFrame(blocks.coupons)
    amortizations_df = pd.DataFrame(blocks.amortizations)
    offers_df = pd.DataFrame(blocks.offers)
    emitents_df = build_emitents_sheet(traded_bonds_df)
    update_emitents_history(emitents_df)
    emitents_df = merge_emitents_incremental(emitents_df)
    merged_bonds_df = build_merged_bonds_sheet(traded_bonds_df, descriptions_wide_df)
    quality_df = build_data_quality_sheet(merged_bonds_df)

    print("[4/7] Сохраняю сырые данные в raw/...")
    save_raw(traded_bonds_df, "traded_bonds.json")
    save_raw(descriptions_df, "bond_descriptions_long.json")
    save_raw(descriptions_wide_df, "bond_descriptions_wide.json")
    save_raw(coupons_df, "bond_coupons.json")
    save_raw(amortizations_df, "bond_amortizations.json")
    save_raw(offers_df, "bond_offers.json")
    save_raw(emitents_df, "emitents.json")
    save_raw(merged_bonds_df, "bonds_traded_merged.json")

    print("[5/7] Формирую основной Excel (облегченный)...")
    write_core_excel(merged_bonds_df, quality_df)

    print("[6/7] Сохраняю справочники в быстрый формат Parquet...")
    save_parquet(emitents_df, cfg.EMITENTS_OUTPUT_FILE)
    save_parquet(coupons_df, cfg.COUPONS_OUTPUT_FILE)
    save_parquet(amortizations_df, cfg.AMORTIZATIONS_OUTPUT_FILE)
    save_parquet(offers_df, cfg.OFFERS_OUTPUT_FILE)

    archive_raw_data()
    clear_checkpoint()
    enforce_cache_soft_limit()
    elapsed = time.perf_counter() - start_time

    print("[7/7] Готово.")
    print(f"Основной Excel сохранен: {cfg.CORE_OUTPUT_FILE}")
    print(f"Время выполнения: {elapsed:.2f} сек.")

    logging.info("Основной Excel сохранен: %s", cfg.CORE_OUTPUT_FILE)
    logging.info(
        "Доп. файлы (Parquet): %s, %s, %s, %s",
        cfg.EMITENTS_OUTPUT_FILE,
        cfg.COUPONS_OUTPUT_FILE,
        cfg.AMORTIZATIONS_OUTPUT_FILE,
        cfg.OFFERS_OUTPUT_FILE,
    )
    logging.info("Время выполнения: %.2f сек.", elapsed)


if __name__ == "__main__":
    main()
