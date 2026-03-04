import os
import re
import json
import time
import random
from datetime import datetime
from urllib.parse import urljoin

from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeoutError, Error as PWError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

LIST_URL = ("https://www.acra-ratings.ru/ratings/issuers/"
            "?text=&sectors[]=&activities[]=&countries[]=&forecasts[]=&on_revision=0"
            "&rating_scale=0&rate_from=0&rate_to=0&page=1&sort=&count=1000&")

BASE = "https://www.acra-ratings.ru"
OUT_DIR = "acra_dump"
OUT_XLSX = "acraraiting.xlsx"
PROGRESS_LOG = os.path.join(OUT_DIR, "progress.jsonl")

# persistent-профиль браузера (важно для "не терять" куки и ручного прохождения защиты)
PROFILE_DIR = "acra_profile"

RU_MONTHS = {
    "янв": 1, "фев": 2, "мар": 3, "апр": 4, "май": 5, "июн": 6,
    "июл": 7, "авг": 8, "сен": 9, "окт": 10, "ноя": 11, "дек": 12,
}

HEADERS = ["Ссылка", "Наименование", "Рейтинг", "Дата", "ИНН"]


def human_sleep(a=0.7, b=1.8):
    time.sleep(random.uniform(a, b))


def ensure_dirs():
    os.makedirs(OUT_DIR, exist_ok=True)
    os.makedirs(os.path.join(OUT_DIR, "issuers"), exist_ok=True)


def safe_filename(name: str, limit: int = 80) -> str:
    s = re.sub(r"[^a-zA-Z0-9а-яА-Я_-]+", "_", name or "").strip("_")
    return (s[:limit] or "issuer")


def normalize_date_ru(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    try:
        return datetime.strptime(s, "%d.%m.%Y").strftime("%Y-%m-%d")
    except ValueError:
        pass
    m = re.match(r"^(\d{1,2})\s+([А-Яа-я]{3})\s+(\d{4})$", s)
    if m:
        d = int(m.group(1))
        mon = RU_MONTHS.get(m.group(2).lower())
        y = int(m.group(3))
        if mon:
            try:
                return datetime(y, mon, d).strftime("%Y-%m-%d")
            except ValueError:
                return s
    return s


def log_progress(obj: dict):
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(PROGRESS_LOG, "a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def save_mhtml(page, path: str):
    # Chromium-only CDP snapshot -> MHTML. Для "chrome" channel работает.
    cdp = page.context.new_cdp_session(page)
    snap = cdp.send("Page.captureSnapshot", {"format": "mhtml"})
    with open(path, "w", encoding="utf-8") as f:
        f.write(snap["data"])


def parse_list(html: str):
    soup = BeautifulSoup(html, "lxml")
    out = []

    for row in soup.select("div.emits-row.search-table-row"):
        a_person = row.select_one('a.emits-row__item[data-type="ratePerson"]')
        if not a_person or not a_person.get("href"):
            continue

        href = a_person["href"].strip()
        url = href if href.startswith("http") else urljoin(BASE, href)
        name = a_person.get_text(" ", strip=True)

        rating_p = row.select_one('div.emits-row__item[data-type="rate"] p')
        rating = rating_p.get_text(" ", strip=True) if rating_p else ""

        pr_a = row.select_one('div.emits-row__item[data-type="pressRelease"] a')
        date_raw = pr_a.get_text(" ", strip=True) if pr_a else ""
        date_norm = normalize_date_ru(date_raw)

        if name and url:
            out.append({
                "url": url,
                "name": name,
                "rating": rating,
                "date": date_norm or date_raw,
                "inn": "",
            })

    return out


def extract_inn_from_issuer_html(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")

    # стабильный путь под твою карточку: div.info -> small == "ИНН" -> p
    for info in soup.select("div.info"):
        small = info.find("small")
        if not small:
            continue
        if small.get_text(" ", strip=True).lower() == "инн":
            p = info.find("p")
            val = p.get_text(" ", strip=True) if p else ""
            digits = re.sub(r"\D+", "", val)
            return digits if digits else ""

    # fallback
    text = soup.get_text("\n", strip=True)
    m = re.search(r"ИНН\D{0,50}(\d[\d\s]{8,14}\d)", text, flags=re.IGNORECASE)
    if m:
        return re.sub(r"\D+", "", m.group(1))

    return ""


# ---------------- Excel helpers ----------------

def build_new_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "acra"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:E1"

    widths = [55, 45, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return wb, ws


def load_or_create_workbook(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        return wb, ws
    return build_new_workbook()


def read_existing_index(ws):
    url_to_row = {}
    for r in range(2, ws.max_row + 1):
        url = (ws.cell(r, 1).value or "").strip()
        if not url:
            continue
        url_to_row[url] = r
    return url_to_row


def upsert_row(ws, url_to_row, rec):
    url = rec["url"]
    if url in url_to_row:
        r = url_to_row[url]
    else:
        ws.append(["", "", "", "", ""])
        r = ws.max_row
        url_to_row[url] = r

    ws.cell(r, 1).value = rec.get("url", "")
    ws.cell(r, 2).value = rec.get("name", "")
    ws.cell(r, 3).value = rec.get("rating", "")
    ws.cell(r, 4).value = rec.get("date", "")
    ws.cell(r, 5).value = rec.get("inn", "")

    for c in range(1, 6):
        ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=True)

    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    return r


def get_cell_str(ws, row, col):
    v = ws.cell(row, col).value
    return (v or "").strip() if isinstance(v, str) else (str(v).strip() if v is not None else "")


# ---------------- Robust navigation ----------------

def goto_with_retries(page, url: str, wait_selector: str = None, attempts: int = 6):
    """
    Устойчивый goto: ретраи + “переждать” антибот-разрывы.
    При ERR_CONNECTION_CLOSED/ERR_CONNECTION_RESET и т.п. просто повторяем.
    """
    last_err = None
    for k in range(1, attempts + 1):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=60_000)
            if wait_selector:
                page.wait_for_selector(wait_selector, timeout=60_000)
            return True
        except (PWTimeoutError, PWError) as e:
            last_err = e
            msg = str(e)
            # типичные “сетевые отстрелы”
            retryable = any(x in msg for x in [
                "ERR_CONNECTION_CLOSED",
                "ERR_CONNECTION_RESET",
                "ERR_EMPTY_RESPONSE",
                "ERR_TIMED_OUT",
                "net::",
            ])
            if not retryable:
                # не сетевое — смысла долбить нет
                raise

            sleep_s = min(3.0 * k + random.uniform(0.5, 2.0), 20.0)
            print(f"[goto retry {k}/{attempts}] {msg[:120]}... sleep {sleep_s:.1f}s")
            time.sleep(sleep_s)

    print("[goto] failed after retries:", last_err)
    return False


def main():
    ensure_dirs()

    wb, ws = load_or_create_workbook(OUT_XLSX)
    url_to_row = read_existing_index(ws)

    with sync_playwright() as p:
        # ВАЖНО: используем установленный Chrome + persistent profile
        # Это значительно уменьшает вероятность ERR_CONNECTION_CLOSED на агрессивных сайтах.
        context = p.chromium.launch_persistent_context(
            user_data_dir=PROFILE_DIR,
            channel="chrome",          # нужен установленный Google Chrome
            headless=False,
            viewport={"width": 1365, "height": 768},
            locale="ru-RU",
            timezone_id="Europe/Moscow",
            args=["--start-maximized"],
        )

        page = context.new_page()

        # 1) Открываем список с ретраями
        ok = goto_with_retries(
            page,
            LIST_URL,
            wait_selector='a.emits-row__item[data-type="ratePerson"]',
            attempts=6
        )
        if not ok:
            print("Не смог открыть страницу списка. Скорее всего IP/сеть режется.")
            print("Проверь: открывается ли этот URL в обычном Chrome с этого же ПК/IP.")
            context.close()
            return

        human_sleep(1.0, 2.0)
        page.mouse.wheel(0, random.randint(500, 1400))
        human_sleep(0.6, 1.2)

        # 2) Сохраняем список
        try:
            save_mhtml(page, os.path.join(OUT_DIR, "issuers_list.mhtml"))
        except Exception as e:
            print("Не удалось сохранить MHTML списка:", e)

        list_html = page.content()
        with open(os.path.join(OUT_DIR, "issuers_list.html"), "w", encoding="utf-8") as f:
            f.write(list_html)

        # 3) Парсим список
        parsed = parse_list(list_html)
        print("Строк в списке:", len(parsed))

        # 4) Уникальные по URL + upsert списка сразу в Excel
        uniq = {}
        for r in parsed:
            uniq.setdefault(r["url"], r)

        for url, rec in uniq.items():
            # если в Excel уже есть ИНН — не теряем
            if url in url_to_row:
                existing_inn = get_cell_str(ws, url_to_row[url], 5)
                if existing_inn:
                    rec["inn"] = existing_inn

            upsert_row(ws, url_to_row, rec)

        wb.save(OUT_XLSX)
        print("Excel обновлён списком:", OUT_XLSX)

        urls = list(uniq.keys())
        print("Уникальных эмитентов:", len(urls))

        # 5) Карточки — только где ИНН пустой
        for i, url in enumerate(urls, 1):
            row_idx = url_to_row.get(url)
            current_inn = get_cell_str(ws, row_idx, 5) if row_idx else ""
            if current_inn:
                continue

            name = uniq[url].get("name", "")
            try:
                ok = goto_with_retries(page, url, wait_selector=None, attempts=5)
                if not ok:
                    log_progress({
                        "url": url, "name": name, "inn": "",
                        "status": "goto_failed",
                        "ts": datetime.utcnow().isoformat() + "Z",
                    })
                    continue

                human_sleep(0.8, 1.8)
                if random.random() < 0.6:
                    page.mouse.wheel(0, random.randint(250, 900))
                    human_sleep(0.3, 0.8)

                html = page.content()

                # сохраняем карточку
                fn = f"{i:04d}_{safe_filename(name)}.html"
                with open(os.path.join(OUT_DIR, "issuers", fn), "w", encoding="utf-8") as f:
                    f.write(html)

                inn = extract_inn_from_issuer_html(html)

                rec = {
                    "url": url,
                    "name": uniq[url].get("name", ""),
                    "rating": uniq[url].get("rating", ""),
                    "date": uniq[url].get("date", ""),
                    "inn": inn,
                }
                upsert_row(ws, url_to_row, rec)
                wb.save(OUT_XLSX)  # checkpoint после каждого эмитента

                log_progress({
                    "url": url, "name": name, "inn": inn,
                    "status": "ok",
                    "ts": datetime.utcnow().isoformat() + "Z",
                })

                print(f"[{i}/{len(urls)}] OK inn={inn} {name}")

            except Exception as e:
                log_progress({
                    "url": url, "name": name, "inn": "",
                    "status": f"error: {type(e).__name__}: {e}",
                    "ts": datetime.utcnow().isoformat() + "Z",
                })
                print(f"[{i}/{len(urls)}] ERROR {name} -> {e}")

            human_sleep(0.6, 1.6)

        print("Готово:", OUT_XLSX)
        print("Дампы:", OUT_DIR)
        print("Профиль Chrome:", PROFILE_DIR)

        context.close()


if __name__ == "__main__":
    main()