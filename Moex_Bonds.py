#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Moex_Bonds.py
1) Downloads MOEX ISS rates.csv
2) Saves to Excel (.xlsx)
3) Applies readability: header styling, filters, freeze panes, column widths, basic formats

Usage:
  python Moex_Bonds.py --out moex_bonds.xlsx
  python Moex_Bonds.py --out moex_bonds.xlsx --url "<custom_url>"
"""

from __future__ import annotations

import argparse
import io
import re
import sys
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import pandas as pd
import requests

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_URL = (
    "https://iss.moex.com/iss/apps/infogrid/emission/rates.csv"
    "?sec_type=stock_ofz_bond,stock_cb_bond,stock_subfederal_bond,"
    "stock_municipal_bond,stock_corporate_bond,stock_exchange_bond"
    "&iss.dp=comma&iss.df=%25d.%25m.%25Y&iss.tf=%25H:%25M:%25S"
    "&iss.dtf=%25d.%25m.%25Y%20%25H:%25M:%25S"
    "&iss.only=rates&limit=unlimited&lang=ru"
)

# Heuristics for column types
DATE_COL_HINTS = (
    "MATDATE", "COUPONDATE", "ISSUEDATE", "STARTDATEMOEX", "OFFERDATE"
)
INT_COL_HINTS = (
    "LISTLEVEL", "FACEVALUE", "ISSUESIZE", "IS_COLLATERAL", "IS_EXTERNAL",
    "IS_RII", "INCLUDEDBYMOEX", "DURATION", "IS_QUALIFIED_INVESTORS",
    "HIGH_RISK", "COUPONFREQUENCY", "EVENINGSESSION", "MORNINGSESSION",
    "WEEKENDSESSION", "SUSPENSION_LISTING", "LOTSIZE", "COUPONDAYSPASSED",
    "COUPONDAYSREMAIN", "COUPONLENGTH", "DAYSTOREDEMPTION", "REPLBOND",
)
FLOAT_COL_HINTS = (
    "ZSPREAD", "WAPRICE", "YIELDATWAP", "COUPONPERCENT", "COUPONVALUE",
    "PRICE", "PRICE_RUB", "RTL1", "RTH1", "RTL2", "RTH2", "RTL3", "RTH3",
    "DISCOUNT1", "DISCOUNT2", "DISCOUNT3", "DISCOUNTL0", "DISCOUNTH0",
    "LIMIT1", "LIMIT2", "LIMIT3", "FULL_COVERED_LIMIT",
)


def download_text(url: str, timeout: int = 60) -> str:
    r = requests.get(url, timeout=timeout, headers={"User-Agent": "Moex_Bonds/1.0"})
    r.raise_for_status()
    r.encoding = r.apparent_encoding or "utf-8"
    return r.text


def _strip_iss_preamble(text: str) -> str:
    """
    MOEX infogrid CSV often starts with:
      rates
      <blank>
      SECID\t...
    We'll remove leading block name + empty lines before header.
    """
    lines = text.splitlines()
    # Drop UTF-8 BOM if any
    if lines and lines[0].startswith("\ufeff"):
        lines[0] = lines[0].lstrip("\ufeff")

    # Remove first line if it's just the block name
    if lines and lines[0].strip().lower() == "rates":
        lines = lines[1:]

    # Remove leading empty lines
    while lines and not lines[0].strip():
        lines = lines[1:]

    return "\n".join(lines) + "\n"


def detect_sep(sample: str) -> str:
    """
    Try to detect separator by looking at the header line.
    Priority: tab -> semicolon -> comma
    """
    first_line = sample.splitlines()[0] if sample.splitlines() else ""
    if "\t" in first_line:
        return "\t"
    if ";" in first_line and first_line.count(";") >= first_line.count(","):
        return ";"
    return ","


def parse_rates_csv(text: str) -> pd.DataFrame:
    clean = _strip_iss_preamble(text)
    sep = detect_sep(clean[:5000])

    # With iss.dp=comma, floats might be like "74,41".
    # If sep is comma too, it would be ambiguous; in practice MOEX often uses tab/; in such exports.
    # We'll still try robust parsing with python engine.
    df = pd.read_csv(
        io.StringIO(clean),
        sep=sep,
        engine="python",
        dtype=str,          # start as strings; we will coerce
        keep_default_na=False,
    )

    # Drop totally empty columns (just in case)
    df = df.loc[:, [c for c in df.columns if str(c).strip() != ""]]

    # Coerce types
    df = coerce_types(df)

    return df


def _coerce_date(series: pd.Series) -> pd.Series:
    # Dates are in dd.mm.yyyy by iss.df; empty -> NaT
    return pd.to_datetime(series.replace("", pd.NA), format="%d.%m.%Y", errors="coerce")


def _coerce_int(series: pd.Series) -> pd.Series:
    s = series.replace("", pd.NA)
    # remove spaces
    s = s.str.replace(" ", "", regex=False)
    return pd.to_numeric(s, errors="coerce").astype("Int64")


def _coerce_float(series: pd.Series) -> pd.Series:
    s = series.replace("", pd.NA)
    s = s.str.replace(" ", "", regex=False)
    # decimal comma -> dot
    s = s.str.replace(",", ".", regex=False)
    return pd.to_numeric(s, errors="coerce")


def coerce_types(df: pd.DataFrame) -> pd.DataFrame:
    cols = list(df.columns)

    for c in cols:
        uc = str(c).upper()

        if any(h in uc for h in DATE_COL_HINTS):
            df[c] = _coerce_date(df[c].astype(str))
            continue

        if any(h == uc for h in INT_COL_HINTS) or uc == "INN":
            # INN is numeric-like but keep as string if it has leading zeros; here it's fine to keep string.
            if uc == "INN":
                df[c] = df[c].astype(str).replace("nan", "")
            else:
                df[c] = _coerce_int(df[c].astype(str))
            continue

        if any(h == uc for h in FLOAT_COL_HINTS):
            df[c] = _coerce_float(df[c].astype(str))
            continue

        # Boolean-like values sometimes come as True/False
        if uc in ("FULLCOVERED", "REPLBOND") or df[c].astype(str).isin(["True", "False"]).mean() > 0.8:
            df[c] = df[c].astype(str).replace({"True": True, "False": False})
            continue

        # Otherwise keep as text
        df[c] = df[c].astype(str)

    return df


def save_to_excel_with_formatting(df: pd.DataFrame, out_path: str, sheet_name: str = "rates") -> None:
    # Write via pandas first
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(out_path)
    ws = wb[sheet_name]

    # Freeze header row
    ws.freeze_panes = "A2"

    # Header styling
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 28

    # AutoFilter over full range (Excel will apply it)
    last_col_letter = get_column_letter(max_col)
    ws.auto_filter.ref = f"A1:{last_col_letter}{max_row}"  # openpyxl filters API :contentReference[oaicite:1]{index=1}

    # Add an Excel "Table" for nicer UX (striped rows)
    table_ref = f"A1:{last_col_letter}{max_row}"
    table_name = "MOEX_RATES"
    # Ensure unique table name if re-run
    existing_names = {t.displayName for t in ws._tables}
    if table_name in existing_names:
        table_name = f"MOEX_RATES_{datetime.now().strftime('%H%M%S')}"

    tab = Table(displayName=table_name, ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)  # tables supported in openpyxl :contentReference[oaicite:2]{index=2}

    # Number formats
    date_fmt = "DD.MM.YYYY"
    int_fmt = "0"
    float_fmt = "0.00"
    price_fmt = "0.00"
    percent_fmt = "0.00"  # keep as numeric (not Excel %), because values are like 10.74 not 0.1074

    headers = [ws.cell(1, c).value for c in range(1, max_col + 1)]
    header_to_col = {str(h): idx + 1 for idx, h in enumerate(headers)}

    def set_col_format(col_name: str, fmt: str):
        col = header_to_col.get(col_name)
        if not col:
            return
        for r in range(2, max_row + 1):
            ws.cell(r, col).number_format = fmt

    # Apply formats by known column names (exact)
    for name in ("MATDATE", "COUPONDATE", "ISSUEDATE", "STARTDATEMOEX", "OFFERDATE"):
        set_col_format(name, date_fmt)

    for name in ("FACEVALUE", "ISSUESIZE", "DURATION", "LOTSIZE", "COUPONDAYSPASSED", "COUPONDAYSREMAIN", "COUPONLENGTH", "DAYSTOREDEMPTION"):
        set_col_format(name, int_fmt)

    for name in ("ZSPREAD", "WAPRICE", "YIELDATWAP", "PRICE", "PRICE_RUB", "RTL1", "RTH1", "RTL2", "RTH2", "RTL3", "RTH3",
                 "DISCOUNT1", "LIMIT1", "DISCOUNT2", "LIMIT2", "DISCOUNT3", "DISCOUNTL0", "DISCOUNTH0", "FULL_COVERED_LIMIT", "COUPONVALUE"):
        set_col_format(name, price_fmt)

    for name in ("COUPONPERCENT",):
        set_col_format(name, percent_fmt)

    # Column widths (simple heuristic: based on max text length in column, capped)
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        # consider header
        h = ws.cell(1, col).value
        if h is not None:
            max_len = max(max_len, len(str(h)))
        # sample some rows for speed
        step = max(1, (max_row - 1) // 200)  # up to ~200 checks
        for r in range(2, max_row + 1, step):
            v = ws.cell(r, col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    # Better alignment for data rows
    for r in range(2, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).alignment = Alignment(vertical="top", wrap_text=False)

    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Download MOEX bonds rates CSV and save to formatted Excel.")
    ap.add_argument("--url", default=DEFAULT_URL, help="MOEX ISS rates.csv URL")
    ap.add_argument("--out", default="moex_bonds.xlsx", help="Output .xlsx file")
    ap.add_argument("--sheet", default="rates", help="Excel sheet name")
    args = ap.parse_args()

    try:
        text = download_text(args.url)
        df = parse_rates_csv(text)
        save_to_excel_with_formatting(df, args.out, sheet_name=args.sheet)
        print(f"OK: saved {len(df):,} rows to {args.out}")
        return 0
    except requests.RequestException as e:
        print(f"HTTP error: {e}", file=sys.stderr)
        return 2
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())