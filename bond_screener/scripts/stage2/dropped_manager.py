from __future__ import annotations

import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from core.db import get_connection, utc_now_iso
from core.excel_debug import export_dataframe, should_export
from core.logging import get_script_logger
from core.progress import progress_iter
from core.settings import AppSettings

DATE_FMT = "%d.%m.%Y"
HEADERS = ["isin", "secid", "reason", "dropped_at", "until", "source", "comment", "updated_at"]


@dataclass(frozen=True)
class DroppedManagerStats:
    loaded_manual_rows: int
    excluded_bonds: int
    remaining_candidates: int
    duration_s: float


class DroppedManager:
    def __init__(self, settings: AppSettings) -> None:
        self.settings = settings
        self.logger = get_script_logger(
            settings.paths.logs_dir / "stage2_dropped_manager.log",
            "stage2.dropped_manager",
        )
        self.ui_file = settings.paths.source_xlsx_dir / settings.stage2.dropped_ui_filename

    def run(self) -> DroppedManagerStats:
        started = time.perf_counter()
        self._ensure_db_objects()
        self._ensure_ui_file()

        manual_rows = self._read_manual_rows_from_excel()
        self._upsert_manual_rows(manual_rows)

        excluded_bonds, remaining_candidates = self._apply_dropped_to_candidates()
        self._export_debug_if_needed()

        duration_s = time.perf_counter() - started
        self.logger.info(
            (
                "DroppedManager завершён: manual_rows=%s, excluded_bonds=%s, "
                "remaining_candidates=%s, duration=%.2fs"
            ),
            len(manual_rows),
            excluded_bonds,
            remaining_candidates,
            duration_s,
        )
        print(
            "[STAGE2][dropped_manager] "
            f"excluded={excluded_bonds} | remaining={remaining_candidates} | manual_rows={len(manual_rows)}"
        )
        return DroppedManagerStats(
            loaded_manual_rows=len(manual_rows),
            excluded_bonds=excluded_bonds,
            remaining_candidates=remaining_candidates,
            duration_s=duration_s,
        )

    def _ensure_db_objects(self) -> None:
        with get_connection(self.settings.paths.db_file) as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS dropped_manual (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    isin TEXT,
                    secid TEXT,
                    reason TEXT,
                    dropped_at TEXT,
                    until TEXT,
                    source TEXT NOT NULL DEFAULT 'manual',
                    updated_at TEXT NOT NULL,
                    UNIQUE (isin, secid)
                );

                CREATE TABLE IF NOT EXISTS dropped_auto (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    isin TEXT,
                    secid TEXT,
                    reason TEXT,
                    dropped_at TEXT,
                    until TEXT,
                    source TEXT NOT NULL DEFAULT 'auto',
                    updated_at TEXT NOT NULL,
                    UNIQUE (isin, secid)
                );

                DROP VIEW IF EXISTS dropped_effective;
                CREATE VIEW dropped_effective AS
                WITH manual_active AS (
                    SELECT *
                    FROM dropped_manual
                    WHERE until IS NULL OR TRIM(until) = ''
                       OR date(substr(until, 7, 4) || '-' || substr(until, 4, 2) || '-' || substr(until, 1, 2)) >= date('now', 'localtime')
                ),
                auto_active AS (
                    SELECT *
                    FROM dropped_auto
                    WHERE until IS NULL OR TRIM(until) = ''
                       OR date(substr(until, 7, 4) || '-' || substr(until, 4, 2) || '-' || substr(until, 1, 2)) >= date('now', 'localtime')
                )
                SELECT isin, secid, reason, dropped_at, until, source, updated_at
                FROM manual_active
                UNION ALL
                SELECT a.isin, a.secid, a.reason, a.dropped_at, a.until, a.source, a.updated_at
                FROM auto_active a
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM manual_active m
                    WHERE COALESCE(m.isin, '') = COALESCE(a.isin, '')
                      AND COALESCE(m.secid, '') = COALESCE(a.secid, '')
                );
                """
            )

    def _ensure_ui_file(self) -> None:
        if not self.ui_file.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Dropped_bonds"
            ws.append(HEADERS)
            self._style_sheet(ws)
            wb.save(self.ui_file)
            self.logger.info("Создан новый UI файл: %s", self.ui_file)

    def _read_manual_rows_from_excel(self) -> list[dict[str, Any]]:
        wb = load_workbook(self.ui_file)
        ws = wb.active

        rows: list[dict[str, Any]] = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=len(HEADERS), values_only=True), start=2):
            isin = self._normalize_text(row[0])
            secid = self._normalize_text(row[1])
            if not isin and not secid:
                if any(value not in (None, "") for value in row):
                    self.logger.warning("Строка %s пропущена: заполните isin или secid.", idx)
                continue

            reason = self._normalize_text(row[2])
            dropped_at = self._normalize_date(row[3], field_name="dropped_at", row_num=idx)
            until = self._normalize_date(row[4], field_name="until", row_num=idx)
            source = self._normalize_text(row[5]).lower() or "manual"
            if source not in {"manual", "auto"}:
                self.logger.warning("Строка %s: source=%s невалиден, заменён на manual.", idx, source)
                source = "manual"
            comment = self._normalize_text(row[6])
            updated_at = utc_now_iso()

            rows.append(
                {
                    "isin": isin,
                    "secid": secid,
                    "reason": reason,
                    "dropped_at": dropped_at,
                    "until": until,
                    "source": source,
                    "comment": comment,
                    "updated_at": updated_at,
                }
            )

        self._rewrite_ui_excel(rows)
        return rows

    def _rewrite_ui_excel(self, rows: list[dict[str, Any]]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dropped_bonds"
        ws.append(HEADERS)

        for row in rows:
            ws.append(
                [
                    row["isin"],
                    row["secid"],
                    row["reason"],
                    self._parse_ddmmyyyy(row["dropped_at"]),
                    self._parse_ddmmyyyy(row["until"]),
                    row["source"],
                    row["comment"],
                    row["updated_at"],
                ]
            )

        self._style_sheet(ws)
        wb.save(self.ui_file)

    def _style_sheet(self, ws: Any) -> None:
        max_row = ws.max_row
        max_col = ws.max_column

        header_fill = PatternFill(fill_type="solid", fgColor="E6E6E6")
        thin = Side(style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        manual_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        auto_fill = PatternFill(fill_type="solid", fgColor="EFEFEF")

        for col in range(1, max_col + 1):
            head = ws.cell(1, col)
            head.font = Font(bold=True)
            head.fill = header_fill
            head.alignment = Alignment(horizontal="center", vertical="center")
            head.border = border

        for r in range(2, max_row + 1):
            ws.cell(r, 4).number_format = "DD.MM.YYYY"
            ws.cell(r, 5).number_format = "DD.MM.YYYY"

            source_value = (ws.cell(r, 6).value or "").strip().lower() if ws.cell(r, 6).value else ""
            if source_value == "manual":
                ws.cell(r, 6).fill = manual_fill
            elif source_value == "auto":
                ws.cell(r, 6).fill = auto_fill

            for c in range(1, max_col + 1):
                ws.cell(r, c).border = border

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{ws.cell(1, max_col).column_letter}{max_row}"

        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if isinstance(cell.value, (datetime, date)):
                    txt = cell.value.strftime(DATE_FMT)
                else:
                    txt = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(txt))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    def _upsert_manual_rows(self, rows: list[dict[str, Any]]) -> None:
        now_iso = utc_now_iso()
        with get_connection(self.settings.paths.db_file) as conn:
            for row in progress_iter(rows, desc="Stage2/DroppedManager", total=len(rows)):
                conn.execute(
                    """
                    INSERT INTO dropped_manual (isin, secid, reason, dropped_at, until, source, updated_at)
                    VALUES (?, ?, ?, ?, ?, 'manual', ?)
                    ON CONFLICT(isin, secid)
                    DO UPDATE SET
                        reason=excluded.reason,
                        dropped_at=excluded.dropped_at,
                        until=excluded.until,
                        source='manual',
                        updated_at=excluded.updated_at
                    """,
                    (
                        row["isin"],
                        row["secid"],
                        row["reason"],
                        row["dropped_at"],
                        row["until"],
                        now_iso,
                    ),
                )

    def _apply_dropped_to_candidates(self) -> tuple[int, int]:
        with get_connection(self.settings.paths.db_file) as conn:
            before = int(conn.execute("SELECT COUNT(*) AS cnt FROM candidate_bonds").fetchone()["cnt"])
            conn.execute(
                """
                DELETE FROM candidate_bonds
                WHERE EXISTS (
                    SELECT 1
                    FROM dropped_effective d
                    WHERE (COALESCE(d.secid, '') <> '' AND d.secid = candidate_bonds.secid)
                       OR (COALESCE(d.isin, '') <> '' AND d.isin = candidate_bonds.isin)
                )
                """
            )
            after = int(conn.execute("SELECT COUNT(*) AS cnt FROM candidate_bonds").fetchone()["cnt"])

        return before - after, after

    def _export_debug_if_needed(self) -> None:
        if not should_export(self.settings, "stage2"):
            return
        with get_connection(self.settings.paths.db_file) as conn:
            dropped = [dict(r) for r in conn.execute("SELECT * FROM dropped_effective ORDER BY source, secid, isin").fetchall()]
            candidates = [dict(r) for r in conn.execute("SELECT * FROM candidate_bonds ORDER BY issuer_key, secid").fetchall()]

        dropped_path = export_dataframe(
            self.settings,
            filename="stage2_debug_dropped_effective.xlsx",
            df=pd.DataFrame(dropped),
            export_name="stage2",
        )
        candidates_path = export_dataframe(
            self.settings,
            filename="stage2_debug_candidate_after_drop.xlsx",
            df=pd.DataFrame(candidates),
            export_name="stage2",
        )
        if dropped_path:
            self.logger.info("Excel debug выгрузка создана: %s", dropped_path)
        if candidates_path:
            self.logger.info("Excel debug выгрузка создана: %s", candidates_path)

    def _normalize_date(self, value: Any, field_name: str, row_num: int) -> str:
        parsed = self._parse_to_date(value)
        if parsed is None:
            if value not in (None, ""):
                self.logger.warning(
                    "Строка %s: поле %s не удалось распарсить (%s). Значение очищено.",
                    row_num,
                    field_name,
                    value,
                )
            return ""
        if isinstance(value, str):
            stripped = value.strip()
            normalized = parsed.strftime(DATE_FMT)
            if stripped and stripped != normalized:
                self.logger.warning(
                    "Строка %s: поле %s нормализовано: '%s' -> '%s'.",
                    row_num,
                    field_name,
                    stripped,
                    normalized,
                )
        return parsed.strftime(DATE_FMT)

    def _parse_to_date(self, value: Any) -> date | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
        return None

    @staticmethod
    def _normalize_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _parse_ddmmyyyy(value: str) -> date | None:
        if not value:
            return None
        try:
            return datetime.strptime(value, DATE_FMT).date()
        except ValueError:
            return None
