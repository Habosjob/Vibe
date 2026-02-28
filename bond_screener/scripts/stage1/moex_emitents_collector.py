from __future__ import annotations

import asyncio
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from tqdm import tqdm

from core.db import get_connection, utc_now_iso
from core.excel_debug import should_export
from core.logging import get_script_logger
from core.settings import AppSettings
from net.cache import HttpCache
from net.http_client import HttpClient, moex_get

ALLOWED_SCORING_FLAGS = {"Greenlist", "Yellowlist", "Redlist"}
DATE_FMT = "%d.%m.%Y"


@dataclass(frozen=True)
class Stage1Stats:
    emitents_count: int
    securities_count: int
    duration_s: float


class MoexEmitentsCollector:
    def __init__(self, settings: AppSettings) -> None:
        self.settings = settings
        self.logger = get_script_logger(
            settings.paths.logs_dir / "stage1_moex_emitents_collector.log",
            "stage1.moex_emitents_collector",
        )
        self.emitents_xlsx_path = settings.paths.source_xlsx_dir / "Emitents.xlsx"

    def run(self) -> Stage1Stats:
        started = time.perf_counter()
        self._ensure_stage1_db_objects()

        existing_manual = self._read_manual_from_excel(self.emitents_xlsx_path)
        skip_network = self._is_fresh_success_run()
        self.logger.info("TTL stage1 ttl_hours=%s, skip_network=%s", self.settings.stage1.ttl_hours, skip_network)

        if not skip_network:
            emitents_rows, securities_rows = asyncio.run(self._fetch_moex_data())
            self._save_raw_data(emitents_rows, securities_rows)
        else:
            self.logger.info("Сетевая загрузка пропущена по TTL. Используем данные из БД.")

        previous_manual = self._load_manual_from_db()
        rows_for_excel = self._build_emitents_rows(existing_manual, previous_manual)
        self._write_emitents_excel(rows_for_excel)
        manual_from_excel = self._read_manual_from_excel(self.emitents_xlsx_path)
        self._sync_manual_table(manual_from_excel)
        self._refresh_effective_view()
        self._export_debug_raw_if_needed()

        counts = self._read_counts()
        duration_s = time.perf_counter() - started
        return Stage1Stats(
            emitents_count=counts["emitents"],
            securities_count=counts["securities"],
            duration_s=duration_s,
        )

    def _ensure_stage1_db_objects(self) -> None:
        with get_connection(self.settings.paths.db_file) as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS emitents_raw (
                    issuer_key TEXT PRIMARY KEY,
                    inn TEXT,
                    name TEXT NOT NULL,
                    name_norm TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS securities_raw (
                    secid TEXT PRIMARY KEY,
                    isin TEXT,
                    issuer_key TEXT NOT NULL,
                    shortname TEXT,
                    matdate TEXT,
                    facevalue REAL,
                    faceunit TEXT,
                    typenm TEXT,
                    status TEXT,
                    updated_at TEXT NOT NULL
                );

                CREATE TABLE IF NOT EXISTS emitents_manual (
                    issuer_key TEXT PRIMARY KEY,
                    scoring_flag TEXT,
                    scoring_date TEXT,
                    comment TEXT,
                    group_hint TEXT
                );
                """
            )

    def _is_fresh_success_run(self) -> bool:
        ttl_hours = max(0, int(self.settings.stage1.ttl_hours))
        if ttl_hours == 0:
            return False

        with get_connection(self.settings.paths.db_file) as conn:
            row = conn.execute(
                """
                SELECT finished_at
                FROM runs
                WHERE stage = 'stage1' AND script = 'run' AND status = 'ok'
                ORDER BY finished_at DESC
                LIMIT 1
                """
            ).fetchone()
        if not row or not row["finished_at"]:
            return False

        finished_at = datetime.fromisoformat(row["finished_at"])
        age_h = (datetime.now(finished_at.tzinfo) - finished_at).total_seconds() / 3600
        return age_h < ttl_hours

    async def _fetch_moex_data(self) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        cache = HttpCache(self.settings.paths.cache_http_dir)
        client = HttpClient(self.settings, cache)
        try:
            securities = await self._fetch_securities(client)
            target_secids = {str(row.get("SECID")) for row in securities if row.get("SECID")}
            emitents_ref = await self._fetch_emitents_reference(client, target_secids)
        finally:
            await client.aclose()

        now_iso = utc_now_iso()
        emitents_rows: dict[str, dict[str, Any]] = {}
        securities_rows: list[dict[str, Any]] = []

        for sec in securities:
            secid = sec.get("SECID")
            if not secid:
                continue
            ref = emitents_ref.get(secid)
            if not ref:
                continue

            emitter_id = ref.get("emitent_id")
            issuer_key = f"moex_emitter_{emitter_id}" if emitter_id else f"moex_secid_{secid}"
            issuer_name = ref.get("emitent_title") or sec.get("SHORTNAME") or secid

            emitents_rows[issuer_key] = {
                "issuer_key": issuer_key,
                "inn": ref.get("emitent_inn") or None,
                "name": issuer_name,
                "name_norm": self._normalize_name(issuer_name),
                "updated_at": now_iso,
            }
            securities_rows.append(
                {
                    "secid": secid,
                    "isin": sec.get("ISIN") or None,
                    "issuer_key": issuer_key,
                    "shortname": sec.get("SHORTNAME") or None,
                    "matdate": sec.get("MATDATE") or None,
                    "facevalue": sec.get("FACEVALUE"),
                    "faceunit": sec.get("FACEUNIT") or None,
                    "typenm": sec.get("TYPENAME") or None,
                    "status": sec.get("STATUS") or None,
                    "updated_at": now_iso,
                }
            )

        self.logger.info("Сеть MOEX: emitents=%s, securities=%s", len(emitents_rows), len(securities_rows))
        return sorted(emitents_rows.values(), key=lambda x: x["issuer_key"]), securities_rows

    async def _fetch_securities(self, client: HttpClient) -> list[dict[str, Any]]:
        params = {
            "iss.meta": "off",
            "iss.only": "securities",
            "securities.columns": "SECID,ISIN,SHORTNAME,MATDATE,FACEVALUE,FACEUNIT,TYPENAME,STATUS",
        }
        with tqdm(total=1, desc="Stage1/MOEX bonds", unit="req", dynamic_ncols=True) as pbar:
            payload = await moex_get(client, "/iss/engines/stock/markets/bonds/securities.json", params=params)
            pbar.update(1)

        block = payload.get("securities", {})
        cols = block.get("columns", [])
        all_rows = [dict(zip(cols, row, strict=False)) for row in block.get("data", [])]

        today = date.today()
        filtered = []
        for row in all_rows:
            matdate_raw = row.get("MATDATE")
            if matdate_raw:
                try:
                    if datetime.strptime(matdate_raw, "%Y-%m-%d").date() < today:
                        continue
                except ValueError:
                    pass
            filtered.append(row)

        unique_by_secid: dict[str, dict[str, Any]] = {}
        for row in filtered:
            secid = row.get("SECID")
            if not secid:
                continue
            unique_by_secid.setdefault(str(secid), row)

        unique_rows = list(unique_by_secid.values())
        self.logger.info(
            "Получено облигаций MOEX: total=%s active=%s unique=%s",
            len(all_rows),
            len(filtered),
            len(unique_rows),
        )
        return unique_rows

    async def _fetch_emitents_reference(
        self,
        client: HttpClient,
        target_secids: set[str],
    ) -> dict[str, dict[str, Any]]:
        params = {
            "iss.meta": "off",
            "iss.only": "securities",
            "engine": "stock",
            "market": "bonds",
            "limit": max(1, self.settings.stage1.emitents_page_size),
            "start": 0,
        }

        remaining = {secid for secid in target_secids if secid}
        out: dict[str, dict[str, Any]] = {}
        seen_page_signatures: set[tuple[int, str]] = set()
        pages_processed = 0

        with tqdm(total=len(remaining), desc="Stage1/MOEX emitents", unit="sec", dynamic_ncols=True) as pbar:
            while remaining:
                if pages_processed >= max(1, self.settings.stage1.emitents_max_pages):
                    self.logger.warning(
                        "Достигнут лимит страниц emitents_max_pages=%s. Останавливаем цикл.",
                        self.settings.stage1.emitents_max_pages,
                    )
                    break

                payload = await moex_get(client, "/iss/securities.json", params=params)
                block = payload.get("securities", {})
                cols = block.get("columns", [])
                rows = [dict(zip(cols, row, strict=False)) for row in block.get("data", [])]
                if not rows:
                    break

                first_secid = str(rows[0].get("secid") or "")
                signature = (int(params["start"]), first_secid)
                if signature in seen_page_signatures:
                    self.logger.warning(
                        "Детектирован повтор страницы MOEX emitents: start=%s secid=%s. Останавливаем цикл.",
                        params["start"],
                        first_secid,
                    )
                    break
                seen_page_signatures.add(signature)

                newly_found = 0
                for row in rows:
                    secid = row.get("secid")
                    if not secid:
                        continue
                    secid_str = str(secid)
                    if secid_str in remaining and row.get("emitent_id"):
                        out[secid_str] = row
                        remaining.remove(secid_str)
                        newly_found += 1

                if newly_found:
                    pbar.update(newly_found)

                params["start"] = int(params["start"]) + len(rows)
                pages_processed += 1
                if len(rows) < int(params["limit"]):
                    break

        if remaining:
            self.logger.warning("Не удалось найти emitent reference для %s бумаг.", len(remaining))
        self.logger.info("Получены reference эмитентов: %s из %s", len(out), len(target_secids))
        return out

    def _save_raw_data(self, emitents_rows: list[dict[str, Any]], securities_rows: list[dict[str, Any]]) -> None:
        with get_connection(self.settings.paths.db_file) as conn:
            conn.execute("DELETE FROM emitents_raw")
            conn.execute("DELETE FROM securities_raw")
            conn.executemany(
                """
                INSERT INTO emitents_raw (issuer_key, inn, name, name_norm, updated_at)
                VALUES (:issuer_key, :inn, :name, :name_norm, :updated_at)
                """,
                emitents_rows,
            )
            conn.executemany(
                """
                INSERT INTO securities_raw (
                    secid, isin, issuer_key, shortname, matdate, facevalue,
                    faceunit, typenm, status, updated_at
                ) VALUES (
                    :secid, :isin, :issuer_key, :shortname, :matdate, :facevalue,
                    :faceunit, :typenm, :status, :updated_at
                )
                """,
                securities_rows,
            )

    def _load_manual_from_db(self) -> dict[str, dict[str, Any]]:
        with get_connection(self.settings.paths.db_file) as conn:
            rows = conn.execute(
                """
                SELECT issuer_key, scoring_flag, scoring_date, comment, group_hint
                FROM emitents_manual
                """
            ).fetchall()
        return {
            row["issuer_key"]: {
                "scoring_flag": row["scoring_flag"] or "",
                "scoring_date": self._parse_excel_date(row["scoring_date"]),
                "comment": row["comment"] or "",
                "group_hint": row["group_hint"] or "",
            }
            for row in rows
        }

    def _build_emitents_rows(
        self,
        manual_map: dict[str, dict[str, Any]],
        previous_manual_map: dict[str, dict[str, Any]],
    ) -> list[dict[str, Any]]:
        with get_connection(self.settings.paths.db_file) as conn:
            rows = conn.execute(
                """
                SELECT e.issuer_key, e.inn, e.name, COUNT(s.secid) AS active_bonds_count
                FROM emitents_raw e
                LEFT JOIN securities_raw s ON s.issuer_key = e.issuer_key
                GROUP BY e.issuer_key, e.inn, e.name, e.name_norm
                ORDER BY e.name_norm, e.name
                """
            ).fetchall()

        out = []
        for row in rows:
            issuer_key = row["issuer_key"]
            manual = manual_map.get(issuer_key, {})
            previous = previous_manual_map.get(issuer_key, {})

            scoring_flag = self._validated_scoring_flag(manual.get("scoring_flag"), issuer_key)
            previous_flag = self._validated_scoring_flag(previous.get("scoring_flag"), issuer_key)
            excel_date = self._parse_excel_date(manual.get("scoring_date"))
            previous_date = self._parse_excel_date(previous.get("scoring_date"))

            if scoring_flag == "":
                scoring_date = None
            elif scoring_flag != previous_flag:
                scoring_date = date.today()
            else:
                scoring_date = excel_date or previous_date

            out.append(
                {
                    "issuer_key": issuer_key,
                    "inn": row["inn"],
                    "name": row["name"],
                    "scoring_flag": scoring_flag,
                    "scoring_date": scoring_date,
                    "comment": manual.get("comment") or "",
                    "group_hint": manual.get("group_hint") or "",
                    "active_bonds_count": int(row["active_bonds_count"] or 0),
                }
            )
        return out

    def _write_emitents_excel(self, rows: list[dict[str, Any]]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Emitents"
        ws.append([
            "issuer_key",
            "inn",
            "name",
            "scoring_flag",
            "scoring_date",
            "comment",
            "group_hint",
            "active_bonds_count",
        ])

        for row in rows:
            ws.append([
                row["issuer_key"],
                row["inn"],
                row["name"],
                row["scoring_flag"],
                row["scoring_date"],
                row["comment"],
                row["group_hint"],
                row["active_bonds_count"],
            ])

        max_row = ws.max_row
        max_col = ws.max_column

        header_fill = PatternFill(fill_type="solid", fgColor="E6E6E6")
        thin = Side(style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        color_map = {
            "Greenlist": PatternFill(fill_type="solid", fgColor="C6EFCE"),
            "Yellowlist": PatternFill(fill_type="solid", fgColor="FFF2CC"),
            "Redlist": PatternFill(fill_type="solid", fgColor="F4CCCC"),
        }

        for col in range(1, max_col + 1):
            head = ws.cell(1, col)
            head.font = Font(bold=True)
            head.fill = header_fill
            head.alignment = Alignment(horizontal="center", vertical="center")
            head.border = border

        for r in range(2, max_row + 1):
            date_cell = ws.cell(r, 5)
            date_cell.number_format = "DD.MM.YYYY"
            ws.cell(r, 8).number_format = "0"

            scoring = ws.cell(r, 4).value
            if scoring in color_map:
                ws.cell(r, 4).fill = color_map[scoring]

            for c in range(1, max_col + 1):
                ws.cell(r, c).border = border

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{max_row}"

        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if isinstance(cell.value, (datetime, date)):
                    txt = cell.value.strftime(DATE_FMT)
                else:
                    txt = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(txt))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        validation = DataValidation(
            type="list",
            formula1='"Greenlist,Yellowlist,Redlist"',
            allowBlank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid value",
            error="Allowed values: Greenlist, Yellowlist, Redlist",
            promptTitle="Scoring",
            prompt="Select one of: Greenlist, Yellowlist, Redlist",
        )
        ws.add_data_validation(validation)
        if max_row >= 2:
            validation.add(f"D2:D{max_row}")

        wb.save(self.emitents_xlsx_path)
        self.logger.info("Emitents.xlsx сохранён: %s", self.emitents_xlsx_path)

    def _read_manual_from_excel(self, file_path: Path) -> dict[str, dict[str, Any]]:
        if not file_path.exists():
            return {}

        wb = load_workbook(file_path)
        ws = wb.active
        out: dict[str, dict[str, Any]] = {}
        for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
            issuer_key = "" if row[0] is None else str(row[0]).strip()
            if not issuer_key:
                continue
            out[issuer_key] = {
                "scoring_flag": self._validated_scoring_flag(row[3], issuer_key),
                "scoring_date": self._parse_excel_date(row[4]),
                "comment": "" if row[5] is None else str(row[5]),
                "group_hint": "" if row[6] is None else str(row[6]),
            }
        return out

    def _sync_manual_table(self, manual_map: dict[str, dict[str, Any]]) -> None:
        rows = []
        for issuer_key, manual in manual_map.items():
            scoring_flag = self._validated_scoring_flag(manual.get("scoring_flag"), issuer_key)
            scoring_date = self._parse_excel_date(manual.get("scoring_date"))
            if scoring_flag == "":
                scoring_date = None
            rows.append(
                {
                    "issuer_key": issuer_key,
                    "scoring_flag": scoring_flag,
                    "scoring_date": scoring_date.strftime(DATE_FMT) if scoring_date else "",
                    "comment": manual.get("comment") or "",
                    "group_hint": manual.get("group_hint") or "",
                }
            )

        with get_connection(self.settings.paths.db_file) as conn:
            conn.execute("DELETE FROM emitents_manual")
            if rows:
                conn.executemany(
                    """
                    INSERT INTO emitents_manual (issuer_key, scoring_flag, scoring_date, comment, group_hint)
                    VALUES (:issuer_key, :scoring_flag, :scoring_date, :comment, :group_hint)
                    """,
                    rows,
                )

    def _refresh_effective_view(self) -> None:
        with get_connection(self.settings.paths.db_file) as conn:
            conn.executescript(
                """
                DROP VIEW IF EXISTS emitents_effective;
                CREATE VIEW emitents_effective AS
                SELECT
                    e.issuer_key,
                    e.inn,
                    e.name,
                    e.name_norm,
                    COALESCE(m.scoring_flag, '') AS scoring_flag,
                    COALESCE(m.scoring_date, '') AS scoring_date,
                    COALESCE(m.comment, '') AS comment,
                    COALESCE(m.group_hint, '') AS group_hint,
                    COALESCE(s.active_bonds_count, 0) AS active_bonds_count,
                    e.updated_at
                FROM emitents_raw e
                LEFT JOIN emitents_manual m ON m.issuer_key = e.issuer_key
                LEFT JOIN (
                    SELECT issuer_key, COUNT(*) AS active_bonds_count
                    FROM securities_raw
                    GROUP BY issuer_key
                ) s ON s.issuer_key = e.issuer_key;
                """
            )

    def _export_debug_raw_if_needed(self) -> None:
        if not should_export(self.settings, "stage1"):
            return
        with get_connection(self.settings.paths.db_file) as conn:
            emitents = [dict(r) for r in conn.execute("SELECT * FROM emitents_raw ORDER BY issuer_key").fetchall()]
            securities = [dict(r) for r in conn.execute("SELECT * FROM securities_raw ORDER BY issuer_key, secid").fetchall()]
        self._export_debug_xlsx(self.settings.paths.source_xlsx_dir / "stage1_debug_emitents_raw.xlsx", pd.DataFrame(emitents))
        self._export_debug_xlsx(self.settings.paths.source_xlsx_dir / "stage1_debug_securities_raw.xlsx", pd.DataFrame(securities))

    def _export_debug_xlsx(self, path: Path, df: pd.DataFrame) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)

        wb = load_workbook(path)
        ws = wb.active
        header_fill = PatternFill(fill_type="solid", fgColor="E6E6E6")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{ws.cell(1, ws.max_column).column_letter}{ws.max_row}"
        for col_cells in ws.columns:
            max_len = 0
            letter = col_cells[0].column_letter
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[letter].width = min(max_len + 2, 60)
        wb.save(path)

    def _read_counts(self) -> dict[str, int]:
        with get_connection(self.settings.paths.db_file) as conn:
            emitents = conn.execute("SELECT COUNT(*) AS cnt FROM emitents_raw").fetchone()["cnt"]
            securities = conn.execute("SELECT COUNT(*) AS cnt FROM securities_raw").fetchone()["cnt"]
        return {"emitents": int(emitents), "securities": int(securities)}

    @staticmethod
    def _normalize_name(name: str) -> str:
        return " ".join(name.lower().strip().split())

    def _validated_scoring_flag(self, value: Any, issuer_key: str) -> str:
        if value is None:
            return ""
        txt = str(value).strip()
        if not txt:
            return ""
        if txt not in ALLOWED_SCORING_FLAGS:
            self.logger.warning("Некорректный scoring_flag для issuer_key=%s: %s. Значение очищено.", issuer_key, txt)
            return ""
        return txt

    @staticmethod
    def _parse_excel_date(value: Any) -> date | None:
        if value in (None, ""):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            txt = value.strip()
            if not txt:
                return None
            try:
                return datetime.strptime(txt, DATE_FMT).date()
            except ValueError:
                return None
        return None
