from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from core.settings import AppSettings


def should_export(settings: AppSettings, export_name: str) -> bool:
    if not settings.excel_debug:
        return False
    allowed = {item.lower() for item in settings.excel_debug_exports}
    return export_name.lower() in allowed


def export_dataframe(settings: AppSettings, filename: str, df: pd.DataFrame, export_name: str = "stage0") -> Path | None:
    if not should_export(settings, export_name):
        return None
    out_path = settings.paths.source_xlsx_dir / filename
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(out_path, index=False)
    return out_path


def export_dataframe_styled(
    settings: AppSettings,
    filename: str,
    df: pd.DataFrame,
    export_name: str,
    date_columns: list[str] | None = None,
) -> Path | None:
    if not should_export(settings, export_name):
        return None

    out_path = settings.paths.source_xlsx_dir / filename
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(out_path, index=False)

    wb = load_workbook(out_path)
    ws = wb.active
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)

    date_columns_lower = {c.lower() for c in (date_columns or [])}
    header_map = {str(c.value).strip().lower(): idx for idx, c in enumerate(ws[1], start=1) if c.value is not None}

    for col_name, col_idx in header_map.items():
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            if value is None:
                continue
            if col_name in date_columns_lower and isinstance(value, str):
                try:
                    dt = datetime.fromisoformat(value)
                    cell.value = dt
                    cell.number_format = "DD.MM.YYYY"
                    value = dt.strftime("%d.%m.%Y")
                except ValueError:
                    pass
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

    wb.save(out_path)
    return out_path
