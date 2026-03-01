from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


def export_screener(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    out = df.copy()

    pct_cols = {c for c in out.columns if any(k in c.lower() for k in ["ytm", "yield", "rate_pct", "price_pct"]) }

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="Screener", index=False)
        ws = writer.book["Screener"]

        header_fill = PatternFill("solid", fgColor="D9D9D9")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for i, c in enumerate(out.columns, start=1):
            max_len = max(len(str(c)), *(len(str(v)) for v in out[c].head(8000).fillna("")))
            ws.column_dimensions[ws.cell(1, i).column_letter].width = min(60, max(10, max_len + 2))
            low = c.lower()
            if "date" in low or low.endswith("_at") or low.endswith("until"):
                for r in range(2, ws.max_row + 1):
                    ws.cell(r, i).number_format = "DD.MM.YYYY"
            if c in pct_cols:
                for r in range(2, ws.max_row + 1):
                    v = ws.cell(r, i).value
                    if isinstance(v, (int, float)):
                        ws.cell(r, i).value = float(v) * 100
                    ws.cell(r, i).number_format = "0.00%"
            if any(x in low for x in ["price", "nkd", "coupon", "nominal", "amt", "value"]):
                for r in range(2, ws.max_row + 1):
                    ws.cell(r, i).number_format = "0.00"

        method_idx = {name: idx + 1 for idx, name in enumerate(out.columns)}
        method_colors = {
            "floater_scenario": "DDEBF7",
            "zero_coupon": "E4DFEC",
            "perpetual_compounded": "FCE4D6",
            "linker_scenario": "E2F0D9",
        }

        for r in range(2, ws.max_row + 1):
            dropped = ws.cell(r, method_idx.get("dropped_flag", 1)).value if "dropped_flag" in method_idx else False
            method = ws.cell(r, method_idx.get("ytm_method", 1)).value if "ytm_method" in method_idx else None
            fill = None
            if dropped:
                fill = PatternFill("solid", fgColor="F8CBAD")
            elif method in method_colors:
                fill = PatternFill("solid", fgColor=method_colors[method])
            if fill:
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = fill
