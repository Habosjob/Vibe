#!/usr/bin/env python3
"""Пробует все доступные endpoint'ы ISS MOEX для облигаций по выбранным SECID и сохраняет ответы в Excel."""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import random
import re
import time
import warnings
from json import JSONDecodeError
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REFERENCE_URL = "https://iss.moex.com/iss/reference"
BASE_URL = "https://iss.moex.com"

LOGGER = logging.getLogger("moex_bond_endpoints_probe")

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

TARGET_ENDPOINT_SLUG = "iss__engines__engine__markets__market__boardgroups__boardgroup__securities__security"
TARGET_ENDPOINT_DROP_COLUMNS = {
    "BOARDID",
    "BOARDNAME",
    "SECNAME",
    "ISIN",
    "LATNAME",
    "REGNUMBER",
    "LISTLEVEL",
}

COLUMN_DESCRIPTIONS: dict[str, str] = {
    "SECID": "Код финансового инструмента (ценной бумаги) в ISS MOEX.",
    "BOARDID": "Идентификатор торгового режима/доски.",
    "BOARDNAME": "Название торгового режима/доски.",
    "SECNAME": "Краткое наименование бумаги.",
    "ISIN": "Международный идентификационный код ценной бумаги.",
    "LATNAME": "Название бумаги латиницей.",
    "REGNUMBER": "Регистрационный номер выпуска.",
    "LISTLEVEL": "Уровень листинга MOEX.",
    "SHORTNAME": "Сокращённое имя инструмента.",
    "PREVPRICE": "Предыдущая цена закрытия.",
    "LOTSIZE": "Размер стандартного лота.",
    "FACEVALUE": "Номинальная стоимость бумаги.",
    "STATUS": "Статус инструмента (например, A — активен).",
    "BOARD_GROUP_ID": "Идентификатор группы торговых режимов.",
    "MARKETCODE": "Код рынка.",
    "INSTRID": "Тип/класс инструмента.",
    "SECTORID": "Идентификатор сектора торгов.",
    "PREVWAPRICE": "Предыдущая средневзвешенная цена.",
    "YIELDATPREVWAPRICE": "Доходность на базе предыдущей средневзвешенной цены.",
    "COUPONVALUE": "Размер купона за период.",
    "NEXTCOUPON": "Дата следующей купонной выплаты.",
    "ACCRUEDINT": "Накопленный купонный доход (НКД).",
    "PREVLEGALCLOSEPRICE": "Предыдущая официальная цена закрытия.",
    "PREVADMITTEDQUOTE": "Предыдущая признанная котировка.",
    "CURRENCYID": "Код валюты расчётов.",
    "REQUEST_URL": "URL endpoint ISS, откуда получены данные.",
}


def setup_logging(log_file: Path, level: str) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    numeric_level = getattr(logging, level.upper(), logging.INFO)
    logging.basicConfig(
        level=numeric_level,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_file, encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )


def load_secids_from_excel(path: Path) -> list[str]:
    LOGGER.info("Читаю SECID из %s", path)
    df = pd.read_excel(path)
    if "SECID" not in df.columns:
        raise ValueError(f"В файле {path} отсутствует колонка SECID")

    secids = [str(value).strip() for value in df["SECID"].dropna().tolist() if str(value).strip()]
    unique_secids = sorted(set(secids))
    LOGGER.info("Найдено уникальных SECID: %s", len(unique_secids))
    return unique_secids


def parse_security_endpoint_templates(reference_html: str) -> list[str]:
    endpoints = re.findall(r'<dt><a href="\.\/\d+">([^<]+)</a></dt>', reference_html)
    templates: list[str] = []
    for endpoint in endpoints:
        if "[security]" not in endpoint:
            continue
        if endpoint.startswith("/iss/cci/"):
            continue
        templates.append(endpoint)

    deduped = sorted(set(templates))
    LOGGER.info("Найдено шаблонов endpoint с [security]: %s", len(deduped))
    return deduped


def normalize_sheet_name(value: str, used_names: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "_", value).strip()
    if not cleaned:
        cleaned = "sheet"

    base = cleaned[:31]
    candidate = base
    idx = 1
    while candidate in used_names:
        suffix = f"_{idx}"
        candidate = f"{base[:31-len(suffix)]}{suffix}"
        idx += 1
    used_names.add(candidate)
    return candidate


def build_context_for_secid(session: requests.Session, secid: str) -> dict[str, list[str]]:
    url = f"{BASE_URL}/iss/securities/{secid}.json"
    response = session.get(url, params={"iss.meta": "off"}, timeout=30)
    response.raise_for_status()
    payload = response.json()

    boards_df = pd.DataFrame(
        payload.get("boards", {}).get("data", []),
        columns=payload.get("boards", {}).get("columns", []),
    )
    if boards_df.empty:
        return {"board": [], "boardgroup": [], "session": ["total"]}

    if {"engine", "market"}.issubset(boards_df.columns):
        boards_df = boards_df[(boards_df["engine"] == "stock") & (boards_df["market"] == "bonds")].copy()

    boards = sorted({str(board).strip() for board in boards_df.get("boardid", pd.Series(dtype=str)).dropna().tolist() if str(board).strip()})
    boardgroups = sorted({str(boardgroup).strip() for boardgroup in boards_df.get("board_group_id", pd.Series(dtype=str)).dropna().tolist() if str(boardgroup).strip()})
    return {
        "board": boards,
        "boardgroup": boardgroups,
        "session": ["total"],
    }


def instantiate_endpoints(template: str, secid: str, context: dict[str, list[str]]) -> list[str]:
    values: list[dict[str, str]] = [
        {
            "[engine]": "stock",
            "[market]": "bonds",
            "[security]": secid,
        }
    ]

    if "[board]" in template:
        values = [
            {**base, "[board]": board}
            for base in values
            for board in context.get("board", [])
        ]
    if "[boardgroup]" in template:
        values = [
            {**base, "[boardgroup]": boardgroup}
            for base in values
            for boardgroup in context.get("boardgroup", [])
        ]
    if "[session]" in template:
        values = [
            {**base, "[session]": session}
            for base in values
            for session in context.get("session", ["total"])
        ]

    if not values:
        return []

    urls: list[str] = []
    for mapping in values:
        endpoint = template
        for key, replacement in mapping.items():
            endpoint = endpoint.replace(key, replacement)
        urls.append(f"{BASE_URL}{endpoint}.json")
    return sorted(set(urls))


def build_cache_key(url: str, params: dict[str, Any]) -> str:
    fingerprint = json.dumps({"url": url, "params": params}, ensure_ascii=False, sort_keys=True)
    return hashlib.sha256(fingerprint.encode("utf-8")).hexdigest()


def fetch_json_with_cache(
    session: requests.Session,
    url: str,
    params: dict[str, Any],
    cache_dir: Path,
    cache_ttl: int,
) -> tuple[dict[str, Any] | None, str]:
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_key = build_cache_key(url, params)
    cache_file = cache_dir / f"{cache_key}.json"

    if cache_file.exists():
        age = time.time() - cache_file.stat().st_mtime
        if age <= cache_ttl:
            with cache_file.open("r", encoding="utf-8") as handle:
                return json.load(handle), "cache"

    try:
        response = session.get(url, params=params, timeout=45)
        if response.status_code != 200:
            LOGGER.debug("Пропуск endpoint %s: HTTP %s", response.url, response.status_code)
            return None, f"http_{response.status_code}"
    except requests.RequestException as exc:
        LOGGER.warning("Ошибка запроса %s: %s", url, exc)
        return None, "error"

    try:
        payload = response.json()
    except (JSONDecodeError, ValueError):
        content_type = response.headers.get("Content-Type", "")
        LOGGER.debug("Endpoint вернул не-JSON %s (content-type=%s)", response.url, content_type)
        return None, "non_json"

    with cache_file.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False)
    return payload, "api"


def payload_to_frames(payload: dict[str, Any], secid: str, request_url: str) -> dict[str, pd.DataFrame]:
    frames: dict[str, pd.DataFrame] = {}
    for block_name, block in payload.items():
        if not isinstance(block, dict):
            continue
        data = block.get("data")
        columns = block.get("columns")
        if not isinstance(data, list) or not isinstance(columns, list):
            continue
        frame = pd.DataFrame(data, columns=columns)
        if "REQUEST_URL" not in frame.columns:
            frame.insert(0, "REQUEST_URL", request_url)
        if "SECID" not in frame.columns:
            frame.insert(0, "SECID", secid)
        frames[block_name] = frame
    return frames


def drop_unwanted_columns_for_endpoint(endpoint_slug: str, frame: pd.DataFrame) -> pd.DataFrame:
    if endpoint_slug != TARGET_ENDPOINT_SLUG:
        return frame
    filtered = frame.drop(columns=[col for col in TARGET_ENDPOINT_DROP_COLUMNS if col in frame.columns], errors="ignore")
    return filtered


def estimate_column_width(series: pd.Series, column_name: str) -> int:
    samples = [column_name]
    samples.extend(str(value) for value in series.dropna().head(250).tolist())
    max_len = max((len(value) for value in samples), default=10)
    return max(10, min(max_len + 2, 60))


def style_endpoint_worksheet(worksheet, frame: pd.DataFrame, table_name: str) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.zoomScale = 110
    worksheet.row_dimensions[1].height = 22

    for idx, column_name in enumerate(frame.columns, start=1):
        column_letter = get_column_letter(idx)
        header_cell = worksheet.cell(row=1, column=idx)
        header_cell.fill = HEADER_FILL
        header_cell.font = HEADER_FONT
        header_cell.border = BORDER
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.column_dimensions[column_letter].width = estimate_column_width(frame[column_name], column_name)

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    if worksheet.max_row >= 2 and worksheet.max_column >= 1:
        table = Table(displayName=table_name, ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)


def add_column_docs_sheet(writer: pd.ExcelWriter, frames: dict[str, pd.DataFrame]) -> None:
    docs_sheet = writer.book.create_sheet(title="COLUMN_DOCS")
    docs_sheet.append(["SHEET_NAME", "COLUMN_NAME", "DESCRIPTION_RU"])

    for sheet_name, frame in frames.items():
        for column_name in frame.columns:
            docs_sheet.append([
                sheet_name,
                column_name,
                COLUMN_DESCRIPTIONS.get(column_name, "Описание не задано"),
            ])

    docs_sheet.freeze_panes = "A2"
    docs_sheet.sheet_view.zoomScale = 110
    docs_sheet.column_dimensions["A"].width = 32
    docs_sheet.column_dimensions["B"].width = 26
    docs_sheet.column_dimensions["C"].width = 90

    for cell in docs_sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in docs_sheet.iter_rows(min_row=2, max_row=docs_sheet.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if docs_sheet.max_row >= 2:
        docs_table = Table(displayName="ENDPOINT_COLUMN_DOCS", ref=docs_sheet.dimensions)
        docs_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        docs_sheet.add_table(docs_table)


def save_endpoint_workbook(endpoint_slug: str, frames: dict[str, pd.DataFrame], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"{endpoint_slug}.xlsx"
    used_sheet_names: set[str] = set()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        rendered_frames: dict[str, pd.DataFrame] = {}
        for sheet_raw_name, frame in frames.items():
            prepared_frame = drop_unwanted_columns_for_endpoint(endpoint_slug, frame)
            sheet_name = normalize_sheet_name(sheet_raw_name, used_sheet_names)
            prepared_frame.to_excel(writer, index=False, sheet_name=sheet_name)
            rendered_frames[sheet_name] = prepared_frame

        for idx, (sheet_name, frame) in enumerate(rendered_frames.items(), start=1):
            worksheet = writer.sheets[sheet_name]
            table_name = f"TBL_{idx:02d}_{endpoint_slug[:18]}"
            table_name = re.sub(r"[^A-Za-z0-9_]", "_", table_name)[:31]
            style_endpoint_worksheet(worksheet=worksheet, frame=frame, table_name=table_name)

        add_column_docs_sheet(writer, rendered_frames)

    return output_path



def endpoint_slug_from_template(template: str) -> str:
    slug = template.strip("/")
    slug = (
        slug.replace("[engine]", "engine")
        .replace("[market]", "market")
        .replace("[security]", "security")
        .replace("[board]", "board")
        .replace("[boardgroup]", "boardgroup")
        .replace("[session]", "session")
    )
    slug = re.sub(r"[^a-zA-Z0-9_\-/]+", "_", slug)
    slug = slug.replace("/", "__")
    return slug[:180]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Собирает данные по endpoint'ам ISS MOEX для облигаций.")
    parser.add_argument("--input", type=Path, default=Path("moex_bonds.xlsx"), help="Excel-файл со списком облигаций (SECID).")
    parser.add_argument("--output-dir", type=Path, default=Path("endpoint_excels"), help="Каталог для Excel-файлов по endpoint'ам.")
    parser.add_argument("--cache-dir", type=Path, default=Path(".cache/moex_endpoint_probe"), help="Каталог кэша HTTP JSON.")
    parser.add_argument("--cache-ttl", type=int, default=1800, help="TTL кэша в секундах.")
    parser.add_argument("--static-secid", type=str, default="SU26238RMFS4", help="Статичный SECID для проверки кэша.")
    parser.add_argument("--seed", type=int, default=42, help="Seed для выбора random SECID.")
    parser.add_argument("--log-file", type=Path, default=Path("logs/moex_bond_endpoints_probe.log"), help="Путь к log-файлу.")
    parser.add_argument("--log-level", type=str, default="INFO", help="Уровень логирования: DEBUG/INFO/WARNING/ERROR.")
    return parser.parse_args()


def main() -> None:
    started_at = time.perf_counter()
    args = parse_args()
    setup_logging(args.log_file, args.log_level)

    secids = load_secids_from_excel(args.input)
    if not secids:
        raise ValueError("Список SECID пуст")

    random.seed(args.seed)
    random_candidates = [secid for secid in secids if secid != args.static_secid]
    random_secid = random.choice(random_candidates) if random_candidates else args.static_secid
    selected_secids = [args.static_secid, random_secid] if random_secid != args.static_secid else [args.static_secid]

    LOGGER.info("Выбранные SECID на период отладки: %s", selected_secids)

    with requests.Session() as session:
        session.headers.update({"User-Agent": "moex-bonds-endpoints-probe/1.0"})

        LOGGER.info("Получаю список endpoint'ов из %s", REFERENCE_URL)
        reference_html = session.get(REFERENCE_URL, timeout=30).text
        templates = parse_security_endpoint_templates(reference_html)

        endpoint_frames: dict[str, dict[str, list[pd.DataFrame]]] = {}
        stats = {"api": 0, "cache": 0, "non_json": 0, "http_skipped": 0, "errors": 0}

        for secid in selected_secids:
            LOGGER.info("Обрабатываю SECID=%s", secid)
            context = build_context_for_secid(session=session, secid=secid)
            LOGGER.debug("Контекст SECID=%s: %s", secid, context)

            secid_urls_total = 0
            for template in templates:
                endpoint_slug = endpoint_slug_from_template(template)
                urls = instantiate_endpoints(template=template, secid=secid, context=context)
                secid_urls_total += len(urls)

                for url in urls:
                    params = {"iss.meta": "off", "limit": 100}
                    payload, source = fetch_json_with_cache(
                        session=session,
                        url=url,
                        params=params,
                        cache_dir=args.cache_dir,
                        cache_ttl=args.cache_ttl,
                    )

                    if payload is None:
                        if source == "non_json":
                            stats["non_json"] += 1
                        elif source.startswith("http_"):
                            stats["http_skipped"] += 1
                        else:
                            stats["errors"] += 1
                        continue

                    stats[source] += 1
                    frames = payload_to_frames(payload=payload, secid=secid, request_url=url)
                    if not frames:
                        continue

                    endpoint_frames.setdefault(endpoint_slug, {})
                    for block_name, frame in frames.items():
                        endpoint_frames[endpoint_slug].setdefault(block_name, []).append(frame)

            LOGGER.info("SECID=%s: сформировано endpoint URL: %s", secid, secid_urls_total)

        LOGGER.info(
            "Статистика запросов: api=%s cache=%s non_json=%s http_skipped=%s errors=%s",
            stats["api"],
            stats["cache"],
            stats["non_json"],
            stats["http_skipped"],
            stats["errors"],
        )

        LOGGER.info("Сохраняю Excel по endpoint-шаблонам в %s", args.output_dir)
        for endpoint_slug, blocks in endpoint_frames.items():
            sheets: dict[str, pd.DataFrame] = {}
            for block_name, frames in blocks.items():
                non_empty_frames = [
                    frame for frame in frames if not frame.empty and not frame.dropna(how="all").empty
                ]
                frames_for_concat = non_empty_frames or frames
                if len(frames_for_concat) == 1:
                    sheets[block_name] = frames_for_concat[0].copy()
                else:
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore", FutureWarning)
                        sheets[block_name] = pd.concat(frames_for_concat, ignore_index=True)
            output_path = save_endpoint_workbook(endpoint_slug=endpoint_slug, frames=sheets, output_dir=args.output_dir)
            LOGGER.info("Сохранён endpoint Excel: %s (листов=%s)", output_path, len(sheets))

    total_seconds = time.perf_counter() - started_at
    LOGGER.info("Готово. Всего endpoint Excel: %s", len(endpoint_frames))
    LOGGER.info("Общее время выполнения: %.2f сек.", total_seconds)


if __name__ == "__main__":
    main()
