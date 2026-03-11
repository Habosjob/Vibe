from __future__ import annotations

import hashlib
import json
import logging
import re
import sqlite3
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from statistics import median
from time import perf_counter
from typing import Any

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from tqdm import tqdm
from urllib3.util import Retry

import config


# =============================
# Базовые утилиты
# =============================
def ensure_directories() -> None:
    for path in (
        config.CACHE_DIR,
        config.RAW_DIR,
        config.DB_DIR,
        config.LOG_DIR,
        config.BASE_SNAPSHOTS_DIR,
    ):
        path.mkdir(parents=True, exist_ok=True)


def setup_logger() -> logging.Logger:
    logger = logging.getLogger("scalp")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    handler = logging.FileHandler(config.LOG_PATH, mode="w", encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(handler)
    return logger


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace("\xa0", "").replace(",", ".")
    if text in {"", "-", "None", "nan", "NaN"}:
        return None
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_str(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_date_any(value: Any) -> datetime | None:
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    txt = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d", "%d-%m-%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(txt, fmt)
        except ValueError:
            continue
    return None


def percent_delta(new_value: float | None, old_value: float | None) -> float | None:
    if new_value is None or old_value in (None, 0):
        return None
    return (new_value - old_value) / old_value * 100


def resolve_column_map(headers: list[str], aliases: dict[str, tuple[str, ...]]) -> dict[str, int]:
    low_map = {h.strip().lower(): idx for idx, h in enumerate(headers)}
    result: dict[str, int] = {}
    for target, options in aliases.items():
        for option in options:
            if option.lower() in low_map:
                result[target] = low_map[option.lower()]
                break
    return result


def nearest_event_days(instrument: dict[str, Any], now_dt: datetime) -> int | None:
    candidates: list[int] = []
    for key in ("next_coupon_date", "offerdate", "amort_start_date"):
        dt = parse_date_any(instrument.get(key))
        if dt:
            candidates.append((dt.date() - now_dt.date()).days)
    return min(candidates) if candidates else None


def make_signal_hash(payload: dict[str, Any]) -> str:
    raw = "|".join(
        [
            normalize_str(payload.get("isin")),
            normalize_str(payload.get("signal_type")),
            normalize_str(payload.get("snapshot_time")),
            f"{parse_number(payload.get('current_clean')) or 0:.4f}",
            f"{parse_number(payload.get('current_dirty')) or 0:.4f}",
        ]
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


# =============================
# SQLite
# =============================
def connect_db() -> sqlite3.Connection:
    conn = sqlite3.connect(config.SQLITE_PATH)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    return conn


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS instruments (
            isin TEXT PRIMARY KEY,
            secid TEXT,
            shortname TEXT,
            issuer_name TEXT,
            inn TEXT,
            scoring TEXT,
            source_sheet TEXT,
            next_coupon_date TEXT,
            offerdate TEXT,
            amort_start_date TEXT,
            ytm TEXT,
            aci TEXT,
            liquidity_score TEXT,
            loaded_at TEXT
        );

        CREATE TABLE IF NOT EXISTS market_snapshots (
            snapshot_id INTEGER PRIMARY KEY AUTOINCREMENT,
            isin TEXT,
            secid TEXT,
            snapshot_time TEXT,
            prev_close_clean REAL,
            open_clean REAL,
            current_clean REAL,
            current_aci REAL,
            current_dirty REAL,
            day_low REAL,
            day_high REAL,
            volume_pieces REAL,
            turnover_rub REAL,
            num_trades REAL,
            source TEXT
        );

        CREATE TABLE IF NOT EXISTS scalp_signals (
            signal_hash TEXT PRIMARY KEY,
            isin TEXT,
            secid TEXT,
            snapshot_time TEXT,
            signal_type TEXT,
            signal_score REAL,
            reason TEXT,
            prev_close_clean REAL,
            open_clean REAL,
            prev_snapshot_clean REAL,
            current_clean REAL,
            current_aci REAL,
            current_dirty REAL,
            delta_open_vs_prevclose_pct REAL,
            delta_current_vs_open_pct REAL,
            delta_current_vs_prevsnapshot_pct REAL,
            delta_dirty_vs_prevclose_pct REAL,
            rebound_from_low_pct REAL,
            volume_pieces REAL,
            turnover_rub REAL,
            num_trades REAL,
            ytm REAL,
            next_coupon_date TEXT,
            offerdate TEXT,
            amort_start_date TEXT,
            days_to_nearest_event INTEGER,
            scoring TEXT,
            liquidity_score TEXT,
            created_at TEXT
        );

        CREATE TABLE IF NOT EXISTS meta (
            key TEXT PRIMARY KEY,
            value TEXT
        );
        """
    )
    conn.commit()


def upsert_instruments(conn: sqlite3.Connection, instruments: list[dict[str, Any]]) -> None:
    now = datetime.now().isoformat(timespec="seconds")
    rows = [
        (
            i.get("isin"), i.get("secid"), i.get("shortname"), i.get("issuer_name"), i.get("inn"), i.get("scoring"),
            i.get("source_sheet"), i.get("next_coupon_date"), i.get("offerdate"), i.get("amort_start_date"),
            i.get("ytm"), i.get("aci"), i.get("liquidity_score"), now,
        )
        for i in instruments
    ]
    conn.executemany(
        """
        INSERT INTO instruments (
            isin, secid, shortname, issuer_name, inn, scoring, source_sheet,
            next_coupon_date, offerdate, amort_start_date, ytm, aci, liquidity_score, loaded_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(isin) DO UPDATE SET
            secid=excluded.secid,
            shortname=excluded.shortname,
            issuer_name=excluded.issuer_name,
            inn=excluded.inn,
            scoring=excluded.scoring,
            source_sheet=excluded.source_sheet,
            next_coupon_date=excluded.next_coupon_date,
            offerdate=excluded.offerdate,
            amort_start_date=excluded.amort_start_date,
            ytm=excluded.ytm,
            aci=excluded.aci,
            liquidity_score=excluded.liquidity_score,
            loaded_at=excluded.loaded_at
        """,
        rows,
    )
    conn.commit()


def insert_market_snapshots(conn: sqlite3.Connection, snapshots: list[dict[str, Any]]) -> None:
    conn.executemany(
        """
        INSERT INTO market_snapshots (
            isin, secid, snapshot_time, prev_close_clean, open_clean, current_clean,
            current_aci, current_dirty, day_low, day_high, volume_pieces, turnover_rub,
            num_trades, source
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                s.get("isin"), s.get("secid"), s.get("snapshot_time"), s.get("prev_close_clean"), s.get("open_clean"),
                s.get("current_clean"), s.get("current_aci"), s.get("current_dirty"), s.get("day_low"), s.get("day_high"),
                s.get("volume_pieces"), s.get("turnover_rub"), s.get("num_trades"), s.get("source"),
            )
            for s in snapshots
        ],
    )
    conn.commit()


def insert_signals(conn: sqlite3.Connection, signals: list[dict[str, Any]]) -> None:
    if not signals:
        return
    conn.executemany(
        """
        INSERT OR IGNORE INTO scalp_signals (
            signal_hash, isin, secid, snapshot_time, signal_type, signal_score, reason,
            prev_close_clean, open_clean, prev_snapshot_clean, current_clean,
            current_aci, current_dirty, delta_open_vs_prevclose_pct, delta_current_vs_open_pct,
            delta_current_vs_prevsnapshot_pct, delta_dirty_vs_prevclose_pct,
            rebound_from_low_pct, volume_pieces, turnover_rub, num_trades, ytm,
            next_coupon_date, offerdate, amort_start_date, days_to_nearest_event,
            scoring, liquidity_score, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                s.get("signal_hash"), s.get("isin"), s.get("secid"), s.get("snapshot_time"), s.get("signal_type"),
                s.get("signal_score"), s.get("reason"), s.get("prev_close_clean"), s.get("open_clean"),
                s.get("prev_snapshot_clean"), s.get("current_clean"), s.get("current_aci"), s.get("current_dirty"),
                s.get("delta_open_vs_prevclose_pct"), s.get("delta_current_vs_open_pct"), s.get("delta_current_vs_prevsnapshot_pct"),
                s.get("delta_dirty_vs_prevclose_pct"), s.get("rebound_from_low_pct"), s.get("volume_pieces"),
                s.get("turnover_rub"), s.get("num_trades"), s.get("ytm"), s.get("next_coupon_date"), s.get("offerdate"),
                s.get("amort_start_date"), s.get("days_to_nearest_event"), s.get("scoring"), s.get("liquidity_score"), s.get("created_at"),
            )
            for s in signals
        ],
    )
    conn.commit()


def load_previous_snapshot_map(conn: sqlite3.Connection) -> dict[str, float]:
    query = """
    SELECT s1.isin, s1.current_clean
    FROM market_snapshots s1
    JOIN (
      SELECT isin, MAX(snapshot_id) AS max_id
      FROM market_snapshots
      GROUP BY isin
    ) x ON x.isin = s1.isin AND x.max_id = s1.snapshot_id
    """
    return {row[0]: row[1] for row in conn.execute(query).fetchall() if row[0]}


# =============================
# Данные входа (Screener / Emitents)
# =============================
def read_emitents_scoring() -> dict[str, dict[str, str]]:
    if not config.INPUT_EMITENTS_PATH.exists():
        return {}
    wb = load_workbook(config.INPUT_EMITENTS_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [normalize_str(h) for h in next(rows)]
    cmap = resolve_column_map(
        headers,
        {
            "inn": ("ИНН", "inn"),
            "issuer_name": ("Эмитент", "Название", "issuer", "issuer_name"),
            "scoring": ("Scoring", "Рейтинг", "score"),
        },
    )
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        inn = normalize_str(row[cmap["inn"]]) if "inn" in cmap else ""
        if not inn:
            continue
        result[inn] = {
            "issuer_name": normalize_str(row[cmap["issuer_name"]]) if "issuer_name" in cmap else "",
            "scoring": normalize_str(row[cmap["scoring"]]) if "scoring" in cmap else "",
        }
    wb.close()
    return result


def load_instruments_from_screener(logger: logging.Logger) -> list[dict[str, Any]]:
    if not config.INPUT_SCREENER_PATH.exists():
        raise FileNotFoundError(f"Screener not found: {config.INPUT_SCREENER_PATH}")
    wb = load_workbook(config.INPUT_SCREENER_PATH, read_only=True, data_only=True)
    emitents = read_emitents_scoring()
    sheets = [name for name in wb.sheetnames if name in {"Green", "Yellow"}]
    instruments: dict[str, dict[str, Any]] = {}

    aliases = {
        "isin": ("ISIN",),
        "secid": ("SECID", "Ticker", "Тикер"),
        "shortname": ("Название", "Name", "shortname"),
        "issuer_name": ("Эмитент", "Issuer", "issuer_name"),
        "inn": ("ИНН", "inn"),
        "scoring": ("Scoring", "Рейтинг", "score"),
        "aci": ("НКД", "ACI"),
        "ytm": ("YTM", "Доходность", "Yield"),
        "next_coupon_date": ("Ближайший купон", "NextCoupon", "coupon_date"),
        "offerdate": ("Offerdate", "Оферта"),
        "amort_start_date": ("AmortStarrtDate", "AmortStartDate", "Амортизация"),
        "liquidity_score": ("Ликвидность", "Liquidity", "liquidity_score"),
    }

    for sheet_name in sheets:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = [normalize_str(h) for h in next(rows)]
        cmap = resolve_column_map(headers, aliases)

        for row in rows:
            isin = normalize_str(row[cmap["isin"]]) if "isin" in cmap else ""
            secid = normalize_str(row[cmap["secid"]]) if "secid" in cmap else ""
            if not isin and not secid:
                continue
            key = isin or secid
            inn = normalize_str(row[cmap["inn"]]) if "inn" in cmap else ""
            em_data = emitents.get(inn, {})
            item = {
                "isin": isin,
                "secid": secid,
                "shortname": normalize_str(row[cmap["shortname"]]) if "shortname" in cmap else "",
                "issuer_name": normalize_str(row[cmap["issuer_name"]]) if "issuer_name" in cmap else em_data.get("issuer_name", ""),
                "inn": inn,
                "scoring": normalize_str(row[cmap["scoring"]]) if "scoring" in cmap else em_data.get("scoring", ""),
                "source_sheet": sheet_name,
                "next_coupon_date": normalize_str(row[cmap["next_coupon_date"]]) if "next_coupon_date" in cmap else "",
                "offerdate": normalize_str(row[cmap["offerdate"]]) if "offerdate" in cmap else "",
                "amort_start_date": normalize_str(row[cmap["amort_start_date"]]) if "amort_start_date" in cmap else "",
                "ytm": normalize_str(row[cmap["ytm"]]) if "ytm" in cmap else "",
                "aci": normalize_str(row[cmap["aci"]]) if "aci" in cmap else "",
                "liquidity_score": normalize_str(row[cmap["liquidity_score"]]) if "liquidity_score" in cmap else "",
            }
            if key not in instruments or (item["source_sheet"] == "Green" and instruments[key].get("source_sheet") != "Green"):
                instruments[key] = item

    wb.close()
    logger.info("Loaded %s instruments from Green/Yellow", len(instruments))
    return list(instruments.values())


# =============================
# MOEX market data
# =============================
def make_session() -> requests.Session:
    retry = Retry(
        total=config.REQUEST_RETRIES,
        backoff_factor=config.REQUEST_BACKOFF_SECONDS,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET"}),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session = requests.Session()
    session.headers.update({"User-Agent": config.USER_AGENT})
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def _cache_path(secid: str) -> Path:
    return config.CACHE_DIR / f"{secid}.json"


def _read_cache(secid: str) -> dict[str, Any] | None:
    path = _cache_path(secid)
    if not path.exists():
        return None
    if time.time() - path.stat().st_mtime > config.MARKET_TTL_SECONDS:
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return None


def _write_cache(secid: str, payload: dict[str, Any]) -> None:
    _cache_path(secid).write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")


def _extract_first_row(block: dict[str, Any]) -> dict[str, Any]:
    columns = block.get("columns", [])
    values = block.get("data", [])
    if not columns or not values:
        return {}
    row = values[0]
    return {columns[idx]: row[idx] for idx in range(min(len(columns), len(row)))}


def fetch_market_snapshot(session: requests.Session, instrument: dict[str, Any]) -> dict[str, Any]:
    secid = instrument.get("secid")
    if not secid:
        return {}

    payload = _read_cache(secid)
    source = "moex_cache"
    if payload is None:
        url = config.MOEX_ISS_SECURITY_ENDPOINT.format(secid=secid)
        response = session.get(url, timeout=config.REQUEST_TIMEOUT_SECONDS)
        response.raise_for_status()
        payload = response.json()
        _write_cache(secid, payload)
        source = "moex"

    market = _extract_first_row(payload.get("marketdata", {}))
    securities = _extract_first_row(payload.get("securities", {}))

    current_clean = parse_number(market.get("LAST") or securities.get("PREVPRICE") or securities.get("PREVWAPRICE"))
    prev_close_clean = parse_number(securities.get("PREVPRICE") or market.get("LCLOSEPRICE"))
    open_clean = parse_number(market.get("OPEN") or securities.get("PREVLEGALCLOSEPRICE") or prev_close_clean)
    day_low = parse_number(market.get("LOW") or current_clean)
    day_high = parse_number(market.get("HIGH") or current_clean)
    volume_pieces = parse_number(market.get("VOLUME") or securities.get("VOLTODAY"))
    turnover_rub = parse_number(market.get("VALUE") or securities.get("VALTODAY_RUR"))
    num_trades = parse_number(market.get("NUMTRADES") or securities.get("NUMTRADES"))

    current_aci = parse_number(securities.get("ACCRUEDINT"))
    if current_aci is None:
        current_aci = parse_number(instrument.get("aci"))
        source = f"{source}+screener_aci"

    current_dirty = current_clean + current_aci if current_clean is not None and current_aci is not None else None

    return {
        "isin": instrument.get("isin"),
        "secid": secid,
        "snapshot_time": datetime.now().isoformat(timespec="seconds"),
        "prev_close_clean": prev_close_clean,
        "open_clean": open_clean,
        "current_clean": current_clean,
        "current_aci": current_aci,
        "current_dirty": current_dirty,
        "day_low": day_low,
        "day_high": day_high,
        "volume_pieces": volume_pieces,
        "turnover_rub": turnover_rub,
        "num_trades": num_trades,
        "source": source,
    }


def collect_market_snapshots(instruments: list[dict[str, Any]], logger: logging.Logger) -> list[dict[str, Any]]:
    session = make_session()
    snapshots: list[dict[str, Any]] = []
    futures = []
    with ThreadPoolExecutor(max_workers=config.MAX_WORKERS) as pool:
        for instrument in instruments:
            if instrument.get("secid"):
                futures.append(pool.submit(fetch_market_snapshot, session, instrument))
        bar = tqdm(total=len(futures), desc="Этап 3", unit="bond", leave=False, position=0, dynamic_ncols=True)
        for f in as_completed(futures):
            try:
                row = f.result()
                if row:
                    snapshots.append(row)
            except Exception as exc:
                logger.exception("Snapshot error: %s", exc)
            finally:
                bar.update(1)
        bar.close()
    return snapshots


# =============================
# Signal engine
# =============================
def _liquidity_ok(snapshot: dict[str, Any]) -> bool:
    return (
        (snapshot.get("num_trades") or 0) >= config.MIN_NUM_TRADES
        and (snapshot.get("turnover_rub") or 0) >= config.MIN_TURNOVER_RUB
        and (snapshot.get("volume_pieces") or 0) >= config.MIN_VOLUME_PIECES
    )


def _quality_bonus(source_sheet: str) -> float:
    return 12 if source_sheet == "Green" else 6


def _event_penalty(days_to_event: int | None) -> float:
    if days_to_event is None:
        return 0
    if days_to_event <= config.EVENT_HARD_BLOCK_DAYS:
        return 40
    if days_to_event <= config.MIN_DAYS_TO_EVENT:
        return 20
    return 0


def _evaluate_signal(row: dict[str, Any], peer_medians: dict[str, float]) -> dict[str, Any] | None:
    days_to_event = row.get("days_to_nearest_event")
    liquidity_ok = _liquidity_ok(row)
    event_penalty = _event_penalty(days_to_event)

    delta_open = row.get("delta_open_vs_prevclose_pct")
    delta_curr_open = row.get("delta_current_vs_open_pct")
    delta_dirty = row.get("delta_dirty_vs_prevclose_pct")
    rebound = row.get("rebound_from_low_pct")

    candidates: list[tuple[str, float, str]] = []

    if delta_open is not None and delta_open <= config.GAPDOWN_PCT_THRESHOLD:
        score = min(100, abs(delta_open) * 19 + _quality_bonus(row.get("source_sheet", "")) - event_penalty)
        reason = f"Открытие {delta_open:.2f}% к prev close, ликвидность {'OK' if liquidity_ok else 'низкая'}"
        candidates.append(("GapDown", score, reason))

    if (
        (delta_curr_open is not None and delta_curr_open <= config.INTRADAY_DUMP_PCT_THRESHOLD)
        or (delta_dirty is not None and delta_dirty <= config.DIRTY_DROP_PCT_THRESHOLD)
    ):
        base = abs(delta_dirty or delta_curr_open or 0)
        score = min(100, base * 22 + _quality_bonus(row.get("source_sheet", "")) - event_penalty)
        reason = (
            f"Dirty {delta_dirty:.2f}% к prev close, оборот {int(row.get('turnover_rub') or 0):,} руб"
            if delta_dirty is not None
            else f"Текущая {delta_curr_open:.2f}% к open"
        )
        candidates.append(("IntradayDump", score, reason))

    low_vs_open = percent_delta(row.get("day_low"), row.get("open_clean"))
    if (
        low_vs_open is not None
        and low_vs_open <= config.REBOUND_REQUIRED_DUMP_PCT
        and rebound is not None
        and rebound >= config.REBOUND_MIN_PCT
    ):
        score = min(100, abs(low_vs_open) * 15 + rebound * 15 + _quality_bonus(row.get("source_sheet", "")) - event_penalty)
        reason = f"Падение до low {low_vs_open:.2f}% и отскок {rebound:.2f}%"
        candidates.append(("ReboundCandidate", score, reason))

    issuer = row.get("issuer_name")
    issuer_med = peer_medians.get(issuer) if issuer else None
    if issuer_med is not None and delta_dirty is not None and delta_dirty <= issuer_med + config.PEER_DISLOCATION_DELTA_PCT:
        score = min(100, abs(delta_dirty - issuer_med) * 24 + _quality_bonus(row.get("source_sheet", "")) - event_penalty)
        reason = f"Хуже peers эмитента на {delta_dirty - issuer_med:.2f} п.п. по dirty-изменению"
        candidates.append(("PeerDislocation", score, reason))

    if not candidates:
        return None

    signal_type, score, reason = sorted(candidates, key=lambda x: x[1], reverse=True)[0]

    if not liquidity_ok:
        score *= 0.6
        reason += "; сигнал ослаблен из-за низкой ликвидности"

    if days_to_event is not None and days_to_event <= config.EVENT_HARD_BLOCK_DAYS:
        return None

    if row.get("delta_dirty_vs_prevclose_pct") is not None and abs(row["delta_dirty_vs_prevclose_pct"]) < 0.35:
        score *= 0.75
        reason += "; dirty-движение слабое"

    if score < config.SIGNAL_MIN_SCORE:
        return None

    payload = {
        "isin": row.get("isin"),
        "secid": row.get("secid"),
        "snapshot_time": row.get("snapshot_time"),
        "signal_type": signal_type,
        "signal_score": round(score, 2),
        "reason": reason,
        "prev_close_clean": row.get("prev_close_clean"),
        "open_clean": row.get("open_clean"),
        "prev_snapshot_clean": row.get("prev_snapshot_clean"),
        "current_clean": row.get("current_clean"),
        "current_aci": row.get("current_aci"),
        "current_dirty": row.get("current_dirty"),
        "delta_open_vs_prevclose_pct": row.get("delta_open_vs_prevclose_pct"),
        "delta_current_vs_open_pct": row.get("delta_current_vs_open_pct"),
        "delta_current_vs_prevsnapshot_pct": row.get("delta_current_vs_prevsnapshot_pct"),
        "delta_dirty_vs_prevclose_pct": row.get("delta_dirty_vs_prevclose_pct"),
        "rebound_from_low_pct": row.get("rebound_from_low_pct"),
        "volume_pieces": row.get("volume_pieces"),
        "turnover_rub": row.get("turnover_rub"),
        "num_trades": row.get("num_trades"),
        "ytm": row.get("ytm_float"),
        "next_coupon_date": row.get("next_coupon_date"),
        "offerdate": row.get("offerdate"),
        "amort_start_date": row.get("amort_start_date"),
        "days_to_nearest_event": row.get("days_to_nearest_event"),
        "scoring": row.get("scoring"),
        "liquidity_score": row.get("liquidity_score"),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "shortname": row.get("shortname"),
        "issuer_name": row.get("issuer_name"),
        "inn": row.get("inn"),
        "day_low": row.get("day_low"),
        "day_high": row.get("day_high"),
    }
    payload["signal_hash"] = make_signal_hash(payload)
    return payload


def build_signals(
    instruments: list[dict[str, Any]],
    snapshots: list[dict[str, Any]],
    previous_snapshot_map: dict[str, float],
    logger: logging.Logger,
) -> list[dict[str, Any]]:
    now_dt = datetime.now()
    by_isin = {i.get("isin"): i for i in instruments if i.get("isin")}
    issuer_dirty_moves: dict[str, list[float]] = {}
    enriched: list[dict[str, Any]] = []

    for snap in snapshots:
        instrument = by_isin.get(snap.get("isin"))
        if not instrument:
            continue
        prev_snapshot_clean = previous_snapshot_map.get(snap.get("isin"))
        row = {
            **instrument,
            **snap,
            "prev_snapshot_clean": prev_snapshot_clean,
            "delta_open_vs_prevclose_pct": percent_delta(snap.get("open_clean"), snap.get("prev_close_clean")),
            "delta_current_vs_open_pct": percent_delta(snap.get("current_clean"), snap.get("open_clean")),
            "delta_current_vs_prevsnapshot_pct": percent_delta(snap.get("current_clean"), prev_snapshot_clean),
            "delta_dirty_vs_prevclose_pct": percent_delta(snap.get("current_dirty"), snap.get("prev_close_clean")),
            "rebound_from_low_pct": percent_delta(snap.get("current_clean"), snap.get("day_low")),
            "days_to_nearest_event": nearest_event_days(instrument, now_dt),
            "ytm_float": parse_number(instrument.get("ytm")),
        }
        enriched.append(row)

        issuer = instrument.get("issuer_name")
        dirty_move = row.get("delta_dirty_vs_prevclose_pct")
        if issuer and dirty_move is not None:
            issuer_dirty_moves.setdefault(issuer, []).append(dirty_move)

    peer_medians = {issuer: median(values) for issuer, values in issuer_dirty_moves.items() if values}

    signals: list[dict[str, Any]] = []
    bar = tqdm(total=len(enriched), desc="Этап 4", unit="bond", leave=False, position=0, dynamic_ncols=True)
    for row in enriched:
        try:
            if row.get("current_clean") is None or row.get("prev_close_clean") is None:
                continue
            signal = _evaluate_signal(row, peer_medians)
            if signal:
                signals.append(signal)
        except Exception as exc:
            logger.exception("Signal error: %s", exc)
        finally:
            bar.update(1)
    bar.close()

    logger.info("Signals produced: %s", len(signals))
    return signals


# =============================
# Excel export
# =============================
HEADERS = [
    "Signal", "SignalScore", "SignalType", "ISIN", "SECID", "Название", "Эмитент", "ИНН", "Scoring", "Ликвидность",
    "Цена закрытия пред. дня", "Цена открытия", "Цена предыдущей выгрузки", "Текущая цена", "Текущий НКД", "Текущая dirty price",
    "Δ Open vs PrevClose, %", "Δ Current vs Open, %", "Δ Current vs PrevSnapshot, %", "Δ Dirty vs PrevClose, %",
    "Day Low", "Day High", "Отскок от low, %", "Объем, шт", "Оборот, руб", "Сделки, шт", "YTM",
    "Ближайший купон", "Offerdate", "AmortStarrtDate", "DaysToNearestEvent", "Reason", "SnapshotTime",
]


def export_scalp_excel(signals: list[dict[str, Any]], snapshots: list[dict[str, Any]], logger: logging.Logger) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scalp"
    ws.append(HEADERS)

    for signal in sorted(signals, key=lambda x: x.get("signal_score", 0), reverse=True):
        ws.append(
            [
                "BUY" if signal.get("signal_score", 0) >= config.SIGNAL_MIN_SCORE else "WATCH",
                signal.get("signal_score"), signal.get("signal_type"), signal.get("isin"), signal.get("secid"),
                signal.get("shortname"), signal.get("issuer_name"), signal.get("inn"), signal.get("scoring"),
                signal.get("liquidity_score"), signal.get("prev_close_clean"), signal.get("open_clean"),
                signal.get("prev_snapshot_clean"), signal.get("current_clean"), signal.get("current_aci"), signal.get("current_dirty"),
                signal.get("delta_open_vs_prevclose_pct"), signal.get("delta_current_vs_open_pct"), signal.get("delta_current_vs_prevsnapshot_pct"),
                signal.get("delta_dirty_vs_prevclose_pct"), signal.get("day_low"), signal.get("day_high"), signal.get("rebound_from_low_pct"),
                signal.get("volume_pieces"), signal.get("turnover_rub"), signal.get("num_trades"), signal.get("ytm"),
                signal.get("next_coupon_date"), signal.get("offerdate"), signal.get("amort_start_date"),
                signal.get("days_to_nearest_event"), signal.get("reason"), signal.get("snapshot_time"),
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if ws.max_row >= 2:
        ws.conditional_formatting.add(
            f"B2:B{ws.max_row}",
            ColorScaleRule(start_type="num", start_value=0, start_color="FFF2CC", mid_type="num", mid_value=50, mid_color="FFD966", end_type="num", end_value=100, end_color="63BE7B"),
        )
    else:
        logger.info("No signal rows for conditional formatting; exported headers only")

    for row in range(2, ws.max_row + 1):
        score_cell = ws[f"I{row}"]
        scoring = str(score_cell.value or "")
        if "Green" in scoring or scoring.startswith("G"):
            ws[f"A{row}"].fill = PatternFill("solid", fgColor="E2F0D9")
        elif "Yellow" in scoring or scoring.startswith("Y"):
            ws[f"A{row}"].fill = PatternFill("solid", fgColor="FFF2CC")

    widths = {"A": 10, "B": 11, "C": 16, "D": 14, "E": 12, "F": 22, "G": 24, "H": 14, "I": 12, "J": 12, "K": 14,
              "L": 14, "M": 16, "N": 12, "O": 12, "P": 16, "Q": 17, "R": 17, "S": 20, "T": 17, "U": 10, "V": 10,
              "W": 14, "X": 14, "Y": 16, "Z": 10, "AA": 8, "AB": 16, "AC": 13, "AD": 14, "AE": 16, "AF": 55, "AG": 21}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    raw = wb.create_sheet("BaseSnapshot")
    raw.append(["isin", "secid", "prev_close_clean", "open_clean", "current_clean", "current_aci", "current_dirty", "volume_pieces", "turnover_rub", "num_trades", "source", "snapshot_time"])
    for item in snapshots[: config.SNAPSHOT_SAMPLE_ROWS]:
        raw.append([
            item.get("isin"), item.get("secid"), item.get("prev_close_clean"), item.get("open_clean"), item.get("current_clean"),
            item.get("current_aci"), item.get("current_dirty"), item.get("volume_pieces"), item.get("turnover_rub"),
            item.get("num_trades"), item.get("source"), item.get("snapshot_time"),
        ])
    raw.freeze_panes = "A2"
    raw.auto_filter.ref = raw.dimensions

    wb.save(config.OUTPUT_EXCEL_PATH)
    wb.save(config.SNAPSHOT_EXCEL_PATH)
    logger.info("Excel exported: %s and %s", config.OUTPUT_EXCEL_PATH, config.SNAPSHOT_EXCEL_PATH)


# =============================
# Orchestration
# =============================
def _run_stage(name: str, fn):
    print(f"=====\n{name}")
    start = perf_counter()
    result = fn()
    return result, perf_counter() - start


def main() -> None:
    total_start = perf_counter()
    timings: list[tuple[str, float]] = []
    logger = None
    conn = None

    try:
        _, t0 = _run_stage("Этап 1: Подготовка окружения", ensure_directories)
        timings.append(("Подготовка окружения", t0))

        logger = setup_logger()
        conn = connect_db()
        init_db(conn)

        instruments, t1 = _run_stage("Этап 2: Загрузка Green/Yellow", lambda: load_instruments_from_screener(logger))
        timings.append(("Загрузка инструментов", t1))

        _, t2 = _run_stage("Этап 2.1: Запись instruments в SQLite", lambda: upsert_instruments(conn, instruments))
        timings.append(("Запись instruments", t2))

        previous_snapshot_map = load_previous_snapshot_map(conn)
        snapshots, t3 = _run_stage("Этап 3: Сбор market snapshots", lambda: collect_market_snapshots(instruments, logger))
        timings.append(("Сбор snapshots", t3))

        _, t4 = _run_stage("Этап 3.1: Запись snapshots в SQLite", lambda: insert_market_snapshots(conn, snapshots))
        timings.append(("Запись snapshots", t4))

        signals, t5 = _run_stage("Этап 4: Расчет сигналов", lambda: build_signals(instruments, snapshots, previous_snapshot_map, logger))
        timings.append(("Расчет сигналов", t5))

        _, t6 = _run_stage("Этап 4.1: Запись сигналов", lambda: insert_signals(conn, signals))
        timings.append(("Запись сигналов", t6))

        _, t7 = _run_stage("Этап 5: Экспорт Excel", lambda: export_scalp_excel(signals, snapshots, logger))
        timings.append(("Экспорт Excel", t7))

        logger.info("Run completed: instruments=%s snapshots=%s signals=%s", len(instruments), len(snapshots), len(signals))
    except Exception as exc:
        if logger:
            logger.exception("Run failed: %s", exc)
        print(f"Ошибка выполнения: {exc}")
    finally:
        if conn:
            conn.close()
        total_elapsed = perf_counter() - total_start
        print("=====\nSummary")
        for stage_name, sec in timings:
            print(f"- {stage_name}: {sec:.2f} сек")
        print(f"- Всего: {total_elapsed:.2f} сек")
        if logger:
            logger.info("Finished at %s", datetime.now().isoformat(timespec="seconds"))


if __name__ == "__main__":
    with tqdm(total=0, desc="Scalp", leave=False, position=0, dynamic_ncols=True):
        main()
