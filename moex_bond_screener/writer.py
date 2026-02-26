"""Сохранение результатов в Excel/CSV."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from functools import lru_cache
from pathlib import Path
from typing import Any

import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import yaml


DEFAULT_FIELDS = ["SECID", "SHORTNAME", "ISIN", "CURRENCYID", "PREVLEGALCLOSEPRICE", "MATDATE"]
UNWANTED_FIELDS = {
    "BOARDID",
    "LOTSIZE",
    "BOARDNAME",
    "STATUS",
    "DECIMALS",
    "PREVDATE",
    "SECNAME",
    "REMARKS",
    "MARKETCODE",
    "INSTRID",
    "LATNAME",
    "REGNUMBER",
    "LISTLEVEL",
    "SECTYPE",
    "SETTLEDATE",
    "MINSTEP",
    "LOTVALUE",
    "FACEVALUEONSETTLEDATE",
    "SECTORID",
}
DEFAULT_GROUP_ORDER = [
    "Служебная информация",
    "Торги и доходность",
    "Купоны и номинал",
    "Даты",
    "Прочее",
]
DEFAULT_FORCED_GROUPS = {
    "AMORTIZATION_START_DATE": "Даты",
    "CURRENCYID": "Купоны и номинал",
    "YTM": "Купоны и номинал",
}
DEFAULT_PRIORITY_FIELDS = ["SHORTNAME", "ISIN", "DATA_STATUS", "MATDATE", "AMORTIZATION_START_DATE", "SECID"]
DEFAULT_COLLAPSED_GROUPS = {"Прочее", "Даты", "Торги и доходность"}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
GROUP_FONT = Font(color="000000", bold=True)
ROW_FILL = PatternFill(fill_type="solid", fgColor="F2F7FF")
GROUP_COLORS = {
    "Служебная информация": "D9E1F2",
    "Торги и доходность": "E2F0D9",
    "Купоны и номинал": "FCE4D6",
    "Даты": "FFF2CC",
    "Прочее": "E4DFEC",
}

SEPARATOR_FIELD = "__GROUP_SEPARATOR__"
SEPARATOR_COLUMN_WIDTH = 18
NUMERIC_STRING_RE = re.compile(r"^[+-]?\d[\d\s\u00A0]*(?:[.,]\d+)?$")
UNICODE_SPACES_RE = re.compile(r"[\s\u00A0\u202F\u2007]+")
APPROX_FILL = PatternFill(fill_type="solid", fgColor="FFF59D")


@lru_cache(maxsize=1)
def _load_excel_layout() -> tuple[list[str], dict[str, str], list[str]]:
    path = Path("excel_layout.yml")
    if not path.exists():
        return DEFAULT_GROUP_ORDER, DEFAULT_FORCED_GROUPS, DEFAULT_PRIORITY_FIELDS

    with path.open("r", encoding="utf-8") as file:
        payload = yaml.safe_load(file) or {}

    group_order = payload.get("group_order") if isinstance(payload, dict) else None
    forced_groups = payload.get("forced_groups") if isinstance(payload, dict) else None
    priority_fields = payload.get("priority_fields") if isinstance(payload, dict) else None

    normalized_group_order = [str(item).strip() for item in (group_order or []) if str(item).strip()] or DEFAULT_GROUP_ORDER
    normalized_forced_groups = {
        str(field).strip().upper(): str(group).strip()
        for field, group in (forced_groups or {}).items()
        if str(field).strip() and str(group).strip()
    }
    if not normalized_forced_groups:
        normalized_forced_groups = DEFAULT_FORCED_GROUPS

    normalized_priority = [str(item).strip() for item in (priority_fields or []) if str(item).strip()]
    if not normalized_priority:
        normalized_priority = DEFAULT_PRIORITY_FIELDS

    return normalized_group_order, normalized_forced_groups, normalized_priority


def _resolve_fields(bonds: list[dict[str, Any]]) -> list[str]:
    if not bonds:
        return DEFAULT_FIELDS.copy()

    fields = list(bonds[0].keys())
    for bond in bonds[1:]:
        for key in bond.keys():
            if key not in fields:
                fields.append(key)

    return fields


def _group_name(field: str) -> str:
    group_order, forced_groups, _ = _load_excel_layout()
    upper = field.upper()
    if upper in forced_groups:
        forced = forced_groups[upper]
        if forced in group_order:
            return forced
    if upper in {"SHORTNAME", "ISIN", "FACEUNIT", "BONDNAME", "EMITTER"}:
        return "Служебная информация"
    if any(token in upper for token in ["PRICE", "YIELD", "WAPRICE", "DURATION", "SPREAD"]):
        return "Торги и доходность"
    if any(token in upper for token in ["COUPON", "ACCRUED", "ACCINT", "FACE", "NOMINAL", "AMORT"]):
        return "Купоны и номинал"
    if "DATE" in upper or any(token in upper for token in ["MAT", "OFFER", "BEGIN", "END"]):
        return "Даты"
    return "Прочее" if "Прочее" in group_order else group_order[-1]


def _apply_priority_fields(fields: list[str]) -> list[str]:
    _, _, priority_fields = _load_excel_layout()
    order_map = {name.upper(): idx for idx, name in enumerate(priority_fields)}
    return sorted(fields, key=lambda field: (order_map.get(field.upper(), len(order_map)), fields.index(field)))


def _is_iso_date(value: str) -> bool:
    try:
        datetime.strptime(value, "%Y-%m-%d")
        return True
    except ValueError:
        return False


def _is_date_field(field: str) -> bool:
    upper = field.upper()
    return "DATE" in upper or any(token in upper for token in ["MAT", "OFFER", "BEGIN", "END", "COUPON"])


def _coerce_excel_date(field: str, value: Any) -> Any:
    """Возвращает значение даты в типе datetime для корректной группировки в фильтрах Excel."""
    if isinstance(value, datetime):
        return value

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    if isinstance(value, str):
        if value == "0000-00-00":
            return ""
        if _is_date_field(field) and _is_iso_date(value):
            return datetime.strptime(value, "%Y-%m-%d")

    return value


def _format_value(field: str, value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return value.strftime("%d.%m.%Y")

    if isinstance(value, str):
        if value == "0000-00-00":
            return ""
        if _is_date_field(field) and _is_iso_date(value):
            return datetime.strptime(value, "%Y-%m-%d").strftime("%d.%m.%Y")

    return value


def _format_excel_value(field: str, value: Any) -> Any:
    if value is None:
        return ""

    excel_date = _coerce_excel_date(field, value)
    if excel_date == "":
        return ""
    if isinstance(excel_date, str):
        numeric = _coerce_numeric_string(excel_date)
        if numeric is not None:
            return numeric
    return excel_date


def _coerce_numeric_string(value: str) -> int | float | None:
    normalized = value.strip().replace("\u00A0", " ").replace("\u202F", " ").replace("\u2007", " ")
    if not normalized or not NUMERIC_STRING_RE.match(normalized):
        return None

    compact = UNICODE_SPACES_RE.sub("", normalized)
    if "," in compact and "." not in compact:
        compact = compact.replace(",", ".")

    try:
        if "." in compact:
            as_float = float(compact)
            if as_float.is_integer():
                return int(as_float)
            return as_float
        return int(compact)
    except ValueError:
        return None


def _is_numeric_like(value: Any) -> bool:
    return isinstance(value, (int, float, Decimal)) and not isinstance(value, bool)


def _excel_number_format(field: str, values: list[Any]) -> str | None:
    numeric_values = [value for value in values if _is_numeric_like(value)]
    if not numeric_values:
        return None

    has_fraction = any(abs(float(value) - int(float(value))) > 1e-9 for value in numeric_values)
    if has_fraction:
        return "#,##0.00"
    return "#,##0"


def _summary_metrics(prepared_rows: list[dict[str, Any]], summary: dict[str, Any] | None) -> dict[str, Any]:
    payload = summary.copy() if summary else {}
    payload.setdefault("bonds_count", len(prepared_rows))
    payload.setdefault("errors_count", 0)
    payload.setdefault("elapsed_seconds", 0.0)
    payload.setdefault("generated_at", datetime.now())
    return payload


def _write_summary_sheet(workbook: Workbook, prepared_rows: list[dict[str, Any]], summary: dict[str, Any] | None) -> None:
    payload = _summary_metrics(prepared_rows, summary)
    summary_sheet = workbook.create_sheet("SUMMARY", 0)
    summary_sheet.append(["Параметр", "Значение"])

    base_rows = [
        ("Дата и время формирования", payload["generated_at"]),
        ("Количество бумаг", int(payload["bonds_count"])),
        ("Количество ошибок", int(payload["errors_count"])),
        ("Время выполнения, сек", float(payload["elapsed_seconds"])),
    ]

    known = {"generated_at", "bonds_count", "errors_count", "elapsed_seconds"}
    extra_rows: list[tuple[str, Any]] = []
    for key, value in payload.items():
        if key in known:
            continue
        title = key.replace("_", " ").strip().capitalize()
        extra_rows.append((title, value))

    for label, value in [*base_rows, *extra_rows]:
        summary_sheet.append([label, value])

    for cell in summary_sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_sheet.column_dimensions["A"].width = 42
    summary_sheet.column_dimensions["B"].width = 24
    summary_sheet["B2"].number_format = "DD.MM.YYYY HH:MM:SS"
    summary_sheet["B5"].number_format = "0.00"
    summary_sheet.freeze_panes = "A2"


def _excel_date_format(field: str, values: list[Any]) -> str | None:
    if not _is_date_field(field):
        return None

    if any(isinstance(value, datetime) for value in values):
        return "DD.MM.YYYY"

    return None


def _prepare_export_data(bonds: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    fields = _resolve_fields(bonds)
    fields = [field for field in fields if field not in UNWANTED_FIELDS]

    prepared_rows: list[dict[str, Any]] = []
    for bond in bonds:
        row = dict(bond)
        row.setdefault("_COUPONPERCENT_APPROX", False)
        if not row.get("CURRENCYID") and row.get("FACEUNIT"):
            row["CURRENCYID"] = row["FACEUNIT"]
        row.pop("FACEUNIT", None)
        prepared_rows.append(row)

    fields = [field for field in fields if field != "FACEUNIT" and not str(field).startswith("_")]
    if "CURRENCYID" not in fields and any(row.get("CURRENCYID") for row in prepared_rows):
        fields.append("CURRENCYID")

    seen_signatures: dict[tuple[Any, ...], str] = {}
    deduplicated_fields: list[str] = []
    for field in fields:
        signature = tuple(prepared.get(field, "") for prepared in prepared_rows)
        if signature in seen_signatures and any(value not in ("", None) for value in signature):
            continue
        seen_signatures[signature] = field
        deduplicated_fields.append(field)

    group_order, _, _ = _load_excel_layout()
    grouped: dict[str, list[str]] = {name: [] for name in group_order}
    for field in deduplicated_fields:
        grouped[_group_name(field)].append(field)

    for group_name in group_order:
        grouped[group_name] = _apply_priority_fields(grouped[group_name])

    if "SECID" in grouped["Служебная информация"]:
        grouped["Служебная информация"].remove("SECID")
        grouped["Прочее"].append("SECID")

    ordered_fields: list[str] = []
    for group_name in group_order:
        ordered_fields.extend(grouped[group_name])

    return ordered_fields, prepared_rows


def _build_columns(fields: list[str]) -> list[tuple[str, str]]:
    group_order, _, _ = _load_excel_layout()
    columns: list[tuple[str, str]] = []
    grouped: dict[str, list[str]] = {name: [] for name in group_order}
    for field in fields:
        grouped[_group_name(field)].append(field)

    for group_name in group_order:
        group_fields = grouped[group_name]
        if not group_fields:
            continue
        columns.append((group_name, SEPARATOR_FIELD))
        for field in group_fields:
            columns.append((group_name, field))

    return columns


def save_bonds_excel(path: str, bonds: list[dict[str, Any]], summary: dict[str, Any] | None = None) -> None:
    """Сохраняет список облигаций в Excel (.xlsx) без проблем с кодировкой."""
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    fields, prepared_rows = _prepare_export_data(bonds)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MOEX_BONDS"

    excel_columns = _write_grouped_headers(sheet, fields)
    for bond in prepared_rows:
        row_values: list[Any] = []
        for field in excel_columns:
            if field == SEPARATOR_FIELD:
                row_values.append("")
                continue
            row_values.append(_format_excel_value(field, bond.get(field, "")))
        sheet.append(row_values)

    _apply_excel_formatting(sheet)
    _highlight_approximate_coupon(sheet, prepared_rows, excel_columns)
    _write_summary_sheet(workbook, prepared_rows, summary)

    workbook.save(target)


def _apply_excel_formatting(sheet: Any) -> None:
    header_row = 2
    max_col = sheet.max_column
    max_row = sheet.max_row

    for cell in sheet[1]:
        if cell.value:
            cell.fill = PatternFill(fill_type="solid", fgColor=GROUP_COLORS.get(cell.value, "7EA6D8"))
            cell.font = GROUP_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, cell in enumerate(sheet[header_row], start=1):
        if not cell.value:
            group_name = sheet.cell(row=1, column=col_idx).value
            if group_name:
                cell.fill = PatternFill(fill_type="solid", fgColor=GROUP_COLORS.get(group_name, "D9E1F2"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                sheet.column_dimensions[get_column_letter(col_idx)].width = SEPARATOR_COLUMN_WIDTH
            continue

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    separator_columns: set[int] = set()
    for col_idx in range(1, max_col + 1):
        field_name = sheet.cell(row=2, column=col_idx).value
        if field_name in ("", None):
            separator_columns.add(col_idx)

    for row_idx in range(3, max_row + 1):
        if row_idx % 2 == 1:
            for col_idx in range(1, max_col + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = ROW_FILL

    for col_idx in separator_columns:
        group_name = sheet.cell(row=1, column=col_idx).value
        fill = PatternFill(fill_type="solid", fgColor=GROUP_COLORS.get(group_name, "D9E1F2"))
        for row_idx in range(1, max_row + 1):
            sheet.cell(row=row_idx, column=col_idx).fill = fill

    for col_idx in range(1, max_col + 1):
        column_letter = get_column_letter(col_idx)
        values = [sheet.cell(row=row_idx, column=col_idx).value for row_idx in range(1, max_row + 1)]
        max_len = max((len(str(value)) for value in values if value is not None), default=0)
        width = min(max(max_len + 2, 10), 28)

        field_name = sheet.cell(row=2, column=col_idx).value
        if not field_name:
            sheet.column_dimensions[column_letter].width = max(width, SEPARATOR_COLUMN_WIDTH)
            continue

        sheet.column_dimensions[column_letter].width = width
        data_values = [sheet.cell(row=row_idx, column=col_idx).value for row_idx in range(3, max_row + 1)]
        number_format = _excel_number_format(str(field_name), data_values)
        if number_format:
            for row_idx in range(3, max_row + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if _is_numeric_like(value):
                    sheet.cell(row=row_idx, column=col_idx).number_format = number_format

        date_format = _excel_date_format(str(field_name), data_values)
        if date_format:
            for row_idx in range(3, max_row + 1):
                value = sheet.cell(row=row_idx, column=col_idx).value
                if isinstance(value, datetime):
                    sheet.cell(row=row_idx, column=col_idx).number_format = date_format

    sheet.row_dimensions[1].height = 42
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"


def _write_grouped_headers(sheet: Any, fields: list[str]) -> list[str]:
    if not fields:
        sheet.append([])
        sheet.append([])
        return []

    columns = _build_columns(fields)
    sheet.append([group if field == SEPARATOR_FIELD else "" for group, field in columns])
    sheet.append(["" if field == SEPARATOR_FIELD else field for _, field in columns])

    current_group: str | None = None
    group_start = 0
    group_end = 0

    for index, (group_name, field_name) in enumerate(columns, start=1):
        column_letter = get_column_letter(index)
        if field_name == SEPARATOR_FIELD:
            sheet.column_dimensions[column_letter].outlineLevel = 0
            if current_group is not None and group_start > 0:
                for col in range(group_start, group_end + 1):
                    letter = get_column_letter(col)
                    sheet.column_dimensions[letter].outlineLevel = 1
                    sheet.column_dimensions[letter].hidden = current_group in DEFAULT_COLLAPSED_GROUPS
            current_group = group_name
            group_start = index + 1
            group_end = index
            continue

        group_end = index

    if current_group is not None and group_start > 0 and group_end >= group_start:
        for col in range(group_start, group_end + 1):
            letter = get_column_letter(col)
            sheet.column_dimensions[letter].outlineLevel = 1
            sheet.column_dimensions[letter].hidden = current_group in DEFAULT_COLLAPSED_GROUPS

    sheet.sheet_properties.outlinePr.summaryRight = True
    return [field for _, field in columns]



def _highlight_approximate_coupon(sheet: Any, prepared_rows: list[dict[str, Any]], excel_columns: list[str]) -> None:
    try:
        secid_col = excel_columns.index("SECID") + 1
        coupon_col = excel_columns.index("COUPONPERCENT") + 1
    except ValueError:
        return

    approx_secids = {str(row.get("SECID") or "") for row in prepared_rows if bool(row.get("_COUPONPERCENT_APPROX"))}
    if not approx_secids:
        return

    for row_idx in range(3, sheet.max_row + 1):
        secid = str(sheet.cell(row=row_idx, column=secid_col).value or "")
        if secid in approx_secids:
            sheet.cell(row=row_idx, column=coupon_col).fill = APPROX_FILL


def save_bonds_csv(path: str, bonds: list[dict[str, Any]]) -> None:
    """Сохраняет список облигаций в CSV (UTF-8 BOM для корректного открытия в Excel)."""
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    fields, prepared_rows = _prepare_export_data(bonds)

    with target.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(
            [
                {field: _format_value(field, row.get(field, "")) for field in fields}
                for row in prepared_rows
            ]
        )


def save_bonds_file(path: str, bonds: list[dict[str, Any]], summary: dict[str, Any] | None = None) -> None:
    """Сохраняет результат в формате по расширению файла.

    По умолчанию поддерживаются `.xlsx` и `.csv`.
    """
    extension = Path(path).suffix.lower()
    if extension == ".csv":
        save_bonds_csv(path, bonds)
        return

    if extension == ".xlsx":
        save_bonds_excel(path, bonds, summary=summary)
        return

    raise ValueError("Поддерживаются только форматы .xlsx и .csv")


def save_emitents_excel(path: str, emitents: list[dict[str, str]]) -> None:
    """Сохраняет справочник эмитентов в отдельный Excel-файл."""
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EMITENTS"

    fields = [
        "Полное наименование",
        "ИНН",
        "Тикеры акций",
        "ISIN облигаций",
        "missing_full_name",
        "missing_inn",
        "Флаг качества",
    ]
    sheet.append(fields)
    for row in emitents:
        sheet.append([row.get(field, "") for field in fields])

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(fields) + 1):
        column_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(sheet.cell(row=row_idx, column=col_idx).value or "")) for row_idx in range(1, sheet.max_row + 1)),
            default=0,
        )
        sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 16), 60)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{sheet.max_row}"
    workbook.save(target)
