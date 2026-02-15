from __future__ import annotations

import argparse
import csv
import io
import json
import logging
import os
import random
import sqlite3
import time
import re
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

MOEX_RATES_URL = (
    "https://iss.moex.com/iss/apps/infogrid/stock/rates.csv?"
    "sec_type=stock_ofz_bond,stock_cb_bond,stock_subfederal_bond,"
    "stock_municipal_bond,stock_corporate_bond,stock_exchange_bond&"
    "iss.dp=comma&iss.df=%25d.%25m.%25Y&iss.tf=%25H:%25M:%25S&"
    "iss.dtf=%25d.%25m.%25Y%20%25H:%25M:%25S&iss.only=rates&"
    "limit=unlimited&lang=ru"
)
MOEX_INTRADAY_QUOTES_URL = (
    "https://iss.moex.com/iss/engines/stock/markets/bonds/securities.json?"
    "iss.meta=off&iss.only=marketdata&marketdata.columns=SECID,BOARDID,TRADINGSTATUS,OPEN,CLOSE,LCLOSE,LAST,NUMTRADES,VOLVALUE,UPDATETIME&"
    "limit=100"
)

BASE_DIR = Path(__file__).resolve().parent
DB_DIR = BASE_DIR / "DB"
LOGS_DIR = BASE_DIR / "logs"
CACHE_DB_PATH = DB_DIR / "moex_cache.sqlite3"
RAW_RESPONSE_PATH = LOGS_DIR / "raw_response_latest.csv"
LOG_PATH = LOGS_DIR / "moex_api.log"
EXCEL_PATH = BASE_DIR / "Moex_Bonds.xlsx"
DETAILS_EXCEL_PATH = BASE_DIR / "Moex_Bonds_Details.xlsx"
FINISH_EXCEL_PATH = BASE_DIR / "Moex_Bonds_Finish.xlsx"
FINISH_PRICE_EXCEL_PATH = BASE_DIR / "MOEX_Bonds_Finish_Price.xlsx"
FINISH_BATCH_DIR = BASE_DIR / "finish_batches"
DETAILS_PARQUET_DIR = BASE_DIR / "details_parquet"
DETAILS_TTL_HOURS = 24 * 7
DEFAULT_DETAILS_WORKER_PROCESSES = min(16, max(4, (os.cpu_count() or 2) * 2))
RANDOM_SECID_SAMPLE_SIZE = 10
DETAILS_SAMPLE_ONLY = os.getenv("DETAILS_SAMPLE_ONLY", "0") == "1"
CB_FAILURE_THRESHOLD = 3
CB_COOLDOWN_SECONDS = 180
HEALTH_RETENTION_DAYS = 14
ROW_COUNT_SPIKE_THRESHOLD = 0.30
DISCOVERY_MAX_ATTEMPTS = 4
DISCOVERY_BACKOFF_BASE_SECONDS = 0.7
DISCOVERY_BACKOFF_MAX_SECONDS = 8.0
INTRADAY_SNAPSHOT_INTERVAL_MINUTES = 10

FINISH_EXCEL_EXCLUDED_COLUMNS = {
    "NAME",
    "TYPENAME",
    "REGNUMBER",
    "LISTLEVEL",
    "IS_COLLATERAL",
    "IS_EXTERNAL",
    "PRIMARY_BOARDID",
    "PRIMARY_BOARD_TITLE",
    "IS_RII",
    "INCLUDEDBYMOEX",
    "SUSPENSION_LISTING",
    "EVENINGSESSION",
    "MORNINGSESSION",
    "WEEKENDSESSION",
    "REGISTRYCLOSEDATE",
    "DIVIDENDVALUE",
    "DIVIDENDYIELD",
    "REGISTRYCLOSETYPE",
    "LOTSIZE",
    "RTL1",
    "RTH1",
    "RTL2",
    "RTH2",
    "RTL3",
    "RTH3",
    "DISCOUNT1",
    "LIMIT1",
    "DISCOUNT2",
    "LIMIT2",
    "DISCOUNT3",
    "DISCOUNTL0",
    "FULLCOVERED",
    "FULL_COVERED_LIMIT",
}

PIPELINE_STAGE_ORDER = ["fetch", "parse", "details", "export"]
PIPELINE_STAGE_TITLES = {
    "fetch": "Загрузка дневного CSV",
    "parse": "Парсинг и baseline DQ",
    "details": "Обогащение деталей и материализация",
    "export": "Финальные отчёты и intraday snapshot",
}

STATIC_SECIDS = [
    "RU000A1021G3",
    "RU000A107GM1",
    "RU000A109BW7",
    "RU000A10B4W8",
    "RU000A10BFQ1",
    "RU000A10BRN3",
    "RU000A10CPQ8",
    "RU000A10D616",
    "RU000A10DKS3",
    "SU26212RMFS9",
]

DETAILS_ENDPOINTS = [
    (
        "security_overview",
        "https://iss.moex.com/iss/securities/{secid}.json?iss.meta=off",
    ),
    (
        "bondization",
        "https://iss.moex.com/iss/securities/{secid}/bondization.json?iss.meta=off",
    ),
    (
        "bondization_stats",
        "https://iss.moex.com/iss/statistics/engines/stock/markets/bonds/bondization/{secid}.json?iss.meta=off",
    ),
    (
        "security_events",
        "https://iss.moex.com/iss/securities/{secid}/events.json?iss.meta=off",
    ),
]


class CacheMissError(RuntimeError):
    """Raised when cache does not contain data for requested day."""


def build_retry_session() -> requests.Session:
    retry = Retry(
        total=5,
        connect=5,
        read=5,
        status=5,
        backoff_factor=0.8,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=20, pool_maxsize=20)
    session = requests.Session()
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update({"User-Agent": "Vibe-MOEX-Client/2.0"})
    return session


def setup_logging() -> logging.Logger:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)

    logger = logging.getLogger("moex_api")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()

    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    file_handler = logging.FileHandler(LOG_PATH, mode="w", encoding="utf-8")
    file_handler.setFormatter(formatter)

    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.WARNING)
    stream_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(stream_handler)

    return logger


def init_db() -> None:
    DB_DIR.mkdir(parents=True, exist_ok=True)

    with sqlite3.connect(CACHE_DB_PATH) as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS bonds_cache (
                fetch_date TEXT PRIMARY KEY,
                csv_data TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS details_cache (
                endpoint TEXT NOT NULL,
                secid TEXT NOT NULL,
                response_json TEXT NOT NULL,
                fetched_at TEXT NOT NULL,
                PRIMARY KEY(endpoint, secid)
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS details_rows (
                fetched_at TEXT NOT NULL,
                endpoint TEXT NOT NULL,
                secid TEXT NOT NULL,
                block_name TEXT NOT NULL,
                row_json TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS details_update_watermark (
                endpoint TEXT NOT NULL,
                secid TEXT NOT NULL,
                fetched_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY(endpoint, secid, fetched_at)
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS endpoint_health_history (
                checked_at TEXT NOT NULL,
                endpoint TEXT NOT NULL,
                secid TEXT NOT NULL,
                status TEXT NOT NULL,
                source TEXT NOT NULL,
                http_status INTEGER,
                latency_ms REAL,
                blocks TEXT,
                error_text TEXT
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS endpoint_circuit_breaker (
                endpoint TEXT PRIMARY KEY,
                failure_count INTEGER NOT NULL DEFAULT 0,
                state TEXT NOT NULL DEFAULT 'closed',
                opened_at TEXT,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS endpoint_health_mv (
                window TEXT NOT NULL,
                endpoint TEXT NOT NULL,
                total_requests INTEGER NOT NULL,
                error_requests INTEGER NOT NULL,
                error_rate REAL NOT NULL,
                avg_latency_ms REAL,
                p95_latency_ms REAL,
                updated_at TEXT NOT NULL,
                PRIMARY KEY(window, endpoint)
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS dq_run_history (
                run_id TEXT PRIMARY KEY,
                run_at TEXT NOT NULL,
                source TEXT NOT NULL,
                row_count INTEGER NOT NULL,
                empty_secid_ratio REAL NOT NULL,
                empty_isin_ratio REAL NOT NULL,
                row_count_delta_ratio REAL,
                notes TEXT
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS bonds_enriched_incremental (
                batch_id TEXT NOT NULL,
                export_date TEXT NOT NULL,
                exported_at TEXT NOT NULL,
                source TEXT NOT NULL,
                row_json TEXT NOT NULL
            )
            """
        )
        connection.execute(
            "CREATE INDEX IF NOT EXISTS idx_endpoint_health_history_checked_endpoint ON endpoint_health_history(checked_at, endpoint)"
        )
        connection.execute(
            "CREATE INDEX IF NOT EXISTS idx_details_rows_endpoint_secid_block ON details_rows(endpoint, secid, block_name)"
        )
        connection.execute(
            "CREATE INDEX IF NOT EXISTS idx_dq_run_history_run_at ON dq_run_history(run_at)"
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS dq_metrics_daily_mv (
                run_day TEXT PRIMARY KEY,
                runs_count INTEGER NOT NULL,
                avg_row_count REAL NOT NULL,
                max_row_count INTEGER NOT NULL,
                min_row_count INTEGER NOT NULL,
                avg_empty_secid_ratio REAL NOT NULL,
                avg_empty_isin_ratio REAL NOT NULL,
                max_row_count_delta_ratio REAL,
                warning_runs_count INTEGER NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        connection.execute(
            "CREATE INDEX IF NOT EXISTS idx_dq_metrics_daily_mv_run_day ON dq_metrics_daily_mv(run_day)"
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS etl_stage_sla (
                run_id TEXT NOT NULL,
                stage TEXT NOT NULL,
                started_at TEXT NOT NULL,
                finished_at TEXT NOT NULL,
                duration_ms REAL NOT NULL,
                status TEXT NOT NULL,
                source TEXT,
                details TEXT,
                PRIMARY KEY(run_id, stage)
            )
            """
        )
        connection.execute(
            "CREATE INDEX IF NOT EXISTS idx_etl_stage_sla_stage_finished ON etl_stage_sla(stage, finished_at)"
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS intraday_quotes_snapshot (
                snapshot_at TEXT NOT NULL,
                secid TEXT NOT NULL,
                boardid TEXT,
                tradingstatus TEXT,
                open REAL,
                close REAL,
                lclose REAL,
                last REAL,
                numtrades REAL,
                volvalue REAL,
                updatetime TEXT,
                PRIMARY KEY(snapshot_at, secid, boardid)
            )
            """
        )
        connection.execute(
            "CREATE INDEX IF NOT EXISTS idx_intraday_quotes_snapshot_snapshot_at ON intraday_quotes_snapshot(snapshot_at)"
        )
        for column_name in ["open", "close", "lclose"]:
            try:
                connection.execute(f"ALTER TABLE intraday_quotes_snapshot ADD COLUMN {column_name} REAL")
            except sqlite3.OperationalError:
                pass
        connection.execute(
            "CREATE INDEX IF NOT EXISTS idx_bonds_enriched_incremental_export_date_batch ON bonds_enriched_incremental(export_date, batch_id)"
        )
        connection.commit()


def cleanup_details_cache() -> None:
    cutoff = (datetime.now() - timedelta(hours=DETAILS_TTL_HOURS)).isoformat(timespec="seconds")
    health_cutoff = (datetime.now() - timedelta(days=HEALTH_RETENTION_DAYS)).isoformat(timespec="seconds")
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        connection.execute("DELETE FROM details_cache WHERE fetched_at < ?", (cutoff,))
        connection.execute("DELETE FROM details_rows WHERE fetched_at < ?", (cutoff,))
        connection.execute("DELETE FROM endpoint_health_history WHERE checked_at < ?", (health_cutoff,))
        connection.commit()


def get_cached_data(target_date: str) -> str:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        row = connection.execute(
            "SELECT csv_data FROM bonds_cache WHERE fetch_date = ?",
            (target_date,),
        ).fetchone()

    if row is None:
        raise CacheMissError(f"No cache for {target_date}")

    return row[0]


def get_latest_cached_data() -> tuple[str, str]:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        row = connection.execute(
            """
            SELECT fetch_date, csv_data
            FROM bonds_cache
            ORDER BY fetch_date DESC
            LIMIT 1
            """
        ).fetchone()

    if row is None:
        raise CacheMissError("No cached data found")

    return row[0], row[1]


def save_to_cache(target_date: str, csv_data: str) -> None:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        connection.execute(
            """
            INSERT OR REPLACE INTO bonds_cache(fetch_date, csv_data, created_at)
            VALUES (?, ?, ?)
            """,
            (target_date, csv_data, datetime.now().isoformat(timespec="seconds")),
        )
        connection.commit()


def fetch_moex_csv(session: requests.Session) -> str:
    response = session.get(MOEX_RATES_URL, timeout=60)
    response.raise_for_status()
    return response.text


def persist_raw_response(csv_data: str) -> None:
    LOGS_DIR.mkdir(parents=True, exist_ok=True)
    RAW_RESPONSE_PATH.write_text(csv_data, encoding="utf-8")


def parse_rates_csv(csv_data: str) -> pd.DataFrame:
    prepared_csv, delimiter = _prepare_csv_for_pandas(csv_data)
    return pd.read_csv(io.StringIO(prepared_csv), sep=delimiter)


def _extract_blocks(payload: dict) -> dict[str, pd.DataFrame]:
    blocks: dict[str, pd.DataFrame] = {}
    for block_name, block_payload in payload.items():
        if not isinstance(block_payload, dict):
            continue
        columns = block_payload.get("columns")
        data = block_payload.get("data")
        if not isinstance(columns, list) or not isinstance(data, list):
            continue
        blocks[block_name] = pd.DataFrame(data, columns=columns)
    return blocks


def _pick_secids(dataframe: pd.DataFrame, random_sample_size: int = RANDOM_SECID_SAMPLE_SIZE) -> list[str]:
    for column in ["SECID", "secid", "SecID"]:
        if column in dataframe.columns:
            series = dataframe[column].dropna().astype(str).str.strip()
            unique_secids = sorted(set(series[series != ""]))
            if not unique_secids:
                return STATIC_SECIDS.copy()
            if DETAILS_SAMPLE_ONLY:
                rng = random.Random(date.today().isoformat())
                random_part = unique_secids if len(unique_secids) <= random_sample_size else rng.sample(unique_secids, random_sample_size)
                return sorted(set(random_part + STATIC_SECIDS))
            return unique_secids
    return STATIC_SECIDS.copy()


def _fetch_endpoint_payload(
    session: requests.Session,
    endpoint_name: str,
    secid: str,
    endpoint_url: str,
    timeout_seconds: int = 30,
) -> tuple[dict, float, int | None]:
    started = time.perf_counter()
    response = session.get(endpoint_url.format(secid=secid), timeout=timeout_seconds)
    latency_ms = (time.perf_counter() - started) * 1000
    status_code = response.status_code
    response.raise_for_status()
    payload = response.json()
    if not isinstance(payload, dict):
        raise ValueError(f"Unexpected JSON shape for {endpoint_name}:{secid}")
    return payload, latency_ms, status_code


def _fetch_endpoint_payload_with_discovery_backoff(
    session: requests.Session,
    endpoint_name: str,
    secid: str,
    endpoint_url: str,
    logger: logging.Logger,
) -> tuple[dict, float, int | None]:
    last_error: Exception | None = None
    for attempt in range(1, DISCOVERY_MAX_ATTEMPTS + 1):
        try:
            return _fetch_endpoint_payload(session, endpoint_name, secid, endpoint_url)
        except Exception as error:  # noqa: BLE001
            last_error = error
            if attempt >= DISCOVERY_MAX_ATTEMPTS:
                break
            backoff = min(DISCOVERY_BACKOFF_BASE_SECONDS * (2 ** (attempt - 1)), DISCOVERY_BACKOFF_MAX_SECONDS)
            jitter = random.uniform(0, backoff * 0.35)
            sleep_for = backoff + jitter
            logger.warning(
                "Discovery call %s for %s failed on attempt %s/%s (%s). Retry in %.2fs",
                endpoint_name,
                secid,
                attempt,
                DISCOVERY_MAX_ATTEMPTS,
                error,
                sleep_for,
            )
            time.sleep(sleep_for)

    assert last_error is not None
    raise last_error


def _precheck_details_endpoints_health(session: requests.Session, secid: str, logger: logging.Logger) -> list[tuple[str, str]]:
    healthy: list[tuple[str, str]] = []
    for endpoint_name, endpoint_url in DETAILS_ENDPOINTS:
        if _is_circuit_open(endpoint_name):
            logger.warning("Endpoint %s precheck skipped: circuit breaker is open", endpoint_name)
            continue
        probe_url = endpoint_url.format(secid=secid)
        started = time.perf_counter()
        try:
            response = session.get(probe_url, timeout=6)
            latency_ms = (time.perf_counter() - started) * 1000
            if response.status_code >= 500:
                _save_endpoint_health(endpoint_name, secid, "error", "precheck", response.status_code, latency_ms, None, "5xx during precheck")
                logger.warning("Endpoint %s precheck failed with HTTP %s", endpoint_name, response.status_code)
                continue
            _save_endpoint_health(endpoint_name, secid, "ok", "precheck", response.status_code, latency_ms, None, None)
            healthy.append((endpoint_name, endpoint_url))
        except requests.RequestException as error:
            _save_endpoint_health(endpoint_name, secid, "error", "precheck", None, None, None, str(error))
            logger.warning("Endpoint %s precheck request failed: %s", endpoint_name, error)
    return healthy


def _save_endpoint_health(endpoint_name: str, secid: str, status: str, source: str, http_status: int | None, latency_ms: float | None, blocks: list[str] | None, error_text: str | None) -> None:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        connection.execute(
            """
            INSERT INTO endpoint_health_history(
                checked_at, endpoint, secid, status, source, http_status, latency_ms, blocks, error_text
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                datetime.now().isoformat(timespec="seconds"),
                endpoint_name,
                secid,
                status,
                source,
                http_status,
                latency_ms,
                ", ".join(blocks) if blocks else None,
                error_text,
            ),
        )
        connection.commit()




def _is_circuit_open(endpoint_name: str) -> bool:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        row = connection.execute(
            "SELECT state, opened_at FROM endpoint_circuit_breaker WHERE endpoint = ?",
            (endpoint_name,),
        ).fetchone()

    if row is None:
        return False

    state, opened_at = row
    if state != "open" or not opened_at:
        return False

    opened_at_dt = datetime.fromisoformat(opened_at)
    if (datetime.now() - opened_at_dt).total_seconds() >= CB_COOLDOWN_SECONDS:
        with sqlite3.connect(CACHE_DB_PATH) as connection:
            connection.execute(
                """
                INSERT OR REPLACE INTO endpoint_circuit_breaker(endpoint, failure_count, state, opened_at, updated_at)
                VALUES (?, 0, 'half_open', NULL, ?)
                """,
                (endpoint_name, datetime.now().isoformat(timespec="seconds")),
            )
            connection.commit()
        return False

    return True


def _update_circuit_breaker(endpoint_name: str, is_success: bool) -> None:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        row = connection.execute(
            "SELECT failure_count, state FROM endpoint_circuit_breaker WHERE endpoint = ?",
            (endpoint_name,),
        ).fetchone()

        failure_count = 0 if row is None else row[0]
        state = "closed" if row is None else row[1]

        now = datetime.now().isoformat(timespec="seconds")
        if is_success:
            connection.execute(
                """
                INSERT OR REPLACE INTO endpoint_circuit_breaker(endpoint, failure_count, state, opened_at, updated_at)
                VALUES (?, 0, 'closed', NULL, ?)
                """,
                (endpoint_name, now),
            )
        else:
            failure_count += 1
            new_state = "open" if failure_count >= CB_FAILURE_THRESHOLD else state
            opened_at = now if new_state == "open" else None
            connection.execute(
                """
                INSERT OR REPLACE INTO endpoint_circuit_breaker(endpoint, failure_count, state, opened_at, updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (endpoint_name, failure_count, new_state, opened_at, now),
            )
        connection.commit()


def refresh_endpoint_health_mv() -> None:
    now = datetime.now().isoformat(timespec="seconds")
    windows = {"1d": datetime.now() - timedelta(days=1), "7d": datetime.now() - timedelta(days=7)}

    with sqlite3.connect(CACHE_DB_PATH) as connection:
        for window_name, cutoff in windows.items():
            df = pd.read_sql_query(
                """
                SELECT endpoint, status, latency_ms
                FROM endpoint_health_history
                WHERE checked_at >= ?
                """,
                connection,
                params=(cutoff.isoformat(timespec="seconds"),),
            )
            if df.empty:
                continue

            rows = []
            for endpoint, group in df.groupby("endpoint"):
                total = len(group)
                errors = int((group["status"] != "ok").sum())
                latencies = group["latency_ms"].dropna()
                rows.append(
                    (
                        window_name,
                        endpoint,
                        total,
                        errors,
                        errors / total if total else 0.0,
                        float(latencies.mean()) if not latencies.empty else None,
                        float(latencies.quantile(0.95)) if not latencies.empty else None,
                        now,
                    )
                )

            connection.executemany(
                """
                INSERT OR REPLACE INTO endpoint_health_mv(
                    window, endpoint, total_requests, error_requests, error_rate, avg_latency_ms, p95_latency_ms, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                rows,
            )
        connection.commit()

def _get_cached_endpoint_record(endpoint_name: str, secid: str) -> tuple[dict, str] | None:
    cutoff = (datetime.now() - timedelta(hours=DETAILS_TTL_HOURS)).isoformat(timespec="seconds")
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        row = connection.execute(
            """
            SELECT response_json, fetched_at
            FROM details_cache
            WHERE endpoint = ? AND secid = ? AND fetched_at >= ?
            """,
            (endpoint_name, secid, cutoff),
        ).fetchone()

    if row is None:
        return None
    return json.loads(row[0]), row[1]


def _get_latest_endpoint_record(endpoint_name: str, secid: str) -> tuple[dict, str] | None:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        row = connection.execute(
            """
            SELECT response_json, fetched_at
            FROM details_cache
            WHERE endpoint = ? AND secid = ?
            ORDER BY fetched_at DESC
            LIMIT 1
            """,
            (endpoint_name, secid),
        ).fetchone()

    if row is None:
        return None
    return json.loads(row[0]), row[1]


def _watermark_exists(endpoint_name: str, secid: str, fetched_at: str) -> bool:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        row = connection.execute(
            """
            SELECT 1
            FROM details_update_watermark
            WHERE endpoint = ? AND secid = ? AND fetched_at = ?
            LIMIT 1
            """,
            (endpoint_name, secid, fetched_at),
        ).fetchone()
    return row is not None


def _save_watermark(endpoint_name: str, secid: str, fetched_at: str) -> None:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        connection.execute(
            """
            INSERT OR IGNORE INTO details_update_watermark(endpoint, secid, fetched_at, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            (endpoint_name, secid, fetched_at, datetime.now().isoformat(timespec="seconds")),
        )
        connection.commit()


def _save_endpoint_payload(endpoint_name: str, secid: str, payload: dict, fetched_at: str | None = None) -> str:
    now = fetched_at or datetime.now().isoformat(timespec="seconds")
    serialized_payload = json.dumps(payload, ensure_ascii=False)
    blocks = _extract_blocks(payload)

    rows_to_insert: list[tuple[str, str, str, str, str]] = []
    for block_name, block_df in blocks.items():
        for row in block_df.to_dict(orient="records"):
            rows_to_insert.append(
                (
                    now,
                    endpoint_name,
                    secid,
                    block_name,
                    json.dumps(row, ensure_ascii=False),
                )
            )

    with sqlite3.connect(CACHE_DB_PATH) as connection:
        connection.execute(
            """
            INSERT OR REPLACE INTO details_cache(endpoint, secid, response_json, fetched_at)
            VALUES (?, ?, ?, ?)
            """,
            (endpoint_name, secid, serialized_payload, now),
        )
        connection.execute(
            "DELETE FROM details_rows WHERE endpoint = ? AND secid = ?",
            (endpoint_name, secid),
        )
        if rows_to_insert:
            connection.executemany(
                """
                INSERT INTO details_rows(fetched_at, endpoint, secid, block_name, row_json)
                VALUES (?, ?, ?, ?, ?)
                """,
                rows_to_insert,
            )
        connection.commit()
    return now


def _discover_working_endpoints(
    session: requests.Session,
    sample_secid: str,
    logger: logging.Logger,
    candidates: list[tuple[str, str]] | None = None,
) -> list[tuple[str, str]]:
    working_endpoints: list[tuple[str, str]] = []
    for endpoint_name, endpoint_url in (candidates or DETAILS_ENDPOINTS):
        if _is_circuit_open(endpoint_name):
            logger.warning("Endpoint %s skipped during discovery: circuit breaker is open", endpoint_name)
            continue
        try:
            payload, latency_ms, status_code = _fetch_endpoint_payload_with_discovery_backoff(
                session,
                endpoint_name,
                sample_secid,
                endpoint_url,
                logger,
            )
            blocks = _extract_blocks(payload)
            if blocks:
                block_names = list(blocks.keys())
                logger.info("Endpoint %s is available (blocks: %s)", endpoint_name, ", ".join(block_names))
                _save_endpoint_health(endpoint_name, sample_secid, "ok", "network", status_code, latency_ms, block_names, None)
                working_endpoints.append((endpoint_name, endpoint_url))
            else:
                logger.warning("Endpoint %s returned no tabular blocks for %s", endpoint_name, sample_secid)
                _save_endpoint_health(endpoint_name, sample_secid, "empty", "network", status_code, latency_ms, [], None)
        except requests.RequestException as error:
            response = getattr(error, "response", None)
            status_code = response.status_code if response is not None else None
            logger.warning("Endpoint %s is unavailable for %s: %s", endpoint_name, sample_secid, error)
            _save_endpoint_health(endpoint_name, sample_secid, "error", "network", status_code, None, None, str(error))
        except Exception as error:  # noqa: BLE001
            logger.warning("Endpoint %s is unavailable for %s: %s", endpoint_name, sample_secid, error)
            _save_endpoint_health(endpoint_name, sample_secid, "error", "network", None, None, None, str(error))
    return working_endpoints


def _normalize_block_frame(secid: str, block_name: str, block_df: pd.DataFrame) -> pd.DataFrame:
    enriched = block_df.copy()
    enriched = _expand_description_pairs(enriched)
    enriched = enriched.drop(columns=["secid", "block_name"], errors="ignore")
    enriched.insert(0, "secid", secid)
    enriched.insert(1, "block_name", block_name)
    non_meta_cols = [column for column in enriched.columns if column not in {"secid", "block_name"}]
    return enriched[["secid", "block_name", *non_meta_cols]]


def _slugify_column(value: str) -> str:
    lowered = value.strip().lower()
    normalized = re.sub(r"[^a-zа-я0-9]+", "_", lowered, flags=re.IGNORECASE)
    return normalized.strip("_")[:60] or "field"


def _expand_description_pairs(frame: pd.DataFrame) -> pd.DataFrame:
    title_col = next((c for c in frame.columns if c.endswith("_description_title")), None)
    value_col = next((c for c in frame.columns if c.endswith("_description_value")), None)
    if title_col is None or value_col is None or frame.empty:
        return frame

    expanded: list[dict[str, str | float | int | None]] = []
    for _, row in frame.iterrows():
        row_dict = row.to_dict()
        titles = [part.strip() for part in str(row_dict.get(title_col, "")).split("|") if part.strip()]
        values = [part.strip() for part in str(row_dict.get(value_col, "")).split("|")]
        for idx, title in enumerate(titles):
            key = f"overview_{_slugify_column(title)}"
            row_dict[key] = values[idx].strip() if idx < len(values) else ""
        expanded.append(row_dict)

    result = pd.DataFrame(expanded)
    return result.loc[:, ~result.columns.duplicated()].copy()


def _fetch_details_worker(secid: str, endpoints: list[tuple[str, str]]) -> list[tuple[str, str, dict | None, float | None, int | None, str | None]]:
    session = build_retry_session()
    results: list[tuple[str, str, dict | None, float | None, int | None, str | None]] = []
    for endpoint_name, endpoint_url in endpoints:
        attempt_timeouts = [15, 30, 45]
        last_error: Exception | None = None
        for timeout_seconds in attempt_timeouts:
            try:
                payload, latency_ms, status_code = _fetch_endpoint_payload(
                    session,
                    endpoint_name,
                    secid,
                    endpoint_url,
                    timeout_seconds=timeout_seconds,
                )
                results.append((endpoint_name, secid, payload, latency_ms, status_code, None))
                last_error = None
                break
            except Exception as error:  # noqa: BLE001
                last_error = error

        if last_error is not None:
            status_code = getattr(getattr(last_error, "response", None), "status_code", None)
            results.append((endpoint_name, secid, None, None, status_code, str(last_error)))
    return results


def _build_sheet_name(endpoint_name: str, block_name: str) -> str:
    base_name = f"{endpoint_name}_{block_name}".replace("/", "_")
    return base_name[:31]


def _autosize_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells[:200]:
            value_length = len(str(cell.value)) if cell.value is not None else 0
            if value_length > max_length:
                max_length = value_length
        worksheet.column_dimensions[column].width = min(max(12, max_length + 2), 60)


def _build_enrichment_frame(endpoint_frames: dict[str, dict[str, pd.DataFrame]]) -> pd.DataFrame:
    secid_features: dict[str, dict[str, str | int | float]] = {}
    for endpoint_name, block_frames in endpoint_frames.items():
        for block_name, frame in block_frames.items():
            if frame.empty:
                continue
            for secid, group in frame.groupby("secid"):
                secid_features.setdefault(secid, {})
                secid_features[secid][f"{endpoint_name}_{block_name}_rows"] = len(group)
                for col in group.columns:
                    if col in {"secid", "block_name"}:
                        continue
                    values = [str(v) for v in group[col].dropna().tolist() if str(v).strip()]
                    if not values:
                        continue
                    unique_values = list(dict.fromkeys(values))
                    secid_features[secid][f"{endpoint_name}_{block_name}_{col}"] = " | ".join(unique_values[:5])

    if not secid_features:
        return pd.DataFrame(columns=["secid"])

    rows = [{"secid": secid, **features} for secid, features in secid_features.items()]
    enrichment = pd.DataFrame(rows)
    enrichment = enrichment.loc[:, ~enrichment.columns.duplicated()].copy()
    return enrichment


def _write_details_excel(endpoint_frames: dict[str, dict[str, pd.DataFrame]], enrichment_df: pd.DataFrame) -> int:
    mode = "a" if DETAILS_EXCEL_PATH.exists() else "w"
    writer_kwargs = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"

    sheet_count = 0
    with pd.ExcelWriter(DETAILS_EXCEL_PATH, **writer_kwargs) as writer:
        summary_rows: list[dict[str, str | int]] = []
        for endpoint_name, block_frames in endpoint_frames.items():
            for block_name, frame in block_frames.items():
                clean_frame = frame.dropna(axis=1, how="all")
                if clean_frame.empty:
                    continue
                sheet_name = _build_sheet_name(endpoint_name, block_name)
                clean_frame.to_excel(writer, sheet_name=sheet_name, index=False)
                summary_rows.append({"endpoint": endpoint_name, "block_name": block_name, "rows": len(clean_frame), "sheet": sheet_name})
                sheet_count += 1

        enrichment_df.to_excel(writer, sheet_name="details_transposed", index=False)
        summary_df = pd.DataFrame(summary_rows).sort_values(["endpoint", "block_name"]) if summary_rows else pd.DataFrame(columns=["endpoint", "block_name", "rows", "sheet"])
        summary_df.to_excel(writer, sheet_name="summary", index=False)

        for worksheet in writer.book.worksheets:
            _autosize_worksheet(worksheet)

    return sheet_count + 1


def _update_details_parquet(endpoint_frames: dict[str, dict[str, pd.DataFrame]]) -> int:
    DETAILS_PARQUET_DIR.mkdir(parents=True, exist_ok=True)
    updated_files = 0

    for endpoint_name, block_frames in endpoint_frames.items():
        flat_frames = []
        for block_name, frame in block_frames.items():
            if frame.empty:
                continue
            prepared = frame.copy()
            prepared["block_name"] = block_name
            hash_series = pd.util.hash_pandas_object(prepared.fillna(""), index=False).astype(str)
            prepared["_row_hash"] = hash_series
            prepared["_updated_at"] = datetime.now().isoformat(timespec="seconds")
            flat_frames.append(prepared)

        if not flat_frames:
            continue

        new_data = pd.concat(flat_frames, ignore_index=True, sort=False)
        parquet_path = DETAILS_PARQUET_DIR / f"{endpoint_name}.parquet"
        if parquet_path.exists():
            existing = pd.read_parquet(parquet_path)
            merged = pd.concat([existing, new_data], ignore_index=True, sort=False)
            merged = merged.drop_duplicates(subset=["secid", "block_name", "_row_hash"], keep="last")
        else:
            merged = new_data

        merged.to_parquet(parquet_path, index=False)
        updated_files += 1

    return updated_files


def _merge_base_with_enrichment(base_df: pd.DataFrame, enrichment_df: pd.DataFrame) -> pd.DataFrame:
    secid_column = next((c for c in ["SECID", "secid", "SecID"] if c in base_df.columns), None)
    if secid_column is None:
        return base_df.copy()

    prepared = base_df.copy()
    prepared["secid"] = prepared[secid_column].astype(str).str.strip().str.upper()
    normalized_enrichment = enrichment_df.copy()
    if "secid" in normalized_enrichment.columns:
        normalized_enrichment["secid"] = normalized_enrichment["secid"].astype(str).str.strip().str.upper()

    merged = prepared.merge(normalized_enrichment, on="secid", how="left")
    merged = merged.drop(columns=["secid"]) if secid_column != "secid" else merged
    merged = merged.loc[:, ~merged.columns.duplicated()].copy()
    return merged


def _normalize_identity_columns(dataframe: pd.DataFrame) -> pd.DataFrame:
    prepared = dataframe.copy()
    secid_col = next((c for c in ["SECID", "secid", "SecID"] if c in prepared.columns), None)
    isin_col = next((c for c in ["ISIN", "isin"] if c in prepared.columns), None)

    if secid_col is None and isin_col is not None:
        prepared["SECID"] = prepared[isin_col]
        secid_col = "SECID"

    if secid_col is not None:
        prepared[secid_col] = prepared[secid_col].astype(str).str.strip()
        if isin_col is not None:
            isin_series = prepared[isin_col].astype(str).str.strip()
            prepared.loc[prepared[secid_col] == "", secid_col] = isin_series

    drop_candidates = [c for c in ["ISIN", "isin"] if c in prepared.columns]
    if secid_col not in drop_candidates:
        prepared = prepared.drop(columns=drop_candidates, errors="ignore")

    prepared = prepared.loc[:, ~prepared.columns.duplicated()].copy()
    return prepared


def _build_finish_excel_view(full_df: pd.DataFrame) -> pd.DataFrame:
    view_df = full_df.drop(columns=[c for c in full_df.columns if c in FINISH_EXCEL_EXCLUDED_COLUMNS], errors="ignore").copy()
    preferred_order = [c for c in ["SECID", "SHORTNAME", "PRICE", "WAPRICE", "YIELDATWAP", "MATDATE", "EMITENTNAME"] if c in view_df.columns]
    tail = [c for c in view_df.columns if c not in preferred_order]
    view_df = view_df[preferred_order + tail]
    return view_df.loc[:, ~view_df.columns.duplicated()].copy()


def _build_price_snapshot_export(base_df: pd.DataFrame, now_snapshot_at: str) -> pd.DataFrame:
    today = date.today().isoformat()
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        now_df = pd.read_sql_query(
            """
            SELECT secid, boardid, last, updatetime, open, close
            FROM intraday_quotes_snapshot
            WHERE snapshot_at = ?
            """,
            connection,
            params=(now_snapshot_at,),
        )
        latest_df = pd.read_sql_query(
            """
            SELECT q.secid, q.last AS last_known
            FROM intraday_quotes_snapshot q
            JOIN (
                SELECT secid, MAX(snapshot_at) AS max_snapshot
                FROM intraday_quotes_snapshot
                WHERE substr(snapshot_at, 1, 10) = ?
                GROUP BY secid
            ) x ON q.secid = x.secid AND q.snapshot_at = x.max_snapshot
            """,
            connection,
            params=(today,),
        )

    if now_df.empty:
        return pd.DataFrame(columns=["SECID", "CLOSE_TODAY", "OPEN_TODAY", "PRICE_NOW", "LAST_PRICE_TODAY", "DAY_CHANGE_PCT"])

    now_df = now_df.sort_values(["secid", "boardid"]).drop_duplicates("secid", keep="first")
    latest_df = latest_df.sort_values("secid").drop_duplicates("secid", keep="first")

    daily_close = base_df[[c for c in ["SECID", "PRICE"] if c in base_df.columns]].copy()
    if not daily_close.empty:
        daily_close = daily_close.rename(columns={"PRICE": "CLOSE_TODAY"})

    result = now_df.merge(latest_df, on="secid", how="left")
    if not daily_close.empty:
        result = result.merge(daily_close.rename(columns={"SECID": "secid"}), on="secid", how="left")

    result["PRICE_NOW"] = pd.to_numeric(result.get("last"), errors="coerce")
    result["OPEN_TODAY"] = pd.to_numeric(result.get("open"), errors="coerce")
    result["LAST_PRICE_TODAY"] = pd.to_numeric(result.get("last_known"), errors="coerce")
    result["CLOSE_TODAY"] = pd.to_numeric(result.get("CLOSE_TODAY"), errors="coerce")
    result["DAY_CHANGE_PCT"] = ((result["PRICE_NOW"] - result["OPEN_TODAY"]) / result["OPEN_TODAY"]) * 100

    result = result.rename(columns={"secid": "SECID"})
    return result[["SECID", "CLOSE_TODAY", "OPEN_TODAY", "PRICE_NOW", "LAST_PRICE_TODAY", "DAY_CHANGE_PCT"]]


def _export_price_workbook(base_df: pd.DataFrame, now_snapshot_at: str) -> None:
    price_df = _build_price_snapshot_export(base_df, now_snapshot_at).sort_values("DAY_CHANGE_PCT", na_position="last")
    top_fall = price_df.nsmallest(10, "DAY_CHANGE_PCT") if not price_df.empty else price_df
    top_rise = price_df.nlargest(10, "DAY_CHANGE_PCT") if not price_df.empty else price_df

    with pd.ExcelWriter(FINISH_PRICE_EXCEL_PATH, engine="openpyxl") as writer:
        price_df.to_excel(writer, sheet_name="prices", index=False)
        top_fall.to_excel(writer, sheet_name="top_fall", index=False)
        top_rise.to_excel(writer, sheet_name="top_rise", index=False)

        for worksheet in writer.book.worksheets:
            _autosize_worksheet(worksheet)

        from openpyxl.chart import BarChart, Reference

        summary_ws = writer.book.create_sheet("top10_charts")
        summary_ws.append(["Top-10 падение", "Изменение %"])
        for _, row in top_fall.iterrows():
            summary_ws.append([row.get("SECID"), row.get("DAY_CHANGE_PCT")])
        start_rise = len(top_fall) + 4
        summary_ws.cell(row=start_rise, column=1, value="Top-10 рост")
        summary_ws.cell(row=start_rise, column=2, value="Изменение %")
        for idx, (_, row) in enumerate(top_rise.iterrows(), start=start_rise + 1):
            summary_ws.cell(row=idx, column=1, value=row.get("SECID"))
            summary_ws.cell(row=idx, column=2, value=row.get("DAY_CHANGE_PCT"))

        chart_fall = BarChart()
        chart_fall.title = "TOP-10 упавших"
        chart_fall.y_axis.title = "%"
        data = Reference(summary_ws, min_col=2, min_row=1, max_row=len(top_fall) + 1)
        cats = Reference(summary_ws, min_col=1, min_row=2, max_row=len(top_fall) + 1)
        chart_fall.add_data(data, titles_from_data=True)
        chart_fall.set_categories(cats)
        summary_ws.add_chart(chart_fall, "D2")

        chart_rise = BarChart()
        chart_rise.title = "TOP-10 выросших"
        chart_rise.y_axis.title = "%"
        data2 = Reference(summary_ws, min_col=2, min_row=start_rise, max_row=start_rise + len(top_rise))
        cats2 = Reference(summary_ws, min_col=1, min_row=start_rise + 1, max_row=start_rise + len(top_rise))
        chart_rise.add_data(data2, titles_from_data=True)
        chart_rise.set_categories(cats2)
        summary_ws.add_chart(chart_rise, "D22")
        _autosize_worksheet(summary_ws)


def _persist_finish_to_sqlite(finish_df: pd.DataFrame, batch_id: str, export_date: str, source: str) -> None:
    exported_at = datetime.now().isoformat(timespec="seconds")
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        finish_df.to_sql("bonds_enriched", connection, if_exists="replace", index=False)

        payload_rows = [
            (
                batch_id,
                export_date,
                exported_at,
                source,
                json.dumps(row, ensure_ascii=False, default=str),
            )
            for row in finish_df.to_dict(orient="records")
        ]
        connection.execute("DELETE FROM bonds_enriched_incremental WHERE batch_id = ?", (batch_id,))
        connection.executemany(
            """
            INSERT INTO bonds_enriched_incremental(batch_id, export_date, exported_at, source, row_json)
            VALUES (?, ?, ?, ?, ?)
            """,
            payload_rows,
        )
        connection.commit()


def _export_finish_incremental(finish_df: pd.DataFrame, export_date: str, batch_id: str) -> Path:
    FINISH_BATCH_DIR.mkdir(parents=True, exist_ok=True)
    batch_path = FINISH_BATCH_DIR / f"Moex_Bonds_Finish_{export_date}_{batch_id}.xlsx"
    finish_excel_df = _build_finish_excel_view(finish_df)
    with pd.ExcelWriter(batch_path, engine="openpyxl") as writer:
        finish_excel_df.to_excel(writer, sheet_name="finish", index=False)
        _autosize_worksheet(writer.book["finish"])
    with pd.ExcelWriter(FINISH_EXCEL_PATH, engine="openpyxl") as writer:
        finish_excel_df.to_excel(writer, sheet_name="finish", index=False)
        _autosize_worksheet(writer.book["finish"])
    return batch_path


def _refresh_dq_metrics_daily_mv() -> None:
    now = datetime.now().isoformat(timespec="seconds")
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        rows = connection.execute(
            """
            SELECT
                substr(run_at, 1, 10) AS run_day,
                COUNT(*) AS runs_count,
                AVG(row_count) AS avg_row_count,
                MAX(row_count) AS max_row_count,
                MIN(row_count) AS min_row_count,
                AVG(empty_secid_ratio) AS avg_empty_secid_ratio,
                AVG(empty_isin_ratio) AS avg_empty_isin_ratio,
                MAX(row_count_delta_ratio) AS max_row_count_delta_ratio,
                SUM(CASE WHEN COALESCE(notes, '') <> '' THEN 1 ELSE 0 END) AS warning_runs_count
            FROM dq_run_history
            GROUP BY run_day
            """
        ).fetchall()
        connection.executemany(
            """
            INSERT OR REPLACE INTO dq_metrics_daily_mv(
                run_day,
                runs_count,
                avg_row_count,
                max_row_count,
                min_row_count,
                avg_empty_secid_ratio,
                avg_empty_isin_ratio,
                max_row_count_delta_ratio,
                warning_runs_count,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [(*row, now) for row in rows],
        )
        connection.commit()


def _persist_intraday_quotes_snapshot(session: requests.Session, logger: logging.Logger) -> tuple[int, str]:
    snapshot_at = datetime.now().isoformat(timespec="seconds")
    offset = 0
    page_size = 100
    rows_to_save: list[tuple[str, str, str | None, str | None, float | None, float | None, float | None, float | None, float | None, float | None, str | None]] = []

    while True:
        url = f"{MOEX_INTRADAY_QUOTES_URL}&start={offset}"
        try:
            response = session.get(url, timeout=20)
            response.raise_for_status()
            payload = response.json()
        except requests.RequestException as error:
            logger.warning("Intraday snapshot skipped due to MOEX error: %s", error)
            return 0, snapshot_at
        marketdata = payload.get("marketdata", {}) if isinstance(payload, dict) else {}
        columns = marketdata.get("columns", []) if isinstance(marketdata, dict) else []
        data = marketdata.get("data", []) if isinstance(marketdata, dict) else []
        if not columns or not data:
            break

        for item in data:
            row = dict(zip(columns, item))
            rows_to_save.append(
                (
                    snapshot_at,
                    str(row.get("SECID", "")).strip(),
                    row.get("BOARDID"),
                    row.get("TRADINGSTATUS"),
                    row.get("OPEN"),
                    row.get("CLOSE"),
                    row.get("LCLOSE"),
                    row.get("LAST"),
                    row.get("NUMTRADES"),
                    row.get("VOLVALUE"),
                    row.get("UPDATETIME"),
                )
            )

        if len(data) < page_size:
            break
        offset += page_size

    rows_to_save = [row for row in rows_to_save if row[1]]
    if rows_to_save:
        with sqlite3.connect(CACHE_DB_PATH) as connection:
            connection.executemany(
                """
                INSERT OR REPLACE INTO intraday_quotes_snapshot(
                    snapshot_at, secid, boardid, tradingstatus, open, close, lclose, last, numtrades, volvalue, updatetime
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                rows_to_save,
            )
            connection.commit()

    logger.info("Intraday quotes snapshot saved: %s rows", len(rows_to_save))
    return len(rows_to_save), snapshot_at


def _humanize_duration(seconds: float) -> str:
    total_seconds = max(int(seconds), 0)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours} ч {minutes} мин {secs} сек"


def _format_eta_timestamp(seconds: float) -> str:
    return (datetime.now() + timedelta(seconds=max(seconds, 0))).strftime("%H:%M:%S")


@dataclass
class StageProgressView:
    total_stages: int
    started_at: float

    def print_stage_start(self, stage: str, stage_index: int) -> None:
        stage_title = PIPELINE_STAGE_TITLES.get(stage, stage)
        print(
            f"\n🔹 Этап {stage_index}/{self.total_stages}: {stage_title}. "
            f"Выполнено этапов: {stage_index - 1}/{self.total_stages}.",
            flush=True,
        )

    def print_stage_finish(self, stage: str, stage_index: int, duration_ms: float) -> None:
        elapsed = time.perf_counter() - self.started_at
        per_stage_avg = elapsed / max(stage_index, 1)
        remaining_stages = max(self.total_stages - stage_index, 0)
        eta_seconds = per_stage_avg * remaining_stages
        stage_title = PIPELINE_STAGE_TITLES.get(stage, stage)
        print(
            f"✅ Этап завершён: {stage_title} за {duration_ms / 1000:.1f} сек. "
            f"Выполнено этапов {stage_index}/{self.total_stages}. "
            f"Ориентировочное время завершения: {_humanize_duration(eta_seconds)} "
            f"(≈ {_format_eta_timestamp(eta_seconds)}).",
            flush=True,
        )


def _print_details_progress(completed: int, total: int, eta_seconds: int | None, stage_started_at: float) -> None:
    percent = (completed / total * 100) if total else 100.0
    elapsed = time.perf_counter() - stage_started_at
    avg_per_item = elapsed / completed if completed else 0.0
    eta_text = f"{eta_seconds} сек" if eta_seconds is not None else "расчёт после первых результатов"
    print(
        "\r"
        f"   ↳ details: прогресс {percent:5.1f}% | выполнено {completed}/{total} | "
        f"до окончания ~ {eta_text} | ср. шаг {avg_per_item:.2f} сек",
        end="",
        flush=True,
    )
    if completed >= total:
        print()


def _print_details_heartbeat(completed: int, total: int, pending: int, stage_started_at: float) -> None:
    elapsed = time.perf_counter() - stage_started_at
    print(
        f"\n   ↳ details: ожидание ответов {completed}/{total}, активных задач {pending}, прошло {elapsed:.0f} сек",
        flush=True,
    )


def _run_data_quality_checks(dataframe: pd.DataFrame, run_id: str, source: str, logger: logging.Logger) -> list[str]:
    row_count = len(dataframe)
    secid_col = next((c for c in ["SECID", "secid", "SecID"] if c in dataframe.columns), None)
    isin_col = next((c for c in ["ISIN", "isin"] if c in dataframe.columns), None)

    empty_secid_ratio = 1.0
    empty_isin_ratio = 1.0
    if secid_col:
        secid_values = dataframe[secid_col].astype(str).str.strip()
        empty_secid_ratio = float((secid_values == "").mean())
    if isin_col:
        isin_values = dataframe[isin_col].astype(str).str.strip()
        empty_isin_ratio = float((isin_values == "").mean())

    notes: list[str] = []
    row_count_delta_ratio = None

    with sqlite3.connect(CACHE_DB_PATH) as connection:
        prev = connection.execute(
            """
            SELECT row_count
            FROM dq_run_history
            ORDER BY run_at DESC
            LIMIT 1
            """
        ).fetchone()

        if prev and prev[0]:
            previous_count = int(prev[0])
            row_count_delta_ratio = abs(row_count - previous_count) / previous_count
            if row_count_delta_ratio >= ROW_COUNT_SPIKE_THRESHOLD:
                notes.append(
                    f"Резкое изменение row_count: {previous_count} -> {row_count} ({row_count_delta_ratio:.1%})"
                )

        if empty_secid_ratio > 0.01:
            notes.append(f"Высокая доля пустого SECID: {empty_secid_ratio:.2%}")
        if empty_isin_ratio > 0.05:
            notes.append(f"Высокая доля пустого ISIN: {empty_isin_ratio:.2%}")

        connection.execute(
            """
            INSERT OR REPLACE INTO dq_run_history(
                run_id, run_at, source, row_count, empty_secid_ratio, empty_isin_ratio, row_count_delta_ratio, notes
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                datetime.now().isoformat(timespec="seconds"),
                source,
                row_count,
                empty_secid_ratio,
                empty_isin_ratio,
                row_count_delta_ratio,
                " | ".join(notes),
            ),
        )
        connection.commit()

    _refresh_dq_metrics_daily_mv()

    for note in notes:
        logger.warning("DQ check: %s", note)
    if not notes:
        logger.info("DQ check: ok (row_count=%s, empty SECID %.2f%%, empty ISIN %.2f%%)", row_count, empty_secid_ratio * 100, empty_isin_ratio * 100)

    return notes


def fetch_and_save_bond_details(
    dataframe: pd.DataFrame,
    logger: logging.Logger,
    session: requests.Session,
    export_date: str,
    source: str,
    details_worker_processes: int,
    debug: bool,
    stage_started_at: float | None = None,
) -> tuple[int, int, int, int, Path]:
    cleanup_details_cache()
    secids = _pick_secids(dataframe)
    if not secids:
        logger.warning("Could not determine SECID list. Skipping details export")
        finish_df = _normalize_identity_columns(dataframe.copy())
        batch_id = datetime.now().strftime("%Y%m%dT%H%M%S")
        batch_path = _export_finish_incremental(finish_df, export_date, batch_id)
        _persist_finish_to_sqlite(finish_df, batch_id=batch_id, export_date=export_date, source=source)
        return 0, 0, 0, len(finish_df), batch_path

    prechecked_endpoints = _precheck_details_endpoints_health(session, secids[0], logger)
    if not prechecked_endpoints:
        logger.warning("All details endpoints failed cold-start precheck. Skip heavy details phase")
        finish_df = _normalize_identity_columns(dataframe.copy())
        batch_id = datetime.now().strftime("%Y%m%dT%H%M%S")
        batch_path = _export_finish_incremental(finish_df, export_date, batch_id)
        _persist_finish_to_sqlite(finish_df, batch_id=batch_id, export_date=export_date, source=source)
        return len(secids), 0, 0, len(finish_df), batch_path

    working_endpoints = _discover_working_endpoints(session, secids[0], logger, candidates=prechecked_endpoints)
    if not working_endpoints:
        logger.warning("No working details endpoints found. Skipping details export")
        finish_df = _normalize_identity_columns(dataframe.copy())
        batch_id = datetime.now().strftime("%Y%m%dT%H%M%S")
        batch_path = _export_finish_incremental(finish_df, export_date, batch_id)
        _persist_finish_to_sqlite(finish_df, batch_id=batch_id, export_date=export_date, source=source)
        return len(secids), 0, 0, len(finish_df), batch_path

    endpoint_frames: dict[str, dict[str, list[pd.DataFrame]]] = {name: {} for name, _ in working_endpoints}

    pending_by_secid: dict[str, list[tuple[str, str]]] = {}
    for secid in secids:
        missing_endpoints: list[tuple[str, str]] = []
        for endpoint_name, endpoint_url in working_endpoints:
            cached_record = _get_cached_endpoint_record(endpoint_name, secid)
            if cached_record is None:
                missing_endpoints.append((endpoint_name, endpoint_url))
                continue
            cached_payload, fetched_at = cached_record
            blocks = _extract_blocks(cached_payload)
            logger.info("Details %s for %s loaded from cache", endpoint_name, secid)
            _save_endpoint_health(endpoint_name, secid, "ok", "cache", None, 0.0, list(blocks.keys()), None)
            _save_watermark(endpoint_name, secid, fetched_at)
            for block_name, block_df in blocks.items():
                if not block_df.empty:
                    endpoint_frames[endpoint_name].setdefault(block_name, []).append(_normalize_block_frame(secid, block_name, block_df))
        if missing_endpoints:
            pending_by_secid[secid] = missing_endpoints

    pending_secids = list(pending_by_secid.keys())

    with ThreadPoolExecutor(max_workers=details_worker_processes) as executor:
        futures = []
        for secid in pending_secids:
            allowed_endpoints = [(n, u) for n, u in pending_by_secid[secid] if not _is_circuit_open(n)]
            if not allowed_endpoints:
                logger.warning("All endpoints are blocked by circuit breaker for %s", secid)
                continue
            futures.append(executor.submit(_fetch_details_worker, secid, allowed_endpoints))

        completed = 0
        total = len(futures)
        progress_started_at = time.perf_counter()
        last_heartbeat_at = progress_started_at
        pending_futures = set(futures)
        if total:
            _print_details_progress(0, total, None, stage_started_at=stage_started_at or progress_started_at)
        else:
            print("   ↳ details: все данные взяты из кэша, сетевых запросов нет", flush=True)
        while pending_futures:
            done_futures, pending_futures = wait(pending_futures, timeout=2.0, return_when=FIRST_COMPLETED)
            if not done_futures:
                _print_details_progress(completed, total, None, stage_started_at=stage_started_at or progress_started_at)
                now = time.perf_counter()
                if now - last_heartbeat_at >= 15:
                    _print_details_heartbeat(
                        completed,
                        total,
                        len(pending_futures),
                        stage_started_at=stage_started_at or progress_started_at,
                    )
                    last_heartbeat_at = now
                continue

            for future in done_futures:
                completed += 1
                elapsed = time.perf_counter() - progress_started_at
                avg_seconds = elapsed / completed if completed else 0.0
                remaining = max(total - completed, 0)
                eta_seconds = int(avg_seconds * remaining)
                _print_details_progress(completed, total, eta_seconds, stage_started_at=stage_started_at or progress_started_at)
                for endpoint_name, secid, payload, latency_ms, status_code, error_text in future.result():
                    if payload is None:
                        stale_cached_record = _get_latest_endpoint_record(endpoint_name, secid)
                        if stale_cached_record is not None:
                            stale_payload, stale_fetched_at = stale_cached_record
                            stale_blocks = _extract_blocks(stale_payload)
                            _save_endpoint_health(endpoint_name, secid, "ok", "stale_cache", status_code, latency_ms, list(stale_blocks.keys()), error_text)
                            _save_watermark(endpoint_name, secid, stale_fetched_at)
                            logger.warning(
                                "Details %s for %s loaded from stale cache (%s) after network error: %s",
                                endpoint_name,
                                secid,
                                stale_fetched_at,
                                error_text,
                            )
                            for block_name, block_df in stale_blocks.items():
                                if not block_df.empty:
                                    endpoint_frames[endpoint_name].setdefault(block_name, []).append(_normalize_block_frame(secid, block_name, block_df))
                            continue

                        _save_endpoint_health(endpoint_name, secid, "error", "network", status_code, latency_ms, None, error_text)
                        _update_circuit_breaker(endpoint_name, is_success=False)
                        logger.warning("Details %s for %s failed: %s", endpoint_name, secid, error_text)
                        continue

                    _update_circuit_breaker(endpoint_name, is_success=True)
                    network_fetched_at = datetime.now().isoformat(timespec="seconds")
                    if _watermark_exists(endpoint_name, secid, network_fetched_at):
                        logger.info("Skip details update for %s:%s by watermark %s", endpoint_name, secid, network_fetched_at)
                        continue

                    persisted_fetched_at = _save_endpoint_payload(
                        endpoint_name,
                        secid,
                        payload,
                        fetched_at=network_fetched_at,
                    )
                    blocks = _extract_blocks(payload)
                    _save_endpoint_health(endpoint_name, secid, "ok", "network", status_code, latency_ms, list(blocks.keys()), None)
                    _save_watermark(endpoint_name, secid, persisted_fetched_at)
                    logger.info("Details %s for %s loaded from network", endpoint_name, secid)
                    for block_name, block_df in blocks.items():
                        if not block_df.empty:
                            endpoint_frames[endpoint_name].setdefault(block_name, []).append(_normalize_block_frame(secid, block_name, block_df))

    merged_frames: dict[str, dict[str, pd.DataFrame]] = {}
    for endpoint_name, blocks in endpoint_frames.items():
        merged_frames[endpoint_name] = {}
        for block_name, frames in blocks.items():
            valid_frames = [frame.dropna(axis=1, how="all") for frame in frames if not frame.empty]
            if valid_frames:
                merged_frames[endpoint_name][block_name] = pd.concat(valid_frames, ignore_index=True, sort=False)

    enrichment_df = _build_enrichment_frame(merged_frames)
    excel_sheets = 0
    if debug:
        excel_sheets = _write_details_excel(merged_frames, enrichment_df)
    else:
        logger.info("Debug mode disabled: skip %s generation", DETAILS_EXCEL_PATH.name)

    parquet_files = _update_details_parquet(merged_frames)

    finish_df = _normalize_identity_columns(_merge_base_with_enrichment(dataframe, enrichment_df))
    batch_id = datetime.now().strftime("%Y%m%dT%H%M%S")
    batch_path = _export_finish_incremental(finish_df, export_date, batch_id)
    _persist_finish_to_sqlite(finish_df, batch_id=batch_id, export_date=export_date, source=source)
    refresh_endpoint_health_mv()
    return len(secids), excel_sheets, parquet_files, len(finish_df), batch_path


def _prepare_csv_for_pandas(csv_data: str) -> tuple[str, str]:
    lines = [line for line in csv_data.splitlines() if line.strip()]
    if not lines:
        raise ValueError("CSV data is empty")

    delimiter = ","
    if sum(";" in line for line in lines[:10]) > sum("," in line for line in lines[:10]):
        delimiter = ";"

    parsed_rows: list[list[str]] = []
    for line in lines:
        parsed_rows.append(next(csv.reader([line], delimiter=delimiter)))

    header_index = next(
        (index for index, row in enumerate(parsed_rows) if len(row) > 1),
        None,
    )
    if header_index is None:
        raise ValueError("Could not find CSV header in response")

    expected_width = len(parsed_rows[header_index])
    filtered_lines = [
        lines[index]
        for index in range(header_index, len(lines))
        if len(parsed_rows[index]) == expected_width
    ]

    if len(filtered_lines) < 2:
        raise ValueError("Could not parse tabular CSV data from response")

    return "\n".join(filtered_lines), delimiter


def _resolve_details_worker_processes(cli_value: int | None, logger: logging.Logger) -> int:
    if cli_value is not None:
        value = cli_value
        source = "cli"
    else:
        raw_env = os.getenv("DETAILS_WORKER_PROCESSES", str(DEFAULT_DETAILS_WORKER_PROCESSES)).strip()
        try:
            value = int(raw_env)
        except ValueError:
            logger.warning(
                "Invalid DETAILS_WORKER_PROCESSES env value '%s'. Fallback to %s",
                raw_env,
                DEFAULT_DETAILS_WORKER_PROCESSES,
            )
            value = DEFAULT_DETAILS_WORKER_PROCESSES
        source = "env/default"

    if value < 1:
        logger.warning("DETAILS_WORKER_PROCESSES must be >= 1, got %s. Fallback to 1", value)
        value = 1

    logger.info("Details worker processes resolved to %s (%s)", value, source)
    return value


def _build_daily_dq_report(logger: logging.Logger, report_day: date) -> Path:
    report_path = LOGS_DIR / f"dq_daily_report_{report_day.isoformat()}.txt"
    day_start = datetime.combine(report_day, datetime.min.time())
    day_end = day_start + timedelta(days=1)

    with sqlite3.connect(CACHE_DB_PATH) as connection:
        daily_rows = connection.execute(
            """
            SELECT run_at, row_count, notes
            FROM dq_run_history
            WHERE run_at >= ? AND run_at < ?
            ORDER BY run_at ASC
            """,
            (day_start.isoformat(timespec="seconds"), day_end.isoformat(timespec="seconds")),
        ).fetchall()
        trend_rows = connection.execute(
            """
            SELECT run_day, avg_row_count, max_row_count
            FROM dq_metrics_daily_mv
            WHERE run_day >= ?
            ORDER BY run_day DESC
            LIMIT 7
            """,
            ((day_start - timedelta(days=30)).date().isoformat(),),
        ).fetchall()

    warning_counts: dict[str, int] = {}
    for _, _, notes in daily_rows:
        if not notes:
            continue
        for note in [part.strip() for part in str(notes).split("|") if part.strip()]:
            warning_counts[note] = warning_counts.get(note, 0) + 1

    top_warnings = sorted(warning_counts.items(), key=lambda item: item[1], reverse=True)[:5]
    trend_rows_desc = trend_rows
    trend_rows_asc = list(reversed(trend_rows_desc))

    lines = [
        f"DQ daily report for {report_day.isoformat()}",
        f"Runs today: {len(daily_rows)}",
        "Top warnings:",
    ]
    if top_warnings:
        lines.extend([f"- {warning} (count={count})" for warning, count in top_warnings])
    else:
        lines.append("- none")

    lines.append("Row count trend (last <=7 days):")
    if trend_rows_asc:
        for run_day, avg_row_count, max_row_count in trend_rows_asc:
            lines.append(f"- {run_day}: avg={avg_row_count:.1f}, max={int(max_row_count)}")
    else:
        lines.append("- no history")

    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    logger.info("Daily DQ report saved to %s", report_path)
    return report_path




def _save_sla_stage_timer(run_id: str, stage: str, started_at: datetime, finished_at: datetime, status: str, source: str | None = None, details: str | None = None) -> float:
    duration_ms = max((finished_at - started_at).total_seconds() * 1000, 0.0)
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        connection.execute(
            """
            INSERT OR REPLACE INTO etl_stage_sla(run_id, stage, started_at, finished_at, duration_ms, status, source, details)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                run_id,
                stage,
                started_at.isoformat(timespec="seconds"),
                finished_at.isoformat(timespec="seconds"),
                duration_ms,
                status,
                source,
                details,
            ),
        )
        connection.commit()
    return duration_ms


def _check_sla_degradation(stage: str, duration_ms: float, logger: logging.Logger) -> None:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        row = connection.execute(
            """
            SELECT AVG(duration_ms), COUNT(*)
            FROM (
                SELECT duration_ms
                FROM etl_stage_sla
                WHERE stage = ? AND status = 'ok'
                ORDER BY finished_at DESC
                LIMIT 30
            )
            """,
            (stage,),
        ).fetchone()

    baseline_ms = float(row[0]) if row and row[0] is not None else None
    samples = int(row[1]) if row and row[1] is not None else 0
    if baseline_ms and samples >= 5 and duration_ms > baseline_ms * 1.5 and (duration_ms - baseline_ms) > 1000:
        logger.warning(
            "SLA degradation detected for %s: %.0fms vs baseline %.0fms (%s samples)",
            stage,
            duration_ms,
            baseline_ms,
            samples,
        )

def _run_sqlite_maintenance(logger: logging.Logger) -> None:
    with sqlite3.connect(CACHE_DB_PATH) as connection:
        connection.execute("ANALYZE")
        connection.execute("PRAGMA optimize")
        connection.commit()
    logger.info("SQLite maintenance completed (ANALYZE + PRAGMA optimize)")


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="MOEX bonds ETL pipeline")
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable debug artifacts (Moex_Bonds.xlsx and Moex_Bonds_Details.xlsx)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preflight MOEX/cache checks without writing Excel/DB payloads",
    )
    return parser.parse_args()


def _load_daily_csv_with_cache_fallback(session: requests.Session, logger: logging.Logger, start_time: float, run_id: str) -> tuple[str, str, str, float]:
    today = date.today().isoformat()
    data_source = "cache"
    fetch_started_dt = datetime.now()
    try:
        csv_data = get_cached_data(today)
        logger.info("Using cached data for %s", today)
    except CacheMissError:
        data_source = "network"
        logger.info("Cache miss for %s. Fetching from MOEX...", today)
        try:
            csv_data = fetch_moex_csv(session)
            save_to_cache(today, csv_data)
            logger.info("Data fetched and cached")
        except requests.RequestException as error:
            try:
                cached_date, csv_data = get_latest_cached_data()
            except CacheMissError:
                elapsed = time.perf_counter() - start_time
                _save_sla_stage_timer(run_id, "fetch", fetch_started_dt, datetime.now(), "error", source=data_source, details="MOEX unavailable and cache empty")
                logger.error("MOEX is unavailable (%s) and cache is empty", error)
                logger.info("Execution time before failure: %.3f seconds", elapsed)
                raise RuntimeError("MOEX unavailable and cache is empty") from error

            data_source = f"cache_fallback:{cached_date}"
            logger.warning(
                "MOEX is unavailable (%s). Falling back to cached data from %s",
                error,
                cached_date,
            )
    fetch_duration_ms = _save_sla_stage_timer(run_id, "fetch", fetch_started_dt, datetime.now(), "ok", source=data_source)
    _check_sla_degradation("fetch", fetch_duration_ms, logger)
    return csv_data, data_source, today, fetch_duration_ms


def _diff_bonds_against_previous_day(current_df: pd.DataFrame, export_date: str) -> tuple[list[str], list[str]]:
    key_col = "ISIN" if "ISIN" in current_df.columns else ("SECID" if "SECID" in current_df.columns else None)
    if key_col is None:
        return [], []

    short_col = "SHORTNAME" if "SHORTNAME" in current_df.columns else None
    today_rows = {
        str(row[key_col]).strip(): str(row.get(short_col, "")).strip()
        for _, row in current_df.iterrows()
        if str(row.get(key_col, "")).strip()
    }

    with sqlite3.connect(CACHE_DB_PATH) as connection:
        prev_row = connection.execute(
            """
            SELECT csv_data
            FROM bonds_cache
            WHERE fetch_date < ?
            ORDER BY fetch_date DESC
            LIMIT 1
            """,
            (export_date,),
        ).fetchone()

    if prev_row is None:
        return sorted([f"{isin}, {name}" for isin, name in today_rows.items() if isin]), []

    prev_df = parse_rates_csv(prev_row[0])
    prev_key_col = "ISIN" if "ISIN" in prev_df.columns else ("SECID" if "SECID" in prev_df.columns else None)
    if prev_key_col is None:
        return [], []
    prev_short_col = "SHORTNAME" if "SHORTNAME" in prev_df.columns else None
    prev_rows = {
        str(row[prev_key_col]).strip(): str(row.get(prev_short_col, "")).strip()
        for _, row in prev_df.iterrows()
        if str(row.get(prev_key_col, "")).strip()
    }

    added = sorted(set(today_rows) - set(prev_rows))
    removed = sorted(set(prev_rows) - set(today_rows))
    added_lines = [f"{isin}, {today_rows.get(isin, '')}" for isin in added]
    removed_lines = [f"{isin}, {prev_rows.get(isin, '')}" for isin in removed]
    return added_lines, removed_lines


def run_rates_ingest(session: requests.Session, logger: logging.Logger, run_id: str, debug: bool, stage_view: StageProgressView | None = None) -> tuple[pd.DataFrame, str, str]:
    if stage_view is not None:
        stage_view.print_stage_start("fetch", 1)
    pipeline_started = stage_view.started_at if stage_view else time.perf_counter()
    csv_data, data_source, export_date, fetch_duration_ms = _load_daily_csv_with_cache_fallback(session, logger, pipeline_started, run_id)
    if stage_view is not None:
        stage_view.print_stage_finish("fetch", 1, fetch_duration_ms)

    if stage_view is not None:
        stage_view.print_stage_start("parse", 2)
    parse_started_dt = datetime.now()
    dataframe = parse_rates_csv(csv_data)
    if debug:
        dataframe.to_excel(EXCEL_PATH, index=False)
    else:
        logger.info("Debug mode disabled: skip %s generation", EXCEL_PATH.name)
    _run_data_quality_checks(dataframe, run_id=run_id, source=data_source, logger=logger)
    persist_raw_response(csv_data)
    parse_duration_ms = _save_sla_stage_timer(run_id, "parse", parse_started_dt, datetime.now(), "ok", source=data_source, details=f"rows={len(dataframe)}")
    _check_sla_degradation("parse", parse_duration_ms, logger)
    if stage_view is not None:
        stage_view.print_stage_finish("parse", 2, parse_duration_ms)

    added_lines, removed_lines = _diff_bonds_against_previous_day(dataframe, export_date)
    logger.info("Добавлено %s облигаций", len(added_lines))
    for line in added_lines[:50]:
        logger.info("  + %s", line)
    logger.info("Погашено %s бумаг", len(removed_lines))
    for line in removed_lines[:50]:
        logger.info("  - %s", line)
    if not added_lines and not removed_lines:
        print("Изменений по бумагам нет", flush=True)

    return dataframe, data_source, export_date


def run_details_enricher(dataframe: pd.DataFrame, session: requests.Session, logger: logging.Logger, run_id: str, source: str, export_date: str, details_worker_processes: int, debug: bool, stage_view: StageProgressView | None = None) -> tuple[int, int, int, int, Path]:
    if stage_view is not None:
        stage_view.print_stage_start("details", 3)
    details_started_dt = datetime.now()
    result = fetch_and_save_bond_details(
        dataframe,
        logger,
        session,
        export_date=export_date,
        source=source,
        details_worker_processes=details_worker_processes,
        debug=debug,
        stage_started_at=time.perf_counter(),
    )
    details_duration_ms = _save_sla_stage_timer(run_id, "details", details_started_dt, datetime.now(), "ok", source=source, details=f"secids={result[0]}")
    _check_sla_degradation("details", details_duration_ms, logger)
    if stage_view is not None:
        stage_view.print_stage_finish("details", 3, details_duration_ms)
    return result


def run_quotes_snapshotter(session: requests.Session, logger: logging.Logger, run_id: str, source: str, stage_view: StageProgressView | None = None) -> tuple[Path, int]:
    if stage_view is not None:
        stage_view.print_stage_start("export", 4)
    export_started_dt = datetime.now()
    dq_daily_report_path = _build_daily_dq_report(logger, report_day=date.today())
    intraday_rows, snapshot_at = _persist_intraday_quotes_snapshot(session, logger)
    try:
        with sqlite3.connect(CACHE_DB_PATH) as connection:
            finish_df = pd.read_sql_query("SELECT * FROM bonds_enriched", connection)
        _export_price_workbook(finish_df, snapshot_at)
    except Exception as error:  # noqa: BLE001
        logger.warning("Price workbook export skipped: %s", error)
    export_duration_ms = _save_sla_stage_timer(run_id, "export", export_started_dt, datetime.now(), "ok", source=source, details=f"intraday_rows={intraday_rows}")
    _check_sla_degradation("export", export_duration_ms, logger)
    if stage_view is not None:
        stage_view.print_stage_finish("export", 4, export_duration_ms)
    return dq_daily_report_path, intraday_rows


def main() -> int:
    args = _parse_args()
    logger = setup_logging()
    init_db()
    session = build_retry_session()
    details_worker_processes = _resolve_details_worker_processes(None, logger)
    _run_sqlite_maintenance(logger)

    if args.dry_run:
        try:
            try:
                csv_data = fetch_moex_csv(session)
                source = "network"
            except Exception:
                try:
                    _, csv_data = get_latest_cached_data()
                    source = "cache"
                except Exception:
                    csv_data = RAW_RESPONSE_PATH.read_text(encoding="utf-8")
                    source = "raw_file"
            dataframe = parse_rates_csv(csv_data)
            secids = _pick_secids(dataframe)
            healthy = _precheck_details_endpoints_health(session, secids[0] if secids else STATIC_SECIDS[0], logger)
            print(f"DRY-RUN OK: source={source}, rows={len(dataframe)}, secids={len(secids)}, healthy_endpoints={len(healthy)}")
            return 0
        except Exception as error:  # noqa: BLE001
            logger.error("Dry-run failed: %s", error)
            return 1

    start_time = time.perf_counter()
    run_id = datetime.now().strftime("%Y%m%dT%H%M%S")
    stage_view = StageProgressView(total_stages=len(PIPELINE_STAGE_ORDER), started_at=start_time)
    logger.info("MOEX API script started")

    try:
        dataframe, data_source, export_date = run_rates_ingest(session, logger, run_id=run_id, debug=args.debug, stage_view=stage_view)
        secids_count, details_sheets, parquet_files, finish_rows, finish_batch_path = run_details_enricher(
            dataframe,
            session,
            logger,
            run_id=run_id,
            source=data_source,
            export_date=export_date,
            details_worker_processes=details_worker_processes,
            debug=args.debug,
            stage_view=stage_view,
        )
        dq_daily_report_path, intraday_rows = run_quotes_snapshotter(session, logger, run_id=run_id, source=data_source, stage_view=stage_view)

        elapsed = time.perf_counter() - start_time
        logger.info("Data source: %s", data_source)
        logger.info("Saved extended details for %s securities into %s endpoint sheets", secids_count, details_sheets)
        logger.info("Updated %s parquet detail files in %s", parquet_files, DETAILS_PARQUET_DIR)
        logger.info("Saved enriched finish dataset with %s rows to %s", finish_rows, FINISH_EXCEL_PATH)
        logger.info("Saved incremental finish batch to %s", finish_batch_path)
        logger.info("DQ daily report path: %s", dq_daily_report_path)
        logger.info("Intraday quotes snapshot rows: %s", intraday_rows)
        logger.info("Execution time: %.3f seconds", elapsed)
        print(f"\n🎉 Готово! Полный пайплайн завершён за {_humanize_duration(elapsed)}.", flush=True)
        return 0

    except Exception as error:
        elapsed = time.perf_counter() - start_time
        _save_sla_stage_timer(run_id, "export", datetime.now(), datetime.now(), "error", source="unknown", details=str(error))
        logger.error("Script failed: %s", error)
        logger.info("Execution time before failure: %.3f seconds", elapsed)
        print(f"MOEX_API failed: {error}")
        return 1


if __name__ == "__main__":
    main()
