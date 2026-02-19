from __future__ import annotations

import json
import logging
import math
import re
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl.styles import Alignment, Font, PatternFill
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter


# =========================
# Базовые настройки скрипта
# =========================
BASE_URL = "https://iss.moex.com/iss/engines/stock/markets/bonds/securities.json"
PAGE_SIZE = 100
REQUEST_TIMEOUT = 20
CACHE_TTL_HOURS = 6
MAX_RETRIES = 4
BACKOFF_FACTOR = 1.2
MAX_WORKERS = 10
DOHOD_MAX_WORKERS = 10
DOHOD_REQUEST_RETRIES = 3
OFFER_CHECK_MAX_WORKERS = 10
OFFER_CHECK_BATCH_SIZE = 120
OFFER_SOURCE_MAX_WORKERS = 10
OFFER_CHECK_REQUEST_TIMEOUT = 8
OFFER_CHECK_REQUEST_RETRIES = 2
OFFER_RECHECK_DAILY_MIN_JOBS = 300
SECID_BATCH_SIZE = 50
EMITTER_BATCH_SIZE = 80

ROOT_DIR = Path(__file__).resolve().parent
LOGS_DIR = ROOT_DIR / "logs"
RAW_DIR = ROOT_DIR / "raw"
OUTPUT_DIR = ROOT_DIR / "output"
CACHE_DIR = ROOT_DIR / "cache"

LOG_FILE = LOGS_DIR / "moex_bonds.log"
RAW_FILE = RAW_DIR / "moex_bonds_raw.json"
CACHE_FILE = CACHE_DIR / "moex_bonds_cache.json"
ISSUER_CACHE_FILE = CACHE_DIR / "issuer_directory_cache.json"
ISSUER_CHECKPOINT_FILE = CACHE_DIR / "issuer_enrichment_checkpoint.json"
DAILY_CACHE_FILE = CACHE_DIR / "daily_security_metrics_cache.json"
OFFER_VERIFICATION_CACHE_FILE = CACHE_DIR / "offer_verification_cache.json"
COUPON_FORMULA_CACHE_FILE = CACHE_DIR / "coupon_formula_cache.json"
COUPON_VALUE_CACHE_FILE = CACHE_DIR / "coupon_value_cache.json"
OUTPUT_FILE = OUTPUT_DIR / "moex_bonds.xlsx"

# Колонки, которые пользователь попросил убрать из итогового файла
REMOVED_COLUMNS = {
    "SECNAME",
    "LISTLEVEL",
    "STATUS",
    "EXCLUDE_BY_AMORTIZATION",
    "AMORTIZATION_EXCLUDE_REASON",
    "EXCLUDE_BY_OFFER_DATE",
    "OFFER_DATE_EXCLUDE_REASON",
    "EXCLUDE_BY_COUPONPERIOD",
    "COUPONPERIOD_EXCLUDE_REASON",
    "EXCLUDE_BY_BOND_TYPE",
    "BOND_TYPE_EXCLUDE_REASON",
}

DEPRECATED_OUTPUT_COLUMNS = {"ISSUER_BOND_CLASS", "ISSUER_SECTOR", "SECTORID", "INSTRID"}
# SECID нужен для технической работы, но в Excel должен быть скрыт
HIDDEN_COLUMN_NAME = "SECID"
ISSUER_COLUMN_NAME = "ISSUER_NAME"
ISSUER_INN_COLUMN_NAME = "ISSUER_INN"
ISSUER_BOND_CLASS_COLUMN_NAME = "ISSUER_BOND_CLASS"
ISSUER_RATING_COLUMN_NAME = "ISSUER_RATING"
FIRST_COLUMN_NAME = "ISIN"
GROUP_SEPARATOR_PREFIX = "GROUP_SEPARATOR__"
QUALIFIED_INVESTOR_COLUMN_NAME = "QUALIFIED_INVESTOR"
MATURITY_DATE_COLUMN_NAME = "MATDATE"
AMORTIZATION_FLAG_COLUMN_NAME = "HAS_AMORTIZATION"
AMORTIZATION_START_DATE_COLUMN_NAME = "AMORTIZATION_START_DATE"
ACCRUED_INT_COLUMN_NAME = "ACCRUEDINT"
TRADE_VOLUME_COLUMN_NAME = "VOLTODAY"
TRADE_VALUE_COLUMN_NAME = "VALTODAY"
YIELD_COLUMN_NAME = "YIELD"
BOND_TYPE_COLUMN_NAME = "BOND_TYPE"
HAS_PUT_CALL_OFFER_COLUMN_NAME = "HAS_PUT_CALL_OFFER"
PUT_CALL_OFFER_DATE_COLUMN_NAME = "PUT_CALL_OFFER_DATE"
COUPON_FORMULA_SOURCE_COLUMN_NAME = "COUPON_FORMULA_SOURCE"
SECURITY_DAILY_CACHE_TTL_HOURS = 24
OFFER_VERIFICATION_CACHE_TTL_DAYS = 7
OFFER_VERIFICATION_CACHE_SCHEMA_VERSION = 8
COUPON_FORMULA_CACHE_TTL_DAYS = 7
MIN_MATURITY_YEARS = 1
DAILY_METRICS_CACHE_SCHEMA_VERSION = 7
DOHOD_BOND_URL = "https://analytics.dohod.ru/bond/{isin}"
CORPBONDS_BOND_URL = "https://corpbonds.ru/bond/{isin}"
DOHOD_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/132.0.0.0 Safari/537.36"
)
CALCULATED_COUPON_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
STRUCTURAL_BOND_TYPE_VALUES = {"структурная облигация", "структурные облигации"}
COUPONPERIOD_EXCLUDE_REASON = "Купонный период не определён (COUPONPERIOD = 0)"
BOND_TYPE_EXCLUDE_REASON = "Структурная облигация исключена из выгрузки"
OFFER_DATE_EXCLUDE_REASON = "Дата оферты наступит менее чем через год"
INSTRID_TO_BOND_CLASS = {
    "GOFZ": "Государственный",
    "YOFZ": "Государственный",
    "EICB": "Корпоративный",
    "EIOD": "Муниципальный/субфедеральный",
    "EIYO": "Иностранный",
    "EIUS": "Иностранный",
}

CORPBONDS_INDEX_ALIASES = {
    "KC": "CBR_RATE",
    "КС": "CBR_RATE",
    "KEYRATE": "CBR_RATE",
    "CBR_RATE": "CBR_RATE",
    "RUONIA": "RUONIA",
    "G_CURVE_RUS": "G_CURVE_RUS",
}




@dataclass
class IssPage:
    """Одна страница ответа ISS API."""

    start: int
    rows: list[dict[str, Any]]


def setup_folders() -> None:
    """Создаёт рабочие директории и очищает папку raw перед запуском."""
    LOGS_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)
    CACHE_DIR.mkdir(exist_ok=True)

    if RAW_DIR.exists():
        shutil.rmtree(RAW_DIR)
    RAW_DIR.mkdir(exist_ok=True)


def setup_logging() -> None:
    """Настраивает логирование в файл, который перезаписывается при каждом запуске."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )


def load_cache(allow_stale: bool = False) -> list[dict[str, Any]] | None:
    """Читает кэш, если он есть; по флагу allow_stale может вернуть и устаревшие данные."""
    if not CACHE_FILE.exists():
        return None

    try:
        payload = json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        created_at = datetime.fromisoformat(payload["created_at"])
        if datetime.now() - created_at > timedelta(hours=CACHE_TTL_HOURS):
            if allow_stale:
                rows = payload.get("rows", [])
                logging.warning("Кэш устарел, но будет использован как резерв: %s записей.", len(rows))
                return rows
            logging.info("Кэш найден, но устарел. Загружаем свежие данные.")
            return None

        rows = payload.get("rows", [])
        logging.info("Данные загружены из кэша: %s записей.", len(rows))
        return rows
    except Exception as exc:
        logging.warning("Не удалось прочитать кэш: %s", exc)
        return None


def save_cache(rows: list[dict[str, Any]]) -> None:
    """Сохраняет кэш в JSON файл."""
    payload = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
    }
    CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_issuer_directory_cache() -> tuple[dict[str, int | None], dict[int, str], dict[int, str], dict[str, str], dict[str, str], dict[str, int], dict[str, bool], dict[str, str]]:
    """Читает пожизненный кэш SECID -> EMITTER_ID/QUALIFIED_INVESTOR/BOND_TYPE/COUPONPERIOD/ISSUER_RATING, EMITTER_ID -> имя и ИНН."""
    if not ISSUER_CACHE_FILE.exists():
        return {}, {}, {}, {}, {}, {}, {}, {}

    try:
        payload = json.loads(ISSUER_CACHE_FILE.read_text(encoding="utf-8"))
        secid_to_emitter_id: dict[str, int | None] = {}
        emitter_id_to_name: dict[int, str] = {}
        emitter_id_to_inn: dict[int, str] = {}
        secid_to_qualified_sign: dict[str, str] = {}
        secid_to_bond_type: dict[str, str] = {}
        secid_to_coupon_period: dict[str, int] = {}
        secid_to_is_structural: dict[str, bool] = {}
        secid_to_issuer_rating: dict[str, str] = {}

        for secid, emitter_id in payload.get("secid_to_emitter_id", {}).items():
            if emitter_id is None:
                secid_to_emitter_id[str(secid)] = None
            else:
                try:
                    secid_to_emitter_id[str(secid)] = int(emitter_id)
                except (TypeError, ValueError):
                    secid_to_emitter_id[str(secid)] = None

        for emitter_id, emitter_name in payload.get("emitter_id_to_name", {}).items():
            if not emitter_name:
                continue
            try:
                emitter_id_to_name[int(emitter_id)] = str(emitter_name)
            except (TypeError, ValueError):
                continue

        for emitter_id, emitter_inn in payload.get("emitter_id_to_inn", {}).items():
            if not emitter_inn:
                continue
            try:
                emitter_id_to_inn[int(emitter_id)] = str(emitter_inn)
            except (TypeError, ValueError):
                continue

        for secid, qualified_sign in payload.get("secid_to_qualified_sign", {}).items():
            if str(qualified_sign) in {"✔", "✖"}:
                secid_to_qualified_sign[str(secid)] = str(qualified_sign)

        for secid, bond_type in payload.get("secid_to_bond_type", {}).items():
            if bond_type:
                secid_to_bond_type[str(secid)] = str(bond_type)

        for secid, coupon_period in payload.get("secid_to_coupon_period", {}).items():
            try:
                secid_to_coupon_period[str(secid)] = int(coupon_period)
            except (TypeError, ValueError):
                continue

        for secid, is_structural in payload.get("secid_to_is_structural", {}).items():
            secid_to_is_structural[str(secid)] = bool(is_structural)

        for secid, issuer_rating in payload.get("secid_to_issuer_rating", {}).items():
            if issuer_rating:
                secid_to_issuer_rating[str(secid)] = str(issuer_rating)

        logging.info(
            "Загружен пожизненный справочник эмитентов: SECID=%s, EMITTER_ID=%s.",
            len(secid_to_emitter_id),
            len(emitter_id_to_name),
        )
        return secid_to_emitter_id, emitter_id_to_name, emitter_id_to_inn, secid_to_qualified_sign, secid_to_bond_type, secid_to_coupon_period, secid_to_is_structural, secid_to_issuer_rating
    except Exception as exc:
        logging.warning("Не удалось прочитать пожизненный кэш эмитентов: %s", exc)
        return {}, {}, {}, {}, {}, {}, {}, {}


def save_issuer_directory_cache(
    secid_to_emitter_id: dict[str, int | None],
    emitter_id_to_name: dict[int, str],
    emitter_id_to_inn: dict[int, str],
    secid_to_qualified_sign: dict[str, str],
    secid_to_bond_type: dict[str, str],
    secid_to_coupon_period: dict[str, int],
    secid_to_is_structural: dict[str, bool],
    secid_to_issuer_rating: dict[str, str],
) -> None:
    """Сохраняет пожизненный справочник эмитентов."""
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "secid_to_emitter_id": secid_to_emitter_id,
        "secid_to_qualified_sign": secid_to_qualified_sign,
        "secid_to_bond_type": secid_to_bond_type,
        "secid_to_coupon_period": secid_to_coupon_period,
        "secid_to_is_structural": secid_to_is_structural,
        "secid_to_issuer_rating": secid_to_issuer_rating,
        "emitter_id_to_name": {str(k): v for k, v in emitter_id_to_name.items()},
        "emitter_id_to_inn": {str(k): v for k, v in emitter_id_to_inn.items()},
    }
    ISSUER_CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_issuer_checkpoint() -> tuple[dict[str, int | None], dict[int, str], dict[int, str], dict[str, str], dict[str, str], dict[str, int], dict[str, bool], dict[str, str]]:
    """Возвращает checkpoint по этапу обогащения эмитентов, если он есть."""
    if not ISSUER_CHECKPOINT_FILE.exists():
        return {}, {}, {}, {}, {}, {}, {}, {}

    try:
        payload = json.loads(ISSUER_CHECKPOINT_FILE.read_text(encoding="utf-8"))
        secid_to_emitter_id: dict[str, int | None] = {}
        emitter_id_to_name: dict[int, str] = {}
        emitter_id_to_inn: dict[int, str] = {}
        secid_to_qualified_sign: dict[str, str] = {}
        secid_to_bond_type: dict[str, str] = {}
        secid_to_coupon_period: dict[str, int] = {}
        secid_to_issuer_rating: dict[str, str] = {}

        for secid, emitter_id in payload.get("secid_to_emitter_id", {}).items():
            if emitter_id is None:
                secid_to_emitter_id[str(secid)] = None
            else:
                try:
                    secid_to_emitter_id[str(secid)] = int(emitter_id)
                except (TypeError, ValueError):
                    secid_to_emitter_id[str(secid)] = None

        for emitter_id, emitter_name in payload.get("emitter_id_to_name", {}).items():
            if not emitter_name:
                continue
            try:
                emitter_id_to_name[int(emitter_id)] = str(emitter_name)
            except (TypeError, ValueError):
                continue

        for emitter_id, emitter_inn in payload.get("emitter_id_to_inn", {}).items():
            if not emitter_inn:
                continue
            try:
                emitter_id_to_inn[int(emitter_id)] = str(emitter_inn)
            except (TypeError, ValueError):
                continue

        for secid, qualified_sign in payload.get("secid_to_qualified_sign", {}).items():
            if str(qualified_sign) in {"✔", "✖"}:
                secid_to_qualified_sign[str(secid)] = str(qualified_sign)

        for secid, bond_type in payload.get("secid_to_bond_type", {}).items():
            if bond_type:
                secid_to_bond_type[str(secid)] = str(bond_type)

        for secid, coupon_period in payload.get("secid_to_coupon_period", {}).items():
            try:
                secid_to_coupon_period[str(secid)] = int(coupon_period)
            except (TypeError, ValueError):
                continue

        for secid, is_structural in payload.get("secid_to_is_structural", {}).items():
            secid_to_is_structural[str(secid)] = bool(is_structural)

        for secid, issuer_rating in payload.get("secid_to_issuer_rating", {}).items():
            if issuer_rating:
                secid_to_issuer_rating[str(secid)] = str(issuer_rating)

        logging.info(
            "Найден checkpoint обогащения: SECID=%s, EMITTER_ID=%s.",
            len(secid_to_emitter_id),
            len(emitter_id_to_name),
        )
        return secid_to_emitter_id, emitter_id_to_name, emitter_id_to_inn, secid_to_qualified_sign, secid_to_bond_type, secid_to_coupon_period, secid_to_is_structural, secid_to_issuer_rating
    except Exception as exc:
        logging.warning("Не удалось прочитать checkpoint эмитентов: %s", exc)
        return {}, {}, {}, {}, {}, {}, {}, {}


def save_issuer_checkpoint(
    secid_to_emitter_id: dict[str, int | None],
    emitter_id_to_name: dict[int, str],
    emitter_id_to_inn: dict[int, str],
    secid_to_qualified_sign: dict[str, str],
    secid_to_bond_type: dict[str, str],
    secid_to_coupon_period: dict[str, int],
    secid_to_is_structural: dict[str, bool],
    secid_to_issuer_rating: dict[str, str],
) -> None:
    """Сохраняет checkpoint обогащения эмитентов после каждого пакета."""
    payload = {
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "secid_to_emitter_id": secid_to_emitter_id,
        "secid_to_qualified_sign": secid_to_qualified_sign,
        "secid_to_bond_type": secid_to_bond_type,
        "secid_to_coupon_period": secid_to_coupon_period,
        "secid_to_is_structural": secid_to_is_structural,
        "secid_to_issuer_rating": secid_to_issuer_rating,
        "emitter_id_to_name": {str(k): v for k, v in emitter_id_to_name.items()},
        "emitter_id_to_inn": {str(k): v for k, v in emitter_id_to_inn.items()},
    }
    ISSUER_CHECKPOINT_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def clear_issuer_checkpoint() -> None:
    """Удаляет checkpoint после успешного завершения этапа обогащения."""
    if ISSUER_CHECKPOINT_FILE.exists():
        ISSUER_CHECKPOINT_FILE.unlink()


def load_daily_security_cache() -> tuple[int, dict[str, dict[str, Any]]]:
    """Читает суточный кэш по полям, которые нужно обновлять раз в день."""
    if not DAILY_CACHE_FILE.exists():
        return DAILY_METRICS_CACHE_SCHEMA_VERSION, {}

    try:
        payload = json.loads(DAILY_CACHE_FILE.read_text(encoding="utf-8"))
        schema_version = int(payload.get("schema_version", 1))
        return schema_version, payload.get("secid_to_metrics", {})
    except Exception as exc:
        logging.warning("Не удалось прочитать суточный кэш бумаг: %s", exc)
        return DAILY_METRICS_CACHE_SCHEMA_VERSION, {}


def save_daily_security_cache(secid_to_metrics: dict[str, dict[str, Any]]) -> None:
    """Сохраняет суточный кэш по полям бумаг (амортизация и НКД)."""
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "schema_version": DAILY_METRICS_CACHE_SCHEMA_VERSION,
        "secid_to_metrics": secid_to_metrics,
    }
    DAILY_CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_offer_verification_cache() -> dict[str, dict[str, Any]]:
    """Читает 7-дневный кэш проверки оферт через Corpbonds."""
    if not OFFER_VERIFICATION_CACHE_FILE.exists():
        return {}

    try:
        payload = json.loads(OFFER_VERIFICATION_CACHE_FILE.read_text(encoding="utf-8"))
        schema_version = int(payload.get("schema_version", 1))
        if schema_version != OFFER_VERIFICATION_CACHE_SCHEMA_VERSION:
            logging.info(
                "Обнаружен старый формат кэша оферт (v%s). Старый кэш будет стёрт и собран заново.",
                schema_version,
            )
            OFFER_VERIFICATION_CACHE_FILE.unlink(missing_ok=True)
            return {}
        rows = payload.get("rows", {})
        if isinstance(rows, dict):
            return {str(isin): value for isin, value in rows.items() if isinstance(value, dict)}
    except Exception as exc:
        logging.warning("Не удалось прочитать кэш проверки оферт: %s", exc)
    return {}


def save_offer_verification_cache(rows: dict[str, dict[str, Any]]) -> None:
    """Сохраняет 7-дневный кэш проверки оферт через Corpbonds."""
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "schema_version": OFFER_VERIFICATION_CACHE_SCHEMA_VERSION,
        "rows": rows,
    }
    OFFER_VERIFICATION_CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_coupon_formula_cache() -> dict[str, dict[str, Any]]:
    """Читает кэш расчётов COUPONPERCENT для бумаг с нулевым COUPONVALUE."""
    if not COUPON_FORMULA_CACHE_FILE.exists():
        return {}

    try:
        payload = json.loads(COUPON_FORMULA_CACHE_FILE.read_text(encoding="utf-8"))
        rows = payload.get("rows", {})
        if isinstance(rows, dict):
            return {str(k): v for k, v in rows.items() if isinstance(v, dict)}
    except Exception as exc:
        logging.warning("Не удалось прочитать кэш формул купона: %s", exc)
    return {}


def save_coupon_formula_cache(rows: dict[str, dict[str, Any]]) -> None:
    """Сохраняет кэш расчётных купонных ставок по ISIN."""
    CACHE_DIR.mkdir(exist_ok=True)
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
    }
    COUPON_FORMULA_CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def load_coupon_value_cache() -> dict[str, dict[str, Any]]:
    """Читает кэш расчётов COUPONVALUE для бумаг с нулевым купоном от MOEX."""
    if not COUPON_VALUE_CACHE_FILE.exists():
        return {}

    try:
        payload = json.loads(COUPON_VALUE_CACHE_FILE.read_text(encoding="utf-8"))
        rows = payload.get("rows", {})
        if isinstance(rows, dict):
            return {str(k): v for k, v in rows.items() if isinstance(v, dict)}
    except Exception as exc:
        logging.warning("Не удалось прочитать кэш COUPONVALUE: %s", exc)
    return {}


def save_coupon_value_cache(rows: dict[str, dict[str, Any]]) -> None:
    """Сохраняет кэш расчётных значений COUPONVALUE."""
    CACHE_DIR.mkdir(exist_ok=True)
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
    }
    COUPON_VALUE_CACHE_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def calculate_coupon_value(face_value: float, coupon_percent: float, coupon_period: int) -> float:
    """Считает COUPONVALUE по формуле пользователя: FACEVALUE*COUPONPERCENT/100/365*COUPONPERIOD."""
    return (face_value * coupon_percent / 100 / 365) * coupon_period


def enrich_coupon_value_from_percent(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], set[tuple[str, str]]]:
    """Если COUPONVALUE=0 и есть COUPONPERCENT, рассчитывает COUPONVALUE и подсвечивает его в Excel."""
    logging.info("Этап 5.7/8: Расчёт COUPONVALUE по COUPONPERCENT для бумаг с нулевым купоном...")
    cache = load_coupon_value_cache()
    calculated_cells: set[tuple[str, str]] = set()
    now = datetime.now()

    for row in rows:
        isin = str(row.get("ISIN") or "").strip()
        secid = str(row.get("SECID") or "").strip()
        if not isin:
            continue

        try:
            coupon_value = float(str(row.get("COUPONVALUE") or 0).replace(",", "."))
        except ValueError:
            coupon_value = 0.0
        if coupon_value != 0:
            continue

        try:
            coupon_percent = float(str(row.get("COUPONPERCENT") or 0).replace(",", "."))
        except ValueError:
            coupon_percent = 0.0
        if coupon_percent <= 0:
            continue

        try:
            face_value = float(str(row.get("FACEVALUE") or 0).replace(",", "."))
        except ValueError:
            face_value = 0.0
        try:
            coupon_period = int(float(str(row.get("COUPONPERIOD") or 0).replace(",", ".")))
        except ValueError:
            coupon_period = 0

        if face_value <= 0 or coupon_period <= 0:
            continue

        cache_entry = cache.get(isin)
        if cache_entry is not None:
            cached_percent = cache_entry.get("coupon_percent")
            cached_period = cache_entry.get("coupon_period")
            cached_face_value = cache_entry.get("face_value")
            cached_coupon_value = cache_entry.get("coupon_value")
            cached_secid = str(cache_entry.get("secid") or "")
            if (
                isinstance(cached_coupon_value, (int, float))
                and isinstance(cached_percent, (int, float))
                and isinstance(cached_period, int)
                and isinstance(cached_face_value, (int, float))
                and cached_secid == secid
                and math.isclose(float(cached_percent), coupon_percent, rel_tol=0, abs_tol=1e-10)
                and cached_period == coupon_period
                and math.isclose(float(cached_face_value), face_value, rel_tol=0, abs_tol=1e-10)
            ):
                row["COUPONVALUE"] = round(float(cached_coupon_value), 6)
                calculated_cells.add((isin, secid))
                continue

        calculated_value = calculate_coupon_value(face_value, coupon_percent, coupon_period)
        row["COUPONVALUE"] = round(calculated_value, 6)
        calculated_cells.add((isin, secid))
        cache[isin] = {
            "secid": secid,
            "coupon_value": round(calculated_value, 6),
            "coupon_percent": coupon_percent,
            "coupon_period": coupon_period,
            "face_value": face_value,
            "updated_at": now.isoformat(timespec="seconds"),
        }

    save_coupon_value_cache(cache)
    logging.info("COUPONVALUE рассчитан для %s выпусков.", len(calculated_cells))
    return rows, calculated_cells


def parse_date_safe(value: Any) -> datetime | None:
    """Преобразует дату формата YYYY-MM-DD в datetime или возвращает None."""
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def format_date_ddmmyyyy(value: Any) -> str:
    """Преобразует дату в формат ДД.ММ.ГГГГ; если не удалось — возвращает исходное значение строкой."""
    parsed = parse_date_safe(value)
    if parsed is None:
        return str(value or "")
    return parsed.strftime("%d.%m.%Y")


def filter_rows_by_maturity(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Оставляет только бумаги с погашением не раньше, чем через 1 год."""
    today = datetime.now().date()
    min_allowed_date = today + timedelta(days=365 * MIN_MATURITY_YEARS)
    filtered_rows: list[dict[str, Any]] = []
    skipped_count = 0

    for row in rows:
        maturity_date = parse_date_safe(row.get(MATURITY_DATE_COLUMN_NAME))
        if maturity_date is None:
            skipped_count += 1
            continue
        if maturity_date.date() < min_allowed_date:
            skipped_count += 1
            continue
        filtered_rows.append(row)

    logging.info(
        "Фильтр по сроку до погашения: исключено %s бумаг, оставлено %s.",
        skipped_count,
        len(filtered_rows),
    )
    return filtered_rows


def chunked(items: list[Any], size: int) -> list[list[Any]]:
    """Разбивает список на пакеты фиксированного размера."""
    return [items[idx : idx + size] for idx in range(0, len(items), size)]


def is_structural_bond_type(raw_bond_type: str) -> bool:
    """Определяет, что тип облигации относится к структурным."""
    value = str(raw_bond_type or "").strip().lower()
    if not value:
        return False
    return value in STRUCTURAL_BOND_TYPE_VALUES or ("структур" in value and "облигац" in value)


def normalize_bond_type(raw_bond_type: str) -> str:
    """Нормализует тип облигации и убирает значения, которые пользователю не нужны."""
    value = str(raw_bond_type or "").strip()
    if not value:
        return "Не указан"
    if is_structural_bond_type(value):
        return "Не указан"
    return value


def coupon_period_from_frequency(raw_frequency: Any) -> int:
    """Пересчитывает COUPONFREQUENCY в примерный период купона в днях."""
    try:
        frequency = int(raw_frequency)
    except (TypeError, ValueError):
        return 0
    if frequency <= 0:
        return 0
    return round(365 / frequency)


def build_session() -> requests.Session:
    """Создаёт HTTP-сессию с ретраями при временных ошибках сети/API."""
    session = requests.Session()
    retry = Retry(
        total=MAX_RETRIES,
        connect=MAX_RETRIES,
        read=MAX_RETRIES,
        backoff_factor=BACKOFF_FACTOR,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=frozenset(["GET"]),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session



def fetch_latest_table_rate(session: requests.Session, url: str) -> float | None:
    """Возвращает первую ставку из HTML-таблицы ЦБ РФ (верхняя строка = самая свежая)."""
    response = session.get(url, timeout=REQUEST_TIMEOUT, headers={"User-Agent": DOHOD_USER_AGENT})
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.find("table", class_="data")
    if table is None:
        return None

    rows = table.find_all("tr")
    for row in rows[1:]:
        cells = row.find_all("td")
        if len(cells) < 2:
            continue
        raw_value = cells[1].get_text(" ", strip=True).replace(" ", " ")
        rate_match = re.search(r"-?\d+[,.]?\d*", raw_value)
        if not rate_match:
            continue
        return float(rate_match.group(0).replace(",", "."))
    return None


def fetch_g_curve_value(session: requests.Session, period_years: float) -> float | None:
    """Возвращает значение кривой бескупонной доходности ОФЗ для указанного срока в годах."""
    response = session.get(
        "https://iss.moex.com/iss/engines/stock/zcyc/securities.json",
        params={"iss.meta": "off"},
        timeout=REQUEST_TIMEOUT,
    )
    response.raise_for_status()
    data = response.json()
    rows = data.get("yearyields", {}).get("data", [])
    columns = data.get("yearyields", {}).get("columns", [])
    yields = [dict(zip(columns, row)) for row in rows]
    if not yields:
        return None

    exact_value: float | None = None
    nearest_value: tuple[float, float] | None = None
    for item in yields:
        try:
            period = float(item.get("period"))
            value = float(item.get("value"))
        except (TypeError, ValueError):
            continue

        if abs(period - period_years) < 1e-9:
            exact_value = value
            break

        distance = abs(period - period_years)
        if nearest_value is None or distance < nearest_value[0]:
            nearest_value = (distance, value)

    if exact_value is not None:
        return exact_value

    if nearest_value is not None:
        logging.warning(
            "G_CURVE_RUS: точный срок %.2f лет не найден, использовано ближайшее значение.",
            period_years,
        )
        return nearest_value[1]

    return None


def extract_dohod_section_value(html: str, label: str) -> str:
    """Извлекает текст значения из блока ДОХОД по заголовку секции."""
    soup = BeautifulSoup(html, "html.parser")
    for p_tag in soup.find_all("p"):
        label_text = re.sub(r"\s+", " ", p_tag.get_text(" ", strip=True))
        if label not in label_text:
            continue

        info_block = p_tag.find_parent("div", class_=re.compile("dohod-description-info"))
        if info_block is None:
            continue

        data_tag = info_block.find("p", class_=re.compile("dohod-description-info_data"))
        if data_tag is None:
            continue

        value = re.sub(r"\s+", " ", data_tag.get_text(" ", strip=True)).strip()
        if value:
            return value

    pattern = (
        rf"<p[^>]*>\s*{re.escape(label)}\s*</p>"
        r'.*?<p[^>]*class="[^"]*dohod-description-info_data[^"]*"[^>]*>(.*?)</p>'
    )
    match = re.search(pattern, html, flags=re.IGNORECASE | re.DOTALL)
    if not match:
        return ""
    raw_value = match.group(1)
    text_value = BeautifulSoup(raw_value, "html.parser").get_text(" ", strip=True)
    return re.sub(r"\s+", " ", text_value).strip()


def fetch_dohod_page_html(session: requests.Session, isin: str) -> str:
    """Возвращает HTML карточки облигации ДОХОД с дополнительными повторами запроса."""
    last_error: Exception | None = None
    url = DOHOD_BOND_URL.format(isin=isin)

    for attempt in range(1, DOHOD_REQUEST_RETRIES + 1):
        try:
            response = session.get(
                url,
                timeout=REQUEST_TIMEOUT,
                headers={"User-Agent": DOHOD_USER_AGENT},
            )
            response.raise_for_status()
            html = response.text or ""
            if len(html) < 500:
                raise ValueError(f"Пустой/слишком короткий HTML (длина={len(html)})")
            return html
        except Exception as exc:  # noqa: PERF203
            last_error = exc
            logging.warning("ДОХОД: попытка %s/%s для %s не удалась: %s", attempt, DOHOD_REQUEST_RETRIES, isin, exc)
            time.sleep(0.4 * attempt)

    raise RuntimeError(f"Не удалось загрузить страницу ДОХОД для {isin}: {last_error}")


def fetch_dohod_offer_page_html(session: requests.Session, isin: str) -> str:
    """Быстрый запрос страницы ДОХОД для проверки оферт (укороченные таймауты/ретраи)."""
    last_error: Exception | None = None
    url = DOHOD_BOND_URL.format(isin=isin)

    for attempt in range(1, OFFER_CHECK_REQUEST_RETRIES + 1):
        try:
            response = session.get(
                url,
                timeout=OFFER_CHECK_REQUEST_TIMEOUT,
                headers={"User-Agent": DOHOD_USER_AGENT},
            )
            response.raise_for_status()
            html = response.text or ""
            if len(html) < 500:
                raise ValueError(f"Пустой/слишком короткий HTML (длина={len(html)})")
            return html
        except Exception as exc:  # noqa: PERF203
            last_error = exc
            if attempt < OFFER_CHECK_REQUEST_RETRIES:
                time.sleep(0.25 * attempt)

    raise RuntimeError(f"Не удалось быстро загрузить страницу ДОХОД для {isin}: {last_error}")


def parse_dohod_formula(session: requests.Session, isin: str) -> tuple[str, str, str, float, float]:
    """Парсит формулу с analytics.dohod.ru: индекс, сдвиг, описание и срок G-curve (если нужен)."""
    html = fetch_dohod_page_html(session, isin)

    formula_text = extract_dohod_section_value(html, "Привязка к индексу")
    description_text = extract_dohod_section_value(html, "Описание формулы изменяемого купона/номинала")

    if not formula_text:
        fallback = re.search(r"(CBR_RATE|RUONIA|G_CURVE_RUS)\s*[+-]\s*\d+[.,]?\d*", html, flags=re.IGNORECASE)
        if fallback:
            formula_text = re.sub(r"\s+", " ", fallback.group(0)).strip()

    formula_match = re.search(r"(CBR_RATE|RUONIA|G_CURVE_RUS)\s*([+-])?\s*(\d+[.,]?\d*)?", formula_text, flags=re.IGNORECASE)
    if not formula_match:
        raise ValueError(f"Не удалось распознать формулу ДОХОД для {isin}: '{formula_text}'")

    index_name = formula_match.group(1).upper()
    operation = formula_match.group(2) or "+"
    spread = float((formula_match.group(3) or "0").replace(",", "."))
    if operation == "-":
        spread *= -1

    g_curve_years = 7.0
    if index_name == "G_CURVE_RUS":
        years_match = re.search(r"сроком\s+погашения\s+(\d+[.,]?\d*)\s*(?:лет|года|год)", description_text, flags=re.IGNORECASE)
        if years_match:
            g_curve_years = float(years_match.group(1).replace(",", "."))

    return index_name, formula_text, description_text, spread, g_curve_years


def fetch_corpbonds_page_html(session: requests.Session, identifier: str) -> str:
    """Возвращает HTML карточки облигации corpbonds.ru с повторами запроса."""
    last_error: Exception | None = None
    url = CORPBONDS_BOND_URL.format(isin=identifier)

    for attempt in range(1, DOHOD_REQUEST_RETRIES + 1):
        try:
            response = session.get(
                url,
                timeout=REQUEST_TIMEOUT,
                headers={"User-Agent": DOHOD_USER_AGENT},
            )
            response.raise_for_status()
            html = response.text or ""
            if len(html) < 500:
                raise ValueError(f"Пустой/слишком короткий HTML (длина={len(html)})")
            return html
        except Exception as exc:  # noqa: PERF203
            last_error = exc
            status_code = getattr(getattr(exc, "response", None), "status_code", None)
            logging.warning("CORPBONDS: попытка %s/%s для %s не удалась: %s", attempt, DOHOD_REQUEST_RETRIES, identifier, exc)
            if status_code == 404:
                break
            time.sleep(0.4 * attempt)

    raise RuntimeError(f"Не удалось загрузить страницу CORPBONDS для {identifier}: {last_error}")


def fetch_corpbonds_offer_page_html(session: requests.Session, identifier: str) -> str:
    """Быстрый запрос страницы Corpbonds для проверки оферт."""
    last_error: Exception | None = None
    url = CORPBONDS_BOND_URL.format(isin=identifier)

    for attempt in range(1, OFFER_CHECK_REQUEST_RETRIES + 1):
        try:
            response = session.get(
                url,
                timeout=OFFER_CHECK_REQUEST_TIMEOUT,
                headers={"User-Agent": DOHOD_USER_AGENT},
            )
            response.raise_for_status()
            html = response.text or ""
            if len(html) < 500:
                raise ValueError(f"Пустой/слишком короткий HTML (длина={len(html)})")
            return html
        except Exception as exc:  # noqa: PERF203
            last_error = exc
            status_code = getattr(getattr(exc, "response", None), "status_code", None)
            if status_code == 404:
                break
            if attempt < OFFER_CHECK_REQUEST_RETRIES:
                time.sleep(0.25 * attempt)

    raise RuntimeError(f"Не удалось быстро загрузить страницу CORPBONDS для {identifier}: {last_error}")


def parse_corpbonds_formula(session: requests.Session, isin: str) -> tuple[str, str, str, float, float]:
    """Парсит формулу купона с corpbonds.ru, если ДОХОД не дал результата."""
    html = fetch_corpbonds_page_html(session, isin)
    soup = BeautifulSoup(html, "html.parser")

    text_chunks: list[str] = []
    for node in soup.find_all(["h1", "h2", "h3", "div", "span", "p", "td", "th", "li"]):
        text = re.sub(r"\s+", " ", node.get_text(" ", strip=True)).strip()
        if text:
            text_chunks.append(text)

    full_text = " | ".join(text_chunks)
    formula_match = re.search(
        r"(?:∑|Σ|SUM)?\s*(KC|КС|KEYRATE|CBR_RATE|RUONIA|G_CURVE_RUS)\s*([+-])\s*(\d+[.,]?\d*)\s*%",
        full_text,
        flags=re.IGNORECASE,
    )
    if formula_match is None:
        raise ValueError(f"Не удалось распознать формулу CORPBONDS для {isin}")

    source_index = formula_match.group(1).upper()
    index_name = CORPBONDS_INDEX_ALIASES.get(source_index, source_index)
    operation = formula_match.group(2)
    spread = float(formula_match.group(3).replace(",", "."))
    if operation == "-":
        spread *= -1

    formula_text = re.sub(r"\s+", " ", formula_match.group(0)).strip()
    g_curve_years = 7.0
    if index_name == "G_CURVE_RUS":
        years_match = re.search(r"(\d+[.,]?\d*)\s*(?:лет|года|год|y|yr)", full_text, flags=re.IGNORECASE)
        if years_match:
            g_curve_years = float(years_match.group(1).replace(",", "."))

    return index_name, formula_text, "Источник corpbonds.ru", spread, g_curve_years


def normalize_offer_type(value: str | None) -> str:
    """Нормализует тип оферты к одному из значений: PUT, Call, ✖."""
    lowered = str(value or "").strip().lower()
    if not lowered:
        return "✖"
    if any(keyword in lowered for keyword in ("погаш", "maturity", "redemption", "амортиз")):
        return "✖"
    if "put" in lowered or "продат" in lowered:
        return "PUT"
    if "call" in lowered or "выкуп" in lowered:
        return "Call"
    if "оферт" in lowered:
        return "PUT"
    return "✖"


def extract_offer_date_from_text(text: str) -> str | None:
    """Извлекает дату оферты из текста в формате YYYY-MM-DD, если удаётся."""
    match = re.search(r"(\d{2}[.]\d{2}[.]\d{4})", text)
    if not match:
        return None

    day, month, year = match.group(1).split(".")
    try:
        parsed = datetime(year=int(year), month=int(month), day=int(day))
    except ValueError:
        return None
    return parsed.strftime("%Y-%m-%d")


def parse_dohod_offer_metrics(session: requests.Session, isin: str) -> tuple[str, str | None]:
    """Парсит тип и дату оферты со страницы ДОХОД."""
    html = fetch_dohod_offer_page_html(session, isin)
    soup = BeautifulSoup(html, "html.parser")
    full_text = re.sub(r"\s+", " ", soup.get_text(" ", strip=True)).strip()

    offer_type = "✖"
    event_match = re.search(r"Событие\s+в\s+ближ\.\s*дату\s*[:\-]?\s*([^|]+)", full_text, flags=re.IGNORECASE)
    if event_match:
        offer_type = normalize_offer_type(event_match.group(1))

    offer_date = None
    for label in (
        "Дата ближайшей оферты",
    ):
        section_match = re.search(rf"{label}\s*[:\-]?\s*([^|]+)", full_text, flags=re.IGNORECASE)
        if not section_match:
            continue
        offer_date = extract_offer_date_from_text(section_match.group(1))
        if offer_date:
            break

    return offer_type, offer_date


def parse_corpbonds_offer_metrics(session: requests.Session, isin: str, secid: str | None = None) -> tuple[str, str | None, str]:
    """Парсит тип и дату оферты со страницы Corpbonds (с fallback: ISIN -> SECID)."""

    def parse_html(html: str) -> tuple[str, str | None]:
        soup = BeautifulSoup(html, "html.parser")
        lines = [re.sub(r"\s+", " ", line).strip() for line in soup.get_text("\n", strip=True).split("\n")]
        lines = [line for line in lines if line]
        full_text = " | ".join(lines)

        offer_date = None

        def find_nearby_date(start_idx: int) -> str | None:
            direct_date = extract_offer_date_from_text(lines[start_idx])
            if direct_date:
                return direct_date
            for shift in range(1, 8):
                probe_idx = start_idx + shift
                if probe_idx >= len(lines):
                    break
                candidate = extract_offer_date_from_text(lines[probe_idx])
                if candidate:
                    return candidate
            return None

        for idx, line in enumerate(lines):
            if line.strip().lower() != "ближайшая дата":
                continue
            offer_date = find_nearby_date(idx)
            if offer_date:
                break

        if offer_date is None:
            for idx, line in enumerate(lines):
                if "дата ближайшей оферты" not in line.lower():
                    continue
                offer_date = find_nearby_date(idx)
                if offer_date:
                    break

        offer_type = "✖"
        for idx, line in enumerate(lines):
            if "наличие call-опциона" not in line.lower():
                continue
            tail = re.split(r"наличие\s+call-опциона", line, flags=re.IGNORECASE, maxsplit=1)
            call_value = tail[1] if len(tail) > 1 else ""
            if not call_value and idx + 1 < len(lines):
                call_value = lines[idx + 1]
            if re.search(r"\bда\b", call_value.lower()):
                offer_type = "Call"
            break

        if offer_type == "✖" and re.search(r"\bput\b|право\s+продать", full_text, flags=re.IGNORECASE):
            offer_type = "PUT"

        return offer_type, offer_date

    try:
        return (*parse_html(fetch_corpbonds_offer_page_html(session, isin)), "isin")
    except Exception as exc:
        status_code = getattr(getattr(exc, "response", None), "status_code", None)
        is_not_found = status_code == 404 or "404" in str(exc)
        if not is_not_found or not secid:
            raise

    logging.info("CORPBONDS: для %s не найдено по ISIN, пробуем SECID=%s.", isin, secid)
    return (*parse_html(fetch_corpbonds_offer_page_html(session, secid)), "secid")


def merge_offer_metrics(dohod_type: str, dohod_date: str | None, corpbonds_type: str, corpbonds_date: str | None) -> tuple[str, str | None]:
    """Объединяет данные оферты из ДОХОД и Corpbonds в единый результат."""
    offer_type = "✖"

    if dohod_type == "PUT" or corpbonds_type == "PUT":
        offer_type = "PUT"
    elif dohod_type == "Call" or corpbonds_type == "Call":
        offer_type = "Call"

    if offer_type == "✖":
        return "✖", None

    candidate_dates = [date for date in [dohod_date, corpbonds_date] if parse_date_safe(date)]
    if not candidate_dates:
        return "✖", None

    nearest = min(candidate_dates, key=lambda value: parse_date_safe(value) or datetime.max)
    return offer_type, nearest


def fetch_reference_index_values(session: requests.Session) -> dict[str, float]:
    """Собирает актуальные значения индексов: CBR_RATE, RUONIA, G_CURVE_RUS(7Y)."""
    index_values: dict[str, float] = {}
    index_values["CBR_RATE"] = fetch_latest_table_rate(
        session,
        "https://www.cbr.ru/hd_base/KeyRate/?UniDbQuery.Posted=True&UniDbQuery.From=01.01.2020&UniDbQuery.To=31.12.2035",
    ) or 0.0
    index_values["RUONIA"] = fetch_latest_table_rate(
        session,
        "https://www.cbr.ru/hd_base/ruonia/dynamics/?UniDbQuery.Posted=True&UniDbQuery.From=01.01.2020&UniDbQuery.To=31.12.2035",
    ) or 0.0
    g_curve_default = fetch_g_curve_value(session, 7.0)
    index_values["G_CURVE_RUS_7.0"] = g_curve_default or 0.0
    return index_values


def normalize_coupon_formula_source(source_name: str | None) -> str:
    """Преобразует технический код источника формулы в понятную подпись для Excel."""
    source = str(source_name or "").strip().lower()
    if source == "dohod":
        return "DOHOD"
    if source == "corpbonds":
        return "CORPBONDS"
    return ""


def enrich_coupon_percent_from_dohod(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], set[tuple[str, str]]]:
    """Считает COUPONPERCENT через формулы ДОХОД, затем fallback через corpbonds.ru."""
    logging.info("Этап 5.6/8: Расчёт COUPONPERCENT через ДОХОД + fallback CORPBONDS для выпусков с нулевым купоном...")
    calculated_cells: set[tuple[str, str]] = set()
    for row in rows:
        row[COUPON_FORMULA_SOURCE_COLUMN_NAME] = ""

    cache = load_coupon_formula_cache()
    now = datetime.now()

    stale_keys: list[str] = []
    for isin, cached_entry in cache.items():
        try:
            cached_at = datetime.fromisoformat(str(cached_entry.get("updated_at")))
        except ValueError:
            stale_keys.append(isin)
            continue
        if now - cached_at > timedelta(days=COUPON_FORMULA_CACHE_TTL_DAYS):
            stale_keys.append(isin)
    for isin in stale_keys:
        cache.pop(isin, None)

    candidates: list[dict[str, Any]] = []
    for row in rows:
        try:
            coupon_value = float(str(row.get("COUPONVALUE") or 0).replace(",", "."))
        except ValueError:
            coupon_value = 0.0
        if coupon_value == 0:
            candidates.append(row)

    if not candidates:
        logging.info("Нет выпусков с COUPONVALUE=0. Доп. расчёт не требуется.")
        return rows, calculated_cells

    logging.info("Найдено выпусков с COUPONVALUE=0: %s", len(candidates))

    with build_session() as session:
        index_values = fetch_reference_index_values(session)

        to_fetch: list[tuple[str, str]] = []
        for row in candidates:
            isin = str(row.get("ISIN") or "").strip()
            secid = str(row.get("SECID") or "").strip()
            if not isin:
                continue

            try:
                moex_coupon = float(str(row.get("COUPONPERCENT") or 0).replace(",", "."))
            except ValueError:
                moex_coupon = 0.0

            if moex_coupon > 0:
                cache.pop(isin, None)
                continue

            cached_entry = cache.get(isin)
            if cached_entry is not None:
                try:
                    cached_at = datetime.fromisoformat(str(cached_entry.get("updated_at")))
                except ValueError:
                    cached_at = now - timedelta(days=COUPON_FORMULA_CACHE_TTL_DAYS + 1)
                if now - cached_at <= timedelta(days=COUPON_FORMULA_CACHE_TTL_DAYS):
                    coupon_percent = cached_entry.get("coupon_percent")
                    if isinstance(coupon_percent, (float, int)):
                        row["COUPONPERCENT"] = round(float(coupon_percent), 4)
                        row[COUPON_FORMULA_SOURCE_COLUMN_NAME] = normalize_coupon_formula_source(cached_entry.get("source"))
                        calculated_cells.add((isin, secid))
                        continue

            to_fetch.append((isin, secid))

        formula_by_isin: dict[str, tuple[str, str, str, float, float, str]] = {}

        def resolve_formula_job(job: tuple[str, str]) -> tuple[str, str, tuple[str, str, str, float, float] | None, str | None, str]:
            isin_job, secid_job = job
            with build_session() as local_session:
                try:
                    parsed = parse_dohod_formula(local_session, isin_job)
                    return isin_job, secid_job, parsed, None, "dohod"
                except Exception as dohod_exc:
                    try:
                        parsed = parse_corpbonds_formula(local_session, isin_job)
                        return isin_job, secid_job, parsed, None, "corpbonds"
                    except Exception as corpbonds_exc:
                        error_text = f"dohod={dohod_exc}; corpbonds={corpbonds_exc}"
                        return isin_job, secid_job, None, error_text, "none"

        if to_fetch:
            logging.info(
                "ДОХОД/CORPBONDS: нужно опросить формулы для %s выпусков (многопоточно, потоков=%s).",
                len(to_fetch),
                DOHOD_MAX_WORKERS,
            )

            with ThreadPoolExecutor(max_workers=DOHOD_MAX_WORKERS) as executor:
                futures = [executor.submit(resolve_formula_job, job) for job in to_fetch]
                for processed, future in enumerate(as_completed(futures), start=1):
                    isin_job, _secid_job, parsed, error_text, source_name = future.result()
                    if parsed is not None:
                        formula_by_isin[isin_job] = (*parsed, source_name)
                        if source_name == "corpbonds":
                            logging.info("CORPBONDS: формула для %s получена через fallback после ДОХОД.", isin_job)
                    else:
                        logging.warning("ДОХОД/CORPBONDS: не удалось получить формулу для %s: %s", isin_job, error_text)

                    if processed % 25 == 0 or processed == len(futures):
                        logging.info("ДОХОД/CORPBONDS: обработано формул %s/%s", processed, len(futures))

        g_curve_cache: dict[float, float] = {}

        for row in candidates:
            isin = str(row.get("ISIN") or "").strip()
            secid = str(row.get("SECID") or "").strip()
            if not isin:
                continue

            try:
                moex_coupon = float(str(row.get("COUPONPERCENT") or 0).replace(",", "."))
            except ValueError:
                moex_coupon = 0.0
            if moex_coupon > 0:
                continue

            cached_entry = cache.get(isin)
            if cached_entry is not None:
                try:
                    cached_at = datetime.fromisoformat(str(cached_entry.get("updated_at")))
                except ValueError:
                    cached_at = now - timedelta(days=COUPON_FORMULA_CACHE_TTL_DAYS + 1)
                if now - cached_at <= timedelta(days=COUPON_FORMULA_CACHE_TTL_DAYS):
                    coupon_percent = cached_entry.get("coupon_percent")
                    if isinstance(coupon_percent, (float, int)):
                        row["COUPONPERCENT"] = round(float(coupon_percent), 4)
                        row[COUPON_FORMULA_SOURCE_COLUMN_NAME] = normalize_coupon_formula_source(cached_entry.get("source"))
                        calculated_cells.add((isin, secid))
                        continue

            parsed = formula_by_isin.get(isin)
            if parsed is None:
                continue

            index_name, formula_text, description_text, spread, g_curve_years, source_name = parsed

            if index_name == "G_CURVE_RUS":
                if g_curve_years in g_curve_cache:
                    base_value = g_curve_cache[g_curve_years]
                else:
                    g_curve_value = fetch_g_curve_value(session, g_curve_years)
                    if g_curve_value is None:
                        logging.warning("ДОХОД: нет значения G_CURVE_RUS для %s (%.2f лет)", isin, g_curve_years)
                        continue
                    base_value = g_curve_value
                    g_curve_cache[g_curve_years] = g_curve_value
                    index_values[f"G_CURVE_RUS_{g_curve_years:.2f}"] = g_curve_value
            else:
                base_value = index_values.get(index_name, 0.0)

            coupon_percent = round(base_value + spread, 4)
            row["COUPONPERCENT"] = coupon_percent
            row[COUPON_FORMULA_SOURCE_COLUMN_NAME] = normalize_coupon_formula_source(source_name)
            calculated_cells.add((isin, secid))
            cache[isin] = {
                "secid": secid,
                "coupon_percent": coupon_percent,
                "index_name": index_name,
                "index_value": round(base_value, 4),
                "spread": spread,
                "formula": formula_text,
                "description": description_text,
                "g_curve_years": g_curve_years,
                "source": source_name,
                "updated_at": now.isoformat(timespec="seconds"),
            }

    save_coupon_formula_cache(cache)
    logging.info("COUPONPERCENT рассчитан через ДОХОД/CORPBONDS для %s выпусков.", len(calculated_cells))
    return rows, calculated_cells


def fetch_page(session: requests.Session, start: int) -> IssPage:
    """Запрашивает одну страницу облигаций с MOEX ISS."""
    params = {
        "iss.meta": "off",
        "securities.columns": "SECID,SHORTNAME,ISIN,MATDATE,FACEVALUE,FACEUNIT,COUPONVALUE,COUPONPERIOD,COUPONPERCENT,PRIMARYBOARDID,PREVLEGALCLOSEPRICE,PREVPRICE,ACCRUEDINT,SECTORID,INSTRID",
        "marketdata.columns": f"SECID,{TRADE_VOLUME_COLUMN_NAME},{TRADE_VALUE_COLUMN_NAME},NUMTRADES,{YIELD_COLUMN_NAME}",
        "start": start,
        "limit": PAGE_SIZE,
    }
    response = session.get(BASE_URL, params=params, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    data = response.json()
    security_rows = data.get("securities", {}).get("data", [])
    security_columns = data.get("securities", {}).get("columns", [])
    market_rows = data.get("marketdata", {}).get("data", [])
    market_columns = data.get("marketdata", {}).get("columns", [])

    market_by_secid: dict[str, dict[str, Any]] = {}
    for market_row in market_rows:
        market_record = dict(zip(market_columns, market_row))
        secid = str(market_record.get("SECID") or "").strip()
        if secid:
            market_by_secid[secid] = market_record

    records: list[dict[str, Any]] = []
    for security_row in security_rows:
        security_record = dict(zip(security_columns, security_row))
        secid = str(security_record.get("SECID") or "").strip()
        market_record = market_by_secid.get(secid, {})
        security_record[TRADE_VOLUME_COLUMN_NAME] = market_record.get(TRADE_VOLUME_COLUMN_NAME)
        security_record[TRADE_VALUE_COLUMN_NAME] = market_record.get(TRADE_VALUE_COLUMN_NAME)
        security_record["NUMTRADES"] = market_record.get("NUMTRADES")
        security_record[YIELD_COLUMN_NAME] = market_record.get(YIELD_COLUMN_NAME)
        records.append(security_record)

    return IssPage(start=start, rows=records)


def fetch_all_bonds() -> list[dict[str, Any]]:
    """Собирает облигации с MOEX ISS API."""
    logging.info("Этап 1/8: Запрос данных с MOEX ISS...")

    with build_session() as session:
        unique: dict[tuple[Any, Any], dict[str, Any]] = {}
        start = 0
        page_number = 0
        while True:
            page = fetch_page(session, start)
            page_number += 1
            before_count = len(unique)
            for row in page.rows:
                unique[(row.get("SECID"), row.get("ISIN"))] = row
            added_count = len(unique) - before_count

            logging.info("Страница %s: получено %s записей, новых ключей %s.", page_number, len(page.rows), added_count)
            if len(page.rows) < PAGE_SIZE or added_count == 0:
                break
            start += PAGE_SIZE

        rows = list(unique.values())
        logging.info("Получено записей (все страницы): %s", len(rows))
        return rows


def extract_issuer_rating_from_description(description_rows: list[list[Any]]) -> str:
    """Пытается извлечь рейтинг эмитента из карточки MOEX (если биржа его передаёт)."""
    for row in description_rows:
        if len(row) < 3:
            continue
        name = str(row[0] or "").strip().upper()
        title = str(row[1] or "").strip().lower()
        value = str(row[2] or "").strip()
        if not value:
            continue

        name_has_rating = "RATING" in name or "RAT" in name
        title_has_rating = "рейтинг" in title or "rating" in title
        if name_has_rating or title_has_rating:
            return value

    return ""


def fetch_emitter_info_for_security(session: requests.Session, secid: str) -> tuple[int | None, str, str, int, bool, str]:
    """Возвращает ID эмитента, квалификацию, тип облигации и период купона по SECID."""
    params = {
        "iss.meta": "off",
        "iss.only": "description",
        "description.columns": "name,title,value",
    }
    response = session.get(f"https://iss.moex.com/iss/securities/{secid}.json", params=params, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    data = response.json()
    rows = data.get("description", {}).get("data", [])
    emitter_id: int | None = None
    qualified_investor_sign = "✖"
    bond_type = "Не указан"
    coupon_period = 0
    is_structural = False
    issuer_rating = extract_issuer_rating_from_description(rows)

    for name, _, value in rows:
        if name == "EMITTER_ID" and value is not None:
            try:
                emitter_id = int(value)
            except (TypeError, ValueError):
                emitter_id = None
        if name == "ISQUALIFIEDINVESTORS":
            qualified_investor_sign = "✔" if str(value) == "1" else "✖"
        if name == "BOND_TYPE":
            is_structural = is_structural_bond_type(str(value or ""))
            bond_type = normalize_bond_type(str(value or ""))
        if name == "COUPONFREQUENCY":
            coupon_period = coupon_period_from_frequency(value)

    return emitter_id, qualified_investor_sign, bond_type, coupon_period, is_structural, issuer_rating

def parse_offer_metrics(offers_data: list[list[Any]], offers_columns: list[str]) -> tuple[str, str | None]:
    """Определяет актуальную оферту MOEX: тип + ближайшая дата (только сегодня/будущее)."""
    if not offers_data or not offers_columns:
        return "✖", None

    today = datetime.now().date()
    active_offer_records: list[tuple[datetime, str]] = []

    for raw_offer in offers_data:
        offer = dict(zip(offers_columns, raw_offer))
        offer_type = normalize_offer_type(str(offer.get("offertype") or offer.get("offer_type") or ""))
        if offer_type == "✖":
            continue

        candidates: list[datetime] = []
        for date_key in ("offerdate", "offerdatestart", "offerdateend"):
            parsed = parse_date_safe(offer.get(date_key))
            if parsed is not None and parsed.date() >= today:
                candidates.append(parsed)

        if not candidates:
            continue

        active_offer_records.append((min(candidates), offer_type))

    if not active_offer_records:
        return "✖", None

    nearest_date, nearest_type = min(active_offer_records, key=lambda item: item[0])
    return nearest_type, nearest_date.strftime("%Y-%m-%d")


def fetch_daily_security_metrics(session: requests.Session, secid: str, maturity_date: str | None) -> tuple[str, str | None, str, str | None]:
    """Получает суточные метрики бумаги: амортизация + Put/Call оферта.

    Важно: финальное погашение (строка с data_source=maturity на дату MATDATE) не считаем амортизацией.
    """
    params = {"iss.meta": "off"}
    response = session.get(f"https://iss.moex.com/iss/securities/{secid}/bondization.json", params=params, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    data = response.json()
    amort_data = data.get("amortizations", {}).get("data", [])
    amort_columns = data.get("amortizations", {}).get("columns", [])
    offers_data = data.get("offers", {}).get("data", [])
    offers_columns = data.get("offers", {}).get("columns", [])
    has_put_call_offer, put_call_offer_date = parse_offer_metrics(offers_data, offers_columns)

    if not amort_data or not amort_columns:
        return "✖", None, has_put_call_offer, put_call_offer_date

    amort_records = [dict(zip(amort_columns, row)) for row in amort_data]

    real_amortization_dates: list[str] = []
    for record in amort_records:
        amort_date = str(record.get("amortdate") or "")
        if not amort_date:
            continue

        data_source = str(record.get("data_source") or "").lower()
        if data_source and data_source != "maturity":
            real_amortization_dates.append(amort_date)
            continue

        if maturity_date and amort_date < str(maturity_date):
            real_amortization_dates.append(amort_date)

    if not real_amortization_dates:
        return "✖", None, has_put_call_offer, put_call_offer_date

    return "✔", min(real_amortization_dates), has_put_call_offer, put_call_offer_date


def fetch_emitter_details(session: requests.Session, emitter_id: int) -> tuple[str | None, str | None]:
    """Возвращает наименование и ИНН эмитента по его ID."""
    params = {"iss.meta": "off"}
    response = session.get(f"https://iss.moex.com/iss/emitters/{emitter_id}.json", params=params, timeout=REQUEST_TIMEOUT)
    response.raise_for_status()
    data = response.json()
    emitter_rows = data.get("emitter", {}).get("data", [])
    if not emitter_rows:
        return None, None

    columns = data.get("emitter", {}).get("columns", [])
    record = dict(zip(columns, emitter_rows[0]))
    emitter_name = record.get("SHORT_TITLE") or record.get("TITLE")
    emitter_inn = record.get("INN")
    return emitter_name, str(emitter_inn) if emitter_inn else None


def validate_rows(rows: list[dict[str, Any]]) -> None:
    """Проверяет качество данных перед выгрузкой и пишет понятный отчёт в лог."""
    logging.info("Этап 5/8: Проверка качества данных...")

    empty_isin_count = sum(1 for row in rows if not row.get("ISIN"))
    duplicate_keys = len(rows) - len({(row.get("SECID"), row.get("ISIN")) for row in rows})
    invalid_coupon_count = sum(1 for row in rows if isinstance(row.get("COUPONPERCENT"), (int, float)) and row.get("COUPONPERCENT") < 0)
    bonds_with_offer_count = 0
    bonds_with_empty_offer_date_count = 0
    bonds_with_offer_on_maturity_count = 0

    for row in rows:
        has_offer_sign = str(row.get(HAS_PUT_CALL_OFFER_COLUMN_NAME) or "").strip()
        if has_offer_sign != "✔":
            continue

        bonds_with_offer_count += 1
        offer_date_raw = str(row.get(PUT_CALL_OFFER_DATE_COLUMN_NAME) or "").strip()
        maturity_date_raw = str(row.get(MATURITY_DATE_COLUMN_NAME) or "").strip()
        if not offer_date_raw:
            bonds_with_empty_offer_date_count += 1
            continue

        if maturity_date_raw and offer_date_raw == maturity_date_raw:
            bonds_with_offer_on_maturity_count += 1

    logging.info(
        "Проверка качества завершена: пустых ISIN=%s, дубликатов ключа (SECID+ISIN)=%s, отрицательных COUPONPERCENT=%s.",
        empty_isin_count,
        duplicate_keys,
        invalid_coupon_count,
    )
    logging.info(
        "Контроль оферт: бумаг с офертами=%s, из них с пустой датой оферты=%s, оферта совпадает с датой погашения=%s.",
        bonds_with_offer_count,
        bonds_with_empty_offer_date_count,
        bonds_with_offer_on_maturity_count,
    )


def resolve_issuer_bond_class(row: dict[str, Any]) -> str:
    """Возвращает тип бумаги/рынка (государственный, корпоративный и т.д.)."""
    sector_id = str(row.get("SECTORID") or "").strip()
    if sector_id:
        return sector_id

    instr_id = str(row.get("INSTRID") or "").strip().upper()
    if instr_id:
        return INSTRID_TO_BOND_CLASS.get(instr_id, f"INSTRID:{instr_id}")

    return "Не указан"



def enrich_with_issuer_names(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Добавляет в каждую строку имя/ИНН эмитента, квалификацию, тип и корректный COUPONPERIOD."""
    logging.info("Этап 5.1/8: Обогащение данных наименованиями эмитентов...")
    secids = sorted({str(row.get("SECID")) for row in rows if row.get("SECID")})

    if not secids:
        for row in rows:
            row[ISSUER_COLUMN_NAME] = ""
            row[ISSUER_INN_COLUMN_NAME] = ""
            row[ISSUER_BOND_CLASS_COLUMN_NAME] = "Не указан"
            row[ISSUER_RATING_COLUMN_NAME] = ""
            row[QUALIFIED_INVESTOR_COLUMN_NAME] = "✖"
            row[BOND_TYPE_COLUMN_NAME] = "Не указан"
        return rows

    cache_secid_to_emitter_id, cache_emitter_id_to_name, cache_emitter_id_to_inn, cache_secid_to_qualified_sign, cache_secid_to_bond_type, cache_secid_to_coupon_period, cache_secid_to_is_structural, cache_secid_to_issuer_rating = load_issuer_directory_cache()
    checkpoint_secid_to_emitter_id, checkpoint_emitter_id_to_name, checkpoint_emitter_id_to_inn, checkpoint_secid_to_qualified_sign, checkpoint_secid_to_bond_type, checkpoint_secid_to_coupon_period, checkpoint_secid_to_is_structural, checkpoint_secid_to_issuer_rating = load_issuer_checkpoint()

    secid_to_emitter_id: dict[str, int | None] = {**cache_secid_to_emitter_id, **checkpoint_secid_to_emitter_id}
    emitter_cache: dict[int, str] = {**cache_emitter_id_to_name, **checkpoint_emitter_id_to_name}
    emitter_inn_cache: dict[int, str] = {**cache_emitter_id_to_inn, **checkpoint_emitter_id_to_inn}
    secid_to_qualified_sign: dict[str, str] = {**cache_secid_to_qualified_sign, **checkpoint_secid_to_qualified_sign}
    secid_to_bond_type: dict[str, str] = {**cache_secid_to_bond_type, **checkpoint_secid_to_bond_type}
    secid_to_coupon_period: dict[str, int] = {**cache_secid_to_coupon_period, **checkpoint_secid_to_coupon_period}
    secid_to_is_structural: dict[str, bool] = {**cache_secid_to_is_structural, **checkpoint_secid_to_is_structural}
    secid_to_issuer_rating: dict[str, str] = {**cache_secid_to_issuer_rating, **checkpoint_secid_to_issuer_rating}

    secids_with_zero_coupon_period: set[str] = set()
    for row in rows:
        secid = str(row.get("SECID") or "").strip()
        if not secid:
            continue
        try:
            coupon_period_raw = int(float(str(row.get("COUPONPERIOD") or 0)))
        except ValueError:
            coupon_period_raw = 0
        if coupon_period_raw == 0:
            secids_with_zero_coupon_period.add(secid)

    missing_secids = [
        secid
        for secid in secids
        if secid not in secid_to_emitter_id
        or secid not in secid_to_qualified_sign
        or secid not in secid_to_bond_type
        or (secid in secids_with_zero_coupon_period and secid not in secid_to_coupon_period)
        or secid not in secid_to_is_structural
        or secid not in secid_to_issuer_rating
    ]
    secid_batches = chunked(missing_secids, SECID_BATCH_SIZE)

    def resolve_emitter_batch(batch: list[str]) -> tuple[dict[str, int | None], dict[str, str], dict[str, str], dict[str, int], dict[str, bool], dict[str, str]]:
        resolved: dict[str, int | None] = {}
        resolved_qualified: dict[str, str] = {}
        resolved_bond_types: dict[str, str] = {}
        resolved_coupon_periods: dict[str, int] = {}
        resolved_is_structural: dict[str, bool] = {}
        resolved_ratings: dict[str, str] = {}
        with build_session() as local_session:
            for secid in batch:
                try:
                    emitter_id, qualified_sign, bond_type, coupon_period, is_structural, issuer_rating = fetch_emitter_info_for_security(local_session, secid)
                    resolved[secid] = emitter_id
                    resolved_qualified[secid] = qualified_sign
                    resolved_bond_types[secid] = bond_type
                    if secid in secids_with_zero_coupon_period:
                        resolved_coupon_periods[secid] = coupon_period
                    resolved_is_structural[secid] = is_structural
                    resolved_ratings[secid] = issuer_rating
                except Exception as exc:
                    logging.warning("Не удалось получить данные эмитента для %s: %s", secid, exc)
                    resolved[secid] = None
                    resolved_qualified[secid] = "✖"
                    resolved_bond_types[secid] = "Не указан"
                    if secid in secids_with_zero_coupon_period:
                        resolved_coupon_periods[secid] = 0
                    resolved_is_structural[secid] = False
                    resolved_ratings[secid] = ""
        return resolved, resolved_qualified, resolved_bond_types, resolved_coupon_periods, resolved_is_structural, resolved_ratings

    if secid_batches:
        logging.info("Пакетный режим: нужно обработать %s пакетов SECID.", len(secid_batches))

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(resolve_emitter_batch, batch): idx for idx, batch in enumerate(secid_batches, start=1)}
        for processed, future in enumerate(as_completed(futures), start=1):
            secid_chunk, qualified_chunk, bond_types_chunk, coupon_period_chunk, structural_chunk, rating_chunk = future.result()
            secid_to_emitter_id.update(secid_chunk)
            secid_to_qualified_sign.update(qualified_chunk)
            secid_to_bond_type.update(bond_types_chunk)
            secid_to_coupon_period.update(coupon_period_chunk)
            secid_to_is_structural.update(structural_chunk)
            secid_to_issuer_rating.update(rating_chunk)
            save_issuer_checkpoint(
                secid_to_emitter_id,
                emitter_cache,
                emitter_inn_cache,
                secid_to_qualified_sign,
                secid_to_bond_type,
                secid_to_coupon_period,
                secid_to_is_structural,
                secid_to_issuer_rating,
            )
            if processed % 5 == 0 or processed == len(secid_batches):
                logging.info("SECID пакеты: %s/%s.", processed, len(secid_batches))

    unique_emitter_ids = sorted({emitter_id for emitter_id in secid_to_emitter_id.values() if emitter_id is not None})
    missing_emitter_ids = [
        emitter_id
        for emitter_id in unique_emitter_ids
        if emitter_id not in emitter_cache or emitter_id not in emitter_inn_cache
    ]
    emitter_batches = chunked(missing_emitter_ids, EMITTER_BATCH_SIZE)

    def resolve_emitter_names_batch(batch: list[int]) -> tuple[dict[int, str], dict[int, str]]:
        resolved_names: dict[int, str] = {}
        resolved_inn: dict[int, str] = {}
        with build_session() as local_session:
            for emitter_id in batch:
                try:
                    emitter_name, emitter_inn = fetch_emitter_details(local_session, emitter_id)
                    if emitter_name:
                        resolved_names[emitter_id] = emitter_name
                    if emitter_inn:
                        resolved_inn[emitter_id] = emitter_inn
                except Exception as exc:
                    logging.warning("Не удалось получить наименование эмитента %s: %s", emitter_id, exc)
        return resolved_names, resolved_inn

    if emitter_batches:
        logging.info("Пакетный режим: нужно обработать %s пакетов эмитентов.", len(emitter_batches))

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(resolve_emitter_names_batch, batch): idx for idx, batch in enumerate(emitter_batches, start=1)}
        for processed, future in enumerate(as_completed(futures), start=1):
            names_chunk, inn_chunk = future.result()
            emitter_cache.update(names_chunk)
            emitter_inn_cache.update(inn_chunk)
            save_issuer_checkpoint(
                secid_to_emitter_id,
                emitter_cache,
                emitter_inn_cache,
                secid_to_qualified_sign,
                secid_to_bond_type,
                secid_to_coupon_period,
                secid_to_is_structural,
                secid_to_issuer_rating,
            )
            if processed % 5 == 0 or processed == len(emitter_batches):
                logging.info("Пакеты эмитентов: %s/%s.", processed, len(emitter_batches))

    for row in rows:
        secid = str(row.get("SECID", ""))
        emitter_id = secid_to_emitter_id.get(secid)
        row[ISSUER_COLUMN_NAME] = emitter_cache.get(emitter_id) or ""
        row[ISSUER_INN_COLUMN_NAME] = emitter_inn_cache.get(emitter_id) or ""
        row[ISSUER_BOND_CLASS_COLUMN_NAME] = resolve_issuer_bond_class(row)
        row[ISSUER_RATING_COLUMN_NAME] = secid_to_issuer_rating.get(secid, "")
        row[QUALIFIED_INVESTOR_COLUMN_NAME] = secid_to_qualified_sign.get(secid, "✖")
        row[BOND_TYPE_COLUMN_NAME] = secid_to_bond_type.get(secid, "Не указан")
        row["EXCLUDE_BY_BOND_TYPE"] = bool(secid_to_is_structural.get(secid, False))
        row["BOND_TYPE_EXCLUDE_REASON"] = BOND_TYPE_EXCLUDE_REASON if row["EXCLUDE_BY_BOND_TYPE"] else ""
        if secid in secids_with_zero_coupon_period and "COUPONPERIOD" in row:
            row["COUPONPERIOD"] = secid_to_coupon_period.get(secid, 0)

    save_issuer_directory_cache(
        secid_to_emitter_id,
        emitter_cache,
        emitter_inn_cache,
        secid_to_qualified_sign,
        secid_to_bond_type,
        secid_to_coupon_period,
        secid_to_is_structural,
        secid_to_issuer_rating,
    )
    clear_issuer_checkpoint()

    return rows


def select_offer_jobs_for_refresh(
    rows: list[dict[str, Any]],
    offer_cache: dict[str, dict[str, Any]],
    daily_cache: dict[str, dict[str, Any]],
    now: datetime,
) -> list[tuple[str, str]]:
    """Выбирает бумаги для внешней проверки оферт (только там, где MOEX не показал оферту)."""
    unique_jobs: dict[str, str] = {}
    for row in rows:
        isin = str(row.get("ISIN") or "").strip()
        secid = str(row.get("SECID") or "").strip()
        if not isin:
            continue

        daily_entry = daily_cache.get(secid, {})
        moex_offer_type = normalize_offer_type(str(daily_entry.get("MOEX_HAS_PUT_CALL_OFFER", "✖")))
        moex_offer_date = str(daily_entry.get("MOEX_PUT_CALL_OFFER_DATE", "") or "").strip()
        moex_has_offer = moex_offer_type in {"PUT", "Call"} and bool(moex_offer_date)
        if moex_has_offer:
            continue

        if isin not in unique_jobs:
            unique_jobs[isin] = secid

    total = len(unique_jobs)
    if total == 0:
        return []

    daily_target = max(OFFER_RECHECK_DAILY_MIN_JOBS, math.ceil(total / max(OFFER_VERIFICATION_CACHE_TTL_DAYS, 1)))

    never_checked: list[tuple[str, str]] = []
    stale_checked: list[tuple[datetime, str, str]] = []

    for isin, secid in unique_jobs.items():
        entry = offer_cache.get(isin)
        if not entry:
            never_checked.append((isin, secid))
            continue

        checked_at_raw = entry.get("checked_at")
        if not checked_at_raw:
            never_checked.append((isin, secid))
            continue

        try:
            checked_at = datetime.fromisoformat(str(checked_at_raw))
        except ValueError:
            never_checked.append((isin, secid))
            continue

        if now - checked_at > timedelta(days=OFFER_VERIFICATION_CACHE_TTL_DAYS):
            stale_checked.append((checked_at, isin, secid))

    stale_checked.sort(key=lambda item: item[0])

    selected: list[tuple[str, str]] = []
    if len(never_checked) >= daily_target:
        selected = never_checked[:daily_target]
    else:
        selected.extend(never_checked)
        remaining_quota = max(daily_target - len(selected), 0)
        if remaining_quota > 0:
            selected.extend([(isin, secid) for _checked_at, isin, secid in stale_checked[:remaining_quota]])

    logging.info(
        "План проверки оферт: всего ISIN=%s, без кэша=%s, просрочено=%s, в этом запуске проверяем=%s.",
        total,
        len(never_checked),
        len(stale_checked),
        len(selected),
    )

    return selected


def enrich_with_daily_metrics(
    rows: list[dict[str, Any]],
    include_daily_metrics: bool = True,
    include_external_offers: bool = True,
) -> list[dict[str, Any]]:
    """Добавляет НКД/амортизацию и при необходимости внешнюю проверку Put/Call оферт с кэшированием."""
    logging.info("Обогащение суточными метриками (амортизация, оферты и НКД)...")
    secids = sorted({str(row.get("SECID")) for row in rows if row.get("SECID")})
    if not secids:
        return rows

    row_by_secid = {str(row.get("SECID")): row for row in rows if row.get("SECID")}

    cache_schema_version, daily_cache = load_daily_security_cache()
    offer_cache = load_offer_verification_cache()
    now = datetime.now()

    if cache_schema_version != DAILY_METRICS_CACHE_SCHEMA_VERSION:
        logging.info("Обнаружен старый формат суточного кэша (v%s). Метрики будут пересчитаны по новому правилу.", cache_schema_version)

    def needs_daily_refresh(secid: str) -> bool:
        if cache_schema_version != DAILY_METRICS_CACHE_SCHEMA_VERSION:
            return True

        entry = daily_cache.get(secid)
        if not entry:
            return True

        updated_at_raw = entry.get("updated_at")
        if not updated_at_raw:
            return True

        try:
            updated_at = datetime.fromisoformat(str(updated_at_raw))
        except ValueError:
            return True

        return now - updated_at > timedelta(hours=SECURITY_DAILY_CACHE_TTL_HOURS)

    missing_secids = [secid for secid in secids if needs_daily_refresh(secid)] if include_daily_metrics else []
    batches = chunked(missing_secids, SECID_BATCH_SIZE)

    def resolve_daily_batch(batch: list[str]) -> dict[str, dict[str, Any]]:
        resolved: dict[str, dict[str, Any]] = {}
        with build_session() as local_session:
            for secid in batch:
                try:
                    has_amortization, amort_start_date, has_put_call_offer, put_call_offer_date = fetch_daily_security_metrics(
                        local_session,
                        secid,
                        str(row_by_secid.get(secid, {}).get(MATURITY_DATE_COLUMN_NAME) or ""),
                    )
                    resolved[secid] = {
                        AMORTIZATION_FLAG_COLUMN_NAME: has_amortization,
                        AMORTIZATION_START_DATE_COLUMN_NAME: amort_start_date or "",
                        "MOEX_HAS_PUT_CALL_OFFER": has_put_call_offer,
                        "MOEX_PUT_CALL_OFFER_DATE": put_call_offer_date or "",
                        HAS_PUT_CALL_OFFER_COLUMN_NAME: "✖",
                        PUT_CALL_OFFER_DATE_COLUMN_NAME: "",
                        "EXCLUDE_BY_AMORTIZATION": False,
                        "AMORTIZATION_EXCLUDE_REASON": "",
                        "updated_at": datetime.now().isoformat(timespec="seconds"),
                    }
                except Exception as exc:
                    logging.warning("Не удалось получить суточные метрики для %s: %s", secid, exc)
        return resolved

    if batches:
        logging.info("Суточные метрики: нужно обработать %s пакетов SECID.", len(batches))

    if include_daily_metrics:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = [executor.submit(resolve_daily_batch, batch) for batch in batches]
            for processed, future in enumerate(as_completed(futures), start=1):
                daily_cache.update(future.result())
                if processed % 5 == 0 or processed == len(batches):
                    logging.info("Суточные пакеты: %s/%s.", processed, len(batches))

        save_daily_security_cache(daily_cache)

    if include_external_offers:
        offer_jobs = select_offer_jobs_for_refresh(rows, offer_cache, daily_cache, now)
    else:
        offer_jobs = []
    offer_batches = chunked(offer_jobs, OFFER_CHECK_BATCH_SIZE)

    if offer_batches:
        logging.info(
            "Проверка оферт через Corpbonds: нужно обработать %s ISIN (%s пакетов, кэш %s дней, потоков=%s).",
            len(offer_jobs),
            len(offer_batches),
            OFFER_VERIFICATION_CACHE_TTL_DAYS,
            OFFER_SOURCE_MAX_WORKERS,
        )

    if offer_jobs:
        source_results: dict[str, dict[str, dict[str, Any]]] = {"corpbonds": {}}

        def resolve_offer_source_job(isin: str, secid: str) -> tuple[str, dict[str, Any] | None, str | None]:
            try:
                with build_session() as local_session:
                    _offer_type, corpbonds_date, corpbonds_lookup = parse_corpbonds_offer_metrics(local_session, isin, secid or None)
                    return isin, {
                        HAS_PUT_CALL_OFFER_COLUMN_NAME: "✔" if corpbonds_date else "✖",
                        PUT_CALL_OFFER_DATE_COLUMN_NAME: corpbonds_date or "",
                        "lookup": corpbonds_lookup,
                    }, None
            except Exception as exc:
                return isin, None, str(exc)

        with ThreadPoolExecutor(max_workers=OFFER_SOURCE_MAX_WORKERS) as executor:
            futures = [executor.submit(resolve_offer_source_job, isin, secid or "") for isin, secid in offer_jobs]

            for processed, future in enumerate(as_completed(futures), start=1):
                isin, payload, error_text = future.result()
                if payload is not None:
                    source_results["corpbonds"][isin] = payload
                elif processed % 30 == 0:
                    logging.info("Проверка оферт: временная ошибка Corpbonds по %s: %s", isin, error_text)

                if processed % 50 == 0 or processed == len(futures):
                    logging.info("Проверка оферт: обработано задач %s/%s.", processed, len(futures))

        checked_at = datetime.now().isoformat(timespec="seconds")
        for isin, secid in offer_jobs:
            corpbonds_row = source_results["corpbonds"].get(isin)

            if corpbonds_row is None:
                logging.warning("Не удалось проверить оферту через Corpbonds для %s.", isin)
                continue

            corpbonds_date = str((corpbonds_row or {}).get(PUT_CALL_OFFER_DATE_COLUMN_NAME, "") or "")
            lookup = str(corpbonds_row.get("lookup") or "isin")
            offer_cache[isin] = {
                HAS_PUT_CALL_OFFER_COLUMN_NAME: "✔" if corpbonds_date else "✖",
                PUT_CALL_OFFER_DATE_COLUMN_NAME: corpbonds_date,
                "source": f"corpbonds:{lookup}",
                "checked_at": checked_at,
            }

    if include_external_offers:
        save_offer_verification_cache(offer_cache)

    for row in rows:
        secid = str(row.get("SECID", ""))
        entry = daily_cache.get(secid, {})
        has_amortization = str(entry.get(AMORTIZATION_FLAG_COLUMN_NAME, "✖"))
        amortization_start_date = str(entry.get(AMORTIZATION_START_DATE_COLUMN_NAME, "") or "")
        moex_offer_type = normalize_offer_type(str(entry.get("MOEX_HAS_PUT_CALL_OFFER", "✖")))
        moex_offer_date = str(entry.get("MOEX_PUT_CALL_OFFER_DATE", "") or "")

        isin = str(row.get("ISIN") or "")
        offer_entry = offer_cache.get(isin, {})
        external_offer_date = str(offer_entry.get(PUT_CALL_OFFER_DATE_COLUMN_NAME, "") or "")

        external_has_offer = bool(parse_date_safe(external_offer_date))
        moex_has_offer = bool(parse_date_safe(moex_offer_date))

        has_put_call_offer = "✖"
        put_call_offer_date = ""
        if external_has_offer:
            put_call_offer_date = external_offer_date
        elif moex_has_offer:
            put_call_offer_date = moex_offer_date
        else:
            put_call_offer_date = ""

        if has_put_call_offer in {"PUT", "Call"} and not put_call_offer_date:
            has_put_call_offer = "✖"

        maturity_date = str(row.get(MATURITY_DATE_COLUMN_NAME) or "")
        has_put_call_offer = "✔" if put_call_offer_date else "✖"
        if has_put_call_offer == "✔" and maturity_date and put_call_offer_date == maturity_date:
            has_put_call_offer = "✖"
            put_call_offer_date = ""

        if has_amortization == "✔" and maturity_date and amortization_start_date == maturity_date:
            has_amortization = "✖"
            amortization_start_date = ""

        row[AMORTIZATION_FLAG_COLUMN_NAME] = has_amortization
        row[AMORTIZATION_START_DATE_COLUMN_NAME] = amortization_start_date
        row[HAS_PUT_CALL_OFFER_COLUMN_NAME] = has_put_call_offer
        row[PUT_CALL_OFFER_DATE_COLUMN_NAME] = put_call_offer_date
        row[ACCRUED_INT_COLUMN_NAME] = row.get(ACCRUED_INT_COLUMN_NAME)

        amort_date = parse_date_safe(amortization_start_date)
        min_allowed_date = datetime.now().date() + timedelta(days=365 * MIN_MATURITY_YEARS)
        is_excluded = False
        exclude_reason = ""
        if amort_date is not None and amort_date.date() < min_allowed_date:
            is_excluded = True
            exclude_reason = "Амортизация уже началась или начнётся менее чем через год"

        offer_date = parse_date_safe(put_call_offer_date)
        is_offer_excluded = False
        offer_exclude_reason = ""
        if offer_date is not None and offer_date.date() < min_allowed_date:
            is_offer_excluded = True
            offer_exclude_reason = OFFER_DATE_EXCLUDE_REASON

        row["EXCLUDE_BY_AMORTIZATION"] = is_excluded
        row["AMORTIZATION_EXCLUDE_REASON"] = exclude_reason
        row["EXCLUDE_BY_OFFER_DATE"] = is_offer_excluded
        row["OFFER_DATE_EXCLUDE_REASON"] = offer_exclude_reason

        entry[AMORTIZATION_FLAG_COLUMN_NAME] = has_amortization
        entry[AMORTIZATION_START_DATE_COLUMN_NAME] = amortization_start_date
        entry["MOEX_HAS_PUT_CALL_OFFER"] = moex_offer_type
        entry["MOEX_PUT_CALL_OFFER_DATE"] = moex_offer_date
        entry[HAS_PUT_CALL_OFFER_COLUMN_NAME] = has_put_call_offer
        entry[PUT_CALL_OFFER_DATE_COLUMN_NAME] = put_call_offer_date
        entry["EXCLUDE_BY_AMORTIZATION"] = is_excluded
        entry["AMORTIZATION_EXCLUDE_REASON"] = exclude_reason
        entry["EXCLUDE_BY_OFFER_DATE"] = is_offer_excluded
        entry["OFFER_DATE_EXCLUDE_REASON"] = offer_exclude_reason
        entry["updated_at"] = datetime.now().isoformat(timespec="seconds")
        daily_cache[secid] = entry

    save_daily_security_cache(daily_cache)

    return rows


def filter_rows_by_amortization_timing(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Исключает бумаги, у которых амортизация уже началась или стартует менее чем через год."""
    filtered_rows: list[dict[str, Any]] = []
    skipped_count = 0
    for row in rows:
        if bool(row.get("EXCLUDE_BY_AMORTIZATION", False)):
            skipped_count += 1
            continue
        filtered_rows.append(row)

    logging.info(
        "Этап 4/8: Фильтр по дате амортизации: исключено %s бумаг, оставлено %s.",
        skipped_count,
        len(filtered_rows),
    )
    return filtered_rows


def filter_rows_by_offer_date(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Исключает бумаги, у которых дата оферты наступит менее чем через год."""
    filtered_rows: list[dict[str, Any]] = []
    skipped_count = 0

    for row in rows:
        if bool(row.get("EXCLUDE_BY_OFFER_DATE", False)):
            skipped_count += 1
            continue
        filtered_rows.append(row)

    logging.info(
        "Этап 4.2/8: Фильтр по дате оферты: исключено %s бумаг, оставлено %s.",
        skipped_count,
        len(filtered_rows),
    )
    return filtered_rows


def filter_rows_by_bond_type(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Исключает структурные облигации из итоговой выгрузки."""
    filtered_rows: list[dict[str, Any]] = []
    skipped_count = 0

    for row in rows:
        if bool(row.get("EXCLUDE_BY_BOND_TYPE", False)):
            skipped_count += 1
            continue
        filtered_rows.append(row)

    logging.info(
        "Этап 5.3/8: Фильтр структурных облигаций: исключено %s бумаг, оставлено %s.",
        skipped_count,
        len(filtered_rows),
    )
    return filtered_rows


def filter_rows_by_coupon_period(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Исключает бумаги, где купонный период не определён (0 или пусто)."""
    filtered_rows: list[dict[str, Any]] = []
    skipped_count = 0

    for row in rows:
        try:
            coupon_period = int(float(str(row.get("COUPONPERIOD") or 0)))
        except ValueError:
            coupon_period = 0

        if coupon_period <= 0:
            row["EXCLUDE_BY_COUPONPERIOD"] = True
            row["COUPONPERIOD_EXCLUDE_REASON"] = COUPONPERIOD_EXCLUDE_REASON
            skipped_count += 1
            continue

        row["EXCLUDE_BY_COUPONPERIOD"] = False
        row["COUPONPERIOD_EXCLUDE_REASON"] = ""
        filtered_rows.append(row)

    logging.info(
        "Этап 5.4/8: Фильтр по COUPONPERIOD: исключено %s бумаг, оставлено %s.",
        skipped_count,
        len(filtered_rows),
    )
    return filtered_rows


def merge_incremental_data(cached_rows: list[dict[str, Any]], fresh_rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int, int]:
    """Инкрементально объединяет кэш и свежие данные (добавляет новые и обновляет изменённые)."""
    merged = {(row.get("SECID"), row.get("ISIN")): row for row in cached_rows}
    new_count = 0
    updated_count = 0

    for row in fresh_rows:
        key = (row.get("SECID"), row.get("ISIN"))
        if key not in merged:
            merged[key] = row
            new_count += 1
        elif merged[key] != row:
            merged[key] = row
            updated_count += 1

    return list(merged.values()), new_count, updated_count


def save_raw(rows: list[dict[str, Any]]) -> None:
    """Сохраняет сырые данные для отладки."""
    RAW_FILE.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    logging.info("Этап 6/8: Сырые данные сохранены: %s", RAW_FILE)


def apply_excel_formatting(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    """Применяет форматирование итогового Excel, добавляет групповые колонки и скрывает SECID."""
    ws = writer.sheets["MOEX_BONDS"]
    ws.insert_rows(1)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryRight = False

    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    even_fill = PatternFill(fill_type="solid", start_color="F5F9FF", end_color="F5F9FF")
    odd_fill = PatternFill(fill_type="solid", start_color="FFFFFF", end_color="FFFFFF")
    group_fill = PatternFill(fill_type="solid", start_color="D9E2F3", end_color="D9E2F3")
    separator_palette = [
        PatternFill(fill_type="solid", start_color="C9D5EA", end_color="C9D5EA"),
        PatternFill(fill_type="solid", start_color="C3D0E8", end_color="C3D0E8"),
        PatternFill(fill_type="solid", start_color="BCCBE5", end_color="BCCBE5"),
        PatternFill(fill_type="solid", start_color="B6C6E2", end_color="B6C6E2"),
    ]
    group_font = Font(color="1F4E78", bold=True)

    separator_titles: dict[str, str] = df.attrs.get("group_separator_titles", {})
    separator_columns = {c for c in df.columns if str(c).startswith(GROUP_SEPARATOR_PREFIX)}
    calculated_coupon_cells: set[tuple[str, str]] = df.attrs.get("calculated_coupon_cells", set())
    calculated_coupon_value_cells: set[tuple[str, str]] = df.attrs.get("calculated_coupon_value_cells", set())

    ws.row_dimensions[1].height = 78
    ws.row_dimensions[2].height = 22

    separator_indexes = [idx for idx, col in enumerate(df.columns, start=1) if col in separator_columns]
    separator_fill_by_idx = {
        sep_idx: separator_palette[pos % len(separator_palette)]
        for pos, sep_idx in enumerate(separator_indexes)
    }

    for col_idx, col_name in enumerate(df.columns, start=1):
        if col_name in separator_columns:
            header_cell = ws.cell(row=2, column=col_idx)
            header_cell.value = ""
            header_cell.fill = separator_fill_by_idx[col_idx]
            header_cell.font = group_font
            header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            continue

        cell = ws.cell(row=2, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    numeric_columns = {
        "FACEVALUE",
        "COUPONVALUE",
        "COUPONPERIOD",
        "COUPONPERCENT",
        "PREVLEGALCLOSEPRICE",
        "PREVPRICE",
        ACCRUED_INT_COLUMN_NAME,
        TRADE_VOLUME_COLUMN_NAME,
        TRADE_VALUE_COLUMN_NAME,
        "NUMTRADES",
        YIELD_COLUMN_NAME,
    }

    for row_idx in range(3, ws.max_row + 1):
        row_fill = even_fill if row_idx % 2 == 0 else odd_fill
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name in separator_columns:
                cell.value = ""
                cell.fill = separator_fill_by_idx[col_idx]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                continue

            cell.fill = row_fill
            cell.alignment = Alignment(vertical="center")
            if col_name in numeric_columns and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

            if col_name == "COUPONPERCENT":
                isin_value = str(ws.cell(row=row_idx, column=1).value or "")
                secid_value = ""
                if HIDDEN_COLUMN_NAME in df.columns:
                    secid_col_idx = list(df.columns).index(HIDDEN_COLUMN_NAME) + 1
                    secid_value = str(ws.cell(row=row_idx, column=secid_col_idx).value or "")
                if (isin_value, secid_value) in calculated_coupon_cells:
                    cell.fill = CALCULATED_COUPON_FILL

            if col_name == "COUPONVALUE":
                isin_value = str(ws.cell(row=row_idx, column=1).value or "")
                secid_value = ""
                if HIDDEN_COLUMN_NAME in df.columns:
                    secid_col_idx = list(df.columns).index(HIDDEN_COLUMN_NAME) + 1
                    secid_value = str(ws.cell(row=row_idx, column=secid_col_idx).value or "")
                if (isin_value, secid_value) in calculated_coupon_value_cells:
                    cell.fill = CALCULATED_COUPON_FILL

    for idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(idx)

        if col_name in separator_columns:
            ws.column_dimensions[col_letter].width = 15
            continue

        max_len = max(
            len(str(col_name)),
            *(len(str(v)) for v in df[col_name].head(500).fillna("")),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 48)

        if col_name == ISSUER_COLUMN_NAME:
            ws.column_dimensions[col_letter].width = 70

        if col_name == ISSUER_INN_COLUMN_NAME:
            ws.column_dimensions[col_letter].width = 16

        if col_name == QUALIFIED_INVESTOR_COLUMN_NAME:
            ws.column_dimensions[col_letter].width = 15

        if col_name == AMORTIZATION_FLAG_COLUMN_NAME:
            ws.column_dimensions[col_letter].width = 17

        if col_name == AMORTIZATION_START_DATE_COLUMN_NAME:
            ws.column_dimensions[col_letter].width = 24

        if col_name == MATURITY_DATE_COLUMN_NAME:
            ws.column_dimensions[col_letter].width = 14

        if col_name == BOND_TYPE_COLUMN_NAME:
            ws.column_dimensions[col_letter].width = 22

        if col_name == HIDDEN_COLUMN_NAME:
            ws.column_dimensions[col_letter].hidden = True

    for sep_pos, sep_idx in enumerate(separator_indexes):
        next_sep_idx = separator_indexes[sep_pos + 1] if sep_pos + 1 < len(separator_indexes) else ws.max_column + 1
        group_indexes = list(range(sep_idx + 1, next_sep_idx))

        group_title = separator_titles.get(df.columns[sep_idx - 1], "Группа")
        top_cell = ws.cell(row=1, column=sep_idx)
        top_cell.value = group_title
        top_cell.fill = separator_fill_by_idx[sep_idx]
        top_cell.font = group_font
        top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        second_row_cell = ws.cell(row=2, column=sep_idx)
        second_row_cell.fill = separator_fill_by_idx[sep_idx]
        second_row_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if not group_indexes:
            continue

        first_group_col = group_indexes[0]
        for grouped_idx in group_indexes:
            ws.column_dimensions[get_column_letter(grouped_idx)].outlineLevel = 1

        ws.column_dimensions[get_column_letter(first_group_col)].collapsed = True

    first_cell = ws.cell(row=1, column=1)
    first_cell.value = "Основные поля"
    first_cell.fill = group_fill
    first_cell.font = group_font
    first_cell.alignment = Alignment(horizontal="center", vertical="center")


def add_info_sheet(writer: pd.ExcelWriter) -> None:
    """Добавляет лист INFO с пояснением полей простым языком."""
    info_rows = [
        {"Поле": "SECID", "Описание": "Технический код бумаги на бирже. В основном листе скрыт, но сохранён для аналитики и связок."},
        {"Поле": "ISIN", "Описание": "Международный идентификатор ценной бумаги."},
        {"Поле": "SHORTNAME", "Описание": "Краткое название облигации."},
        {"Поле": "ISSUER_NAME", "Описание": "Наименование эмитента облигации (компании или организации, которая выпустила бумагу)."},
        {"Поле": "ISSUER_INN", "Описание": "ИНН эмитента для быстрой сверки компании в ваших внутренних системах и документах."},
        {"Поле": "ISSUER_BOND_CLASS", "Описание": "Тип бумаги на рынке облигаций (например: государственный, корпоративный, муниципальный, иностранный)."},
        {"Поле": "ISSUER_RATING", "Описание": "Рейтинг эмитента по данным карточки бумаги на MOEX. Если поле пустое, биржа не передала рейтинг для этого выпуска."},
        {"Поле": "QUALIFIED_INVESTOR", "Описание": "Показывает, предназначена ли облигация только для квалифицированных инвесторов: ✔ — да, ✖ — нет."},
        {"Поле": "BOND_TYPE", "Описание": "Тип облигации по купону (например: фиксированная, флоатер и т.д.)."},
        {"Поле": "HAS_PUT_CALL_OFFER", "Описание": "Есть ли оферта: ✔ (есть) или ✖ (нет)."},
        {"Поле": "PUT_CALL_OFFER_DATE", "Описание": "Ближайшая дата оферты (MOEX + проверка Corpbonds для бумаг без оферты на MOEX)."},
        {"Поле": "HAS_AMORTIZATION", "Описание": "Есть ли у бумаги амортизация номинала: ✔ — да, ✖ — нет."},
        {"Поле": "AMORTIZATION_START_DATE", "Описание": "Дата начала амортизации (если есть)."},
                {"Поле": "MATDATE", "Описание": "Дата погашения облигации (когда эмитент должен вернуть номинал)."},
        {"Поле": "FACEVALUE", "Описание": "Номинал облигации."},
        {"Поле": "FACEUNIT", "Описание": "Валюта номинала (например, RUB)."},
        {"Поле": "COUPONVALUE", "Описание": "Размер купонной выплаты. Если ячейка жёлтая, значение рассчитано по формуле из FACEVALUE, COUPONPERCENT и COUPONPERIOD (когда MOEX вернул 0)."},
        {"Поле": "ACCRUEDINT", "Описание": "Накопленный купонный доход (НКД) на текущую дату."},
        {"Поле": "COUPONPERIOD", "Описание": "Период выплаты купона в днях."},
        {"Поле": "COUPONPERCENT", "Описание": "Купонная ставка в процентах. Если ячейка жёлтая, ставка рассчитана по формуле (analytics.dohod.ru и, при необходимости, fallback corpbonds.ru). Это приближённое значение."},
        {"Поле": "COUPON_FORMULA_SOURCE", "Описание": "Источник формулы для расчётного COUPONPERCENT: DOHOD или CORPBONDS. Для обычных бумаг поле пустое."},
        {"Поле": "PRIMARYBOARDID", "Описание": "Основной торговый режим/секция."},
        {"Поле": "PREVLEGALCLOSEPRICE", "Описание": "Предыдущая официальная цена закрытия."},
        {"Поле": "PREVPRICE", "Описание": "Предыдущая рыночная цена."},
        {"Поле": "VOLTODAY", "Описание": "Объём торгов за текущий день (в штуках/бумагах)."},
        {"Поле": "VALTODAY", "Описание": "Денежный оборот торгов за текущий день."},
        {"Поле": "NUMTRADES", "Описание": "Количество сделок за текущий день."},
        {"Поле": "YIELD", "Описание": "Доходность к погашению по данным биржи (в процентах годовых)."},
    ]

    info_df = pd.DataFrame(info_rows)
    info_df.to_excel(writer, index=False, sheet_name="INFO")
    info_ws = writer.sheets["INFO"]
    info_ws.freeze_panes = "A2"

    for idx, col_name in enumerate(info_df.columns, start=1):
        col_letter = get_column_letter(idx)
        max_len = max(len(str(col_name)), *(len(str(v)) for v in info_df[col_name]))
        info_ws.column_dimensions[col_letter].width = min(max_len + 2, 80)


def drop_deprecated_output_columns(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Удаляет устаревшие поля из строк, чтобы старый кэш не возвращал удалённые колонки в Excel."""
    for row in rows:
        for old_col in DEPRECATED_OUTPUT_COLUMNS:
            row.pop(old_col, None)
    return rows


def save_excel(
    rows: list[dict[str, Any]],
    calculated_coupon_cells: set[tuple[str, str]] | None = None,
    calculated_coupon_value_cells: set[tuple[str, str]] | None = None,
) -> None:
    """Сохраняет итоговый набор в Excel."""
    logging.info("Этап 7/8: Подготовка итогового Excel...")
    df = pd.DataFrame(rows)

    for col in REMOVED_COLUMNS:
        if col in df.columns:
            df = df.drop(columns=[col])

    sort_columns = [col for col in ["SECID", "ISIN"] if col in df.columns]
    if sort_columns:
        df = df.sort_values(by=sort_columns).reset_index(drop=True)

    for date_col in [MATURITY_DATE_COLUMN_NAME, AMORTIZATION_START_DATE_COLUMN_NAME, PUT_CALL_OFFER_DATE_COLUMN_NAME]:
        if date_col in df.columns:
            df[date_col] = df[date_col].map(format_date_ddmmyyyy)

    if FIRST_COLUMN_NAME in df.columns:
        group_layout = [
            ("Эмитент", [ISSUER_COLUMN_NAME, ISSUER_INN_COLUMN_NAME, ISSUER_BOND_CLASS_COLUMN_NAME, ISSUER_RATING_COLUMN_NAME, "SHORTNAME"]),
            ("Квалификация и тип", [QUALIFIED_INVESTOR_COLUMN_NAME, BOND_TYPE_COLUMN_NAME]),
            ("Оферты", [HAS_PUT_CALL_OFFER_COLUMN_NAME, PUT_CALL_OFFER_DATE_COLUMN_NAME]),
            (
                "Торги, погашение и амортизация",
                [AMORTIZATION_FLAG_COLUMN_NAME, AMORTIZATION_START_DATE_COLUMN_NAME, MATURITY_DATE_COLUMN_NAME],
            ),
            ("Купоны", ["COUPONVALUE", ACCRUED_INT_COLUMN_NAME, "COUPONPERIOD", "COUPONPERCENT", COUPON_FORMULA_SOURCE_COLUMN_NAME]),
            (
                "Рынок",
                [
                    "FACEVALUE",
                    "FACEUNIT",
                    "PRIMARYBOARDID",
                    "PREVLEGALCLOSEPRICE",
                    "PREVPRICE",
                    TRADE_VOLUME_COLUMN_NAME,
                    TRADE_VALUE_COLUMN_NAME,
                    "NUMTRADES",
                    YIELD_COLUMN_NAME,
                ],
            ),
        ]

        ordered_columns = [FIRST_COLUMN_NAME]
        separator_titles: dict[str, str] = {}
        used_columns = {FIRST_COLUMN_NAME}

        for group_idx, (group_title, group_columns) in enumerate(group_layout, start=1):
            existing_columns = [col for col in group_columns if col in df.columns and col not in used_columns]
            if not existing_columns:
                continue

            separator_col = f"{GROUP_SEPARATOR_PREFIX}{group_idx}"
            df[separator_col] = ""
            ordered_columns.append(separator_col)
            separator_titles[separator_col] = group_title

            ordered_columns.extend(existing_columns)
            used_columns.update(existing_columns)

        remaining_columns = [col for col in df.columns if col not in used_columns and not str(col).startswith(GROUP_SEPARATOR_PREFIX)]
        if remaining_columns:
            separator_col = f"{GROUP_SEPARATOR_PREFIX}MISC"
            df[separator_col] = ""
            ordered_columns.append(separator_col)
            separator_titles[separator_col] = "Прочие поля"
            ordered_columns.extend(remaining_columns)

        df = df[ordered_columns]
        df.attrs["group_separator_titles"] = separator_titles

    if calculated_coupon_cells is not None:
        df.attrs["calculated_coupon_cells"] = calculated_coupon_cells
    if calculated_coupon_value_cells is not None:
        df.attrs["calculated_coupon_value_cells"] = calculated_coupon_value_cells

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="MOEX_BONDS")
        apply_excel_formatting(writer, df)
        add_info_sheet(writer)

    logging.info("Этап 8/8: Excel файл сохранён: %s", OUTPUT_FILE)


def main() -> None:
    started_at = time.perf_counter()
    setup_folders()
    setup_logging()

    logging.info("Старт скрипта: сбор облигаций MOEX")
    cached_rows = load_cache()

    try:
        if cached_rows is None:
            rows = fetch_all_bonds()
        else:
            fresh_rows = fetch_all_bonds()
            rows, new_count, updated_count = merge_incremental_data(cached_rows, fresh_rows)
            logging.info(
                "Этап 1.1/8: Инкрементальное обновление завершено. Новых: %s, обновлённых: %s, всего: %s.",
                new_count,
                updated_count,
                len(rows),
            )
    except Exception as exc:
        logging.error("Не удалось получить свежие данные: %s", exc)
        rows = cached_rows or load_cache(allow_stale=True)
        if not rows:
            raise
        logging.warning("Используем резервный кэш из-за недоступности API.")

    # 1) Базовый сбор и первичная фильтрация по данным MOEX
    logging.info("Этап 2/8: Фильтр по сроку до погашения...")
    rows = filter_rows_by_maturity(rows)
    logging.info("Этап 3/8: Обогащение суточными метриками (MOEX)...")
    rows = enrich_with_daily_metrics(rows, include_daily_metrics=True, include_external_offers=False)
    rows = filter_rows_by_amortization_timing(rows)
    rows = filter_rows_by_offer_date(rows)

    # 2) Обогащение из альтернативных источников (после первичной фильтрации)
    logging.info("Этап 5/8: Обогащение из внешних источников и данных эмитента...")
    rows = enrich_with_daily_metrics(rows, include_daily_metrics=False, include_external_offers=True)
    rows = enrich_with_issuer_names(rows)
    rows = filter_rows_by_bond_type(rows)
    rows = filter_rows_by_coupon_period(rows)
    rows, calculated_coupon_cells = enrich_coupon_percent_from_dohod(rows)
    rows, calculated_coupon_value_cells = enrich_coupon_value_from_percent(rows)

    # 3) Повторная фильтрация после обогащения новыми данными
    rows = filter_rows_by_offer_date(rows)
    rows = filter_rows_by_amortization_timing(rows)
    validate_rows(rows)
    rows = drop_deprecated_output_columns(rows)
    save_cache(rows)
    save_raw(rows)
    save_excel(rows, calculated_coupon_cells, calculated_coupon_value_cells)

    elapsed = time.perf_counter() - started_at
    logging.info("Готово. Всего облигаций: %s", len(rows))
    logging.info("Время выполнения: %.2f сек.", elapsed)


if __name__ == "__main__":
    main()
