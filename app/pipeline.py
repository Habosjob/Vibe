from __future__ import annotations

import asyncio
import html
import json
import logging
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
import re
from pathlib import Path
from typing import Any

import aiohttp
from aiohttp import ClientResponseError
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from tqdm import tqdm
from bs4 import BeautifulSoup

import config
from app.bootstrap import load_state, save_state
from app.database import Database

LOGGER = logging.getLogger(__name__)


@dataclass
class RunSummary:
    fetched_count: int
    selected_count: int
    saved_count: int
    errors_count: int
    from_cache_count: int
    duration_total: float
    duration_load: float
    duration_calc: float
    duration_save: float
    moex_output_path: Path | None
    corpbonds_output_path: Path | None
    merged_output_path: Path | None
    emitents_output_path: Path | None
    screener_output_path: Path | None


@dataclass
class BondRow:
    secid: str
    isin: str
    short_name: str
    emitter_name: str
    emitter_inn: str
    current_price: float | None
    previous_price: float | None
    price_change_percent: float | None
    volume_today: float
    volume_20d: float
    maturity_date: date | None
    offer_date: date | None
    amortization_start: date | None
    qualified_investor: str
    default_flag: str
    technical_default_flag: str
    bond_type: str
    sec_sub_type: str
    coupon_period: int | None
    accrued_int: float | None
    coupon_percent: float | None
    ytm: float | None


@dataclass
class CorpBondRow:
    secid: str
    price: str
    credit_rating: str
    coupon_type: str
    coupon_formula: str
    nearest_offer_date: str
    ladder_coupon: str


class MoexClient:
    def __init__(self) -> None:
        self._session: aiohttp.ClientSession | None = None

    async def __aenter__(self) -> MoexClient:
        timeout = aiohttp.ClientTimeout(
            total=config.REQUEST_CONNECT_TIMEOUT_SEC + config.REQUEST_READ_TIMEOUT_SEC,
            connect=config.REQUEST_CONNECT_TIMEOUT_SEC,
            sock_read=config.REQUEST_READ_TIMEOUT_SEC,
        )
        connector = aiohttp.TCPConnector(
            limit=max(32, config.CORPBONDS_MAX_CONCURRENT_TASKS * 2, config.MAX_CONCURRENT_TASKS * 2),
            ttl_dns_cache=300,
        )
        self._session = aiohttp.ClientSession(timeout=timeout, connector=connector)
        return self

    async def __aexit__(self, exc_type: Any, exc: Any, tb: Any) -> None:
        await self.close()

    async def close(self) -> None:
        if self._session is not None:
            await self._session.close()
            self._session = None

    async def get_json(self, url: str) -> dict[str, Any]:
        if self._session is None:
            raise RuntimeError("HTTP-сессия не инициализирована")
        last_exc: Exception | None = None
        for attempt in range(1, config.REQUEST_RETRIES + 1):
            try:
                async with self._session.get(url) as response:
                    response.raise_for_status()
                    return await response.json(content_type=None)
            except ClientResponseError as exc:
                last_exc = exc
                if 400 <= exc.status < 500 and exc.status != 429:
                    break
                if attempt == config.REQUEST_RETRIES:
                    break
                await asyncio.sleep(config.REQUEST_BACKOFF_SEC * (2 ** (attempt - 1)))
            except Exception as exc:
                last_exc = exc
                if attempt == config.REQUEST_RETRIES:
                    break
                await asyncio.sleep(config.REQUEST_BACKOFF_SEC * (2 ** (attempt - 1)))
        raise RuntimeError(f"Не удалось получить данные: {url}. Ошибка: {last_exc}")

    async def get_text(self, url: str, headers: dict[str, str] | None = None) -> str:
        if self._session is None:
            raise RuntimeError("HTTP-сессия не инициализирована")
        last_exc: Exception | None = None
        for attempt in range(1, config.REQUEST_RETRIES + 1):
            try:
                async with self._session.get(url, headers=headers) as response:
                    response.raise_for_status()
                    return await response.text()
            except ClientResponseError as exc:
                last_exc = exc
                if 400 <= exc.status < 500 and exc.status != 429:
                    break
                if attempt == config.REQUEST_RETRIES:
                    break
                await asyncio.sleep(config.REQUEST_BACKOFF_SEC * (2 ** (attempt - 1)))
            except Exception as exc:
                last_exc = exc
                if attempt == config.REQUEST_RETRIES:
                    break
                await asyncio.sleep(config.REQUEST_BACKOFF_SEC * (2 ** (attempt - 1)))
        raise RuntimeError(f"Не удалось получить страницу: {url}. Ошибка: {last_exc}")


def _as_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    text = str(value)
    try:
        return datetime.fromisoformat(text).date()
    except Exception:
        try:
            return datetime.strptime(text, "%d.%m.%Y").date()
        except Exception:
            return None


def _to_yes_no(value: Any) -> str:
    normalized = str(value).strip().lower()
    return "Да" if normalized in {"1", "true"} else "Нет"


def _pick_price(row: dict[str, Any]) -> float | None:
    md = row.get("marketdata", {})
    for key in ("LAST", "LCURRENTPRICE", "LCLOSEPRICE", "PREVPRICE"):
        value = md.get(key)
        if value is not None:
            return float(value)
    prev = row.get("PREVPRICE")
    return float(prev) if prev is not None else None


def _normalize_text(value: str) -> str:
    compact = " ".join(value.replace("\xa0", " ").split())
    if compact.lower() == "нет данных":
        return ""
    return compact


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(" ", "").replace("%", "")
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except Exception:
        return None


def _parse_corp_price(value: str) -> float | None:
    if not value:
        return None
    candidate = value.split()[0].replace("\xa0", " ")
    return _parse_float(candidate)


def _select_last_price(moex_price: float | None, corp_price: str) -> float | None:
    if moex_price is not None:
        return moex_price
    return _parse_corp_price(corp_price)


def _forecast_value(forecast: dict[int, float], years_ahead: int) -> float:
    if years_ahead <= 0:
        return float(forecast[0])
    if years_ahead in forecast:
        return float(forecast[years_ahead])
    return float(forecast[max(forecast.keys())])


def _extract_formula_rate(
    formula: str,
    key_rate: float,
    ruonia: float,
    gcurve_5y: float,
    gcurve_7y: float,
) -> float | None:
    if not formula:
        return None
    normalized = formula.upper().replace(" ", "").replace(",", ".")
    numbers = [float(x) for x in re.findall(r"[+-]?\d+(?:\.\d+)?", normalized)]
    spread = numbers[-1] if numbers else 0.0
    base = None
    if "RUONIA" in normalized:
        base = ruonia
    elif "G-CURVE7Y" in normalized or "G-CURVE7" in normalized:
        base = gcurve_7y
    elif "G-CURVE5Y" in normalized or "G-CURVE5" in normalized:
        base = gcurve_5y
    elif "КС" in normalized or "KC" in normalized:
        base = key_rate
    if base is None:
        return None
    rate = base + spread
    max_match = re.search(r"MAX\(?([\d.]+)%?\)?", normalized)
    if max_match:
        try:
            rate = min(rate, float(max_match.group(1)))
        except Exception:
            pass
    return rate


def _extract_last_row_numbers(page_html: str) -> list[float]:
    soup = BeautifulSoup(page_html, "html.parser")
    rows = soup.find_all("tr")
    for row in reversed(rows):
        cells = row.find_all("td")
        if not cells:
            continue
        numbers: list[float] = []
        for cell in cells:
            match = re.search(r"-?\d+[,.]\d+", cell.get_text(" ", strip=True))
            if match:
                numbers.append(float(match.group(0).replace(",", ".")))
        if numbers:
            return numbers
    return []


async def _fetch_cbr_curve(client: MoexClient) -> tuple[float, float, float, float]:
    urls = {
        "key": "https://cbr.ru/hd_base/KeyRate/",
        "ruonia": "https://cbr.ru/hd_base/ruonia/",
        "zcyc": "https://cbr.ru/hd_base/zcyc_params/",
    }
    pages = await asyncio.gather(*(client.get_text(url) for url in urls.values()), return_exceptions=True)
    key_rate = config.YTM_KEY_RATE_FORECAST[0]
    ruonia = key_rate - config.YTM_RUONIA_KEY_SPREAD_DEFAULT
    gcurve_5y = key_rate
    gcurve_7y = key_rate

    if not isinstance(pages[0], Exception):
        try:
            key_values = _extract_last_row_numbers(str(pages[0]))
            if key_values:
                key_rate = key_values[-1]
        except Exception:
            LOGGER.warning("Не удалось распарсить ключевую ставку ЦБ, используем прогноз Year0")

    if not isinstance(pages[1], Exception):
        try:
            ruonia_values = _extract_last_row_numbers(str(pages[1]))
            if ruonia_values:
                ruonia = ruonia_values[-1]
        except Exception:
            LOGGER.warning("Не удалось распарсить RUONIA, используем спред к КС")
    else:
        ruonia = key_rate - config.YTM_RUONIA_KEY_SPREAD_DEFAULT

    if not isinstance(pages[2], Exception):
        try:
            zcyc_values = _extract_last_row_numbers(str(pages[2]))
            if len(zcyc_values) >= 7:
                gcurve_5y = zcyc_values[4]
                gcurve_7y = zcyc_values[6]
        except Exception:
            LOGGER.warning("Не удалось распарсить КБД ОФЗ, используем КС")
    else:
        gcurve_5y = key_rate
        gcurve_7y = key_rate

    return key_rate, ruonia, gcurve_5y, gcurve_7y


def _calc_ytm_percent(
    clean_price: float | None,
    accrued_int: float | None,
    coupon_percent: float | None,
    coupon_period_days: int | None,
    coupon_type: str,
    coupon_formula: str,
    secid: str,
    end_date: date | None,
    key_rate: float,
    ruonia: float,
    gcurve_5y: float,
    gcurve_7y: float,
) -> float | None:
    if clean_price is None or end_date is None:
        return None
    today = date.today()
    if end_date <= today:
        return None
    total_days = (end_date - today).days
    years = max(total_days / 365.0, 0.01)
    dirty_price = clean_price + max(accrued_int or 0.0, 0.0)
    if dirty_price <= 0:
        return None
    if coupon_period_days is None or coupon_period_days <= 0:
        coupon_period_days = 182
    lower_coupon_type = coupon_type.lower()

    if not (20.0 <= dirty_price <= 200.0):
        return None

    spread_key_ruonia = key_rate - ruonia
    if abs(spread_key_ruonia) > config.YTM_MAX_ABS_SPREAD_SANITY:
        spread_key_ruonia = config.YTM_RUONIA_KEY_SPREAD_DEFAULT
    spread_g5 = gcurve_5y - key_rate
    spread_g7 = gcurve_7y - key_rate
    if abs(spread_g5) > config.YTM_MAX_ABS_SPREAD_SANITY:
        spread_g5 = 0.0
    if abs(spread_g7) > config.YTM_MAX_ABS_SPREAD_SANITY:
        spread_g7 = 0.0

    def resolve_annual_coupon_rate(days_from_today: int) -> float:
        year_index = max(days_from_today - 1, 0) // 365
        annual_rate = coupon_percent if coupon_percent is not None else 0.0
        if "флоатер" in lower_coupon_type or secid.startswith("SU50"):
            projected_key = _forecast_value(config.YTM_KEY_RATE_FORECAST, year_index)
            projected_infl = _forecast_value(config.YTM_INFLATION_FORECAST, year_index)
            projected_ruonia = projected_key - spread_key_ruonia
            projected_g5 = projected_key + spread_g5
            projected_g7 = projected_key + spread_g7
            if secid.startswith("SU50"):
                annual_rate = projected_infl
            parsed = _extract_formula_rate(coupon_formula, projected_key, projected_ruonia, projected_g5, projected_g7)
            if parsed is not None:
                annual_rate = parsed
        return annual_rate

    payment_days: list[int] = []
    elapsed = coupon_period_days
    while elapsed < total_days:
        payment_days.append(elapsed)
        elapsed += coupon_period_days
    payment_days.append(total_days)

    cashflows: list[tuple[int, float]] = []
    previous_day = 0
    for payment_day in payment_days:
        period_days = payment_day - previous_day
        annual_coupon_rate = resolve_annual_coupon_rate(payment_day)
        coupon_cash = 100.0 * (annual_coupon_rate / 100.0) * (period_days / 365.0)
        principal_cash = 100.0 if payment_day == total_days else 0.0
        cashflows.append((payment_day, coupon_cash + principal_cash))
        previous_day = payment_day

    def npv(rate: float) -> float:
        total = -dirty_price
        for payment_day, amount in cashflows:
            discount = (1.0 + rate) ** (payment_day / 365.0)
            total += amount / discount
        return total

    left, right = -0.95, 3.0
    npv_left = npv(left)
    npv_right = npv(right)
    while npv_left * npv_right > 0 and right < 20.0:
        right *= 2.0
        npv_right = npv(right)
    if npv_left * npv_right > 0:
        return None

    for _ in range(120):
        mid = (left + right) / 2.0
        npv_mid = npv(mid)
        if abs(npv_mid) < 1e-10:
            left = right = mid
            break
        if npv_left * npv_mid <= 0:
            right = mid
            npv_right = npv_mid
        else:
            left = mid
            npv_left = npv_mid

    ytm = ((left + right) / 2.0) * 100.0
    if ytm < -95.0 or ytm > 250.0:
        return None
    return round(ytm, 2)


def _parse_corpbonds_html(page_html: str) -> dict[str, str]:
    result = {
        "price": "",
        "credit_rating": "",
        "coupon_type": "",
        "coupon_formula": "",
        "nearest_offer_date": "",
        "ladder_coupon": "",
    }
    mapping = {
        "Цена последняя": "price",
        "Кредитный рейтинг": "credit_rating",
        "Тип купона": "coupon_type",
        "Формула купона": "coupon_formula",
        "Дата ближайшей оферты": "nearest_offer_date",
        "Купон лесенкой": "ladder_coupon",
    }

    rows = re.findall(r"<tr[^>]*>\s*<td[^>]*>(.*?)</td>\s*<td[^>]*>(.*?)</td>", page_html, flags=re.S | re.I)
    for left_raw, right_raw in rows:
        left_text = re.sub(r"<[^>]+>", " ", left_raw)
        right_text = re.sub(r"<[^>]+>", " ", right_raw)
        left = _normalize_text(html.unescape(left_text)).replace(" ?", "")
        right = _normalize_text(html.unescape(right_text))
        for key, target in mapping.items():
            if left.startswith(key):
                result[target] = right
                break
    return result


async def _fetch_all_traded_bonds(client: MoexClient) -> list[dict[str, Any]]:
    url = (
        "https://iss.moex.com/iss/engines/stock/markets/bonds/securities.json"
        "?iss.meta=off&is_traded=1"
        "&iss.only=securities,marketdata"
        "&securities.columns=SECID,ISIN,SHORTNAME,SECNAME,MATDATE,OFFERDATE,COUPONPERIOD,ACCRUEDINT,"
        "COUPONPERCENT,BONDTYPE,BONDSUBTYPE,PREVPRICE"
        "&marketdata.columns=SECID,LAST,LCURRENTPRICE,LCLOSEPRICE,PREVPRICE,VOLTODAY,VALTODAY,YIELD"
    )
    payload = await client.get_json(url)
    sec_cols = payload["securities"]["columns"]
    md_cols = payload["marketdata"]["columns"]
    sec_data = [dict(zip(sec_cols, row)) for row in payload["securities"]["data"]]
    md_data = {row[0]: dict(zip(md_cols, row)) for row in payload["marketdata"]["data"]}

    uniq: dict[str, dict[str, Any]] = {}
    for row in sec_data:
        secid = str(row.get("SECID") or "")
        if not secid:
            continue
        row["marketdata"] = md_data.get(secid, {})
        current = _pick_price(row)
        prev = row.get("PREVPRICE")
        score = (current is not None, prev is not None, float(row.get("marketdata", {}).get("VALTODAY") or 0.0))
        if secid not in uniq or score > uniq[secid]["_score"]:
            uniq[secid] = row | {"_score": score}

    return [{k: v for k, v in row.items() if k != "_score"} for row in uniq.values()]


async def _fetch_volume_20d(client: MoexClient) -> dict[str, float]:
    start_dt = (date.today() - timedelta(days=20)).isoformat()
    end_dt = date.today().isoformat()
    aggregated: dict[str, float] = {}
    start = 0
    progress = tqdm(desc="История объема 20д", unit="стр", dynamic_ncols=True)
    while True:
        url = (
            "https://iss.moex.com/iss/history/engines/stock/markets/bonds/securities.json"
            f"?iss.meta=off&iss.only=history&from={start_dt}&till={end_dt}"
            "&history.columns=SECID,VOLUME"
            f"&start={start}"
        )
        payload = await client.get_json(url)
        cols = payload["history"]["columns"]
        rows = payload["history"]["data"]
        if not rows:
            break
        for row in rows:
            data = dict(zip(cols, row))
            secid = str(data.get("SECID") or "")
            if secid:
                aggregated[secid] = aggregated.get(secid, 0.0) + float(data.get("VOLUME") or 0.0)
        progress.update(len(rows))
        start += len(rows)
        if len(rows) < 100:
            break
    progress.close()
    return aggregated


async def _fetch_descriptions_with_cache(
    client: MoexClient,
    db: Database,
    secids: list[str],
    semaphore: asyncio.Semaphore,
) -> tuple[dict[str, dict[str, Any]], int, int]:
    min_ts = int(time.time()) - config.CACHE_TTL_SEC
    cached_raw, missing = db.get_cached_descriptions(secids, min_ts)
    result: dict[str, dict[str, Any]] = {secid: json.loads(payload) for secid, payload in cached_raw.items()}

    async def fetch_one(secid: str) -> tuple[str, dict[str, Any]]:
        async with semaphore:
            url = f"https://iss.moex.com/iss/securities/{secid}.json?iss.meta=off&iss.only=description"
            payload = await client.get_json(url)
            cols = payload["description"]["columns"]
            values: dict[str, Any] = {}
            for row in payload["description"]["data"]:
                item = dict(zip(cols, row))
                key = str(item.get("name") or "")
                if key:
                    values[key] = item.get("value")
            return secid, values

    cache_hits = len(result)
    upserts: list[tuple[str, str, int]] = []
    errors = 0
    tasks = [asyncio.create_task(fetch_one(secid)) for secid in missing]
    progress = tqdm(total=len(tasks), desc="Описание облигаций", unit="обл", dynamic_ncols=True)
    for task in asyncio.as_completed(tasks):
        try:
            secid, desc = await task
            result[secid] = desc
            upserts.append((secid, json.dumps(desc, ensure_ascii=False), int(time.time())))
        except Exception as exc:
            errors += 1
            LOGGER.warning("Ошибка загрузки описания: %s", exc)
        progress.update(1)
    progress.close()
    db.upsert_descriptions(upserts)
    return result, cache_hits, errors


async def _fetch_emitters_with_cache(
    client: MoexClient,
    db: Database,
    emitter_ids: set[int],
    semaphore: asyncio.Semaphore,
) -> tuple[dict[int, dict[str, str]], int, int]:
    emitter_list = sorted(emitter_ids)
    min_ts = int(time.time()) - config.CACHE_TTL_SEC
    cached, missing = db.get_cached_emitters(emitter_list, min_ts)
    result = dict(cached)

    async def fetch_one(emitter_id: int) -> tuple[int, dict[str, str]]:
        async with semaphore:
            payload = await client.get_json(f"https://iss.moex.com/iss/emitters/{emitter_id}.json?iss.meta=off")
            cols = payload["emitter"]["columns"]
            data_rows = payload["emitter"]["data"]
            if not data_rows:
                return emitter_id, {"name": "", "inn": ""}
            data = dict(zip(cols, data_rows[0]))
            return emitter_id, {"name": str(data.get("TITLE") or ""), "inn": str(data.get("INN") or "")}

    tasks = [asyncio.create_task(fetch_one(emitter_id)) for emitter_id in missing]
    upserts: list[tuple[int, str, str, int]] = []
    errors = 0
    progress = tqdm(total=len(tasks), desc="Загрузка эмитентов", unit="эмит", dynamic_ncols=True)
    for task in asyncio.as_completed(tasks):
        try:
            emitter_id, emitter_data = await task
            result[emitter_id] = emitter_data
            upserts.append((emitter_id, emitter_data["name"], emitter_data["inn"], int(time.time())))
        except Exception as exc:
            errors += 1
            LOGGER.warning("Не удалось загрузить эмитента: %s", exc)
        progress.update(1)
    progress.close()
    db.upsert_emitters(upserts)
    return result, len(cached), errors


async def _fetch_amortizations_with_cache(
    client: MoexClient,
    db: Database,
    secids: list[str],
    semaphore: asyncio.Semaphore,
) -> tuple[dict[str, date | None], int, int]:
    min_ts = int(time.time()) - config.CACHE_TTL_SEC
    cached_raw, missing = db.get_cached_amortizations(secids, min_ts)
    result: dict[str, date | None] = {secid: _as_date(value) for secid, value in cached_raw.items()}
    cache_hits = len(result)

    async def fetch_one(secid: str) -> tuple[str, str]:
        async with semaphore:
            url = f"https://iss.moex.com/iss/securities/{secid}/bondization.json?iss.meta=off&iss.only=amortizations"
            payload = await client.get_json(url)
            rows = payload.get("amortizations", {}).get("data", [])
            cols = payload.get("amortizations", {}).get("columns", [])
            if not rows or not cols:
                return secid, ""
            earliest: date | None = None
            for row in rows:
                data = dict(zip(cols, row))
                parsed = _as_date(data.get("amortdate"))
                if parsed is None:
                    continue
                has_amort = False
                face_value = data.get("facevalue")
                initial_face_value = data.get("initialfacevalue")
                value_prc = data.get("valueprc")
                if face_value is not None and initial_face_value is not None:
                    try:
                        has_amort = float(face_value) < float(initial_face_value)
                    except Exception:
                        has_amort = False
                if not has_amort and value_prc is not None:
                    try:
                        has_amort = float(value_prc) < 100.0
                    except Exception:
                        has_amort = False
                if has_amort and (earliest is None or parsed < earliest):
                    earliest = parsed
            return secid, earliest.isoformat() if earliest else ""

    tasks = [asyncio.create_task(fetch_one(secid)) for secid in missing]
    upserts: list[tuple[str, str, int]] = []
    errors = 0
    progress = tqdm(total=len(tasks), desc="Амортизация", unit="обл", dynamic_ncols=True)
    for task in asyncio.as_completed(tasks):
        try:
            secid, amort = await task
            result[secid] = _as_date(amort)
            upserts.append((secid, amort, int(time.time())))
        except Exception as exc:
            errors += 1
            LOGGER.warning("Не удалось загрузить амортизацию: %s", exc)
        progress.update(1)
    progress.close()
    db.upsert_amortizations(upserts)
    return result, cache_hits, errors


async def _fetch_corpbonds_with_cache(
    client: MoexClient,
    db: Database,
    secids: list[str],
    semaphore: asyncio.Semaphore,
) -> tuple[dict[str, dict[str, str]], int, int]:
    min_ts = int(time.time()) - config.CORPBONDS_CACHE_TTL_SEC
    cached, missing = db.get_cached_corpbonds(secids, min_ts)
    result = dict(cached)

    async def fetch_one(secid: str) -> tuple[str, dict[str, str]]:
        async with semaphore:
            url = config.CORPBONDS_BOND_URL_TEMPLATE.format(secid=secid)
            text = await client.get_text(url, headers={"User-Agent": config.CORPBONDS_USER_AGENT})
            return secid, _parse_corpbonds_html(text)

    tasks = [asyncio.create_task(fetch_one(secid)) for secid in missing]
    upserts: list[tuple[str, str, str, str, str, str, str, int]] = []
    errors = 0
    progress = tqdm(total=len(tasks), desc="CorpBonds", unit="обл", dynamic_ncols=True)
    for task in asyncio.as_completed(tasks):
        try:
            secid, data = await task
            result[secid] = data
            upserts.append(
                (
                    secid,
                    data["price"],
                    data["credit_rating"],
                    data["coupon_type"],
                    data["coupon_formula"],
                    data["nearest_offer_date"],
                    data["ladder_coupon"],
                    int(time.time()),
                )
            )
        except Exception as exc:
            errors += 1
            LOGGER.warning("Не удалось загрузить CorpBonds: %s", exc)
        progress.update(1)
    progress.close()
    db.upsert_corpbonds(upserts)
    return result, len(cached), errors


def _format_ws_base(ws: Any) -> None:
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _save_moex_excel(rows: list[BondRow], output_path: Path) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = config.EXCEL_SHEET_NAME

    headers = [
        "Secid",
        "ISIN",
        "Короткое название",
        "Наименование Эмитента",
        "ИНН эмитента",
        "Актуальная Цена сейчас",
        "Предыдущая цена выгрузки",
        "Динамика цены, %",
        "Объем сделок по бумаге",
        "Объем сделок за 20 дней",
        "Дата погашения",
        "Дата оферты",
        "Дата начала амортизации",
        "Квалифицированный инвестор",
        "Дефолт",
        "Технический дефолт",
        "BOND_TYPE",
        "SECSUBTYPE",
        "Купонный период",
        "НКД",
        "% купона",
    ]

    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(
            [
                row.secid,
                row.isin,
                row.short_name,
                row.emitter_name,
                row.emitter_inn,
                row.current_price,
                row.previous_price,
                row.price_change_percent,
                row.volume_today,
                row.volume_20d,
                row.maturity_date,
                row.offer_date,
                row.amortization_start,
                row.qualified_investor,
                row.default_flag,
                row.technical_default_flag,
                row.bond_type,
                row.sec_sub_type,
                row.coupon_period,
                row.accrued_int,
                row.coupon_percent,
            ]
        )

    for row_idx in range(2, ws.max_row + 1):
        for col in (11, 12, 13):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"

    _format_ws_base(ws)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(f"H2:H{ws.max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill))
    ws.conditional_formatting.add(f"H2:H{ws.max_row}", CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))

    wb.save(output_path)
    return len(rows)


def _save_corpbonds_excel(rows: list[CorpBondRow], output_path: Path) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = config.CORPBONDS_EXCEL_SHEET_NAME

    headers = [
        "Secid",
        "Цена",
        "Рейтинг",
        "Тип купона",
        "Формула купона",
        "Дата ближайшей оферты",
        "Купон лесенкой",
    ]

    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(
            [
                row.secid,
                row.price,
                row.credit_rating,
                row.coupon_type,
                row.coupon_formula,
                row.nearest_offer_date,
                row.ladder_coupon,
            ]
        )

    _format_ws_base(ws)
    wb.save(output_path)
    return len(rows)


def _save_merged_excel(moex_rows: list[BondRow], corpbonds_rows: list[CorpBondRow], output_path: Path) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = config.MERGED_EXCEL_SHEET_NAME

    headers = [
        "Secid",
        "ISIN",
        "Короткое название",
        "Наименование Эмитента",
        "ИНН эмитента",
        "Актуальная Цена сейчас",
        "Предыдущая цена выгрузки",
        "Динамика цены, %",
        "Объем сделок по бумаге",
        "Объем сделок за 20 дней",
        "Дата погашения",
        "Дата оферты",
        "Дата начала амортизации",
        "Квалифицированный инвестор",
        "Дефолт",
        "Технический дефолт",
        "BOND_TYPE",
        "SECSUBTYPE",
        "Купонный период",
        "НКД",
        "% купона",
        "YTM, %",
        "Цена CorpBonds",
        "Рейтинг",
        "Тип купона",
        "Формула купона",
        "Купон лесенкой",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    corpbonds_by_secid = {row.secid: row for row in corpbonds_rows}
    for row in moex_rows:
        corp = corpbonds_by_secid.get(row.secid)
        corp_offer_date = _as_date(corp.nearest_offer_date) if corp else None
        merged_offer_date = row.offer_date if row.offer_date is not None else corp_offer_date
        ws.append(
            [
                row.secid,
                row.isin,
                row.short_name,
                row.emitter_name,
                row.emitter_inn,
                row.current_price,
                row.previous_price,
                row.price_change_percent,
                row.volume_today,
                row.volume_20d,
                row.maturity_date,
                merged_offer_date,
                row.amortization_start,
                row.qualified_investor,
                row.default_flag,
                row.technical_default_flag,
                row.bond_type,
                row.sec_sub_type,
                row.coupon_period,
                row.accrued_int,
                row.coupon_percent,
                row.ytm,
                corp.price if corp else "",
                corp.credit_rating if corp else "",
                corp.coupon_type if corp else "",
                corp.coupon_formula if corp else "",
                corp.ladder_coupon if corp else "",
            ]
        )

    for row_idx in range(2, ws.max_row + 1):
        for col in (11, 12, 13):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell.value, date):
                cell.number_format = "yyyy-mm-dd"

    _format_ws_base(ws)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    ws.conditional_formatting.add(f"H2:H{ws.max_row}", CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill))
    ws.conditional_formatting.add(f"H2:H{ws.max_row}", CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))

    wb.save(output_path)
    return len(moex_rows)


def _read_existing_emitents_scores(path: Path) -> dict[tuple[str, str], tuple[str, str]]:
    if not path.exists():
        return {}
    wb = load_workbook(path)
    if config.EMITENTS_EXCEL_SHEET_NAME not in wb.sheetnames:
        return {}
    ws = wb[config.EMITENTS_EXCEL_SHEET_NAME]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    indexes = {name: idx for idx, name in enumerate(headers)}
    required = ["Наименование Эмитента", "ИНН эмитента", "ScoreList", "DateScoreList"]
    if any(name not in indexes for name in required):
        return {}
    result: dict[tuple[str, str], tuple[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        emitter = str(row[indexes["Наименование Эмитента"]] or "").strip()
        inn = str(row[indexes["ИНН эмитента"]] or "").strip()
        score = str(row[indexes["ScoreList"]] or "").strip()
        score_date = str(row[indexes["DateScoreList"]] or "").strip()
        if emitter or inn:
            result[(emitter, inn)] = (score, score_date)
    return result


def _save_emitents_excel(moex_rows: list[BondRow], corpbonds_rows: list[CorpBondRow], output_path: Path) -> int:
    wb = Workbook()
    ws = wb.active
    ws.title = config.EMITENTS_EXCEL_SHEET_NAME

    headers = ["Наименование Эмитента", "ИНН эмитента", "Рейтинг", "ScoreList", "DateScoreList"]
    ws.append(headers)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    existing_scores = _read_existing_emitents_scores(output_path)
    rating_by_secid = {row.secid: row.credit_rating for row in corpbonds_rows}
    grouped: dict[tuple[str, str], set[str]] = {}
    for row in moex_rows:
        key = (row.emitter_name.strip(), row.emitter_inn.strip())
        rating = rating_by_secid.get(row.secid, "").strip()
        grouped.setdefault(key, set())
        if rating:
            grouped[key].add(rating)

    today = datetime.now().date().isoformat()
    for emitter_key in sorted(grouped.keys()):
        emitter_name, emitter_inn = emitter_key
        ratings = "; ".join(sorted(grouped[emitter_key]))
        old_score, old_date = existing_scores.get(emitter_key, ("", ""))
        score_value = old_score if old_score in config.SCORE_LIST_ALLOWED_VALUES else ""
        date_value = old_date
        if score_value and not date_value:
            date_value = today
        ws.append([emitter_name, emitter_inn, ratings, score_value, date_value])

    if ws.max_row >= 2:
        score_column = "D"
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(config.SCORE_LIST_ALLOWED_VALUES) + '"',
            allow_blank=True,
            showErrorMessage=True,
        )
        dv.errorTitle = "Недопустимое значение"
        dv.error = "Выберите значение из списка: GreenList, YellowList, RedList"
        ws.add_data_validation(dv)
        dv.add(f"{score_column}2:{score_column}{ws.max_row}")

        date_column_index = 5
        for row_idx in range(2, ws.max_row + 1):
            score = str(ws.cell(row=row_idx, column=4).value or "").strip()
            date_cell = ws.cell(row=row_idx, column=date_column_index)
            if score and not date_cell.value:
                date_cell.value = today

    _format_ws_base(ws)
    wb.save(output_path)
    return max(0, ws.max_row - 1)


def _days_to(target: date | None, today: date) -> int | None:
    if target is None:
        return None
    return (target - today).days


def _filter_bonds_for_screener(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    filters = config.SCREENER_FILTERS
    today = datetime.now().date()
    result: list[dict[str, Any]] = []
    for row in rows:
        amort_days = _days_to(_as_date(row.get("Дата начала амортизации")), today)
        maturity_days = _days_to(_as_date(row.get("Дата погашения")), today)
        offer_days = _days_to(_as_date(row.get("Дата оферты")), today)

        exclude = False
        amort_filter = filters.get("exclude_amortization_started_or_soon", {})
        if amort_filter.get("enabled", False) and amort_days is not None and amort_days <= int(amort_filter.get("days", 365)):
            exclude = True

        structural_filter = filters.get("exclude_structural_bonds", {})
        if structural_filter.get("enabled", False) and str(row.get("BOND_TYPE") or "").strip() == "Структурная облигация":
            exclude = True

        default_filter = filters.get("exclude_defaults", {})
        if default_filter.get("enabled", False) and str(row.get("Дефолт") or "").strip() == "Да":
            exclude = True

        maturity_filter = filters.get("exclude_maturity_soon", {})
        if maturity_filter.get("enabled", False) and maturity_days is not None and maturity_days < int(maturity_filter.get("days", 365)):
            exclude = True

        offer_filter = filters.get("exclude_offer_soon", {})
        if offer_filter.get("enabled", False) and offer_days is not None and offer_days < int(offer_filter.get("days", 365)):
            exclude = True

        qualified_filter = filters.get("exclude_qualified_investors", {})
        if qualified_filter.get("enabled", False) and str(row.get("Квалифицированный инвестор") or "").strip() == "Да":
            exclude = True

        if not exclude:
            result.append(row)
    return result


def _save_screener_excel(moex_rows: list[BondRow], corpbonds_rows: list[CorpBondRow], output_path: Path, emitents_path: Path) -> int:
    corpbonds_by_secid = {row.secid: row for row in corpbonds_rows}
    emitent_scores = _read_existing_emitents_scores(emitents_path)

    merged_rows: list[dict[str, Any]] = []
    for row in moex_rows:
        corp = corpbonds_by_secid.get(row.secid)
        corp_offer_date = _as_date(corp.nearest_offer_date) if corp else None
        merged_offer_date = row.offer_date if row.offer_date is not None else corp_offer_date
        emitter_key = (row.emitter_name.strip(), row.emitter_inn.strip())
        score, score_date = emitent_scores.get(emitter_key, ("", ""))
        merged_rows.append(
            {
                "Secid": row.secid,
                "ISIN": row.isin,
                "Короткое название": row.short_name,
                "Наименование Эмитента": row.emitter_name,
                "ИНН эмитента": row.emitter_inn,
                "Актуальная Цена сейчас": row.current_price,
                "Предыдущая цена выгрузки": row.previous_price,
                "Динамика цены, %": row.price_change_percent,
                "Объем сделок по бумаге": row.volume_today,
                "Объем сделок за 20 дней": row.volume_20d,
                "Дата погашения": row.maturity_date,
                "Дата оферты": merged_offer_date,
                "Дата начала амортизации": row.amortization_start,
                "Квалифицированный инвестор": row.qualified_investor,
                "Дефолт": row.default_flag,
                "Технический дефолт": row.technical_default_flag,
                "BOND_TYPE": row.bond_type,
                "SECSUBTYPE": row.sec_sub_type,
                "Купонный период": row.coupon_period,
                "НКД": row.accrued_int,
                "% купона": row.coupon_percent,
                "YTM, %": row.ytm,
                "Цена CorpBonds": corp.price if corp else "",
                "Рейтинг": corp.credit_rating if corp else "",
                "Тип купона": corp.coupon_type if corp else "",
                "Формула купона": corp.coupon_formula if corp else "",
                "Купон лесенкой": corp.ladder_coupon if corp else "",
                "ScoreList": score,
                "DateScoreList": score_date,
            }
        )

    filtered_rows = _filter_bonds_for_screener(merged_rows)
    default_headers = list(merged_rows[0].keys()) if merged_rows else []
    headers = config.SCREENER_INCLUDE_COLUMNS[:] if config.SCREENER_INCLUDE_COLUMNS else default_headers
    headers = [col for col in headers if col not in set(config.SCREENER_EXCLUDE_COLUMNS)]

    grouped = {"Green": [], "Yellow": [], "Red": [], "Unsorted": []}
    for row in filtered_rows:
        score = str(row.get("ScoreList") or "").strip()
        if score == "GreenList":
            grouped["Green"].append(row)
        elif score == "YellowList":
            grouped["Yellow"].append(row)
        elif score == "RedList":
            grouped["Red"].append(row)
        else:
            grouped["Unsorted"].append(row)

    wb = Workbook()
    first_sheet = True
    for sheet_name in ("Green", "Yellow", "Red", "Unsorted"):
        ws = wb.active if first_sheet else wb.create_sheet(title=sheet_name)
        first_sheet = False
        ws.title = sheet_name
        ws.append(headers)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in grouped[sheet_name]:
            ws.append([row.get(h) for h in headers])

        date_columns = [idx + 1 for idx, h in enumerate(headers) if h in {"Дата погашения", "Дата оферты", "Дата начала амортизации"}]
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in date_columns:
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, date):
                    cell.number_format = "yyyy-mm-dd"
        _format_ws_base(ws)

    wb.save(output_path)
    return len(filtered_rows)



async def run_pipeline(db: Database) -> RunSummary:
    total_start = time.perf_counter()
    state = load_state()
    load_start = time.perf_counter()
    errors_count = 0
    from_cache_count = 0

    semaphore = asyncio.Semaphore(config.MAX_CONCURRENT_TASKS)
    corpbonds_semaphore = asyncio.Semaphore(max(config.MAX_CONCURRENT_TASKS, config.CORPBONDS_MAX_CONCURRENT_TASKS))
    bonds: list[dict[str, Any]] = []
    descriptions: dict[str, dict[str, Any]] = {}
    emitters: dict[int, dict[str, str]] = {}
    amortizations: dict[str, date | None] = {}
    corpbonds_data: dict[str, dict[str, str]] = {}
    key_rate = config.YTM_KEY_RATE_FORECAST[0]
    ruonia = key_rate
    gcurve_5y = key_rate
    gcurve_7y = key_rate

    async with MoexClient() as client:
        LOGGER.info("Этап 1/5: загрузка списка облигаций")
        bonds = await _fetch_all_traded_bonds(client)

        LOGGER.info("Этап 2/5: загрузка истории объемов")
        volume_20d = await _fetch_volume_20d(client)

        secids = [str(bond.get("SECID")) for bond in bonds if bond.get("SECID")]

        LOGGER.info("Этап 3/5: загрузка описаний и эмитентов")
        descriptions, cache_desc_count, desc_errors = await _fetch_descriptions_with_cache(client, db, secids, semaphore)
        from_cache_count += cache_desc_count
        errors_count += desc_errors

        emitter_ids: set[int] = set()
        for desc in descriptions.values():
            emitter_id = desc.get("EMITTER_ID")
            if emitter_id is not None:
                try:
                    emitter_ids.add(int(emitter_id))
                except Exception:
                    continue
        emitters, cache_emit_count, emit_errors = await _fetch_emitters_with_cache(client, db, emitter_ids, semaphore)
        from_cache_count += cache_emit_count
        errors_count += emit_errors
        save_state({"processed_ids": list(descriptions.keys()), "last_stage": "description"})

        LOGGER.info("Этап 4/5: загрузка амортизаций")
        amortizations, cache_amort_count, amort_errors = await _fetch_amortizations_with_cache(client, db, secids, semaphore)
        from_cache_count += cache_amort_count
        errors_count += amort_errors

        LOGGER.info("Этап 5/5: загрузка данных CorpBonds")
        corpbonds_data, cache_corp_count, corp_errors = await _fetch_corpbonds_with_cache(
            client, db, secids, corpbonds_semaphore
        )
        from_cache_count += cache_corp_count
        errors_count += corp_errors

        LOGGER.info("Загрузка индикаторов ЦБ для YTM")
        key_rate, ruonia, gcurve_5y, gcurve_7y = await _fetch_cbr_curve(client)

    duration_load = time.perf_counter() - load_start

    calc_start = time.perf_counter()
    previous_prices = db.fetch_previous_prices()
    prepared_rows: list[BondRow] = []
    corpbonds_rows: list[CorpBondRow] = []

    progress_calc = tqdm(total=len(bonds), desc="Подготовка строк", unit="обл", dynamic_ncols=True)
    for bond in bonds:
        secid = str(bond.get("SECID") or "")
        if not secid:
            progress_calc.update(1)
            continue

        desc = descriptions.get(secid, {})
        emitter_id_raw = desc.get("EMITTER_ID")
        emitter_info = {"name": "", "inn": ""}
        if emitter_id_raw is not None:
            try:
                emitter_info = emitters.get(int(emitter_id_raw), emitter_info)
            except Exception:
                pass

        current_price = _pick_price(bond)
        previous_price = previous_prices.get(secid)
        if previous_price not in (None, 0) and current_price is not None:
            price_change = ((current_price - previous_price) / previous_price) * 100
        else:
            price_change = None

        corp = corpbonds_data.get(secid, {})
        accrued_int = float(bond.get("ACCRUEDINT")) if bond.get("ACCRUEDINT") is not None else None
        coupon_percent = float(bond.get("COUPONPERCENT")) if bond.get("COUPONPERCENT") is not None else None

        prepared_rows.append(
            BondRow(
                secid=secid,
                isin=str(bond.get("ISIN") or ""),
                short_name=str(bond.get("SHORTNAME") or ""),
                emitter_name=emitter_info["name"],
                emitter_inn=emitter_info["inn"],
                current_price=current_price,
                previous_price=previous_price,
                price_change_percent=round(price_change, 4) if price_change is not None else None,
                volume_today=float(bond.get("marketdata", {}).get("VOLTODAY") or 0.0),
                volume_20d=round(volume_20d.get(secid, 0.0), 2),
                maturity_date=_as_date(bond.get("MATDATE")),
                offer_date=_as_date(bond.get("OFFERDATE")),
                amortization_start=amortizations.get(secid),
                qualified_investor=_to_yes_no(desc.get("ISQUALIFIEDINVESTORS")),
                default_flag=_to_yes_no(desc.get("HASDEFAULT")),
                technical_default_flag=_to_yes_no(desc.get("HASTECHNICALDEFAULT")),
                bond_type=str(desc.get("BOND_TYPE") or bond.get("BONDTYPE") or ""),
                sec_sub_type=str(desc.get("BOND_SUBTYPE") or bond.get("BONDSUBTYPE") or ""),
                coupon_period=int(bond.get("COUPONPERIOD")) if bond.get("COUPONPERIOD") is not None else None,
                accrued_int=accrued_int,
                coupon_percent=coupon_percent,
                ytm=_calc_ytm_percent(
                    clean_price=_select_last_price(current_price, corp.get("price", "")),
                    accrued_int=accrued_int,
                    coupon_percent=coupon_percent,
                    coupon_period_days=int(bond.get("COUPONPERIOD")) if bond.get("COUPONPERIOD") is not None else None,
                    coupon_type=corp.get("coupon_type", ""),
                    coupon_formula=corp.get("coupon_formula", ""),
                    secid=secid,
                    end_date=_as_date(bond.get("OFFERDATE")) or _as_date(corp.get("nearest_offer_date")) or _as_date(bond.get("MATDATE")),
                    key_rate=key_rate,
                    ruonia=ruonia,
                    gcurve_5y=gcurve_5y,
                    gcurve_7y=gcurve_7y,
                ),
            )
        )

        corpbonds_rows.append(
            CorpBondRow(
                secid=secid,
                price=corp.get("price", ""),
                credit_rating=corp.get("credit_rating", ""),
                coupon_type=corp.get("coupon_type", ""),
                coupon_formula=corp.get("coupon_formula", ""),
                nearest_offer_date=corp.get("nearest_offer_date", ""),
                ladder_coupon=corp.get("ladder_coupon", ""),
            )
        )
        progress_calc.update(1)
    progress_calc.close()

    prepared_rows.sort(key=lambda x: x.secid)
    corpbonds_rows.sort(key=lambda x: x.secid)
    db.upsert_snapshot([(row.secid, row.current_price, datetime.now(timezone.utc).isoformat()) for row in prepared_rows])
    save_state({"processed_ids": [x.secid for x in prepared_rows], "last_stage": "calc", "prev_stage": state.get("last_stage", "init")})
    duration_calc = time.perf_counter() - calc_start

    save_start = time.perf_counter()
    saved_count = 0
    moex_output_path: Path | None = None
    corpbonds_output_path: Path | None = None
    merged_output_path: Path | None = None
    emitents_output_path: Path | None = None
    screener_output_path: Path | None = None

    if config.EXPORT_MOEX_TO_EXCEL:
        moex_output_path = config.get_moex_output_file_path()
        saved_count += _save_moex_excel(prepared_rows, moex_output_path)
    if config.EXPORT_CORPBONDS_TO_EXCEL:
        corpbonds_output_path = config.get_corpbonds_output_file_path()
        _save_corpbonds_excel(corpbonds_rows, corpbonds_output_path)
    if config.EXPORT_MOEX_TO_EXCEL and config.EXPORT_CORPBONDS_TO_EXCEL:
        merged_output_path = config.get_merged_output_file_path()
        _save_merged_excel(prepared_rows, corpbonds_rows, merged_output_path)

    emitents_output_path = config.get_emitents_output_file_path()
    _save_emitents_excel(prepared_rows, corpbonds_rows, emitents_output_path)

    screener_output_path = config.get_screener_output_file_path()
    saved_count = _save_screener_excel(prepared_rows, corpbonds_rows, screener_output_path, emitents_output_path)

    save_state({"processed_ids": [x.secid for x in prepared_rows], "last_stage": "done"})
    duration_save = time.perf_counter() - save_start

    return RunSummary(
        fetched_count=len(bonds),
        selected_count=len(prepared_rows),
        saved_count=saved_count,
        errors_count=errors_count,
        from_cache_count=from_cache_count,
        duration_total=time.perf_counter() - total_start,
        duration_load=duration_load,
        duration_calc=duration_calc,
        duration_save=duration_save,
        moex_output_path=moex_output_path,
        corpbonds_output_path=corpbonds_output_path,
        merged_output_path=merged_output_path,
        emitents_output_path=emitents_output_path,
        screener_output_path=screener_output_path,
    )
