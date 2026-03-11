from __future__ import annotations

import csv
import base64
import html
import io
import json
import logging
import os
import random
import re
import sqlite3
import tempfile
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta, timezone
from pathlib import Path
from time import perf_counter
from urllib.parse import urljoin

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from bs4 import BeautifulSoup
from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import DataBar, FormatObject, Rule
from playwright.sync_api import Error as PWError
from playwright.sync_api import TimeoutError as PWTimeoutError
from playwright.sync_api import sync_playwright
from tqdm import tqdm

import config


def progress(total: int, desc: str, unit: str, position: int = 0):
    return tqdm(total=total, desc=desc, unit=unit, position=position, leave=False, dynamic_ncols=True)


def setup_logging() -> logging.Logger:
    config.LOGS_DIR.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("bonds_main")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    handler = logging.FileHandler(config.LOGS_DIR / config.LOG_FILENAME, mode="w", encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(handler)
    return logger


def create_http_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": config.NRA_REQUEST_USER_AGENT})
    return session


def create_resilient_http_session(pool_size: int = 64) -> requests.Session:
    retry = Retry(
        total=2,
        connect=2,
        read=2,
        backoff_factor=0.2,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET"}),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(pool_connections=pool_size, pool_maxsize=pool_size, max_retries=retry)
    session = requests.Session()
    session.headers.update({
        "User-Agent": config.NRA_REQUEST_USER_AGENT,
        "Accept": "application/json,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
    })
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


def ensure_directories() -> None:
    for folder in (
        config.RAW_DIR,
        config.CACHE_DIR,
        config.DB_DIR,
        config.BASE_SNAPSHOTS_DIR,
        config.LOGS_DIR,
        config.DOCS_DIR,
    ):
        folder.mkdir(parents=True, exist_ok=True)


def migrate_legacy_db_if_needed() -> None:
    legacy_path = config.DB_DIR / config.LEGACY_DB_FILENAME
    target_path = config.DB_DIR / config.DB_FILENAME
    if target_path.exists() or not legacy_path.exists():
        return
    legacy_path.replace(target_path)


def migrate_legacy_rates_table_if_needed(conn: sqlite3.Connection) -> None:
    if config.RATES_TABLE_NAME == config.LEGACY_RATES_TABLE_NAME:
        return
    table_exists = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name = ?",
        (config.RATES_TABLE_NAME,),
    ).fetchone()
    if table_exists:
        return
    legacy_exists = conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name = ?",
        (config.LEGACY_RATES_TABLE_NAME,),
    ).fetchone()
    if legacy_exists:
        conn.execute(
            f'ALTER TABLE "{config.LEGACY_RATES_TABLE_NAME}" RENAME TO "{config.RATES_TABLE_NAME}"'
        )
        conn.commit()


def connect_db(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    conn.execute("PRAGMA temp_store=MEMORY;")
    conn.execute("PRAGMA cache_size=-200000;")
    conn.execute("PRAGMA mmap_size=1073741824;")
    conn.execute("PRAGMA foreign_keys=ON;")
    return conn


def init_meta_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {config.META_TABLE_NAME} (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )
    conn.commit()


def get_meta_value(conn: sqlite3.Connection, key: str) -> str | None:
    row = conn.execute(
        f"SELECT value FROM {config.META_TABLE_NAME} WHERE key = ?",
        (key,),
    ).fetchone()
    return row[0] if row else None


def set_meta_value(conn: sqlite3.Connection, key: str, value: str) -> None:
    conn.execute(
        f"""
        INSERT INTO {config.META_TABLE_NAME}(key, value)
        VALUES(?, ?)
        ON CONFLICT(key) DO UPDATE SET value=excluded.value
        """,
        (key, value),
    )
    conn.commit()


def should_refresh_cache(conn: sqlite3.Connection, now_utc: datetime) -> bool:
    last_refresh_raw = get_meta_value(conn, "last_refresh_utc")
    rows_count_raw = get_meta_value(conn, "last_rows_count")

    if not last_refresh_raw or not rows_count_raw:
        return True

    try:
        last_refresh = datetime.fromisoformat(last_refresh_raw)
        rows_count = int(rows_count_raw)
    except ValueError:
        return True

    if rows_count <= 0:
        return True

    ttl = timedelta(hours=config.CACHE_TTL_HOURS)
    return now_utc - last_refresh >= ttl


def download_csv(url: str, timeout_seconds: int | float) -> str:
    response = requests.get(url, timeout=timeout_seconds)
    response.raise_for_status()
    for encoding in ("utf-8-sig", "cp1251", response.encoding or "utf-8"):
        try:
            return response.content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return response.text


def parse_moex_rates_csv(raw_text: str) -> tuple[list[str], list[list[str]]]:
    lines = raw_text.splitlines()
    header_index = None
    delimiter = ";"

    for idx, line in enumerate(lines):
        if line.startswith("SECID;"):
            header_index = idx
            delimiter = ";"
            break
        if line.startswith("SECID\t"):
            header_index = idx
            delimiter = "\t"
            break

    if header_index is None:
        raise ValueError("Не удалось найти строку заголовков (SECID...) в CSV.")

    data_lines = [line for line in lines[header_index:] if line.strip()]
    reader = csv.reader(data_lines, delimiter=delimiter)

    rows = list(reader)
    headers = [h.strip() for h in rows[0]]
    values: list[list[str]] = []

    for row in rows[1:]:
        if not any(cell.strip() for cell in row):
            continue
        normalized = row + [""] * (len(headers) - len(row))
        values.append([cell.strip() for cell in normalized[: len(headers)]])

    return headers, values


def replace_rates_table(conn: sqlite3.Connection, headers: list[str], rows: list[list[str]]) -> None:
    quoted_columns = ", ".join([f'"{col}" TEXT' for col in headers])
    quoted_headers = ", ".join([f'"{col}"' for col in headers])
    placeholders = ", ".join(["?"] * len(headers))

    conn.execute(f'DROP TABLE IF EXISTS "{config.RATES_TABLE_NAME}"')
    conn.execute(f'CREATE TABLE "{config.RATES_TABLE_NAME}" ({quoted_columns})')

    conn.execute("BEGIN")
    conn.executemany(
        f'INSERT INTO "{config.RATES_TABLE_NAME}" ({quoted_headers}) VALUES ({placeholders})',
        rows,
    )
    conn.commit()


def ensure_emitents_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.EMITENTS_TABLE_NAME}" (
            "INN" TEXT PRIMARY KEY,
            "EMITENTNAME" TEXT NOT NULL,
            "Scoring" TEXT,
            "DateScoring" TEXT,
            "NRA_Rate" TEXT,
            "Acra_Rate" TEXT,
            "NKR_Rate" TEXT,
            "RAEX_Rate" TEXT
        )
        '''
    )
    conn.execute(
        f'ALTER TABLE "{config.EMITENTS_TABLE_NAME}" ADD COLUMN "NRA_Rate" TEXT'
    ) if _column_absent(conn, config.EMITENTS_TABLE_NAME, "NRA_Rate") else None
    conn.execute(
        f'ALTER TABLE "{config.EMITENTS_TABLE_NAME}" ADD COLUMN "Acra_Rate" TEXT'
    ) if _column_absent(conn, config.EMITENTS_TABLE_NAME, "Acra_Rate") else None
    conn.execute(
        f'ALTER TABLE "{config.EMITENTS_TABLE_NAME}" ADD COLUMN "NKR_Rate" TEXT'
    ) if _column_absent(conn, config.EMITENTS_TABLE_NAME, "NKR_Rate") else None
    conn.execute(
        f'ALTER TABLE "{config.EMITENTS_TABLE_NAME}" ADD COLUMN "RAEX_Rate" TEXT'
    ) if _column_absent(conn, config.EMITENTS_TABLE_NAME, "RAEX_Rate") else None
    conn.commit()


def _column_absent(conn: sqlite3.Connection, table_name: str, column_name: str) -> bool:
    rows = conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
    return column_name not in {row[1] for row in rows}


def ensure_nra_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.NRA_TABLE_NAME}" (
            "id" TEXT,
            "organization_name" TEXT,
            "inn" TEXT,
            "press_release_title" TEXT,
            "press_release_date" TEXT,
            "rating" TEXT,
            "rating_status" TEXT,
            "forecast" TEXT,
            "rating_type" TEXT,
            "organization_sector" TEXT,
            "industry" TEXT,
            "osk" TEXT,
            "isin" TEXT,
            "press_release_link" TEXT,
            "under_watch" TEXT,
            "source_file_name" TEXT,
            "loaded_at_utc" TEXT,
            UNIQUE("id")
        )
        '''
    )
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.NRA_LATEST_TABLE_NAME}" (
            "inn" TEXT PRIMARY KEY,
            "organization_name" TEXT,
            "press_release_date" TEXT,
            "rating" TEXT,
            "rating_status" TEXT,
            "forecast" TEXT
        )
        '''
    )
    conn.commit()


def ensure_nkr_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.NKR_TABLE_NAME}" (
            "id" TEXT,
            "issuer_name" TEXT,
            "rating_date" TEXT,
            "rating" TEXT,
            "outlook" TEXT,
            "tin" TEXT,
            "loaded_at_utc" TEXT,
            UNIQUE("tin", "rating_date", "rating", "outlook")
        )
        '''
    )
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.NKR_LATEST_TABLE_NAME}" (
            "tin" TEXT PRIMARY KEY,
            "issuer_name" TEXT,
            "rating_date" TEXT,
            "rating" TEXT,
            "outlook" TEXT
        )
        '''
    )
    conn.commit()


def ensure_raex_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.RAEX_TABLE_NAME}" (
            "inn" TEXT,
            "company_name" TEXT,
            "rating" TEXT,
            "forecast" TEXT,
            "rating_date" TEXT,
            "company_url" TEXT,
            "loaded_at_utc" TEXT,
            UNIQUE("inn", "rating_date", "rating", "forecast")
        )
        '''
    )
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.RAEX_LATEST_TABLE_NAME}" (
            "inn" TEXT PRIMARY KEY,
            "company_name" TEXT,
            "rating" TEXT,
            "forecast" TEXT,
            "rating_date" TEXT,
            "company_url" TEXT,
            "loaded_at_utc" TEXT
        )
        '''
    )
    conn.commit()


def ensure_dohod_table(conn: sqlite3.Connection, headers: list[str] | None = None) -> None:
    base_columns = ['"ISIN" TEXT PRIMARY KEY']
    if headers:
        for header in headers:
            normalized = header.strip()
            if not normalized or normalized == "ISIN":
                continue
            base_columns.append(f'"{normalized}" TEXT')
    base_columns.append('"loaded_at_utc" TEXT')
    conn.execute(
        f'CREATE TABLE IF NOT EXISTS "{config.DOHOD_TABLE_NAME}" ({", ".join(base_columns)})'
    )
    conn.commit()


def ensure_table_has_columns(conn: sqlite3.Connection, table_name: str, headers: list[str]) -> None:
    existing = {row[1] for row in conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()}
    for header in headers:
        normalized = header.strip()
        if not normalized or normalized in existing:
            continue
        conn.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{normalized}" TEXT')
    conn.commit()


NRA_HEADERS_MAP = {
    "id": "id",
    "Название организации": "organization_name",
    "ИНН": "inn",
    "Название пресс-релиза": "press_release_title",
    "Дата опубликования пресс-релиза": "press_release_date",
    "Рейтинг": "rating",
    "Статус рейтинга": "rating_status",
    "Прогноз": "forecast",
    "Вид рейтинга": "rating_type",
    "Сектор организации": "organization_sector",
    "Отрасль": "industry",
    "ОСК": "osk",
    "ISIN": "isin",
    "Ссылка на пресс релиз": "press_release_link",
    "Под наблюдением": "under_watch",
}

NKR_HEADERS_MAP = {
    "ID": "id",
    "Issuer Name": "issuer_name",
    "Date": "rating_date",
    "Rating": "rating",
    "Outlook": "outlook",
    "TIN": "tin",
}

RU_MONTHS = {
    "янв": 1,
    "фев": 2,
    "мар": 3,
    "апр": 4,
    "май": 5,
    "июн": 6,
    "июл": 7,
    "авг": 8,
    "сен": 9,
    "окт": 10,
    "ноя": 11,
    "дек": 12,
}


def normalize_date_ru(value: str) -> str:
    date_str = (value or "").strip()
    if not date_str:
        return ""
    if re.match(r"^\d{2}\.\d{2}\.\d{4}$", date_str):
        return f"{date_str[6:10]}-{date_str[3:5]}-{date_str[0:2]}"
    match = re.match(r"^(\d{1,2})\s+([А-Яа-я]{3})\s+(\d{4})$", date_str)
    if not match:
        return date_str
    day = int(match.group(1))
    month = RU_MONTHS.get(match.group(2).lower())
    year = int(match.group(3))
    if not month:
        return date_str
    try:
        return datetime(year, month, day).strftime("%Y-%m-%d")
    except ValueError:
        return date_str


def find_nra_excel_link(page_html: str) -> str:
    marker = re.search(r'Выгрузить\s*в\s*Excel', page_html, flags=re.IGNORECASE)
    if marker:
        idx = marker.start()
        tag_start = page_html.rfind('<a ', 0, idx)
        tag_end = page_html.find('>', tag_start)
        if tag_start != -1 and tag_end != -1:
            tag = page_html[tag_start:tag_end]
            href_match = re.search(r'href=["\']([^"\']+)["\']', tag, flags=re.IGNORECASE)
            if href_match:
                return html.unescape(href_match.group(1))

    links = re.findall(r'href=["\']([^"\']+\.(?:xlsx|xls)[^"\']*)["\']', page_html, flags=re.IGNORECASE)
    if not links:
        raise ValueError("На странице НРА не найдена ссылка на Excel-файл.")
    return html.unescape(links[0])


def _normalize_cell(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_nra_excel(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_cell(cell) for cell in rows[0]]
    indexes: dict[str, int] = {}
    for idx, header in enumerate(headers):
        if header in NRA_HEADERS_MAP:
            indexes[NRA_HEADERS_MAP[header]] = idx

    missing = [column for column in NRA_HEADERS_MAP.values() if column not in indexes]
    if missing:
        raise ValueError(f"В выгрузке НРА отсутствуют колонки: {', '.join(missing)}")

    parsed_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        row_dict: dict[str, str] = {}
        for key, idx in indexes.items():
            row_dict[key] = _normalize_cell(row[idx]) if idx < len(row) else ""
        if any(row_dict.values()):
            parsed_rows.append(row_dict)
    return parsed_rows


def _normalize_tin(value: object) -> str:
    raw = _normalize_cell(value)
    if not raw:
        return ""

    compact = raw.replace(" ", "").replace("\u00a0", "").replace(",", ".")
    if re.fullmatch(r"\d+", compact):
        return compact.zfill(10) if len(compact) < 10 else compact

    try:
        as_int = str(int(Decimal(compact)))
        return as_int.zfill(10) if len(as_int) < 10 else as_int
    except (InvalidOperation, ValueError):
        digits = re.sub(r"\D+", "", compact)
        return digits.zfill(10) if digits and len(digits) < 10 else digits


def parse_nkr_excel(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_cell(cell) for cell in rows[0]]
    indexes: dict[str, int] = {}
    for idx, header in enumerate(headers):
        if header in NKR_HEADERS_MAP:
            indexes[NKR_HEADERS_MAP[header]] = idx

    missing = [column for column in NKR_HEADERS_MAP.values() if column not in indexes]
    if missing:
        raise ValueError(f"В выгрузке НКР отсутствуют колонки: {', '.join(missing)}")

    parsed_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        row_dict: dict[str, str] = {}
        for key, idx in indexes.items():
            row_dict[key] = _normalize_cell(row[idx]) if idx < len(row) else ""
        row_dict["tin"] = _normalize_tin(row[indexes["tin"]] if indexes["tin"] < len(row) else "")
        row_dict["rating_date"] = normalize_date_ru(row_dict.get("rating_date", ""))
        if any(row_dict.values()):
            parsed_rows.append(row_dict)
    return parsed_rows


def parse_dohod_excel(content: bytes) -> tuple[list[str], list[dict[str, str]]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [_normalize_cell(cell) for cell in rows[0]]
    if "ISIN" not in headers:
        raise ValueError("В выгрузке Доходъ отсутствует колонка ISIN.")

    parsed_rows: list[dict[str, str]] = []
    for row in rows[1:]:
        row_dict: dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row_dict[header] = _normalize_cell(row[idx]) if idx < len(row) else ""
        if any(row_dict.values()):
            parsed_rows.append(row_dict)
    return headers, parsed_rows


def _normalize_isin(value: str | None) -> str:
    return str(value or "").strip().upper()


def _deduplicate_dohod_rows(rows: list[dict[str, str]], headers: list[str]) -> list[dict[str, str]]:
    """Схлопывает дубли по ISIN, объединяя непустые значения по колонкам."""
    normalized_headers = [header for header in headers if header]
    grouped: dict[str, dict[str, str]] = {}

    for row in rows:
        isin = _normalize_isin(row.get("ISIN", ""))
        if not isin:
            continue

        target = grouped.setdefault(isin, {"ISIN": isin})
        for header in normalized_headers:
            if header == "ISIN":
                continue
            current_value = str(target.get(header, "")).strip()
            next_value = str(row.get(header, "")).strip()
            if not current_value and next_value:
                target[header] = next_value

    return list(grouped.values())


def should_refresh_dohod(conn: sqlite3.Connection, now_utc: datetime) -> bool:
    last_refresh_raw = get_meta_value(conn, "dohod_last_refresh_utc")
    rows_count_raw = get_meta_value(conn, "dohod_last_rows_count")
    if not last_refresh_raw or not rows_count_raw:
        return True
    try:
        last_refresh = datetime.fromisoformat(last_refresh_raw)
        rows_count = int(rows_count_raw)
    except ValueError:
        return True
    if rows_count <= 0:
        return True
    return now_utc - last_refresh >= timedelta(hours=config.DOHOD_CACHE_TTL_HOURS)


def download_dohod_excel_via_playwright(logger: logging.Logger) -> bytes:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=config.DOHOD_HEADLESS, channel=config.DOHOD_BROWSER_CHANNEL)
        context = browser.new_context(locale="ru-RU", timezone_id="Europe/Moscow", accept_downloads=True)
        try:
            page = context.new_page()
            page.goto(config.DOHOD_BONDS_PAGE_URL, wait_until="domcontentloaded", timeout=int(config.REQUEST_TIMEOUT_SECONDS * 1000))
            button = page.get_by_text("Скачать Excel", exact=False).first
            page.wait_for_timeout(700)

            try:
                with page.expect_download(timeout=12_000) as download_info:
                    button.click(timeout=8_000)
                download = download_info.value
                download_path = download.path()
                bytes_data = Path(download_path).read_bytes() if download_path else b""
                if bytes_data:
                    logger.info("Доходъ: файл получен через expect_download")
                    return bytes_data
            except Exception as exc:
                logger.info("Доходъ: expect_download не сработал (%s)", exc)

            blob_url = page.evaluate(
                """() => {
                    const anchors = [...document.querySelectorAll('a[href^="blob:"]')];
                    if (anchors.length) return anchors[0].getAttribute('href');
                    const anyBlob = [...document.querySelectorAll('[href],[data-href]')]
                        .map(el => el.getAttribute('href') || el.getAttribute('data-href') || '')
                        .find(v => v && v.startsWith('blob:'));
                    return anyBlob || '';
                }"""
            )
            if blob_url:
                content_b64 = page.evaluate(
                    '''async (blobHref) => {
                        const response = await fetch(blobHref);
                        const buffer = await response.arrayBuffer();
                        let binary = '';
                        const bytes = new Uint8Array(buffer);
                        const chunk = 0x8000;
                        for (let i = 0; i < bytes.length; i += chunk) {
                            binary += String.fromCharCode(...bytes.slice(i, i + chunk));
                        }
                        return btoa(binary);
                    }''',
                    blob_url,
                )
                if content_b64:
                    logger.info("Доходъ: файл получен через blob-ссылку")
                    return base64.b64decode(content_b64)

            raise ValueError("На странице Доходъ не удалось получить Excel-файл.")
        finally:
            context.close()
            browser.close()


def should_refresh_nkr(conn: sqlite3.Connection, now_utc: datetime) -> bool:
    last_refresh_raw = get_meta_value(conn, "nkr_last_refresh_utc")
    rows_count_raw = get_meta_value(conn, "nkr_last_rows_count")
    if not last_refresh_raw or not rows_count_raw:
        return True
    try:
        last_refresh = datetime.fromisoformat(last_refresh_raw)
        rows_count = int(rows_count_raw)
    except ValueError:
        return True
    if rows_count <= 0:
        return True
    return now_utc - last_refresh >= timedelta(hours=config.NKR_CACHE_TTL_HOURS)


def should_refresh_nra(conn: sqlite3.Connection, now_utc: datetime) -> bool:
    last_refresh_raw = get_meta_value(conn, "nra_last_refresh_utc")
    rows_count_raw = get_meta_value(conn, "nra_last_rows_count")
    if not last_refresh_raw or not rows_count_raw:
        return True
    try:
        last_refresh = datetime.fromisoformat(last_refresh_raw)
        rows_count = int(rows_count_raw)
    except ValueError:
        return True
    if rows_count <= 0:
        return True
    return now_utc - last_refresh >= timedelta(hours=config.NRA_CACHE_TTL_HOURS)


def should_refresh_raex(conn: sqlite3.Connection, now_utc: datetime) -> bool:
    latest_exists = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (config.RAEX_LATEST_TABLE_NAME,),
    ).fetchone()
    if latest_exists:
        latest_rows_count = conn.execute(f'SELECT COUNT(*) FROM "{config.RAEX_LATEST_TABLE_NAME}"').fetchone()
        if not latest_rows_count or int(latest_rows_count[0]) <= 0:
            return True

    last_refresh_raw = get_meta_value(conn, "raex_last_refresh_utc")
    rows_count_raw = get_meta_value(conn, "raex_last_rows_count")
    if not last_refresh_raw or not rows_count_raw:
        return True
    try:
        last_refresh = datetime.fromisoformat(last_refresh_raw)
        rows_count = int(rows_count_raw)
    except ValueError:
        return True
    if rows_count <= 0:
        return True
    return now_utc - last_refresh >= timedelta(hours=config.RAEX_CACHE_TTL_HOURS)


def get_emitents_inns_for_raex(main_db_path: Path) -> list[str]:
    with connect_db(main_db_path) as main_conn:
        rows = main_conn.execute(
            f'''SELECT DISTINCT TRIM("INN") FROM "{config.EMITENTS_TABLE_NAME}" WHERE TRIM(COALESCE("INN", '')) <> '' '''
        ).fetchall()
    return [row[0] for row in rows if row and row[0]]


def get_raex_company_urls_by_inn(conn: sqlite3.Connection) -> dict[str, str]:
    rows = conn.execute(
        f'''
        SELECT "inn", "company_url"
        FROM "{config.RAEX_LATEST_TABLE_NAME}"
        WHERE TRIM(COALESCE("inn", '')) <> ''
          AND TRIM(COALESCE("company_url", '')) <> ''
        '''
    ).fetchall()
    return {str(inn).strip(): str(company_url).strip() for inn, company_url in rows if inn and company_url}


def _extract_raex_csrf_token(html_text: str) -> str:
    soup = BeautifulSoup(html_text, "lxml")
    token_input = soup.select_one('input[name="CSRFToken"]')
    return (token_input.get("value") or "").strip() if token_input else ""


def _extract_raex_company_url(html_text: str) -> str:
    match = re.search(r"href=[\"'](?P<url>/database/companies/[^\"']+/?)[\"']", html_text, flags=re.IGNORECASE)
    return match.group("url").strip() if match else ""


def _looks_like_raex_revoked(value: str) -> bool:
    normalized = (value or "").strip().lower()
    if not normalized:
        return True
    return "отозван" in normalized or normalized in {"-", "—", "n/a"}


def parse_raex_company_page(html_text: str) -> dict[str, str] | None:
    soup = BeautifulSoup(html_text, "lxml")
    heading = soup.find(lambda tag: tag.get_text(" ", strip=True) == "Рейтинги компании")
    if heading is None:
        return None

    rating_header_aliases = (
        "национальная шкала",
        "шкала эксперт ра",
        "шкала эксперт\xa0ра",
    )
    forecast_header_aliases = ("прогноз",)
    date_header_aliases = ("дата",)
    for node in heading.find_all_next():
        node_text = node.get_text(" ", strip=True)
        if node is not heading and "Архив рейтингов" in node_text:
            break
        if getattr(node, "name", "") != "table":
            continue

        headers = [th.get_text(" ", strip=True).lower() for th in node.select("thead th")]
        if not headers:
            first_row = node.select_one("tr")
            headers = [cell.get_text(" ", strip=True).lower() for cell in first_row.select("th, td")] if first_row else []
        rating_index = next((idx for idx, value in enumerate(headers) if value in rating_header_aliases), -1)
        forecast_index = next((idx for idx, value in enumerate(headers) if value in forecast_header_aliases), -1)
        date_index = next((idx for idx, value in enumerate(headers) if value in date_header_aliases), -1)
        if rating_index < 0 or forecast_index < 0 or date_index < 0:
            continue

        for tr in node.select("tbody tr") or node.select("tr")[1:]:
            cells = [td.get_text(" ", strip=True) for td in tr.select("td")]
            if not cells:
                continue
            rating_raw = cells[rating_index] if rating_index < len(cells) else ""
            forecast_raw = cells[forecast_index] if forecast_index < len(cells) else ""
            date_raw = cells[date_index] if date_index < len(cells) else ""
            if _looks_like_raex_revoked(rating_raw):
                return None
            rating_clean = re.sub(r"\s+", " ", rating_raw.strip())
            rating_head_match = re.match(r"[A-Za-zА-Яа-я0-9+\-.]+", rating_clean)
            rating = re.sub(r"^ru", "", rating_head_match.group(0) if rating_head_match else rating_clean, flags=re.IGNORECASE)
            if _looks_like_raex_revoked(rating):
                return None
            return {
                "rating": rating,
                "forecast": forecast_raw.strip(),
                "rating_date": date_raw.strip(),
            }
    return None


def _get_raex_thread_local_session() -> requests.Session:
    if not hasattr(fetch_raex_rating_by_inn, "_thread_local"):
        fetch_raex_rating_by_inn._thread_local = threading.local()  # type: ignore[attr-defined]
    thread_local = fetch_raex_rating_by_inn._thread_local  # type: ignore[attr-defined]
    if not hasattr(thread_local, "session"):
        pool_size = max(16, int(config.RAEX_MAX_WORKERS) * 2)
        retry = Retry(
            total=2,
            connect=2,
            read=2,
            backoff_factor=0.2,
            status_forcelist=(429, 500, 502, 503, 504),
            allowed_methods=frozenset({"GET", "POST"}),
            raise_on_status=False,
        )
        adapter = HTTPAdapter(pool_connections=pool_size, pool_maxsize=pool_size, max_retries=retry)
        session = requests.Session()
        session.headers.update({
            "User-Agent": config.NRA_REQUEST_USER_AGENT,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Encoding": "gzip, deflate, br",
            "Connection": "keep-alive",
        })
        session.mount("https://", adapter)
        session.mount("http://", adapter)
        thread_local.session = session
        thread_local.csrf_token = ""
        thread_local.csrf_fetched_at = 0.0
    return thread_local.session


def _get_raex_cached_csrf_token(session: requests.Session, timeout_seconds: int | float, force_refresh: bool = False) -> str:
    thread_local = fetch_raex_rating_by_inn._thread_local  # type: ignore[attr-defined]
    now_ts = time.monotonic()
    cached_token = str(getattr(thread_local, "csrf_token", "") or "").strip()
    fetched_at = float(getattr(thread_local, "csrf_fetched_at", 0.0) or 0.0)
    if (not force_refresh) and cached_token and (now_ts - fetched_at) < 1800:
        return cached_token

    search_page = session.get(config.RAEX_SEARCH_URL, timeout=timeout_seconds)
    search_page.raise_for_status()
    csrf_token = _extract_raex_csrf_token(search_page.text)
    thread_local.csrf_token = csrf_token
    thread_local.csrf_fetched_at = now_ts
    return csrf_token


def fetch_raex_rating_by_inn(
    inn: str,
    timeout_seconds: int | float,
    known_company_url: str = "",
) -> dict[str, str] | None:
    session = _get_raex_thread_local_session()

    company_url = (known_company_url or "").strip()
    if company_url:
        company_response = session.get(company_url, timeout=timeout_seconds)
        company_response.raise_for_status()
        parsed = parse_raex_company_page(company_response.text)
        if parsed:
            return {
                "inn": inn,
                "company_name": "",
                "rating": parsed["rating"],
                "forecast": parsed["forecast"],
                "rating_date": parsed["rating_date"],
                "company_url": company_url,
                "used_known_url": "1",
                "used_search": "0",
            }

    csrf_token = _get_raex_cached_csrf_token(session, timeout_seconds)
    payload = {"search": inn}
    if csrf_token:
        payload["CSRFToken"] = csrf_token

    search_response = session.post(config.RAEX_SEARCH_URL, data=payload, timeout=timeout_seconds)
    if search_response.status_code >= 400:
        search_response.raise_for_status()

    company_relative_url = _extract_raex_company_url(search_response.text)
    if not company_relative_url:
        csrf_token = _get_raex_cached_csrf_token(session, timeout_seconds, force_refresh=True)
        payload = {"search": inn}
        if csrf_token:
            payload["CSRFToken"] = csrf_token
        search_response = session.post(config.RAEX_SEARCH_URL, data=payload, timeout=timeout_seconds)
        search_response.raise_for_status()
        company_relative_url = _extract_raex_company_url(search_response.text)
        if not company_relative_url:
            return None

    company_url = urljoin(config.RAEX_SEARCH_URL, company_relative_url)
    company_response = session.get(company_url, timeout=timeout_seconds)
    company_response.raise_for_status()
    parsed = parse_raex_company_page(company_response.text)
    if not parsed:
        return None

    return {
        "inn": inn,
        "company_name": "",
        "rating": parsed["rating"],
        "forecast": parsed["forecast"],
        "rating_date": parsed["rating_date"],
        "company_url": company_url,
        "used_known_url": "0",
        "used_search": "1",
    }


def refresh_raex_data_if_needed(
    ratings_conn: sqlite3.Connection,
    main_db_path: Path,
    logger: logging.Logger,
    now_utc: datetime,
) -> tuple[bool, int, int, int]:
    ensure_raex_tables(ratings_conn)
    if not should_refresh_raex(ratings_conn, now_utc):
        logger.info("RAEX: кэш актуален, загрузка из сети пропущена.")
        row = ratings_conn.execute(f'SELECT COUNT(*) FROM "{config.RAEX_LATEST_TABLE_NAME}"').fetchone()
        return False, int(row[0]) if row else 0, 0, 0

    inns = get_emitents_inns_for_raex(main_db_path)
    if not inns:
        set_meta_value(ratings_conn, "raex_last_refresh_utc", now_utc.isoformat())
        set_meta_value(ratings_conn, "raex_last_rows_count", "0")
        logger.info("RAEX: в emitents нет ИНН для обработки.")
        return True, 0, 0, 0

    loaded_at = now_utc.isoformat()
    parsed_rows: list[dict[str, str]] = []
    errors_count = 0
    known_company_urls = get_raex_company_urls_by_inn(ratings_conn)
    known_url_hits = 0
    search_requests = 0
    with progress(total=len(inns), desc="RAEX INN", unit="inn") as raex_pbar:
        with ThreadPoolExecutor(max_workers=max(1, int(config.RAEX_MAX_WORKERS))) as executor:
            futures = {
                executor.submit(
                    fetch_raex_rating_by_inn,
                    inn,
                    config.REQUEST_TIMEOUT_SECONDS,
                    known_company_urls.get(inn, ""),
                ): inn
                for inn in inns
            }
            for future in as_completed(futures):
                inn = futures[future]
                try:
                    parsed = future.result()
                    if parsed:
                        parsed_rows.append(parsed)
                        if parsed.get("used_known_url") == "1":
                            known_url_hits += 1
                        if parsed.get("used_search") == "1":
                            search_requests += 1
                except Exception as exc:
                    errors_count += 1
                    logger.warning("RAEX: ошибка для INN=%s: %s", inn, exc)
                finally:
                    raex_pbar.update(1)

    payload = [
        (
            row["inn"],
            row.get("company_name", ""),
            row["rating"],
            row.get("forecast", ""),
            row.get("rating_date", ""),
            row.get("company_url", ""),
            loaded_at,
        )
        for row in parsed_rows
        if row.get("inn", "").strip() and row.get("rating", "").strip()
    ]

    if payload:
        ratings_conn.executemany(
            f'''
            INSERT INTO "{config.RAEX_TABLE_NAME}" (
                "inn", "company_name", "rating", "forecast", "rating_date", "company_url", "loaded_at_utc"
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT("inn", "rating_date", "rating", "forecast") DO UPDATE SET
                "company_name"=excluded."company_name",
                "company_url"=excluded."company_url",
                "loaded_at_utc"=excluded."loaded_at_utc"
            ''',
            payload,
        )

    ratings_conn.execute(f'DELETE FROM "{config.RAEX_LATEST_TABLE_NAME}"')
    ratings_conn.execute(
        f'''
        INSERT INTO "{config.RAEX_LATEST_TABLE_NAME}" (
            "inn", "company_name", "rating", "forecast", "rating_date", "company_url", "loaded_at_utc"
        )
        SELECT src."inn", src."company_name", src."rating", src."forecast", src."rating_date", src."company_url", src."loaded_at_utc"
        FROM "{config.RAEX_TABLE_NAME}" src
        JOIN (
            SELECT "inn", MAX(
                (CASE
                    WHEN instr("rating_date", '.') > 0 THEN substr("rating_date", 7, 4) || '-' || substr("rating_date", 4, 2) || '-' || substr("rating_date", 1, 2)
                    ELSE "rating_date"
                END) || '|' || COALESCE("loaded_at_utc", '')
            ) AS max_key
            FROM "{config.RAEX_TABLE_NAME}"
            WHERE TRIM(COALESCE("inn", '')) <> ''
            GROUP BY "inn"
        ) latest ON latest."inn" = src."inn"
        WHERE
            (CASE
                WHEN instr(src."rating_date", '.') > 0 THEN substr(src."rating_date", 7, 4) || '-' || substr(src."rating_date", 4, 2) || '-' || substr(src."rating_date", 1, 2)
                ELSE src."rating_date"
            END) || '|' || COALESCE(src."loaded_at_utc", '') = latest.max_key
        '''
    )
    ratings_conn.commit()

    set_meta_value(ratings_conn, "raex_last_refresh_utc", loaded_at)
    set_meta_value(ratings_conn, "raex_last_rows_count", str(len(payload)))
    logger.info(
        "RAEX обновление завершено. INN=%s, актуальных=%s, hits_known_url=%s, search_requests=%s, ошибок=%s",
        len(inns),
        len(payload),
        known_url_hits,
        search_requests,
        errors_count,
    )
    return True, len(payload), len(inns), errors_count


def ensure_acra_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.ACRA_TABLE_NAME}" (
            "issuer_url" TEXT NOT NULL,
            "issuer_name" TEXT,
            "rating" TEXT,
            "forecast" TEXT,
            "rating_date" TEXT,
            "inn" TEXT,
            "loaded_at_utc" TEXT,
            UNIQUE("issuer_url", "rating_date", "rating")
        )
        '''
    )
    conn.execute(
        f'CREATE INDEX IF NOT EXISTS "idx_{config.ACRA_TABLE_NAME}_url" ON "{config.ACRA_TABLE_NAME}"("issuer_url")'
    )
    conn.execute(
        f'ALTER TABLE "{config.ACRA_TABLE_NAME}" ADD COLUMN "forecast" TEXT'
    ) if _column_absent(conn, config.ACRA_TABLE_NAME, "forecast") else None
    conn.commit()


def should_refresh_acra(conn: sqlite3.Connection, now_utc: datetime) -> bool:
    last_refresh_raw = get_meta_value(conn, "acra_last_refresh_utc")
    if not last_refresh_raw:
        return True
    try:
        last_refresh = datetime.fromisoformat(last_refresh_raw)
    except ValueError:
        return True
    return now_utc - last_refresh >= timedelta(hours=config.ACRA_CACHE_TTL_HOURS)


def backfill_acra_forecast_from_local_dump(conn: sqlite3.Connection, logger: logging.Logger) -> int:
    list_dump_path = config.ACRA_DUMP_DIR / config.ACRA_LIST_HTML_FILENAME
    if not list_dump_path.exists():
        return 0

    try:
        html_text = list_dump_path.read_text(encoding="utf-8")
    except OSError as exc:
        logger.warning("АКРА backfill прогноза пропущен: не удалось прочитать %s (%s)", list_dump_path, exc)
        return 0

    parsed_rows = parse_acra_list(html_text)
    forecast_by_url = {
        row["issuer_url"]: (row.get("forecast") or "").strip()
        for row in parsed_rows
        if (row.get("issuer_url") or "").strip() and (row.get("forecast") or "").strip()
    }
    if not forecast_by_url:
        return 0

    updated = 0
    for issuer_url, forecast in forecast_by_url.items():
        cursor = conn.execute(
            f'''
            UPDATE "{config.ACRA_TABLE_NAME}"
            SET "forecast" = ?
            WHERE "issuer_url" = ? AND TRIM(COALESCE("forecast", '')) = ''
            ''',
            (forecast, issuer_url),
        )
        if cursor.rowcount and cursor.rowcount > 0:
            updated += int(cursor.rowcount)

    if updated:
        conn.commit()
        logger.info("АКРА backfill прогноза из локального дампа: обновлено строк=%s", updated)
    return updated


def parse_acra_list(html_text: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html_text, "lxml")
    parsed_rows: list[dict[str, str]] = []
    for row in soup.select("div.emits-row.search-table-row"):
        issuer_link = row.select_one('a.emits-row__item[data-type="ratePerson"]')
        if issuer_link is None:
            continue
        href = (issuer_link.get("href") or "").strip()
        if not href:
            continue
        issuer_url = href if href.startswith("http") else urljoin(config.ACRA_RATINGS_LIST_URL, href)
        issuer_name = issuer_link.get_text(" ", strip=True)

        rating_container = row.select_one('div.emits-row__item[data-type="rate"]')
        rating_raw = rating_container.get_text("\n", strip=True) if rating_container else ""
        rating, forecast = split_acra_rating_and_forecast(rating_raw)

        forecast = ""
        forecast_container = row.select_one('div.emits-row__item[data-type="forecast"]')
        if forecast_container:
            forecast = forecast_container.get_text(" ", strip=True)

        rating_container = row.select_one('div.emits-row__item[data-type="rate"]')
        rating_raw = rating_container.get_text("\n", strip=True) if rating_container else ""
        fallback_rating, fallback_forecast = split_acra_rating_and_forecast(rating_raw)
        if not rating:
            rating = fallback_rating
        if not forecast:
            forecast = fallback_forecast

        date_node = row.select_one('div.emits-row__item[data-type="pressRelease"] a')
        date_raw = date_node.get_text(" ", strip=True) if date_node else ""
        parsed_rows.append(
            {
                "issuer_url": issuer_url,
                "issuer_name": issuer_name,
                "rating": rating,
                "forecast": forecast,
                "rating_date": normalize_date_ru(date_raw) or date_raw,
            }
        )
    return parsed_rows


def split_acra_rating_and_forecast(raw_value: str) -> tuple[str, str]:
    normalized_lines = [line.strip() for line in re.split(r"[\r\n]+", raw_value or "") if line.strip()]
    if not normalized_lines:
        return "", ""

    forecast_line = ""
    rating_line = ""

    for line in normalized_lines:
        if not forecast_line and is_acra_forecast_value(line):
            forecast_line = line
            continue
        if not rating_line and is_acra_rating_value(line):
            rating_line = line

    if rating_line:
        return rating_line, forecast_line

    one_line = normalized_lines[0]
    match = re.match(r"^(.*?)\s*[;,]\s*([^;,]+)$", one_line)
    if match:
        left = match.group(1).strip()
        right = match.group(2).strip()
        if is_acra_forecast_value(right):
            return left, right
    return one_line, forecast_line


def is_acra_rating_value(value: str) -> bool:
    text = (value or "").strip()
    if not text:
        return False
    normalized = text.replace(" ", "")
    return bool(re.search(r"[A-ZА-Я][+\-]?(?:\([A-ZА-Я]{2}\))", normalized))


def is_acra_forecast_value(value: str) -> bool:
    normalized = (value or "").strip().lower()
    return normalized in {
        "стабильный",
        "позитивный",
        "негативный",
        "развивающийся",
    }


def extract_inn_from_acra_card(html_text: str) -> str:
    soup = BeautifulSoup(html_text, "lxml")
    for info in soup.select("div.info"):
        label = info.find("small")
        if label and label.get_text(" ", strip=True).lower() == "инн":
            value_node = info.find("p")
            raw_value = value_node.get_text(" ", strip=True) if value_node else ""
            return re.sub(r"\D+", "", raw_value)

    fallback = re.search(r"ИНН\D{0,50}(\d[\d\s]{8,14}\d)", soup.get_text(" ", strip=True), flags=re.IGNORECASE)
    return re.sub(r"\D+", "", fallback.group(1)) if fallback else ""


def acra_human_sleep(min_seconds: float = 0.7, max_seconds: float = 1.8) -> None:
    time.sleep(random.uniform(min_seconds, max_seconds))


def acra_ensure_dirs() -> None:
    config.ACRA_DUMP_DIR.mkdir(parents=True, exist_ok=True)
    config.ACRA_ISSUERS_DUMP_DIR.mkdir(parents=True, exist_ok=True)


def acra_safe_filename(name: str, limit: int = 80) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9а-яА-Я_-]+", "_", name or "").strip("_")
    return cleaned[:limit] or "issuer"


def acra_log_progress(payload: dict[str, str]) -> None:
    acra_ensure_dirs()
    progress_log_path = config.ACRA_DUMP_DIR / config.ACRA_PROGRESS_LOG_FILENAME
    with progress_log_path.open("a", encoding="utf-8") as file_obj:
        file_obj.write(json.dumps(payload, ensure_ascii=False) + "\n")


def acra_save_mhtml(page, file_path: Path) -> None:
    cdp = page.context.new_cdp_session(page)
    snapshot = cdp.send("Page.captureSnapshot", {"format": "mhtml"})
    file_path.write_text(snapshot["data"], encoding="utf-8")


def acra_goto_with_retries(page, url: str, logger: logging.Logger, wait_selector: str | None = None, attempts: int = 5) -> bool:
    last_error: Exception | None = None
    for retry in range(1, attempts + 1):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=int(config.REQUEST_TIMEOUT_SECONDS * 1000))
            if wait_selector:
                page.wait_for_selector(wait_selector, timeout=int(config.REQUEST_TIMEOUT_SECONDS * 1000))
            return True
        except (PWTimeoutError, PWError) as exc:
            last_error = exc
            message = str(exc)
            retryable = any(
                marker in message
                for marker in ("ERR_CONNECTION_CLOSED", "ERR_CONNECTION_RESET", "ERR_EMPTY_RESPONSE", "ERR_TIMED_OUT", "net::")
            )
            if not retryable:
                logger.warning("АКРА goto non-retryable error for %s: %s", url, message)
                return False
            sleep_seconds = min((3.0 * retry) + random.uniform(0.5, 2.0), 20.0)
            logger.warning(
                "АКРА goto retry %s/%s for %s due to %s; sleep %.1fs",
                retry,
                attempts,
                url,
                message[:200],
                sleep_seconds,
            )
            time.sleep(sleep_seconds)
    logger.warning("АКРА goto failed for %s after retries: %s", url, last_error)
    return False


def collect_acra_rows_via_playwright(
    logger: logging.Logger,
    known_urls: set[str],
    inn_cache_by_url: dict[str, str],
) -> tuple[dict[str, dict[str, str]], int, int]:
    acra_ensure_dirs()
    unique_rows: dict[str, dict[str, str]] = {}
    card_fetch_count = 0
    known_url_skip_count = 0
    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=str(config.ACRA_PROFILE_DIR),
            channel=config.ACRA_BROWSER_CHANNEL,
            headless=config.ACRA_HEADLESS,
            viewport={"width": 1365, "height": 768},
            locale="ru-RU",
            timezone_id="Europe/Moscow",
            args=["--start-maximized"],
        )
        try:
            page = context.new_page()
            list_ok = acra_goto_with_retries(
                page,
                config.ACRA_RATINGS_LIST_URL,
                logger,
                wait_selector='a.emits-row__item[data-type="ratePerson"]',
                attempts=config.ACRA_LIST_GOTO_ATTEMPTS,
            )
            if not list_ok:
                return unique_rows, card_fetch_count, known_url_skip_count

            acra_human_sleep(1.0, 2.0)
            page.mouse.wheel(0, random.randint(500, 1400))
            acra_human_sleep(0.6, 1.2)

            try:
                acra_save_mhtml(page, config.ACRA_DUMP_DIR / config.ACRA_LIST_MHTML_FILENAME)
            except Exception as exc:
                logger.warning("Не удалось сохранить MHTML списка АКРА: %s", exc)

            list_html = page.content()
            (config.ACRA_DUMP_DIR / config.ACRA_LIST_HTML_FILENAME).write_text(list_html, encoding="utf-8")

            parsed_rows = parse_acra_list(list_html)
            for row_data in parsed_rows:
                unique_rows[row_data["issuer_url"]] = row_data

            for index, row_data in enumerate(
                tqdm(list(unique_rows.values()), desc="АКРА карточки", unit="эмитент", leave=False, dynamic_ncols=True),
                start=1,
            ):
                issuer_url = row_data["issuer_url"]
                cached_inn = (inn_cache_by_url.get(issuer_url) or "").strip()
                if cached_inn:
                    row_data["inn"] = cached_inn
                if issuer_url in known_urls:
                    known_url_skip_count += 1
                    continue
                card_ok = acra_goto_with_retries(
                    page,
                    row_data["issuer_url"],
                    logger,
                    wait_selector=None,
                    attempts=config.ACRA_CARD_GOTO_ATTEMPTS,
                )
                if not card_ok:
                    acra_log_progress(
                        {
                            "url": row_data["issuer_url"],
                            "name": row_data.get("issuer_name", ""),
                            "inn": "",
                            "status": "goto_failed",
                            "ts": datetime.now(timezone.utc).isoformat(),
                        }
                    )
                    continue
                acra_human_sleep(0.8, 1.8)
                if random.random() < 0.6:
                    page.mouse.wheel(0, random.randint(250, 900))
                    acra_human_sleep(0.3, 0.8)
                card_html = page.content()
                card_filename = f"{index:04d}_{acra_safe_filename(row_data.get('issuer_name', ''))}.html"
                (config.ACRA_ISSUERS_DUMP_DIR / card_filename).write_text(card_html, encoding="utf-8")
                row_data["inn"] = extract_inn_from_acra_card(card_html)
                acra_log_progress(
                    {
                        "url": row_data["issuer_url"],
                        "name": row_data.get("issuer_name", ""),
                        "inn": row_data.get("inn", ""),
                        "status": "ok",
                        "ts": datetime.now(timezone.utc).isoformat(),
                    }
                )
                card_fetch_count += 1
                acra_human_sleep(0.6, 1.6)
        finally:
            context.close()
    return unique_rows, card_fetch_count, known_url_skip_count


def refresh_acra_data_if_needed(conn: sqlite3.Connection, logger: logging.Logger, now_utc: datetime) -> tuple[bool, int, int]:
    ensure_acra_tables(conn)
    current = conn.execute(f'SELECT COUNT(*) FROM "{config.ACRA_TABLE_NAME}"').fetchone()
    current_total = int(current[0]) if current else 0
    if not should_refresh_acra(conn, now_utc):
        backfill_acra_forecast_from_local_dump(conn, logger)
        return False, current_total, 0

    known_urls = {
        str(row[0] or "").strip()
        for row in conn.execute(
            f'''
            SELECT DISTINCT "issuer_url"
            FROM "{config.ACRA_TABLE_NAME}"
            WHERE TRIM(COALESCE("issuer_url", '')) <> ''
            '''
        ).fetchall()
        if str(row[0] or "").strip()
    }

    inn_cache_by_url = {
        row[0]: row[1]
        for row in conn.execute(
            f'''
            SELECT "issuer_url", "inn"
            FROM "{config.ACRA_TABLE_NAME}"
            WHERE TRIM(COALESCE("inn", '')) <> ''
            '''
        ).fetchall()
    }

    try:
        unique_rows, card_requests, known_skipped = collect_acra_rows_via_playwright(logger, known_urls, inn_cache_by_url)

        loaded_at = now_utc.isoformat()
        changed_rows = 0

        for row_data in unique_rows.values():
            cursor = conn.execute(
                    f'''
                    INSERT INTO "{config.ACRA_TABLE_NAME}" (
                        "issuer_url", "issuer_name", "rating", "forecast", "rating_date", "inn", "loaded_at_utc"
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT("issuer_url", "rating_date", "rating") DO UPDATE SET
                        "issuer_name" = excluded."issuer_name",
                        "forecast" = excluded."forecast",
                        "inn" = CASE
                            WHEN TRIM(COALESCE("{config.ACRA_TABLE_NAME}"."inn", '')) = '' THEN excluded."inn"
                            ELSE "{config.ACRA_TABLE_NAME}"."inn"
                        END,
                        "loaded_at_utc" = excluded."loaded_at_utc"
                    ''',
                    (
                        row_data["issuer_url"],
                        row_data.get("issuer_name", ""),
                        row_data.get("rating", ""),
                        row_data.get("forecast", ""),
                        row_data.get("rating_date", ""),
                        row_data.get("inn", ""),
                        loaded_at,
                    ),
                )
            if cursor.rowcount and cursor.rowcount > 0:
                changed_rows += 1
    except Exception as exc:
        logger.warning("АКРА обновление пропущено из-за сетевой ошибки: %s", exc)
        return False, current_total, 0

    conn.commit()
    backfill_acra_forecast_from_local_dump(conn, logger)
    set_meta_value(conn, "acra_last_refresh_utc", now_utc.isoformat())
    set_meta_value(conn, "acra_last_rows_count", str(len(unique_rows)))
    new_urls_count = len([url for url in unique_rows if url not in known_urls])
    logger.info(
        "АКРА обновление завершено. issuer_url в списке=%s, known_urls=%s, новых=%s, вставлено/обновлено=%s, карточек запрошено=%s, карточек пропущено_known=%s",
        len(unique_rows),
        len(known_urls),
        new_urls_count,
        changed_rows,
        card_requests,
        known_skipped,
    )

    total = conn.execute(f'SELECT COUNT(*) FROM "{config.ACRA_TABLE_NAME}"').fetchone()
    return True, int(total[0]) if total else 0, card_requests


def refresh_nra_data_if_needed(conn: sqlite3.Connection, logger: logging.Logger, now_utc: datetime) -> tuple[bool, int]:
    ensure_nra_tables(conn)
    if not should_refresh_nra(conn, now_utc):
        row = conn.execute(f'SELECT COUNT(*) FROM "{config.NRA_TABLE_NAME}"').fetchone()
        return False, int(row[0]) if row else 0

    with create_http_session() as session:
        page_response = session.get(config.NRA_RATINGS_PAGE_URL, timeout=config.REQUEST_TIMEOUT_SECONDS)
        page_response.raise_for_status()
        excel_link = find_nra_excel_link(page_response.text)
        excel_url = requests.compat.urljoin(config.NRA_RATINGS_PAGE_URL, excel_link)
        excel_response = session.get(excel_url, timeout=config.REQUEST_TIMEOUT_SECONDS)
        excel_response.raise_for_status()
        content = excel_response.content

    raw_path = config.RAW_DIR / config.NRA_RAW_FILENAME
    if raw_path.exists():
        raw_path.unlink()
    raw_path.write_bytes(content)

    parsed_rows = parse_nra_excel(content)
    loaded_at = now_utc.isoformat()
    payload = [
        (
            row["id"],
            row["organization_name"],
            row["inn"],
            row["press_release_title"],
            row["press_release_date"],
            row["rating"],
            row["rating_status"],
            row["forecast"],
            row["rating_type"],
            row["organization_sector"],
            row["industry"],
            row["osk"],
            row["isin"],
            row["press_release_link"],
            row["under_watch"],
            config.NRA_RAW_FILENAME,
            loaded_at,
        )
        for row in parsed_rows
    ]
    conn.executemany(
        f'''
        INSERT INTO "{config.NRA_TABLE_NAME}" (
            "id", "organization_name", "inn", "press_release_title", "press_release_date", "rating",
            "rating_status", "forecast", "rating_type", "organization_sector", "industry", "osk",
            "isin", "press_release_link", "under_watch", "source_file_name", "loaded_at_utc"
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT("id") DO UPDATE SET
            "organization_name"=excluded."organization_name",
            "inn"=excluded."inn",
            "press_release_title"=excluded."press_release_title",
            "press_release_date"=excluded."press_release_date",
            "rating"=excluded."rating",
            "rating_status"=excluded."rating_status",
            "forecast"=excluded."forecast",
            "rating_type"=excluded."rating_type",
            "organization_sector"=excluded."organization_sector",
            "industry"=excluded."industry",
            "osk"=excluded."osk",
            "isin"=excluded."isin",
            "press_release_link"=excluded."press_release_link",
            "under_watch"=excluded."under_watch",
            "source_file_name"=excluded."source_file_name",
            "loaded_at_utc"=excluded."loaded_at_utc"
        ''',
        payload,
    )

    conn.execute(f'DELETE FROM "{config.NRA_LATEST_TABLE_NAME}"')
    conn.execute(
        f'''
        INSERT INTO "{config.NRA_LATEST_TABLE_NAME}" (
            "inn", "organization_name", "press_release_date", "rating", "rating_status", "forecast"
        )
        SELECT src."inn", src."organization_name", src."press_release_date", src."rating", src."rating_status", src."forecast"
        FROM "{config.NRA_TABLE_NAME}" src
        JOIN (
            SELECT "inn", MAX(
                (CASE
                    WHEN instr("press_release_date", '.') > 0 THEN substr("press_release_date", 7, 4) || '-' || substr("press_release_date", 4, 2) || '-' || substr("press_release_date", 1, 2)
                    ELSE "press_release_date"
                END) || '|' || printf('%012d', CAST(COALESCE("id", '0') AS INTEGER))
            ) AS max_key
            FROM "{config.NRA_TABLE_NAME}"
            WHERE TRIM(COALESCE("inn", '')) <> ''
            GROUP BY "inn"
        ) latest ON latest."inn" = src."inn"
        WHERE
            (CASE
                WHEN instr(src."press_release_date", '.') > 0 THEN substr(src."press_release_date", 7, 4) || '-' || substr(src."press_release_date", 4, 2) || '-' || substr(src."press_release_date", 1, 2)
                ELSE src."press_release_date"
            END) || '|' || printf('%012d', CAST(COALESCE(src."id", '0') AS INTEGER)) = latest.max_key
        '''
    )
    conn.commit()

    set_meta_value(conn, "nra_last_refresh_utc", loaded_at)
    set_meta_value(conn, "nra_last_rows_count", str(len(parsed_rows)))
    logger.info("NRA обновление завершено. Загружено строк: %s", len(parsed_rows))
    return True, len(parsed_rows)


def download_nkr_excel_via_playwright(logger: logging.Logger) -> bytes:
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=config.NKR_HEADLESS, channel=config.NKR_BROWSER_CHANNEL)
        context = browser.new_context(locale="ru-RU", timezone_id="Europe/Moscow", accept_downloads=True)
        page = context.new_page()
        try:
            page.goto(config.NKR_RATINGS_PAGE_URL, wait_until="domcontentloaded", timeout=int(config.REQUEST_TIMEOUT_SECONDS * 1000))
            page.wait_for_selector(config.NKR_EXPORT_BUTTON_SELECTOR, timeout=int(config.REQUEST_TIMEOUT_SECONDS * 1000))

            for attempt in range(1, config.NKR_DOWNLOAD_ATTEMPTS + 1):
                logger.info("НКР: попытка скачивания %s/%s", attempt, config.NKR_DOWNLOAD_ATTEMPTS)

                try:
                    with page.expect_download(timeout=8_000) as download_info:
                        page.locator(config.NKR_EXPORT_BUTTON_SELECTOR).first.click(timeout=5_000)
                    download = download_info.value
                    download_path = download.path()
                    bytes_data = Path(download_path).read_bytes() if download_path else b""
                    if bytes_data:
                        logger.info("НКР: файл получен через expect_download")
                        return bytes_data
                except Exception as exc:
                    logger.info("НКР: expect_download не сработал (%s)", exc)

                href_value = page.locator(config.NKR_EXPORT_BUTTON_SELECTOR).first.get_attribute("href") or ""
                if href_value and not href_value.lower().startswith("blob:"):
                    direct_url = urljoin(config.NKR_RATINGS_PAGE_URL, href_value)
                    response = context.request.get(direct_url, timeout=int(config.REQUEST_TIMEOUT_SECONDS * 1000))
                    if response.ok and response.body():
                        logger.info("НКР: файл получен по прямой ссылке %s", direct_url)
                        return response.body()

                blob_url = page.evaluate(
                    '''(selector) => {
                        const node = document.querySelector(selector);
                        const href = (node?.getAttribute('href') || '').trim();
                        if (href.startsWith('blob:')) return href;

                        const asText = (el) => (el?.textContent || '').replace(/\s+/g, ' ').trim().toLowerCase();
                        const linkByText = [...document.querySelectorAll('a')]
                            .find((el) => asText(el).includes('выгрузить в excel'));
                        const textHref = (linkByText?.getAttribute('href') || '').trim();
                        if (textHref.startsWith('blob:')) return textHref;

                        const links = [...document.querySelectorAll('a[href^="blob:"]')];
                        return links.length > 0 ? links[0].getAttribute('href') : '';
                    }''',
                    config.NKR_EXPORT_BUTTON_SELECTOR,
                )

                if blob_url:
                    payload_b64 = page.evaluate(
                        '''async (blobHref) => {
                            const response = await fetch(blobHref);
                            const buffer = await response.arrayBuffer();
                            const bytes = new Uint8Array(buffer);
                            const chunkSize = 0x8000;
                            let binary = '';
                            for (let i = 0; i < bytes.length; i += chunkSize) {
                                const chunk = bytes.subarray(i, i + chunkSize);
                                binary += String.fromCharCode(...chunk);
                            }
                            return btoa(binary);
                        }''',
                        blob_url,
                    )
                    bytes_data = base64.b64decode(payload_b64)
                    if bytes_data:
                        logger.info("НКР: файл получен через blob-ссылку")
                        return bytes_data

                page.wait_for_timeout(1200)

            raise ValueError("На странице НКР не удалось получить Excel-файл (ни download, ни href, ни blob).")
        except Exception:
            logger.exception("Ошибка скачивания НКР через Playwright")
            raise
        finally:
            context.close()
            browser.close()


def refresh_nkr_data_if_needed(conn: sqlite3.Connection, logger: logging.Logger, now_utc: datetime) -> tuple[bool, int]:
    ensure_nkr_tables(conn)
    current = conn.execute(f'SELECT COUNT(*) FROM "{config.NKR_TABLE_NAME}"').fetchone()
    current_total = int(current[0]) if current else 0
    if not should_refresh_nkr(conn, now_utc):
        return False, current_total

    try:
        content = download_nkr_excel_via_playwright(logger)
    except Exception as exc:
        logger.warning("НКР обновление пропущено из-за сетевой ошибки: %s", exc)
        return False, current_total
    raw_path = config.RAW_DIR / config.NKR_RAW_FILENAME
    if raw_path.exists():
        raw_path.unlink()
    raw_path.write_bytes(content)

    parsed_rows = parse_nkr_excel(content)
    loaded_at = now_utc.isoformat()
    payload = [
        (
            row.get("id", ""),
            row.get("issuer_name", ""),
            row.get("rating_date", ""),
            row.get("rating", ""),
            row.get("outlook", ""),
            row.get("tin", ""),
            loaded_at,
        )
        for row in parsed_rows
        if row.get("tin", "").strip()
    ]

    conn.executemany(
        f'''
        INSERT INTO "{config.NKR_TABLE_NAME}" (
            "id", "issuer_name", "rating_date", "rating", "outlook", "tin", "loaded_at_utc"
        ) VALUES (?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT("tin", "rating_date", "rating", "outlook") DO UPDATE SET
            "issuer_name"=excluded."issuer_name",
            "id"=excluded."id",
            "loaded_at_utc"=excluded."loaded_at_utc"
        ''',
        payload,
    )

    conn.execute(f'DELETE FROM "{config.NKR_LATEST_TABLE_NAME}"')
    conn.execute(
        f'''
        INSERT INTO "{config.NKR_LATEST_TABLE_NAME}" (
            "tin", "issuer_name", "rating_date", "rating", "outlook"
        )
        SELECT src."tin", src."issuer_name", src."rating_date", src."rating", src."outlook"
        FROM "{config.NKR_TABLE_NAME}" src
        JOIN (
            SELECT "tin", MAX(
                (CASE
                    WHEN instr("rating_date", '.') > 0 THEN substr("rating_date", 7, 4) || '-' || substr("rating_date", 4, 2) || '-' || substr("rating_date", 1, 2)
                    ELSE "rating_date"
                END) || '|' || COALESCE("loaded_at_utc", '')
            ) AS max_key
            FROM "{config.NKR_TABLE_NAME}"
            WHERE TRIM(COALESCE("tin", '')) <> ''
            GROUP BY "tin"
        ) latest ON latest."tin" = src."tin"
        WHERE
            (CASE
                WHEN instr(src."rating_date", '.') > 0 THEN substr(src."rating_date", 7, 4) || '-' || substr(src."rating_date", 4, 2) || '-' || substr(src."rating_date", 1, 2)
                ELSE src."rating_date"
            END) || '|' || COALESCE(src."loaded_at_utc", '') = latest.max_key
        '''
    )
    conn.commit()

    set_meta_value(conn, "nkr_last_refresh_utc", loaded_at)
    set_meta_value(conn, "nkr_last_rows_count", str(len(parsed_rows)))
    logger.info("НКР обновление завершено. Загружено строк: %s", len(parsed_rows))
    return True, len(parsed_rows)


def sync_nra_rate_to_emitents(main_conn: sqlite3.Connection, nra_conn: sqlite3.Connection, logger: logging.Logger) -> int:
    rows = nra_conn.execute(
        f'''
        SELECT "inn", "rating", "forecast"
        FROM "{config.NRA_LATEST_TABLE_NAME}"
        WHERE TRIM(COALESCE("inn", '')) <> ''
        '''
    ).fetchall()

    updates: list[tuple[str, str]] = []
    for inn, rating, forecast in rows:
        rating_text = (rating or "").strip()
        forecast_text = (forecast or "").strip().lower()
        if not rating_text:
            continue
        updates.append((f"{rating_text}({forecast_text})" if forecast_text else rating_text, inn.strip()))

    main_conn.executemany(
        f'UPDATE "{config.EMITENTS_TABLE_NAME}" SET "NRA_Rate" = ? WHERE "INN" = ?',
        updates,
    )
    main_conn.commit()
    logger.info("NRA_Rate синхронизирован для INN: %s", len(updates))
    return len(updates)


def sync_acra_rate_to_emitents(main_conn: sqlite3.Connection, ratings_conn: sqlite3.Connection, logger: logging.Logger) -> int:
    rows = ratings_conn.execute(
        f'''
        SELECT src."inn", src."rating", src."forecast"
        FROM "{config.ACRA_TABLE_NAME}" src
        JOIN (
            SELECT "inn", MAX(
                (CASE
                    WHEN instr("rating_date", '.') > 0 THEN substr("rating_date", 7, 4) || '-' || substr("rating_date", 4, 2) || '-' || substr("rating_date", 1, 2)
                    ELSE "rating_date"
                END) || '|' || "loaded_at_utc"
            ) AS max_key
            FROM "{config.ACRA_TABLE_NAME}"
            WHERE TRIM(COALESCE("inn", '')) <> ''
            GROUP BY "inn"
        ) latest ON latest."inn" = src."inn"
        WHERE
            (CASE
                WHEN instr(src."rating_date", '.') > 0 THEN substr(src."rating_date", 7, 4) || '-' || substr(src."rating_date", 4, 2) || '-' || substr(src."rating_date", 1, 2)
                ELSE src."rating_date"
            END) || '|' || src."loaded_at_utc" = latest.max_key
        '''
    ).fetchall()

    updates: list[tuple[str, str]] = []
    for inn, rating, forecast in rows:
        rating_text = (rating or "").strip()
        forecast_text = (forecast or "").strip().lower()
        if not rating_text:
            continue

        base_rating = rating_text
        if not forecast_text:
            base_rating, forecast_text = split_acra_rating_and_forecast(rating_text)
            forecast_text = forecast_text.lower()
        if not base_rating:
            continue

        rate_for_showcase = f"{base_rating}({forecast_text})" if forecast_text else base_rating
        updates.append((rate_for_showcase, inn.strip()))

    main_conn.executemany(
        f'UPDATE "{config.EMITENTS_TABLE_NAME}" SET "Acra_Rate" = ? WHERE "INN" = ?',
        updates,
    )
    main_conn.commit()
    logger.info("Acra_Rate синхронизирован для INN: %s", len(updates))
    return len(updates)


def sync_nkr_rate_to_emitents(main_conn: sqlite3.Connection, ratings_conn: sqlite3.Connection, logger: logging.Logger) -> int:
    rows = ratings_conn.execute(
        f'''
        SELECT "tin", "rating", "outlook"
        FROM "{config.NKR_LATEST_TABLE_NAME}"
        WHERE TRIM(COALESCE("tin", '')) <> ''
        '''
    ).fetchall()

    updates: list[tuple[str, str]] = []
    for tin, rating, outlook in rows:
        rating_text = (rating or "").strip()
        outlook_text = (outlook or "").strip().lower()
        if not rating_text:
            continue
        updates.append((f"{rating_text}({outlook_text})" if outlook_text else rating_text, tin.strip()))

    main_conn.executemany(
        f'UPDATE "{config.EMITENTS_TABLE_NAME}" SET "NKR_Rate" = ? WHERE "INN" = ?',
        updates,
    )
    main_conn.commit()
    logger.info("NKR_Rate синхронизирован для INN: %s", len(updates))
    return len(updates)


def sync_raex_rate_to_emitents(main_conn: sqlite3.Connection, ratings_conn: sqlite3.Connection, logger: logging.Logger) -> int:
    rows = ratings_conn.execute(
        f''' 
        SELECT "inn", "rating", "forecast"
        FROM "{config.RAEX_LATEST_TABLE_NAME}"
        WHERE TRIM(COALESCE("inn", '')) <> ''
        '''
    ).fetchall()

    updates: list[tuple[str, str]] = []
    for inn, rating, forecast in rows:
        rating_text = (rating or "").strip()
        forecast_text = (forecast or "").strip().lower()
        if not rating_text:
            continue
        updates.append((f"{rating_text}({forecast_text})" if forecast_text else rating_text, inn.strip()))

    main_conn.executemany(
        f'UPDATE "{config.EMITENTS_TABLE_NAME}" SET "RAEX_Rate" = ? WHERE "INN" = ?',
        updates,
    )
    main_conn.commit()
    logger.info("RAEX_Rate синхронизирован для INN: %s", len(updates))
    return len(updates)


def sync_emitents_from_rates(conn: sqlite3.Connection, logger: logging.Logger) -> int:
    cursor = conn.execute(
        f'''
        INSERT INTO "{config.EMITENTS_TABLE_NAME}" ("INN", "EMITENTNAME")
        SELECT DISTINCT TRIM("INN"), TRIM("EMITENTNAME")
        FROM "{config.RATES_TABLE_NAME}"
        WHERE TRIM(COALESCE("INN", '')) <> ''
          AND TRIM(COALESCE("EMITENTNAME", '')) <> ''
        ON CONFLICT("INN") DO UPDATE SET
            "EMITENTNAME" = excluded."EMITENTNAME"
        '''
    )
    conn.commit()
    affected = cursor.rowcount if cursor.rowcount is not None else 0
    logger.info("Синхронизация эмитентов из moex_bonds завершена. Затронуто строк: %s", affected)
    return max(affected, 0)


def pull_scoring_from_excel(conn: sqlite3.Connection, logger: logging.Logger, today_str: str) -> int:
    excel_path = config.BASE_DIR / config.EMITENTS_XLSX_FILENAME
    if not excel_path.exists():
        logger.info("Файл витрины %s пока отсутствует — перенос оценок из Excel пропущен.", excel_path)
        return 0

    wb = load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    required = {"INN", "Scoring", "DateScoring"}
    if not required.issubset(set(headers)):
        logger.warning("В %s не найдены обязательные колонки INN/Scoring/DateScoring.", excel_path)
        return 0

    inn_idx = headers.index("INN")
    scoring_idx = headers.index("Scoring")
    date_idx = headers.index("DateScoring")
    allowed_scoring = set(config.SCORING_ALLOWED_VALUES)

    updates: list[tuple[str | None, str | None, str]] = []
    for row in rows[1:]:
        inn = str(row[inn_idx]).strip() if inn_idx < len(row) and row[inn_idx] is not None else ""
        if not inn:
            continue
        scoring_val = ""
        if scoring_idx < len(row) and row[scoring_idx] is not None:
            scoring_val = str(row[scoring_idx]).strip()
        if scoring_val and scoring_val not in allowed_scoring:
            logger.warning(
                "Пропущено некорректное значение Scoring='%s' для INN=%s. Допустимые значения: %s",
                scoring_val,
                inn,
                ", ".join(config.SCORING_ALLOWED_VALUES),
            )
            scoring_val = ""
        date_val = ""
        if date_idx < len(row) and row[date_idx] is not None:
            date_val = str(row[date_idx]).strip()

        scoring_db = scoring_val or None
        date_db = date_val or (today_str if scoring_db else None)
        updates.append((scoring_db, date_db, inn))

    if not updates:
        return 0

    conn.executemany(
        f'''
        UPDATE "{config.EMITENTS_TABLE_NAME}"
        SET "Scoring" = ?,
            "DateScoring" = ?
        WHERE "INN" = ?
        ''',
        updates,
    )
    conn.commit()
    logger.info("Перенос ручных Scoring из Excel в SQL: обработано строк=%s", len(updates))
    return len(updates)


def ensure_scoring_dates(conn: sqlite3.Connection, logger: logging.Logger, today_str: str) -> int:
    cursor = conn.execute(
        f'''
        UPDATE "{config.EMITENTS_TABLE_NAME}"
        SET "DateScoring" = ?
        WHERE TRIM(COALESCE("Scoring", '')) <> ''
          AND TRIM(COALESCE("DateScoring", '')) = ''
        ''',
        (today_str,),
    )
    conn.commit()
    fixed = cursor.rowcount if cursor.rowcount is not None else 0
    logger.info("Автозаполнение DateScoring выполнено. Добавлено дат: %s", fixed)
    return max(fixed, 0)


def export_emitents_excel(conn: sqlite3.Connection) -> int:
    cursor = conn.execute(
        f'''
        SELECT "EMITENTNAME", "INN", "Scoring", "DateScoring", "NRA_Rate", "Acra_Rate", "NKR_Rate", "RAEX_Rate"
        FROM "{config.EMITENTS_TABLE_NAME}"
        ORDER BY "EMITENTNAME", "INN"
        '''
    )
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = "emitents"
    ws.append(headers)

    for row in rows:
        ws.append(list(row))

    header_fill = PatternFill(fill_type="solid", fgColor=config.EMITENTS_HEADER_FILL_COLOR)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    scoring_column_index = headers.index("Scoring") + 1
    scoring_column_letter = ws.cell(row=1, column=scoring_column_index).column_letter
    validation_values = ",".join(config.SCORING_ALLOWED_VALUES)
    scoring_validation = DataValidation(
        type="list",
        formula1=f'"{validation_values}"',
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="Недопустимое значение",
        error=f"Выберите одно из значений: {validation_values}",
        promptTitle="Scoring",
        prompt=f"Доступные значения: {validation_values}",
    )
    ws.add_data_validation(scoring_validation)
    scoring_validation.add(f"{scoring_column_letter}2:{scoring_column_letter}1048576")

    for column_cells in ws.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[column_letter].width = min(max_len + 2, 80)

    for rating_column_name in ("NRA_Rate", "Acra_Rate", "NKR_Rate", "RAEX_Rate"):
        if rating_column_name not in headers:
            continue
        rating_column_index = headers.index(rating_column_name) + 1
        rating_column_letter = ws.cell(row=1, column=rating_column_index).column_letter
        ws.column_dimensions[rating_column_letter].width = config.EMITENTS_RATINGS_COLUMN_WIDTH

    excel_path = config.BASE_DIR / config.EMITENTS_XLSX_FILENAME
    wb.save(excel_path)
    return len(rows)


def export_emitents_snapshot(conn: sqlite3.Connection) -> int:
    cursor = conn.execute(
        f'''
        SELECT "EMITENTNAME", "INN", "Scoring", "DateScoring", "NRA_Rate", "Acra_Rate", "NKR_Rate", "RAEX_Rate"
        FROM "{config.EMITENTS_TABLE_NAME}"
        ORDER BY RANDOM()
        LIMIT 5
        '''
    )
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = "emitents_snapshot"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / config.EMITENTS_SNAPSHOT_FILENAME
    wb.save(snapshot_path)
    return len(rows)


def export_nra_snapshot(conn: sqlite3.Connection) -> int:
    query = f'''
    SELECT "organization_name", "inn", "press_release_date", "rating", "rating_status", "forecast"
    FROM "{config.NRA_LATEST_TABLE_NAME}"
    ORDER BY
        CASE
            WHEN instr("press_release_date", '.') > 0 THEN substr("press_release_date", 7, 4) || '-' || substr("press_release_date", 4, 2) || '-' || substr("press_release_date", 1, 2)
            ELSE "press_release_date"
        END DESC,
        "inn" ASC
    LIMIT 5
    '''
    cursor = conn.execute(query)
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = "nra_snapshot"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / config.NRA_SNAPSHOT_FILENAME
    wb.save(snapshot_path)
    return len(rows)


def export_acra_snapshot(conn: sqlite3.Connection) -> int:
    cursor = conn.execute(
        f'''
        SELECT "issuer_name", "issuer_url", "rating", "forecast", "rating_date", "inn"
        FROM "{config.ACRA_TABLE_NAME}"
        ORDER BY
            CASE
                WHEN instr("rating_date", '.') > 0 THEN substr("rating_date", 7, 4) || '-' || substr("rating_date", 4, 2) || '-' || substr("rating_date", 1, 2)
                ELSE "rating_date"
            END DESC,
            "loaded_at_utc" DESC
        LIMIT 5
        '''
    )
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = "acra_snapshot"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / config.ACRA_SNAPSHOT_FILENAME
    wb.save(snapshot_path)
    return len(rows)


def export_nkr_snapshot(conn: sqlite3.Connection) -> int:
    cursor = conn.execute(
        f'''
        SELECT "issuer_name", "tin", "rating_date", "rating", "outlook"
        FROM "{config.NKR_LATEST_TABLE_NAME}"
        ORDER BY
            CASE
                WHEN instr("rating_date", '.') > 0 THEN substr("rating_date", 7, 4) || '-' || substr("rating_date", 4, 2) || '-' || substr("rating_date", 1, 2)
                ELSE "rating_date"
            END DESC,
            "tin" ASC
        LIMIT 5
        '''
    )
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = "nkr_snapshot"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / config.NKR_SNAPSHOT_FILENAME
    wb.save(snapshot_path)
    return len(rows)


def export_raex_snapshot(conn: sqlite3.Connection) -> int:
    cursor = conn.execute(
        f'''
        SELECT "inn", "company_name", "rating_date", "rating", "forecast", "company_url"
        FROM "{config.RAEX_LATEST_TABLE_NAME}"
        ORDER BY
            CASE
                WHEN instr("rating_date", '.') > 0 THEN substr("rating_date", 7, 4) || '-' || substr("rating_date", 4, 2) || '-' || substr("rating_date", 1, 2)
                ELSE "rating_date"
            END DESC,
            "inn" ASC
        LIMIT 5
        '''
    )
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = "raex_snapshot"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / config.RAEX_SNAPSHOT_FILENAME
    wb.save(snapshot_path)
    return len(rows)


def save_text_file(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8")


def refresh_data_if_needed(conn: sqlite3.Connection, logger: logging.Logger, now_utc: datetime) -> tuple[bool, int]:
    raw_path = config.RAW_DIR / config.RAW_FILENAME
    cache_path = config.CACHE_DIR / config.CACHE_FILENAME

    if not should_refresh_cache(conn, now_utc):
        logger.info("Кэш актуален: загрузка из сети пропущена.")
        row = conn.execute(f'SELECT COUNT(*) FROM "{config.RATES_TABLE_NAME}"').fetchone()
        return False, int(row[0]) if row else 0

    raw_text = download_csv(config.SOURCE_CSV_URL, config.REQUEST_TIMEOUT_SECONDS)

    if raw_path.exists():
        raw_path.unlink()
    if cache_path.exists():
        cache_path.unlink()

    save_text_file(raw_path, raw_text)
    save_text_file(cache_path, raw_text)

    headers, rows = parse_moex_rates_csv(raw_text)
    replace_rates_table(conn, headers, rows)

    set_meta_value(conn, "last_refresh_utc", now_utc.isoformat())
    set_meta_value(conn, "last_rows_count", str(len(rows)))
    set_meta_value(conn, "last_headers", "|".join(headers))

    logger.info("Данные обновлены: строк=%s, колонок=%s", len(rows), len(headers))
    return True, len(rows)


def refresh_dohod_data_if_needed(conn: sqlite3.Connection, logger: logging.Logger, now_utc: datetime) -> tuple[bool, int]:
    raw_path = config.RAW_DIR / config.DOHOD_RAW_FILENAME

    ensure_dohod_table(conn)
    if not should_refresh_dohod(conn, now_utc):
        logger.info("Доходъ: кэш актуален, загрузка из сети пропущена.")
        row = conn.execute(f'SELECT COUNT(*) FROM "{config.DOHOD_TABLE_NAME}"').fetchone()
        return False, int(row[0]) if row else 0

    content = download_dohod_excel_via_playwright(logger)
    if raw_path.exists():
        raw_path.unlink()
    raw_path.write_bytes(content)

    headers, rows = parse_dohod_excel(content)
    deduplicated_rows = _deduplicate_dohod_rows(rows, headers)
    ensure_dohod_table(conn, headers)
    ensure_table_has_columns(conn, config.DOHOD_TABLE_NAME, headers)

    insert_columns = [header for header in headers if header]
    quoted_cols = ", ".join([f'"{column}"' for column in insert_columns] + ['"loaded_at_utc"'])
    placeholders = ", ".join(["?"] * (len(insert_columns) + 1))
    update_cols = [column for column in insert_columns if column != "ISIN"]
    update_expr = ", ".join([f'"{column}"=excluded."{column}"' for column in update_cols] + ['"loaded_at_utc"=excluded."loaded_at_utc"'])
    payload = [
        tuple(((_normalize_isin(row.get(column, "")) if column == "ISIN" else row.get(column, "")) for column in insert_columns))
        + (now_utc.isoformat(),)
        for row in deduplicated_rows
        if _normalize_isin(row.get("ISIN", ""))
    ]

    conn.execute("BEGIN")
    conn.executemany(
        f'''
        INSERT INTO "{config.DOHOD_TABLE_NAME}" ({quoted_cols})
        VALUES ({placeholders})
        ON CONFLICT("ISIN") DO UPDATE SET {update_expr}
        ''',
        payload,
    )
    conn.commit()

    set_meta_value(conn, "dohod_last_refresh_utc", now_utc.isoformat())
    set_meta_value(conn, "dohod_last_rows_count", str(len(payload)))
    set_meta_value(conn, "dohod_last_headers", "|".join(insert_columns))
    logger.info(
        "Доходъ: данные обновлены, исходных строк=%s, после дедупликации по ISIN=%s, колонок=%s",
        len(rows),
        len(payload),
        len(insert_columns),
    )
    return True, len(payload)


def export_random_snapshot(conn: sqlite3.Connection) -> int:
    query = f"""
    SELECT *
    FROM "{config.RATES_TABLE_NAME}"
    WHERE rowid IN (
        SELECT MIN(rowid)
        FROM "{config.RATES_TABLE_NAME}"
        GROUP BY "SECID"
        ORDER BY RANDOM()
        LIMIT 5
    )
    """

    cursor = conn.execute(query)
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = "snapshot"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / config.SNAPSHOT_FILENAME
    wb.save(snapshot_path)
    return len(rows)


def export_dohod_snapshot(conn: sqlite3.Connection) -> int:
    query = f'''
    SELECT *
    FROM "{config.DOHOD_TABLE_NAME}"
    WHERE rowid IN (
        SELECT MIN(rowid)
        FROM "{config.DOHOD_TABLE_NAME}"
        GROUP BY "ISIN"
        ORDER BY RANDOM()
        LIMIT 5
    )
    '''
    cursor = conn.execute(query)
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = "dohod_snapshot"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / config.DOHOD_SNAPSHOT_FILENAME
    wb.save(snapshot_path)
    return len(rows)


MERGE_MOEX_COLUMNS = [
    "SECID",
    "ISIN",
    "FACEVALUE",
    "FACEUNIT",
    "MATDATE",
    "IS_QUALIFIED_INVESTORS",
    "BOND_TYPE",
    "BOND_SUBTYPE",
    "YIELDATWAP",
    "PRICE",
]

MERGE_DOHOD_COLUMNS = [
    "Название",
    "Ближайшая дата погашения/оферты (Дата)",
    "Событие в дату",
    "Коэф. Ликвидности (max=100)",
    "Медиана дневного оборота (млн в валюте торгов)",
    "Цена Доход",
    "НКД",
    "Размер купона",
    "Текущий купон, %",
    "Тип купона",
    "Купон (раз/год)",
    "Субординированная (да/нет)",
]

SCREENER_TABLE_NAME = config.SCREENER_TABLE_NAME

SCREENER_SOURCE_TABLES = (
    (config.MERGE_GREEN_TABLE_NAME, "Green", 1),
    (config.MERGE_YELLOW_TABLE_NAME, "Yellow", 0),
)

BOND_OVERRIDES_SHEET_NAME = "BondOverrides"
BOND_OVERRIDES_HEADERS = ["ISIN", "Enabled", "Drop", "Квал", "Суборд", "CouponFormulaOverride", "Тип купона"]

SCREENER_COLUMNS = [
    "ISIN",
    "Название",
    "QUALIFIED",
    "Субординированная (да/нет)",
    "Corpbonds_Наличие амортизации",
    "Corpbonds_Купон лесенкой",
    "AmortStarrtDate",
    "MATDATE",
    "Offerdate",
    "Corpbonds_Дата ближайшего купона",
    "Corpbonds_Тип купона",
    "Smartlab_Длительность купона, дней",
    "НКД",
    "Текущий купон, %",
    "YTM",
    "Corpbonds_Формула купона",
    "FACEVALUE",
    "FACEUNIT",
    "Коэф. Ликвидности (max=100)",
    "Corpbonds_Цена последняя",
    "Цена Доход",
    "Smartlab_Котировка облигации, %",
    "PRICE",
    "Score",
    "SourceList",
]

SCREENER_EXPORT_COLUMNS = [
    ("ISIN", "ISIN"),
    ("Название", "Название"),
    ("QUALIFIED", "Квал"),
    ("Субординированная (да/нет)", "Суборд"),
    ("Corpbonds_Наличие амортизации", "Аморт"),
    ("Corpbonds_Купон лесенкой", "Лесенка"),
    ("AmortStarrtDate", "AmortStarrtDate"),
    ("MATDATE", "MATDATE"),
    ("Offerdate", "Offerdate"),
    ("Corpbonds_Дата ближайшего купона", "Ближайший купон"),
    ("Corpbonds_Тип купона", "Тип купона"),
    ("Smartlab_Длительность купона, дней", "КупонПериод"),
    ("НКД", "НКД"),
    ("Текущий купон, %", "Купон, %"),
    ("YTM", "YTM"),
    ("Corpbonds_Формула купона", "Формула купона"),
    ("FACEVALUE", "FACEVALUE"),
    ("FACEUNIT", "FACEUNIT"),
    ("Коэф. Ликвидности (max=100)", "Ликвидность"),
    ("Corpbonds_Цена последняя", "Цена Corpbonds"),
    ("Цена Доход", "Цена Доход"),
    ("Smartlab_Котировка облигации, %", "Цена Smartlab"),
    ("PRICE", "Цена MOEX"),
]

CORPBONDS_COLUMNS_MAP = {
    'Corpbonds_Цена последняя': 'Цена последняя',
    'Corpbonds_Тип купона': 'Тип купона',
    'Corpbonds_Ставка купона': 'Ставка купона',
    'Corpbonds_НКД': 'НКД',
    'Corpbonds_Формула купона': 'Формула купона',
    'Corpbonds_Дата ближайшего купона': 'Дата ближайшего купона',
    'Corpbonds_Дата ближайшей оферты': 'Дата ближайшей оферты',
    'Corpbonds_Наличие амортизации': 'Наличие амортизации',
    'Corpbonds_Купон лесенкой': 'Купон лесенкой',
}

SMARTLAB_COLUMNS_MAP = {
    'Smartlab_Котировка облигации, %': 'Котировка облигации, %',
    'Smartlab_Изм за день, %': 'Изм за день, %',
    'Smartlab_Объем день, млн. руб': 'Объем день, млн. руб',
    'Smartlab_Объем день, штук': 'Объем день, штук',
    'Smartlab_Дата оферты': 'Дата оферты',
    'Smartlab_Только для квалов?': 'Только для квалов?',
    'Smartlab_Длительность купона, дней': 'Длительность купона, дней',
}

AMORTIZATION_START_COLUMN = "AmortStarrtDate"



def ensure_table_columns(conn: sqlite3.Connection, table_name: str, columns: list[str]) -> None:
    existing_columns = {
        str(row[1]) for row in conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
    }
    for column in columns:
        if column not in existing_columns:
            conn.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{column}" TEXT')


def ensure_merge_table(conn: sqlite3.Connection, table_name: str) -> None:
    columns_sql = ['"ISIN" TEXT PRIMARY KEY']
    for column in MERGE_MOEX_COLUMNS:
        if column == "ISIN":
            continue
        columns_sql.append(f'"{column}" TEXT')
    for column in MERGE_DOHOD_COLUMNS:
        columns_sql.append(f'"{column}" TEXT')
    for column in CORPBONDS_COLUMNS_MAP:
        columns_sql.append(f'"{column}" TEXT')
    for column in SMARTLAB_COLUMNS_MAP:
        columns_sql.append(f'"{column}" TEXT')
    columns_sql.append(f'"{AMORTIZATION_START_COLUMN}" TEXT')
    conn.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({", ".join(columns_sql)})')
    # Обратная совместимость: в старых БД колонка называлась
    # "Цена, % от номинала". Новый пайплайн использует "Цена Доход".
    existing_columns = {
        str(row[1]) for row in conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
    }
    if "Цена Доход" not in existing_columns and "Цена, % от номинала" in existing_columns:
        conn.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "Цена Доход" TEXT')
        conn.execute(
            f'''
            UPDATE "{table_name}"
            SET "Цена Доход" = COALESCE(NULLIF(TRIM(COALESCE("Цена Доход", '')), ''), "Цена, % от номинала")
            '''
        )
    ensure_table_columns(
        conn,
        table_name,
        [column for column in MERGE_MOEX_COLUMNS if column != "ISIN"]
        + MERGE_DOHOD_COLUMNS
        + list(CORPBONDS_COLUMNS_MAP.keys())
        + list(SMARTLAB_COLUMNS_MAP.keys())
        + [AMORTIZATION_START_COLUMN],
    )
    conn.commit()


def rebuild_merge_table_by_scoring(conn: sqlite3.Connection, table_name: str, scoring: str) -> int:
    ensure_merge_table(conn, table_name)

    dohod_join_type = "INNER" if getattr(config, "MERGE_REQUIRE_DOHOD_ISIN_MATCH", True) else "LEFT"

    insert_columns = [f'"{column}"' for column in MERGE_MOEX_COLUMNS if column != "ISIN"]
    insert_columns = ['"ISIN"'] + insert_columns + [f'"{column}"' for column in MERGE_DOHOD_COLUMNS]
    insert_columns += [f'"{column}"' for column in CORPBONDS_COLUMNS_MAP]
    insert_columns += [f'"{column}"' for column in SMARTLAB_COLUMNS_MAP]
    insert_columns.append(f'"{AMORTIZATION_START_COLUMN}"')

    selected_columns = [f'm."ISIN"']
    selected_columns.extend(
        f'm."{column}"' for column in MERGE_MOEX_COLUMNS if column != "ISIN"
    )
    for column in MERGE_DOHOD_COLUMNS:
        if column == "Цена Доход":
            selected_columns.append('d."Цена, % от номинала" AS "Цена Доход"')
            continue
        selected_columns.append(f'd."{column}"')
    selected_columns.extend(f"'' AS \"{column}\"" for column in CORPBONDS_COLUMNS_MAP)
    selected_columns.extend(f"'' AS \"{column}\"" for column in SMARTLAB_COLUMNS_MAP)
    selected_columns.append(f"'' AS \"{AMORTIZATION_START_COLUMN}\"")
    placeholders = ", ".join(["?"] * len(insert_columns))

    rows = conn.execute(
        f'''
        SELECT {", ".join(selected_columns)}
        FROM "{config.RATES_TABLE_NAME}" m
        INNER JOIN "{config.EMITENTS_TABLE_NAME}" e
            ON TRIM(COALESCE(m."INN", '')) = TRIM(COALESCE(e."INN", ''))
        {dohod_join_type} JOIN "{config.DOHOD_TABLE_NAME}" d
            ON TRIM(COALESCE(m."ISIN", '')) = TRIM(COALESCE(d."ISIN", ''))
        WHERE LOWER(TRIM(COALESCE(e."Scoring", ''))) = LOWER(TRIM(?))
          AND TRIM(COALESCE(m."ISIN", '')) <> ''
        ''',
        (scoring,),
    ).fetchall()

    conn.execute("BEGIN")
    conn.execute(f'DELETE FROM "{table_name}"')
    if rows:
        conn.executemany(
            f'INSERT OR REPLACE INTO "{table_name}" ({", ".join(insert_columns)}) VALUES ({placeholders})',
            rows,
        )
    conn.commit()
    return len(rows)


def export_merge_snapshot(conn: sqlite3.Connection, table_name: str, snapshot_filename: str, sheet_name: str) -> int:
    query = f'''
    SELECT *
    FROM "{table_name}"
    WHERE rowid IN (
        SELECT MIN(rowid)
        FROM "{table_name}"
        GROUP BY "ISIN"
        ORDER BY RANDOM()
        LIMIT 5
    )
    '''
    cursor = conn.execute(query)
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / snapshot_filename
    wb.save(snapshot_path)
    return len(rows)


def _parse_bond_date(raw_value: str | None) -> datetime | None:
    if raw_value is None:
        return None
    value = str(raw_value).strip()
    if not value:
        return None

    date_part = value.split()[0].replace("/", ".")
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%Y.%m.%d"):
        try:
            return datetime.strptime(date_part, fmt)
        except ValueError:
            continue
    return None


def _normalize_date_to_iso(raw_value: str | None) -> str:
    parsed = _parse_bond_date(raw_value)
    return parsed.strftime("%Y-%m-%d") if parsed else ""


def _to_binary_flag(raw_value: str | None) -> int:
    value = str(raw_value or "").strip().casefold()
    if value in {"1", "да", "yes", "true"}:
        return 1
    return 0


def _merge_qualified(is_qualified_investors: str | None, smartlab_qualified: str | None) -> int:
    if _to_binary_flag(is_qualified_investors) == 1:
        return 1
    return 1 if _to_binary_flag(smartlab_qualified) == 1 else 0


def _pick_offer_date(corpbonds_offer: str | None, smartlab_offer: str | None) -> str:
    corpbonds_date = _normalize_date_to_iso(corpbonds_offer)
    if corpbonds_date:
        return corpbonds_date
    return _normalize_date_to_iso(smartlab_offer)


def _parse_decimal_value(raw_value: object) -> float | None:
    if raw_value is None:
        return None
    value = (
        str(raw_value)
        .replace("\xa0", " ")
        .replace("\u202f", " ")
        .replace("\u2009", " ")
        .replace("−", "-")
        .strip()
    )
    if not value:
        return None

    cleaned = re.sub(r"[^0-9,\.\-]", "", value.replace(" ", ""))
    if not cleaned or cleaned in {"-", ".", ",", "-.", "-,"}:
        return None

    # Поддерживаем форматы: 101,2 / 101.2 / 1 234,56% / 1,234.56
    last_comma = cleaned.rfind(",")
    last_dot = cleaned.rfind(".")
    decimal_pos = max(last_comma, last_dot)

    if decimal_pos >= 0:
        int_part = re.sub(r"[^0-9\-]", "", cleaned[:decimal_pos])
        frac_part = re.sub(r"[^0-9]", "", cleaned[decimal_pos + 1 :])
        if not int_part or int_part == "-":
            int_part = "0" if int_part == "" else int_part
        normalized = f"{int_part}.{frac_part}" if frac_part else int_part
    else:
        normalized = re.sub(r"[^0-9\-]", "", cleaned)

    try:
        return float(normalized)
    except ValueError:
        return None


def _is_false_like(raw_value: object) -> bool:
    value = str(raw_value or "").strip().casefold()
    return value in {"false", "нет", "0", "❌", "x", "х", "крестик"}


def _is_true_like(raw_value: object) -> bool:
    value = str(raw_value or "").strip().casefold()
    return value in {"true", "да", "1", "✅", "галочка", "yes"}


def _is_fixed_coupon_type(raw_value: object) -> bool:
    value = str(raw_value or "").strip().casefold()
    return value.startswith("фикс")


def _pick_price_for_ytm(*prices: object) -> float | None:
    for raw_price in prices:
        parsed = _parse_decimal_value(raw_price)
        if parsed is not None and parsed > 0:
            return parsed
    return None


def _normalize_purchase_price(price_percent: float, facevalue: float, nkd: float) -> float:
    return (facevalue * (price_percent / 100.0)) + nkd


def _format_ytm_percent(ytm_decimal: float) -> str:
    precision = max(0, int(getattr(config, "YTM_OUTPUT_PRECISION", 2)))
    return f"{ytm_decimal * 100:.{precision}f}"


def _nominal_periodic_to_effective_annual(rate_nominal: float, periods_per_year: float) -> float:
    if periods_per_year <= 0:
        return rate_nominal
    base = 1.0 + rate_nominal / periods_per_year
    if base <= 0:
        return rate_nominal
    return (base ** periods_per_year) - 1.0


def _is_ruble_faceunit(faceunit: object) -> bool:
    return str(faceunit or "").strip().upper() in {"RUB", "RUR", "SUR"}


def _resolve_nkd_for_dirty_price(
    *,
    faceunit: object,
    facevalue: float,
    corpbonds_nkd: object,
    fallback_nkd: object,
    secid: str,
    logger: logging.Logger,
) -> tuple[float, str]:
    candidates: list[tuple[str, float | None]] = []
    if _is_ruble_faceunit(faceunit):
        candidates.append(("corpbonds", _parse_decimal_value(corpbonds_nkd)))
    candidates.append(("fallback", _parse_decimal_value(fallback_nkd)))
    candidates.append(("zero", 0.0))

    threshold = facevalue * float(getattr(config, "NCD_FACEVALUE_SANITY_RATIO", 0.2))
    for source, value in candidates:
        if value is None:
            continue
        if value < 0:
            continue
        if facevalue > 0 and value > threshold:
            logger.warning(
                "YTM: suspicious NKD SECID=%s source=%s faceunit=%s nkd=%s facevalue=%s threshold=%s",
                secid,
                source,
                faceunit,
                value,
                facevalue,
                threshold,
            )
            continue
        return value, source
    return 0.0, "zero"


def _resolve_coupon_frequency_per_year(raw_coupon_period: object) -> float | None:
    parsed_value = _parse_decimal_value(raw_coupon_period)
    if parsed_value is None or parsed_value <= 0:
        return None

    # В колонке «КупонПериод» обычно приходят дни (например 182, 91, 30).
    # Для редких источников, где уже указано «раз в год», оставляем как есть.
    if parsed_value > 12:
        return 365.25 / parsed_value
    return parsed_value


def _load_amortization_schedule(
    conn: sqlite3.Connection,
    secids: set[str],
    facevalues: dict[str, float],
) -> dict[str, list[tuple[datetime, float]]]:
    schedule: dict[str, list[tuple[datetime, float]]] = {secid: [] for secid in secids}
    if not secids:
        return schedule

    placeholders = ", ".join(["?"] * len(secids))
    rows = conn.execute(
        f'''
        SELECT "secid", "amortdate", "value", "valueprc"
        FROM "{config.MOEX_AMORTIZATION_TABLE_NAME}"
        WHERE TRIM(COALESCE("secid", '')) IN ({placeholders})
        ORDER BY "secid", "amortdate"
        ''',
        tuple(secids),
    ).fetchall()

    for secid_raw, amortdate_raw, value_raw, valueprc_raw in rows:
        secid = str(secid_raw or "").strip()
        amortdate = _parse_bond_date(str(amortdate_raw or ""))
        if not secid or amortdate is None:
            continue

        payment = _parse_decimal_value(value_raw)
        if payment is None or payment <= 0:
            valueprc = _parse_decimal_value(valueprc_raw)
            facevalue = facevalues.get(secid)
            if valueprc is not None and valueprc > 0 and facevalue is not None and facevalue > 0:
                payment = facevalue * (valueprc / 100.0)

        if payment is None or payment <= 0:
            continue

        schedule.setdefault(secid, []).append((amortdate, float(payment)))

    return schedule


def _dates_match_with_tolerance(date_a: datetime.date, date_b: datetime.date, tolerance_days: int = 1) -> bool:
    return abs((date_a - date_b).days) <= tolerance_days


def _resolve_coupon_period_days(coupon_period_days: object, coupon_frequency: float) -> int:
    period_days = _parse_decimal_value(coupon_period_days)
    if period_days is None or period_days <= 0:
        period_days = 365.25 / coupon_frequency
    return max(1, int(round(period_days)))


def _build_coupon_dates(
    *,
    target_date: datetime,
    coupon_frequency: float,
    coupon_period_days: object,
    next_coupon_date: object,
) -> list[datetime.date]:
    today = datetime.now().date()
    target = target_date.date()
    if target <= today:
        return []

    period_days = _resolve_coupon_period_days(coupon_period_days, coupon_frequency)
    first_coupon_dt = _parse_bond_date(str(next_coupon_date or ""))
    first_coupon = first_coupon_dt.date() if first_coupon_dt is not None else None

    coupon_dates: list[datetime.date] = []
    if first_coupon is not None and first_coupon > today:
        current = first_coupon
        safety_limit = 2000
        while current <= target and len(coupon_dates) < safety_limit:
            coupon_dates.append(current)
            current = current + timedelta(days=period_days)
    else:
        current = today + timedelta(days=period_days)
        safety_limit = 2000
        while current <= target and len(coupon_dates) < safety_limit:
            coupon_dates.append(current)
            current = current + timedelta(days=period_days)

    return coupon_dates


def _build_cashflow_times_years(
    *,
    target_date: datetime,
    coupon_frequency: float,
    coupon_period_days: object,
    next_coupon_date: object,
    period_coupon: float,
    facevalue: float,
) -> list[tuple[float, float]]:
    today = datetime.now().date()
    target = target_date.date()
    if target <= today:
        return []

    period_days = _resolve_coupon_period_days(coupon_period_days, coupon_frequency)
    first_coupon_dt = _parse_bond_date(str(next_coupon_date or ""))
    first_coupon = first_coupon_dt.date() if first_coupon_dt is not None else None

    coupon_dates: list[datetime.date] = []
    if first_coupon is not None and first_coupon > today:
        current = first_coupon
    else:
        current = today + timedelta(days=period_days)

    safety_limit = 2000
    while current <= (target + timedelta(days=1)) and len(coupon_dates) < safety_limit:
        coupon_dates.append(current)
        current = current + timedelta(days=period_days)

    target_has_coupon = any(_dates_match_with_tolerance(coupon_date, target) for coupon_date in coupon_dates)

    cashflows: list[tuple[float, float]] = []
    for coupon_date in coupon_dates:
        if coupon_date > target:
            continue
        if _dates_match_with_tolerance(coupon_date, target):
            continue
        years = (coupon_date - today).days / 365.25
        if years > 0:
            cashflows.append((years, period_coupon))

    target_years = (target - today).days / 365.25
    if target_years <= 0:
        return []

    target_amount = facevalue
    if target_has_coupon:
        target_amount += period_coupon
    cashflows.append((target_years, target_amount))
    return cashflows


def _build_amortized_cashflows(
    *,
    target_date: datetime,
    coupon_frequency: float,
    coupon_period_days: object,
    next_coupon_date: object,
    facevalue: float,
    coupon_rate: float,
    amortization_schedule: list[tuple[datetime, float]] | None,
) -> list[tuple[float, float]]:
    today = datetime.now().date()
    target = target_date.date()
    if target <= today:
        return []

    coupon_dates = _build_coupon_dates(
        target_date=target_date,
        coupon_frequency=coupon_frequency,
        coupon_period_days=coupon_period_days,
        next_coupon_date=next_coupon_date,
    )
    amort_map: dict[datetime.date, float] = {}
    for amort_dt, amount in (amortization_schedule or []):
        amort_date = amort_dt.date()
        if today < amort_date <= target and amount > 0:
            amort_map[amort_date] = amort_map.get(amort_date, 0.0) + float(amount)

    event_dates = sorted(set(coupon_dates) | set(amort_map.keys()) | {target})
    if not event_dates:
        return []

    outstanding = facevalue
    period_coupon_rate = coupon_rate / coupon_frequency
    cashflows: list[tuple[float, float]] = []

    for event_date in event_dates:
        if outstanding <= 0:
            break

        amount = 0.0
        if event_date in coupon_dates:
            amount += outstanding * period_coupon_rate

        amort_payment = amort_map.get(event_date, 0.0)
        if amort_payment > 0:
            principal_payment = min(outstanding, amort_payment)
            amount += principal_payment
            outstanding -= principal_payment

        if event_date == target and outstanding > 0:
            amount += outstanding
            outstanding = 0.0

        years = (event_date - today).days / 365.25
        if amount > 0 and years > 0:
            cashflows.append((years, amount))

    return cashflows


def _solve_nominal_periodic_ytm_bisection(
    dirty_price: float,
    coupon_frequency: float,
    cashflows: list[tuple[float, float]],
) -> float | None:
    if not cashflows:
        return None

    def npv(rate: float) -> float:
        discount = 1.0 + rate / coupon_frequency
        if discount <= 0:
            return float("inf")
        total = 0.0
        for years, amount in cashflows:
            total += amount / (discount ** (years * coupon_frequency))
        return total - dirty_price

    left = -0.95
    right = 5.0
    left_val = npv(left)
    right_val = npv(right)

    if left_val * right_val > 0:
        return None

    for _ in range(120):
        mid = (left + right) / 2
        mid_val = npv(mid)
        if abs(mid_val) < 1e-8:
            return mid
        if left_val * mid_val <= 0:
            right = mid
        else:
            left = mid
            left_val = mid_val
    return (left + right) / 2


def _calculate_perpetual_subord_effective_current_yield(*, annual_coupon: float, dirty_price: float, compounding_frequency: float) -> float | None:
    if dirty_price <= 0 or compounding_frequency <= 0:
        return None
    current_coupon_yield = annual_coupon / dirty_price
    return _nominal_periodic_to_effective_annual(current_coupon_yield, compounding_frequency)


def _calculate_fixed_coupon_ytm(
    *,
    subord_flag: object,
    amort_flag: object,
    coupon_type: object,
    coupon_percent: object,
    coupon_frequency: object,
    coupon_period_days: object,
    next_coupon_date: object,
    faceunit: object,
    corpbonds_nkd: object,
    fallback_nkd: object,
    facevalue: object,
    matdate: object,
    offerdate: object,
    corpbonds_price: object,
    dohod_price: object,
    smartlab_price: object,
    moex_price: object,
    amortization_schedule: list[tuple[datetime, float]] | None,
    logger: logging.Logger,
    secid: str,
) -> str:
    if not _is_fixed_coupon_type(coupon_type):
        return ""

    price = _pick_price_for_ytm(corpbonds_price, dohod_price, smartlab_price, moex_price)
    coupon_rate_percent = _parse_decimal_value(coupon_percent)
    coupon_freq = _resolve_coupon_frequency_per_year(coupon_frequency)
    if coupon_freq is None:
        coupon_freq = _resolve_coupon_frequency_per_year(coupon_period_days)
    facevalue_value = _parse_decimal_value(facevalue)
    target_date = _parse_bond_date(str(offerdate or "")) or _parse_bond_date(str(matdate or ""))

    if price is None or facevalue_value is None or target_date is None:
        return ""
    coupon_rate_percent = coupon_rate_percent or 0.0
    is_zero_coupon = (
        coupon_rate_percent <= 0
        and (
            coupon_freq is None
            or coupon_freq <= 0
            or _parse_bond_date(str(next_coupon_date or "")) is None
        )
    )

    if facevalue_value <= 0:
        return ""
    if not is_zero_coupon and (coupon_freq is None or coupon_freq <= 0):
        return ""

    nkd_value, _ = _resolve_nkd_for_dirty_price(
        faceunit=faceunit,
        facevalue=facevalue_value,
        corpbonds_nkd=corpbonds_nkd,
        fallback_nkd=fallback_nkd,
        secid=secid,
        logger=logger,
    )

    days_to_redemption = (target_date.date() - datetime.now().date()).days
    if days_to_redemption <= 0:
        return ""

    dirty_price = _normalize_purchase_price(price, facevalue_value, nkd_value)

    coupon_rate = coupon_rate_percent / 100.0
    annual_coupon = facevalue_value * coupon_rate
    solver_frequency = coupon_freq if coupon_freq is not None and coupon_freq > 0 else 1.0

    is_subord = _is_true_like(subord_flag)
    is_amort = _is_true_like(amort_flag)
    if is_subord:
        effective_yield = _calculate_perpetual_subord_effective_current_yield(
            annual_coupon=annual_coupon,
            dirty_price=dirty_price,
            compounding_frequency=solver_frequency,
        )
        if effective_yield is None:
            return ""
        return _format_ytm_percent(effective_yield)

    if is_zero_coupon:
        cashflows = _build_amortized_cashflows(
            target_date=target_date,
            coupon_frequency=1.0,
            coupon_period_days=coupon_period_days,
            next_coupon_date=next_coupon_date,
            facevalue=facevalue_value,
            coupon_rate=0.0,
            amortization_schedule=amortization_schedule,
        )
        if not cashflows:
            years = (target_date.date() - datetime.now().date()).days / 365.25
            if years <= 0:
                return ""
            cashflows = [(years, facevalue_value)]
    elif is_amort:
        cashflows = _build_amortized_cashflows(
            target_date=target_date,
            coupon_frequency=solver_frequency,
            coupon_period_days=coupon_period_days,
            next_coupon_date=next_coupon_date,
            facevalue=facevalue_value,
            coupon_rate=coupon_rate,
            amortization_schedule=amortization_schedule,
        )
    else:
        period_coupon = annual_coupon / solver_frequency
        cashflows = _build_cashflow_times_years(
            target_date=target_date,
            coupon_frequency=solver_frequency,
            coupon_period_days=coupon_period_days,
            next_coupon_date=next_coupon_date,
            period_coupon=period_coupon,
            facevalue=facevalue_value,
        )

    if not cashflows:
        return ""

    ytm_nominal = _solve_nominal_periodic_ytm_bisection(
        dirty_price=dirty_price,
        coupon_frequency=solver_frequency,
        cashflows=cashflows,
    )
    if ytm_nominal is None:
        return ""
    ytm_effective = _nominal_periodic_to_effective_annual(ytm_nominal, solver_frequency)
    return _format_ytm_percent(ytm_effective)


def _run_ytm_self_check() -> list[str]:
    errors: list[str] = []

    coupon_freq = 4.0
    period_coupon = 40.0
    facevalue = 1000.0

    target_coupon = datetime.now() + timedelta(days=180)
    cashflows_coupon = _build_cashflow_times_years(
        target_date=target_coupon,
        coupon_frequency=coupon_freq,
        coupon_period_days=90,
        next_coupon_date=datetime.now() + timedelta(days=90),
        period_coupon=period_coupon,
        facevalue=facevalue,
    )
    ytm_1 = _solve_nominal_periodic_ytm_bisection(dirty_price=980.0, coupon_frequency=coupon_freq, cashflows=cashflows_coupon)
    if ytm_1 is None:
        errors.append("Self-check #1: не удалось решить YTM для базового фиксированного кейса.")
    else:
        pv_1 = sum(amount / ((1 + ytm_1 / coupon_freq) ** (years * coupon_freq)) for years, amount in cashflows_coupon)
        if abs(pv_1 - 980.0) > 1e-6:
            errors.append("Self-check #1: PV(cashflows, ytm) не совпадает с dirty price.")
        expected_effective_1 = (1.0 + ytm_1 / coupon_freq) ** coupon_freq - 1.0
        helper_effective_1 = _nominal_periodic_to_effective_annual(ytm_1, coupon_freq)
        if abs(expected_effective_1 - helper_effective_1) > 1e-12:
            errors.append("Self-check #1b: annualization helper mismatch для nominal solver.")

    target_non_coupon = datetime.now() + timedelta(days=150)
    cashflows_non_coupon = _build_cashflow_times_years(
        target_date=target_non_coupon,
        coupon_frequency=coupon_freq,
        coupon_period_days=91,
        next_coupon_date=datetime.now() + timedelta(days=90),
        period_coupon=period_coupon,
        facevalue=facevalue,
    )
    if len(cashflows_non_coupon) > 1 and abs(cashflows_non_coupon[-1][1] - facevalue) > 1e-9:
        errors.append("Self-check #2: на некупонной target_date появился лишний купон.")

    amort_schedule = [
        (datetime.now() + timedelta(days=120), 300.0),
        (datetime.now() + timedelta(days=240), 300.0),
    ]
    cashflows_amort = _build_amortized_cashflows(
        target_date=datetime.now() + timedelta(days=360),
        coupon_frequency=2.0,
        coupon_period_days=182,
        next_coupon_date=datetime.now() + timedelta(days=182),
        facevalue=1000.0,
        coupon_rate=0.12,
        amortization_schedule=amort_schedule,
    )
    ytm_3 = _solve_nominal_periodic_ytm_bisection(dirty_price=960.0, coupon_frequency=2.0, cashflows=cashflows_amort)
    if ytm_3 is None:
        errors.append("Self-check #3: не удалось решить YTM для амортизационного кейса.")
    else:
        pv_3 = sum(amount / ((1 + ytm_3 / 2.0) ** (years * 2.0)) for years, amount in cashflows_amort)
        if abs(pv_3 - 960.0) > 1e-6:
            errors.append("Self-check #3: PV(cashflows, ytm) для амортизации не совпадает с dirty price.")

    regression_logger = logging.getLogger("bonds_main")
    ytm_monthly = _calculate_fixed_coupon_ytm(
        subord_flag="Нет",
        amort_flag="Нет",
        coupon_type="Фиксированный",
        coupon_percent=17,
        coupon_frequency=None,
        coupon_period_days=30,
        next_coupon_date="2026-04-02",
        faceunit="RUB",
        corpbonds_nkd=2.79,
        fallback_nkd=2.79,
        facevalue=1000,
        matdate="2029-02-15",
        offerdate="",
        corpbonds_price=100.3,
        dohod_price="",
        smartlab_price="",
        moex_price="",
        amortization_schedule=None,
        logger=regression_logger,
        secid="RU000A10EES4",
    )
    ytm_monthly_value = _parse_decimal_value(ytm_monthly)
    if ytm_monthly_value is None or ytm_monthly_value < 18.0:
        errors.append("Self-check #4: monthly fixed RU000A10EES4-like case должен давать displayed YTM в районе 18.x.")

    ytm_other = _calculate_other_coupon_ytm(
        subord_flag="Нет",
        coupon_percent=13,
        coupon_frequency=None,
        coupon_period_days=30,
        next_coupon_date="2026-03-31",
        amort_flag="Нет",
        faceunit="RUB",
        corpbonds_nkd=142.47,
        fallback_nkd=142.47,
        facevalue=50000,
        matdate="2028-12-15",
        offerdate="",
        corpbonds_price=90.71,
        dohod_price="",
        smartlab_price="",
        moex_price="",
        amortization_schedule=None,
        logger=regression_logger,
        secid="RU000A10DCM3",
    )
    ytm_other_value = _parse_decimal_value(ytm_other)
    if ytm_other_value is None or ytm_other_value < 15.0:
        errors.append("Self-check #5: other coupon RU000A10DCM3-like case должен давать high-teens, не 3-4%.")

    return errors

def _is_floater_coupon_type(raw_value: object) -> bool:
    return "флоат" in str(raw_value or "").strip().casefold()


def _is_other_coupon_type(raw_value: object) -> bool:
    return "проч" in str(raw_value or "").strip().casefold()


def parse_floater_terms(formula_raw: object) -> tuple[str, float | None, float] | None:
    formula = str(formula_raw or "")
    if not formula.strip():
        return None

    normalized = formula.casefold().replace("ё", "е")
    normalized = normalized.replace("∑", "").replace("Σ", "").replace("σ", "").replace("?", "")
    normalized = normalized.replace(",", ".")
    normalized = normalized.translate(str.maketrans({"к": "k", "с": "c"}))
    normalized = re.sub(r"\s+", " ", normalized).strip()

    index_type = ""
    tenor_years: float | None = None

    zcyc_match = re.search(r"(?:g\s*[- ]?curve|gcurve|zcyc)\s*([0-9]+(?:\.[0-9]+)?)\s*(?:y|yr|year|лет|год(?:а|ов)?)", normalized)
    if zcyc_match:
        index_type = "zcyc"
        tenor_years = float(zcyc_match.group(1))
    elif "ruonia" in normalized:
        index_type = "ruonia"
    elif "ключев" in normalized or re.search(r"\bkc\b", normalized):
        index_type = "key"

    if not index_type:
        return None

    premium = 0.0
    premium_match = re.search(r"([+-])\s*([0-9]+(?:\.[0-9]+)?)\s*%", normalized)
    if premium_match:
        sign = premium_match.group(1)
        number = float(premium_match.group(2))
        premium = number if sign == "+" else -number
    return (index_type, tenor_years, premium)


def ensure_cbr_cache_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        '''
        CREATE TABLE IF NOT EXISTS "CBRIndicatorsCache" (
            "cache_key" TEXT PRIMARY KEY,
            "payload_json" TEXT NOT NULL,
            "updated_at_utc" TEXT NOT NULL
        )
        '''
    )
    conn.commit()


def _fetch_cbr_key_rate() -> float | None:
    session = create_resilient_http_session(pool_size=4)
    response = session.get(config.CBR_KEY_RATE_URL, timeout=30)
    response.raise_for_status()
    parser = etree.HTMLParser(recover=True)
    root = etree.HTML(response.text, parser=parser)
    if root is None:
        return None

    latest_date: datetime | None = None
    latest_rate: float | None = None
    for row in root.xpath("//table//tr[td]"):
        cells = [" ".join("".join(x.itertext()).split()) for x in row.xpath("./td")]
        if len(cells) < 2:
            continue
        row_date = _parse_bond_date(cells[0])
        rate = _parse_decimal_value(cells[1])
        if row_date is None or rate is None:
            continue
        if latest_date is None or row_date > latest_date:
            latest_date = row_date
            latest_rate = rate
    return latest_rate


def _fetch_cbr_ruonia() -> float | None:
    session = create_resilient_http_session(pool_size=4)
    response = session.get(config.CBR_RUONIA_URL, timeout=30)
    response.raise_for_status()
    parser = etree.HTMLParser(recover=True)
    root = etree.HTML(response.text, parser=parser)
    if root is None:
        return None

    for row in root.xpath("//table//tr[td]"):
        cells = [" ".join("".join(x.itertext()).split()) for x in row.xpath("./td")]
        if len(cells) < 2:
            continue
        if "ruonia" not in _normalize_label(cells[0]):
            continue
        values = [_parse_decimal_value(value) for value in cells[1:]]
        values = [value for value in values if value is not None]
        if values:
            return float(values[-1])
    return None


def _fetch_cbr_zcyc_curve() -> dict[float, float]:
    session = create_resilient_http_session(pool_size=4)
    response = session.get(config.CBR_ZCYC_URL, timeout=30)
    response.raise_for_status()
    parser = etree.HTMLParser(recover=True)
    root = etree.HTML(response.text, parser=parser)
    curve: dict[float, float] = {}
    if root is None:
        return curve

    tenors: list[float] = []
    latest_row_date: datetime | None = None
    latest_curve: dict[float, float] = {}
    for row in root.xpath("//table//tr"):
        header_cells = [" ".join("".join(cell.itertext()).split()) for cell in row.xpath("./th")]
        if header_cells and not tenors:
            parsed_tenors = [_parse_decimal_value(cell) for cell in header_cells[1:]]
            parsed_tenors = [float(x) for x in parsed_tenors if x is not None]
            if len(parsed_tenors) >= 10 and any(abs(x - 7.0) < 1e-9 for x in parsed_tenors):
                tenors = parsed_tenors
                continue
        cells = [" ".join("".join(cell.itertext()).split()) for cell in row.xpath("./td")]
        if len(cells) < 13:
            continue
        row_date = _parse_bond_date(cells[0])
        if row_date is None:
            continue
        if not tenors:
            tenors = [0.25, 0.5, 0.75, 1.0, 2.0, 3.0, 5.0, 7.0, 10.0, 15.0, 20.0, 30.0]
        yields = [_parse_decimal_value(value) for value in cells[1 : len(tenors) + 1]]
        if any(v is None for v in yields):
            continue
        if latest_row_date is None or row_date > latest_row_date:
            latest_row_date = row_date
            latest_curve = {tenor: float(yld or 0.0) for tenor, yld in zip(tenors, yields)}
    curve = latest_curve
    return curve


def get_cbr_reference_data(conn: sqlite3.Connection, logger: logging.Logger, now_utc: datetime) -> dict[str, object]:
    ensure_cbr_cache_table(conn)
    cache_key = "cbr_reference_v1"
    if getattr(config, "FORCE_CBR_TTL_RESET", False):
        conn.execute('UPDATE "CBRIndicatorsCache" SET "updated_at_utc" = ?', ("",))
        conn.commit()
        logger.info("CBR cache TTL reset forced via config.FORCE_CBR_TTL_RESET")
    row = conn.execute('SELECT "payload_json", "updated_at_utc" FROM "CBRIndicatorsCache" WHERE "cache_key"=?', (cache_key,)).fetchone()
    if row:
        try:
            updated_at = datetime.fromisoformat(str(row[1]))
            cached_payload = json.loads(str(row[0]))
            has_cached_data = bool(cached_payload.get("zcyc")) or _parse_decimal_value(cached_payload.get("key_rate")) is not None
            if now_utc - updated_at < timedelta(hours=getattr(config, "CBR_CACHE_TTL_HOURS", 12)) and has_cached_data:
                return cached_payload
        except Exception:
            pass

    payload: dict[str, object] = {"key_rate": None, "ruonia": None, "zcyc": {}}
    try:
        payload["key_rate"] = _fetch_cbr_key_rate()
        payload["ruonia"] = _fetch_cbr_ruonia()
        payload["zcyc"] = _fetch_cbr_zcyc_curve()
    except Exception as exc:
        logger.warning("CBR cache update error: %s", exc)

    has_fresh_data = bool(payload.get("zcyc")) or _parse_decimal_value(payload.get("key_rate")) is not None
    if has_fresh_data:
        conn.execute(
            'INSERT OR REPLACE INTO "CBRIndicatorsCache" ("cache_key", "payload_json", "updated_at_utc") VALUES (?, ?, ?)',
            (cache_key, json.dumps(payload, ensure_ascii=False), now_utc.isoformat()),
        )
        conn.commit()
    else:
        logger.warning("CBR cache payload is empty, TTL update skipped")
    return payload


def is_linker(secid: str, bond_type: object, bond_subtype: object) -> bool:
    secid_norm = str(secid or "").strip().upper()
    type_norm = f"{bond_type or ''} {bond_subtype or ''}".casefold()
    if any(token in type_norm for token in getattr(config, "LINKER_BOND_TYPE_PATTERNS", ())):
        return True
    return any(secid_norm.startswith(prefix.upper()) for prefix in getattr(config, "LINKER_SECID_PREFIXES", ()))


def _forecast_by_bucket(forecast_cfg: dict[int, float], bucket: int) -> float:
    return float(forecast_cfg.get(bucket, forecast_cfg.get(2, 0.0)))


def _year_bucket(event_date: datetime.date, valuation_date: datetime.date) -> int:
    return max(0, min(2, int((event_date - valuation_date).days / 365.25)))


def pick_zcyc_point(zcyc: dict[object, object], tenor_years: float | None) -> float | None:
    if tenor_years is None or not zcyc:
        return None
    normalized: dict[float, float] = {}
    for key, value in zcyc.items():
        key_num = _parse_decimal_value(key)
        val_num = _parse_decimal_value(value)
        if key_num is None or val_num is None:
            continue
        normalized[float(key_num)] = float(val_num)
    if not normalized:
        return None
    points = sorted(normalized.items())
    if tenor_years <= points[0][0]:
        return points[0][1]
    if tenor_years >= points[-1][0]:
        return points[-1][1]
    for idx in range(1, len(points)):
        left_tenor, left_yield = points[idx - 1]
        right_tenor, right_yield = points[idx]
        if abs(left_tenor - tenor_years) < 1e-9:
            return left_yield
        if abs(right_tenor - tenor_years) < 1e-9:
            return right_yield
        if left_tenor < tenor_years < right_tenor:
            alpha = (tenor_years - left_tenor) / (right_tenor - left_tenor)
            return left_yield + alpha * (right_yield - left_yield)
    nearest_key = min(normalized.keys(), key=lambda k: abs(k - tenor_years))
    return normalized.get(nearest_key)


def _calculate_floater_ytm(*, coupon_formula: object, subord_flag: object, coupon_frequency: object, coupon_period_days: object, next_coupon_date: object, faceunit: object, corpbonds_nkd: object, fallback_nkd: object, facevalue: object, matdate: object, offerdate: object, corpbonds_price: object, dohod_price: object, smartlab_price: object, moex_price: object, amortization_schedule: list[tuple[datetime, float]] | None, cbr_data: dict[str, object], logger: logging.Logger, secid: str) -> str:
    terms = parse_floater_terms(coupon_formula)
    if not terms:
        logger.info("Floater YTM skipped SECID=%s: parse formula fail", secid)
        _write_ytm_debug_record({"secid": secid, "coupon_formula": str(coupon_formula or ""), "reason": "parse_fail"})
        return ""

    index_type, tenor_years, premium = terms
    key_current = _parse_decimal_value(cbr_data.get("key_rate"))
    if key_current is None:
        logger.info("Floater YTM skipped SECID=%s: key rate unavailable", secid)
        _write_ytm_debug_record({"secid": secid, "coupon_formula": str(coupon_formula or ""), "reason": "no_keyrate"})
        return ""

    if index_type == "key":
        spread_to_key = 0.0
    elif index_type == "ruonia":
        idx = _parse_decimal_value(cbr_data.get("ruonia"))
        if idx is None:
            logger.info("Floater YTM skipped SECID=%s: RUONIA unavailable", secid)
            _write_ytm_debug_record({"secid": secid, "coupon_formula": str(coupon_formula or ""), "reason": "no_ruonia"})
            return ""
        spread_to_key = key_current - idx
    else:
        zcyc = cbr_data.get("zcyc") if isinstance(cbr_data.get("zcyc"), dict) else {}
        idx = pick_zcyc_point(zcyc, tenor_years)
        if idx is None:
            logger.info("Floater YTM skipped SECID=%s: G-Curve unavailable", secid)
            _write_ytm_debug_record({"secid": secid, "coupon_formula": str(coupon_formula or ""), "reason": "no_zcyc"})
            return ""
        spread_to_key = key_current - idx

    price = _pick_price_for_ytm(corpbonds_price, dohod_price, smartlab_price, moex_price)
    coupon_freq = _resolve_coupon_frequency_per_year(coupon_frequency) or _resolve_coupon_frequency_per_year(coupon_period_days)
    facevalue_value = _parse_decimal_value(facevalue)
    target_date = _parse_bond_date(str(offerdate or "")) or _parse_bond_date(str(matdate or ""))
    if price is None or coupon_freq is None or facevalue_value is None or target_date is None:
        _write_ytm_debug_record({"secid": secid, "coupon_formula": str(coupon_formula or ""), "reason": "no_price_or_dates"})
        return ""

    nkd_value, _ = _resolve_nkd_for_dirty_price(
        faceunit=faceunit,
        facevalue=facevalue_value,
        corpbonds_nkd=corpbonds_nkd,
        fallback_nkd=fallback_nkd,
        secid=secid,
        logger=logger,
    )

    dirty_price = _normalize_purchase_price(price, facevalue_value, nkd_value)
    coupon_dates = _build_coupon_dates(target_date=target_date, coupon_frequency=coupon_freq, coupon_period_days=coupon_period_days, next_coupon_date=next_coupon_date)
    if not coupon_dates:
        logger.info("Floater YTM skipped SECID=%s: no coupon periods", secid)
        _write_ytm_debug_record({"secid": secid, "coupon_formula": str(coupon_formula or ""), "reason": "no_dates"})
        return ""

    valuation_date = datetime.now().date()
    principal = facevalue_value
    amort_map = {d.date(): p for d, p in (amortization_schedule or []) if p > 0}
    cashflows: list[tuple[float, float]] = []
    event_dates = sorted(set(coupon_dates) | set(amort_map.keys()) | {target_date.date()})
    for dt in event_dates:
        if dt <= valuation_date:
            continue
        amount = 0.0
        if dt in coupon_dates:
            bucket = _year_bucket(dt, valuation_date)
            index_forecast = _forecast_by_bucket(getattr(config, "KEY_RATE_FORECAST", {0: key_current, 2: key_current}), bucket) - spread_to_key
            coupon_rate = index_forecast + premium
            amount += principal * (coupon_rate / 100.0) / coupon_freq
        if dt in amort_map:
            principal_payment = min(principal, amort_map[dt])
            amount += principal_payment
            principal -= principal_payment
        if dt == target_date.date() and principal > 0:
            amount += principal
        years = (dt - valuation_date).days / 365.25
        if amount > 0:
            cashflows.append((years, amount))

    if _is_true_like(subord_flag):
        rate0 = _forecast_by_bucket(getattr(config, "KEY_RATE_FORECAST", {0: key_current, 2: key_current}), 0) - spread_to_key + premium
        current_coupon = facevalue_value * (rate0 / 100.0)
        effective = _calculate_perpetual_subord_effective_current_yield(
            annual_coupon=current_coupon,
            dirty_price=dirty_price,
            compounding_frequency=coupon_freq,
        )
        if effective is None:
            return ""
        return _format_ytm_percent(effective)

    ytm_nominal = _solve_nominal_periodic_ytm_bisection(dirty_price=dirty_price, coupon_frequency=coupon_freq, cashflows=cashflows)
    if ytm_nominal is None:
        return ""
    ytm_effective = _nominal_periodic_to_effective_annual(ytm_nominal, coupon_freq)
    return _format_ytm_percent(ytm_effective)


def _calculate_linker_ytm(*, secid: str, bond_type: object, bond_subtype: object, coupon_percent: object, coupon_frequency: object, coupon_period_days: object, next_coupon_date: object, faceunit: object, corpbonds_nkd: object, fallback_nkd: object, facevalue: object, matdate: object, offerdate: object, corpbonds_price: object, dohod_price: object, smartlab_price: object, moex_price: object, logger: logging.Logger) -> str:
    if not is_linker(secid, bond_type, bond_subtype):
        return ""
    price = _pick_price_for_ytm(corpbonds_price, dohod_price, smartlab_price, moex_price)
    coupon_rate_percent = _parse_decimal_value(coupon_percent)
    coupon_freq = _resolve_coupon_frequency_per_year(coupon_frequency) or _resolve_coupon_frequency_per_year(coupon_period_days)
    facevalue_value = _parse_decimal_value(facevalue)
    target_date = _parse_bond_date(str(offerdate or "")) or _parse_bond_date(str(matdate or ""))
    if price is None or coupon_rate_percent is None or coupon_freq is None or facevalue_value is None or target_date is None:
        return ""
    nkd_value, _ = _resolve_nkd_for_dirty_price(
        faceunit=faceunit,
        facevalue=facevalue_value,
        corpbonds_nkd=corpbonds_nkd,
        fallback_nkd=fallback_nkd,
        secid=secid,
        logger=logger,
    )
    dirty_price = _normalize_purchase_price(price, facevalue_value, nkd_value)
    coupon_dates = _build_coupon_dates(target_date=target_date, coupon_frequency=coupon_freq, coupon_period_days=coupon_period_days, next_coupon_date=next_coupon_date)
    valuation_date = datetime.now().date()
    cashflows: list[tuple[float, float]] = []
    for dt in coupon_dates:
        if dt <= valuation_date:
            continue
        years = (dt - valuation_date).days / 365.25
        bucket = _year_bucket(dt, valuation_date)
        inf = _forecast_by_bucket(getattr(config, "INFLATION_FORECAST", {0: 5.4, 2: 4.0}), bucket) / 100.0
        principal_t = facevalue_value * ((1 + inf) ** years)
        amount = principal_t * (coupon_rate_percent / 100.0) / coupon_freq
        cashflows.append((years, amount))
    years_target = (target_date.date() - valuation_date).days / 365.25
    if years_target <= 0:
        return ""
    bucket_t = _year_bucket(target_date.date(), valuation_date)
    inf_t = _forecast_by_bucket(getattr(config, "INFLATION_FORECAST", {0: 5.4, 2: 4.0}), bucket_t) / 100.0
    principal_target = facevalue_value * ((1 + inf_t) ** years_target)
    cashflows.append((years_target, principal_target))
    ytm_nominal = _solve_nominal_periodic_ytm_bisection(dirty_price=dirty_price, coupon_frequency=coupon_freq, cashflows=cashflows)
    if ytm_nominal is None:
        return ""
    ytm_effective = _nominal_periodic_to_effective_annual(ytm_nominal, coupon_freq)
    return _format_ytm_percent(ytm_effective)


def _calculate_other_coupon_ytm(*, subord_flag: object, coupon_percent: object, coupon_frequency: object, coupon_period_days: object, next_coupon_date: object, amort_flag: object, faceunit: object, corpbonds_nkd: object, fallback_nkd: object, facevalue: object, matdate: object, offerdate: object, corpbonds_price: object, dohod_price: object, smartlab_price: object, moex_price: object, amortization_schedule: list[tuple[datetime, float]] | None, logger: logging.Logger, secid: str) -> str:
    price = _pick_price_for_ytm(corpbonds_price, dohod_price, smartlab_price, moex_price)
    coupon_rate_percent = _parse_decimal_value(coupon_percent) or 0.0
    coupon_freq = _resolve_coupon_frequency_per_year(coupon_frequency) or _resolve_coupon_frequency_per_year(coupon_period_days)
    next_coupon_dt = _parse_bond_date(str(next_coupon_date or ""))
    tiny_coupon_threshold = float(getattr(config, "YTM_OTHER_COUPON_TINY_THRESHOLD", 1e-8))
    facevalue_value = _parse_decimal_value(facevalue)
    target_date = _parse_bond_date(str(offerdate or "")) or _parse_bond_date(str(matdate or ""))
    if price is None or facevalue_value is None or facevalue_value <= 0 or target_date is None:
        return ""

    nkd_value, _ = _resolve_nkd_for_dirty_price(
        faceunit=faceunit,
        facevalue=facevalue_value,
        corpbonds_nkd=corpbonds_nkd,
        fallback_nkd=fallback_nkd,
        secid=secid,
        logger=logger,
    )
    dirty_price = _normalize_purchase_price(price, facevalue_value, nkd_value)
    use_principal_only = (
        coupon_rate_percent <= tiny_coupon_threshold
        or coupon_freq is None
        or coupon_freq <= 0
        or next_coupon_dt is None
    )

    cashflows: list[tuple[float, float]]
    solver_frequency: float
    if use_principal_only:
        solver_frequency = coupon_freq if coupon_freq is not None and coupon_freq > 0 else 1.0
        valuation_date = datetime.now().date()
        principal = facevalue_value
        amort_map = {d.date(): p for d, p in (amortization_schedule or []) if p > 0}
        event_dates = sorted(set(amort_map.keys()) | {target_date.date()})
        cashflows = []
        for dt in event_dates:
            if dt <= valuation_date:
                continue
            amount = 0.0
            if dt in amort_map:
                principal_payment = min(principal, amort_map[dt])
                amount += principal_payment
                principal -= principal_payment
            if dt == target_date.date() and principal > 0:
                amount += principal
            years = (dt - valuation_date).days / 365.25
            if amount > 0 and years > 0:
                cashflows.append((years, amount))
    else:
        solver_frequency = coupon_freq
        coupon_rate = coupon_rate_percent / 100.0
        if _is_true_like(amort_flag):
            cashflows = _build_amortized_cashflows(
                target_date=target_date,
                coupon_frequency=solver_frequency,
                coupon_period_days=coupon_period_days,
                next_coupon_date=next_coupon_date,
                facevalue=facevalue_value,
                coupon_rate=coupon_rate,
                amortization_schedule=amortization_schedule,
            )
        else:
            period_coupon = facevalue_value * coupon_rate / solver_frequency
            cashflows = _build_cashflow_times_years(
                target_date=target_date,
                coupon_frequency=solver_frequency,
                coupon_period_days=coupon_period_days,
                next_coupon_date=next_coupon_date,
                period_coupon=period_coupon,
                facevalue=facevalue_value,
            )

    if not cashflows:
        logger.info("Other YTM skipped SECID=%s: reason=unsupported_other_coupon_cashflows", secid)
        return ""

    ytm_nominal = _solve_nominal_periodic_ytm_bisection(dirty_price=dirty_price, coupon_frequency=solver_frequency, cashflows=cashflows)
    if ytm_nominal is None:
        return ""
    ytm_effective = _nominal_periodic_to_effective_annual(ytm_nominal, solver_frequency)
    return _format_ytm_percent(ytm_effective)


def _screener_sort_key(row_values: tuple[object, ...]) -> tuple[int, datetime, str, str]:
    amort_raw = row_values[6] if len(row_values) > 6 else None
    amort_dt = _parse_bond_date(None if amort_raw is None else str(amort_raw))
    empty_amort = 1 if amort_dt is None else 0
    fallback_dt = datetime.max if amort_dt is None else amort_dt
    name = str(row_values[1] or "").strip() if len(row_values) > 1 else ""
    isin = str(row_values[0] or "").strip() if len(row_values) > 0 else ""
    return (empty_amort, fallback_dt, name, isin)


def _prepare_screener_export_row(headers: list[str], row_values: list[object]) -> list[object]:
    date_columns = {"AmortStarrtDate", "MATDATE", "Offerdate", "Ближайший купон"}
    float_columns = {
        "НКД",
        "Купон, %",
        "YTM",
        "FACEVALUE",
        "Цена Corpbonds",
        "Цена Доход",
        "Цена Smartlab",
        "Цена MOEX",
        "Ликвидность",
    }

    for index, header in enumerate(headers):
        value = row_values[index]
        if header in date_columns:
            parsed = _parse_bond_date(None if value is None else str(value))
            row_values[index] = parsed.date() if parsed else None
            continue
        if header in float_columns:
            row_values[index] = _parse_decimal_value(value)
            continue
        if header == "КупонПериод":
            parsed = _parse_decimal_value(value)
            row_values[index] = int(parsed) if parsed is not None else None

    return row_values


def _write_ytm_debug_record(payload: dict[str, object]) -> None:
    if not getattr(config, "ENABLE_YTM_DEBUG_LOG", True):
        return
    config.LOGS_DIR.mkdir(parents=True, exist_ok=True)
    path = config.LOGS_DIR / getattr(config, "YTM_DEBUG_FILENAME", "ytm_debug.jsonl")
    with path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(payload, ensure_ascii=False) + "\n")


def _normalize_bond_type(raw_value: str | None) -> str:
    value = str(raw_value or "").replace("\xa0", " ")
    return " ".join(value.split()).casefold()


def presort_merge_table(conn: sqlite3.Connection, table_name: str) -> dict[str, int]:
    min_days = config.PRESORTER_MIN_DAYS_TO_EVENT
    excluded_bond_type = _normalize_bond_type(config.PRESORTER_EXCLUDED_BOND_TYPE)
    use_dohod_nearest_date = bool(getattr(config, "PRESORTER_USE_DOHOD_NEAREST_DATE", True))
    today = datetime.now().date()

    rows_before = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
    rows = conn.execute(
        f'SELECT "ISIN", "MATDATE", "Ближайшая дата погашения/оферты (Дата)", "BOND_TYPE", "Corpbonds_Дата ближайшей оферты", "Smartlab_Дата оферты", "{AMORTIZATION_START_COLUMN}" FROM "{table_name}"'
    ).fetchall()

    matdate_rule_isins: set[str] = set()
    dohod_nearest_rule_isins: set[str] = set()
    bond_type_rule_isins: set[str] = set()
    offerdate_rule_isins: set[str] = set()
    amortstartdate_rule_isins: set[str] = set()

    for isin, matdate, nearest_date, bond_type, corpbonds_offer, smartlab_offer, amort_start in rows:
        isin_value = str(isin or "").strip()
        if not isin_value:
            continue

        mat_dt = _parse_bond_date(matdate)
        if mat_dt is not None and (mat_dt.date() - today).days < min_days:
            matdate_rule_isins.add(isin_value)

        if use_dohod_nearest_date:
            nearest_dt = _parse_bond_date(nearest_date)
            if nearest_dt is not None and (nearest_dt.date() - today).days < min_days:
                dohod_nearest_rule_isins.add(isin_value)

        if excluded_bond_type and _normalize_bond_type(bond_type) == excluded_bond_type:
            bond_type_rule_isins.add(isin_value)

        offer_dt = _parse_bond_date(_pick_offer_date(corpbonds_offer, smartlab_offer))
        if offer_dt is not None and (offer_dt.date() - today).days < min_days:
            offerdate_rule_isins.add(isin_value)

        amort_dt = _parse_bond_date(amort_start)
        if amort_dt is not None and (amort_dt.date() - today).days < min_days:
            amortstartdate_rule_isins.add(isin_value)

    isins_to_delete = (
        matdate_rule_isins
        | dohod_nearest_rule_isins
        | bond_type_rule_isins
        | offerdate_rule_isins
        | amortstartdate_rule_isins
    )
    if isins_to_delete:
        placeholders = ", ".join(["?"] * len(isins_to_delete))
        conn.execute("BEGIN")
        conn.execute(f'DELETE FROM "{table_name}" WHERE "ISIN" IN ({placeholders})', tuple(isins_to_delete))
        conn.commit()

    rows_after = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
    return {
        "rows_before": rows_before,
        "rows_after": rows_after,
        "excluded_by_matdate_rule": len(matdate_rule_isins),
        "excluded_by_dohod_nearest_rule": len(dohod_nearest_rule_isins),
        "excluded_by_bond_type_rule": len(bond_type_rule_isins),
        "excluded_by_offerdate_rule": len(offerdate_rule_isins),
        "excluded_by_amortstartdate_rule": len(amortstartdate_rule_isins),
        "excluded_total": len(isins_to_delete),
    }


def ensure_corpbonds_table(conn: sqlite3.Connection) -> bool:
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.CORPBONDS_TABLE_NAME}" (
            "SECID" TEXT PRIMARY KEY,
            "Цена последняя" TEXT,
            "Тип купона" TEXT,
            "Ставка купона" TEXT,
            "НКД" TEXT,
            "Формула купона" TEXT,
            "Дата ближайшего купона" TEXT,
            "Дата ближайшей оферты" TEXT,
            "Наличие амортизации" TEXT,
            "Купон лесенкой" TEXT,
            "source_url" TEXT,
            "updated_at_utc" TEXT NOT NULL
        )
        '''
    )
    existing_columns = {
        str(row[1]) for row in conn.execute(f'PRAGMA table_info("{config.CORPBONDS_TABLE_NAME}")').fetchall()
    }
    if "Ставка купона" not in existing_columns:
        conn.execute(f'ALTER TABLE "{config.CORPBONDS_TABLE_NAME}" ADD COLUMN "Ставка купона" TEXT')
    if "НКД" not in existing_columns:
        conn.execute(f'ALTER TABLE "{config.CORPBONDS_TABLE_NAME}" ADD COLUMN "НКД" TEXT')

    force_refresh_done = False
    force_refresh_meta_key = "corpbonds_schema_v3_applied"
    if get_meta_value(conn, force_refresh_meta_key) != "1":
        conn.execute(f'UPDATE "{config.CORPBONDS_TABLE_NAME}" SET "updated_at_utc" = ? ', ("",))
        set_meta_value(conn, force_refresh_meta_key, "1")
        force_refresh_done = True
    if getattr(config, "FORCE_CORPBONDS_TTL_RESET", False):
        conn.execute(f'UPDATE "{config.CORPBONDS_TABLE_NAME}" SET "updated_at_utc" = ?', ("",))
        force_refresh_done = True
    conn.commit()
    return force_refresh_done


def collect_merge_secids(conn: sqlite3.Connection) -> list[str]:
    rows = conn.execute(
        f'''
        SELECT DISTINCT TRIM(COALESCE("SECID", ''))
        FROM "{config.MERGE_GREEN_TABLE_NAME}"
        WHERE TRIM(COALESCE("SECID", '')) <> ''
        UNION
        SELECT DISTINCT TRIM(COALESCE("SECID", ''))
        FROM "{config.MERGE_YELLOW_TABLE_NAME}"
        WHERE TRIM(COALESCE("SECID", '')) <> ''
        ORDER BY 1
        '''
    ).fetchall()
    return [str(row[0]).strip() for row in rows if row and str(row[0]).strip()]


def _normalize_label(raw_label: str) -> str:
    cleaned = str(raw_label or "").replace("\xa0", " ").replace("?", " ")
    return " ".join(cleaned.split()).casefold()


def parse_corpbonds_page_fields(raw_html: str) -> dict[str, str]:
    parser = etree.HTMLParser(recover=True)
    root = etree.HTML(raw_html, parser=parser)
    parsed = {
        "Цена последняя": "",
        "Тип купона": "",
        "Ставка купона": "",
        "НКД": "",
        "Формула купона": "",
        "Дата ближайшего купона": "",
        "Дата ближайшей оферты": "",
        "Наличие амортизации": "",
        "Купон лесенкой": "",
    }
    if root is None:
        return parsed

    target_tables = root.xpath(
        "//h1[contains(translate(normalize-space(string(.)), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'доходность')]/following::table[1]"
        " | //*[contains(translate(normalize-space(string(.)), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'купонные выплаты')]/following::table[1]"
        " | //*[contains(translate(normalize-space(string(.)), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'доходность и цена')]/following::table[1]"
    )
    if not target_tables:
        target_tables = root.xpath("//table")

    for row in [r for table in target_tables for r in table.xpath('.//tr[td]')]:
        tds = row.xpath("./td")
        if len(tds) < 2:
            continue
        label = " ".join(" ".join(tds[0].itertext()).split())
        value = " ".join(" ".join(tds[-1].itertext()).split())
        if not label:
            continue
        normalized = _normalize_label(label)
        if normalized.startswith("цена последняя"):
            parsed["Цена последняя"] = value
        elif normalized.startswith("тип купона"):
            parsed["Тип купона"] = value
        elif normalized.startswith("ставка купона"):
            parsed["Ставка купона"] = value
        elif normalized.startswith("накопленный купонный доход (нкд)") or normalized == "нкд":
            parsed["НКД"] = value
        elif normalized.startswith("формула купона") or normalized.startswith("формула флоатера"):
            parsed["Формула купона"] = value
        elif normalized.startswith("дата ближайшего купона"):
            parsed["Дата ближайшего купона"] = value
        elif normalized.startswith("дата ближайшей оферты"):
            parsed["Дата ближайшей оферты"] = value
        elif normalized.startswith("наличие амортизации"):
            parsed["Наличие амортизации"] = value
        elif normalized.startswith("купон лесенкой"):
            parsed["Купон лесенкой"] = value
    return parsed


def _is_corpbonds_stale(updated_at_raw: str | None, now_utc: datetime) -> bool:
    if not updated_at_raw:
        return True
    try:
        updated_at = datetime.fromisoformat(updated_at_raw)
    except ValueError:
        return True
    return now_utc - updated_at >= timedelta(hours=config.CORPBONDS_CACHE_TTL_HOURS)


def get_stale_corpbonds_secids(conn: sqlite3.Connection, secids: list[str], now_utc: datetime) -> list[str]:
    if not secids:
        return []

    stale: list[str] = []
    chunk_size = 500
    for start in range(0, len(secids), chunk_size):
        chunk = secids[start : start + chunk_size]
        placeholders = ", ".join(["?"] * len(chunk))
        rows = conn.execute(
            f'SELECT "SECID", "updated_at_utc" FROM "{config.CORPBONDS_TABLE_NAME}" WHERE "SECID" IN ({placeholders})',
            tuple(chunk),
        ).fetchall()
        actual = {str(row[0]): str(row[1] or "") for row in rows}
        for secid in chunk:
            if _is_corpbonds_stale(actual.get(secid), now_utc):
                stale.append(secid)

    return stale


def fetch_corpbonds_payload(secid: str) -> dict[str, str]:
    if not hasattr(fetch_corpbonds_payload, "_thread_local"):
        fetch_corpbonds_payload._thread_local = threading.local()  # type: ignore[attr-defined]
    thread_local = fetch_corpbonds_payload._thread_local  # type: ignore[attr-defined]
    if not hasattr(thread_local, "session"):
        thread_local.session = create_resilient_http_session(pool_size=max(16, config.CORPBONDS_MAX_WORKERS * 2))

    thread_local.session.headers.update({
        "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.8",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
    })
    url = f"{config.CORPBONDS_BOND_URL_PREFIX}{secid}"
    response = thread_local.session.get(url, timeout=config.CORPBONDS_REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    data = parse_corpbonds_page_fields(response.text)
    data["SECID"] = secid
    data["source_url"] = url
    return data


def refresh_corpbonds_data_if_needed(
    conn: sqlite3.Connection, logger: logging.Logger, now_utc: datetime
) -> tuple[int, int, int]:
    started_at = time.monotonic()
    force_refresh_done = ensure_corpbonds_table(conn)
    if force_refresh_done:
        logger.info("Corpbonds: force refresh cache (schema_v3) выполнен одноразово.")
    secids = collect_merge_secids(conn)
    if not secids:
        return 0, 0, 0

    stale_secids = get_stale_corpbonds_secids(conn, secids, now_utc)
    fetched_rows: list[dict[str, str]] = []
    errors = 0

    if stale_secids:
        with ThreadPoolExecutor(max_workers=config.CORPBONDS_MAX_WORKERS) as executor:
            futures = {executor.submit(fetch_corpbonds_payload, secid): secid for secid in stale_secids}
            with progress(total=len(stale_secids), desc="Corpbonds fetch", unit="бумаг", position=1) as pbar:
                for future in as_completed(futures):
                    secid = futures[future]
                    try:
                        fetched_rows.append(future.result())
                    except Exception as exc:
                        errors += 1
                        logger.warning("Corpbonds: ошибка SECID=%s: %s", secid, exc)
                    pbar.update(1)

    if fetched_rows:
        now_iso = now_utc.isoformat()
        payload = [
            (
                row.get("SECID", ""),
                row.get("Цена последняя", ""),
                row.get("Тип купона", ""),
                row.get("Ставка купона", ""),
                row.get("НКД", ""),
                row.get("Формула купона", ""),
                row.get("Дата ближайшего купона", ""),
                row.get("Дата ближайшей оферты", ""),
                row.get("Наличие амортизации", ""),
                row.get("Купон лесенкой", ""),
                row.get("source_url", ""),
                now_iso,
            )
            for row in fetched_rows
            if row.get("SECID", "")
        ]
        conn.executemany(
            f'''
            INSERT INTO "{config.CORPBONDS_TABLE_NAME}" (
                "SECID", "Цена последняя", "Тип купона", "Ставка купона", "НКД",
                "Формула купона",
                "Дата ближайшего купона", "Дата ближайшей оферты", "Наличие амортизации",
                "Купон лесенкой", "source_url", "updated_at_utc"
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT("SECID") DO UPDATE SET
                "Цена последняя"=excluded."Цена последняя",
                "Тип купона"=excluded."Тип купона",
                "Ставка купона"=excluded."Ставка купона",
                "НКД"=excluded."НКД",
                "Формула купона"=excluded."Формула купона",
                "Дата ближайшего купона"=excluded."Дата ближайшего купона",
                "Дата ближайшей оферты"=excluded."Дата ближайшей оферты",
                "Наличие амортизации"=excluded."Наличие амортизации",
                "Купон лесенкой"=excluded."Купон лесенкой",
                "source_url"=excluded."source_url",
                "updated_at_utc"=excluded."updated_at_utc"
            ''',
            payload,
        )
        conn.commit()

    duration_seconds = time.monotonic() - started_at
    logger.info(
        "Corpbonds: SECID в Merge=%s, запрошено=%s, сохранено=%s, ошибок=%s, TTL=%s ч, duration=%.2fs",
        len(secids),
        len(stale_secids),
        len(fetched_rows),
        errors,
        config.CORPBONDS_CACHE_TTL_HOURS,
        duration_seconds,
    )
    return len(secids), len(stale_secids), len(fetched_rows)


def ensure_smartlab_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.SMARTLAB_TABLE_NAME}" (
            "SECID" TEXT PRIMARY KEY,
            "Котировка облигации, %" TEXT,
            "Изм за день, %" TEXT,
            "Объем день, млн. руб" TEXT,
            "Объем день, штук" TEXT,
            "Дата оферты" TEXT,
            "Только для квалов?" TEXT,
            "Длительность купона, дней" TEXT,
            "source_url" TEXT,
            "updated_at_utc" TEXT NOT NULL
        )
        '''
    )
    conn.commit()


def parse_smartlab_page_fields(raw_html: str) -> dict[str, str]:
    parsed = {
        "Котировка облигации, %": "",
        "Изм за день, %": "",
        "Объем день, млн. руб": "",
        "Объем день, штук": "",
        "Дата оферты": "",
        "Только для квалов?": "",
        "Длительность купона, дней": "",
    }

    try:
        root = etree.HTML(raw_html)
    except Exception:
        return parsed
    if root is None:
        return parsed

    for row in root.xpath('//div[contains(@class, "quotes-simple-table__row")]'):
        items = row.xpath('./div[contains(@class, "quotes-simple-table__item")]')
        if len(items) < 2:
            continue
        label = " ".join(part.strip() for part in items[0].itertext() if str(part).strip())
        value = " ".join(part.strip() for part in items[1].itertext() if str(part).strip())
        normalized = _normalize_label(label)
        if normalized.startswith("котировка облигации"):
            parsed["Котировка облигации, %"] = value
        elif normalized.startswith("изм за день"):
            parsed["Изм за день, %"] = value
        elif normalized.startswith("объем день, млн. руб"):
            parsed["Объем день, млн. руб"] = value
        elif normalized.startswith("объем день, штук"):
            parsed["Объем день, штук"] = value
        elif normalized.startswith("дата оферты"):
            parsed["Дата оферты"] = value
        elif normalized.startswith("только для квалов"):
            parsed["Только для квалов?"] = value
        elif normalized.startswith("длительность купона"):
            parsed["Длительность купона, дней"] = value

    return parsed


def _is_smartlab_stale(updated_at_raw: str | None, now_utc: datetime) -> bool:
    if not updated_at_raw:
        return True
    try:
        updated_at = datetime.fromisoformat(updated_at_raw)
    except ValueError:
        return True
    return now_utc - updated_at >= timedelta(hours=config.SMARTLAB_CACHE_TTL_HOURS)


def get_stale_smartlab_secids(conn: sqlite3.Connection, secids: list[str], now_utc: datetime) -> list[str]:
    if not secids:
        return []

    threshold_iso = (now_utc - timedelta(hours=config.SMARTLAB_CACHE_TTL_HOURS)).isoformat()
    conn.execute('DROP TABLE IF EXISTS "tmp_smartlab_secids"')
    conn.execute('CREATE TEMP TABLE "tmp_smartlab_secids" ("secid" TEXT PRIMARY KEY)')
    conn.executemany(
        'INSERT OR IGNORE INTO "tmp_smartlab_secids" ("secid") VALUES (?)',
        [(secid,) for secid in secids if secid],
    )
    rows = conn.execute(
        f"""
        SELECT t."secid"
        FROM "tmp_smartlab_secids" t
        LEFT JOIN "{config.SMARTLAB_TABLE_NAME}" s
            ON s."SECID" = t."secid"
        WHERE s."SECID" IS NULL
           OR TRIM(COALESCE(s."updated_at_utc", '')) = ''
           OR s."updated_at_utc" < ?
        """,
        (threshold_iso,),
    ).fetchall()
    return [str(row[0]) for row in rows if row and row[0]]


def fetch_smartlab_payload(secid: str) -> dict[str, str]:
    if not hasattr(fetch_smartlab_payload, "_thread_local"):
        fetch_smartlab_payload._thread_local = threading.local()  # type: ignore[attr-defined]
    thread_local = fetch_smartlab_payload._thread_local  # type: ignore[attr-defined]
    if not hasattr(thread_local, "session"):
        thread_local.session = create_resilient_http_session(pool_size=max(16, config.SMARTLAB_MAX_WORKERS))

    url = f"{config.SMARTLAB_BOND_URL_PREFIX}{secid}/"
    response = thread_local.session.get(url, timeout=config.SMARTLAB_REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    data = parse_smartlab_page_fields(response.text)
    data["SECID"] = secid
    data["source_url"] = url
    return data


def refresh_smartlab_data_if_needed(
    conn: sqlite3.Connection, logger: logging.Logger, now_utc: datetime
) -> tuple[int, int, int]:
    ensure_smartlab_table(conn)
    secids = collect_merge_secids(conn)
    if not secids:
        return 0, 0, 0

    stale_secids = get_stale_smartlab_secids(conn, secids, now_utc)
    fetched_rows: list[dict[str, str]] = []
    errors = 0

    if stale_secids:
        with ThreadPoolExecutor(max_workers=config.SMARTLAB_MAX_WORKERS) as executor:
            futures = {executor.submit(fetch_smartlab_payload, secid): secid for secid in stale_secids}
            with progress(total=len(stale_secids), desc="Smartlab fetch", unit="бумаг", position=1) as pbar:
                for future in as_completed(futures):
                    secid = futures[future]
                    try:
                        fetched_rows.append(future.result())
                    except Exception as exc:
                        errors += 1
                        logger.warning("Smartlab: ошибка SECID=%s: %s", secid, exc)
                    pbar.update(1)

    if fetched_rows:
        now_iso = now_utc.isoformat()
        payload = [
            (
                row.get("SECID", ""),
                row.get("Котировка облигации, %", ""),
                row.get("Изм за день, %", ""),
                row.get("Объем день, млн. руб", ""),
                row.get("Объем день, штук", ""),
                row.get("Дата оферты", ""),
                row.get("Только для квалов?", ""),
                row.get("Длительность купона, дней", ""),
                row.get("source_url", ""),
                now_iso,
            )
            for row in fetched_rows
            if row.get("SECID", "")
        ]
        conn.executemany(
            f'''
            INSERT INTO "{config.SMARTLAB_TABLE_NAME}" (
                "SECID", "Котировка облигации, %", "Изм за день, %", "Объем день, млн. руб",
                "Объем день, штук", "Дата оферты", "Только для квалов?", "Длительность купона, дней",
                "source_url", "updated_at_utc"
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT("SECID") DO UPDATE SET
                "Котировка облигации, %"=excluded."Котировка облигации, %",
                "Изм за день, %"=excluded."Изм за день, %",
                "Объем день, млн. руб"=excluded."Объем день, млн. руб",
                "Объем день, штук"=excluded."Объем день, штук",
                "Дата оферты"=excluded."Дата оферты",
                "Только для квалов?"=excluded."Только для квалов?",
                "Длительность купона, дней"=excluded."Длительность купона, дней",
                "source_url"=excluded."source_url",
                "updated_at_utc"=excluded."updated_at_utc"
            ''',
            payload,
        )
        conn.commit()

    logger.info(
        "Smartlab: SECID в Merge=%s, запрошено=%s, сохранено=%s, ошибок=%s, TTL=%s ч",
        len(secids),
        len(stale_secids),
        len(fetched_rows),
        errors,
        config.SMARTLAB_CACHE_TTL_HOURS,
    )
    return len(secids), len(stale_secids), len(fetched_rows)


def ensure_moex_amortizations_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        f'''
        CREATE TABLE IF NOT EXISTS "{config.MOEX_AMORTIZATION_TABLE_NAME}" (
            "secid" TEXT NOT NULL,
            "isin" TEXT,
            "name" TEXT,
            "issuevalue" TEXT,
            "amortdate" TEXT,
            "facevalue" TEXT,
            "initialfacevalue" TEXT,
            "faceunit" TEXT,
            "valueprc" TEXT,
            "value" TEXT,
            "value_rub" TEXT,
            "data_source" TEXT,
            "primary_boardid" TEXT,
            "source_url" TEXT NOT NULL,
            "updated_at_utc" TEXT NOT NULL,
            PRIMARY KEY ("secid", "amortdate")
        )
        '''
    )
    conn.commit()


def collect_amortization_secids(conn: sqlite3.Connection) -> list[str]:
    # Важно: в стандартной SQLite функция LOWER/UPPER без ICU корректно работает
    # в основном для ASCII. Для кириллицы (например "Да") фильтрация через SQL
    # может не сработать. Поэтому нормализуем значение в Python.
    expected = str(getattr(config, "MOEX_AMORTIZATION_REQUIRED_FLAG", "Да"))

    def normalize_flag(value: str | None) -> str:
        raw = str(value or "").replace("\xa0", " ")
        return " ".join(raw.split()).casefold()

    expected_normalized = normalize_flag(expected)
    rows = conn.execute(
        f'''
        SELECT TRIM(COALESCE("SECID", '')) AS secid,
               COALESCE("Corpbonds_Наличие амортизации", '') AS has_amort
        FROM "{config.MERGE_GREEN_TABLE_NAME}"
        WHERE TRIM(COALESCE("SECID", '')) <> ''
        UNION ALL
        SELECT TRIM(COALESCE("SECID", '')) AS secid,
               COALESCE("Corpbonds_Наличие амортизации", '') AS has_amort
        FROM "{config.MERGE_YELLOW_TABLE_NAME}"
        WHERE TRIM(COALESCE("SECID", '')) <> ''
        '''
    ).fetchall()

    secids: set[str] = set()
    for secid, has_amort in rows:
        secid_value = str(secid or "").strip()
        if not secid_value:
            continue
        if normalize_flag(str(has_amort or "")) == expected_normalized:
            secids.add(secid_value)

    return sorted(secids)


def get_stale_moex_amortization_secids(conn: sqlite3.Connection, secids: list[str], now_utc: datetime) -> list[str]:
    if not secids:
        return []

    ttl = timedelta(hours=config.MOEX_AMORTIZATION_CACHE_TTL_HOURS)
    stale: list[str] = []
    chunk_size = 500
    for start in range(0, len(secids), chunk_size):
        chunk = secids[start : start + chunk_size]
        placeholders = ", ".join(["?"] * len(chunk))
        rows = conn.execute(
            f'''
            SELECT "secid", MAX("updated_at_utc")
            FROM "{config.MOEX_AMORTIZATION_TABLE_NAME}"
            WHERE "secid" IN ({placeholders})
            GROUP BY "secid"
            ''',
            tuple(chunk),
        ).fetchall()
        actual = {}
        for secid, updated_at_raw in rows:
            try:
                actual[str(secid)] = datetime.fromisoformat(str(updated_at_raw))
            except ValueError:
                continue

        for secid in chunk:
            updated_at = actual.get(secid)
            if not updated_at or now_utc - updated_at >= ttl:
                stale.append(secid)

    return stale


def fetch_moex_amortization_payload(secid: str) -> list[dict[str, str]]:
    if not hasattr(fetch_moex_amortization_payload, "_thread_local"):
        fetch_moex_amortization_payload._thread_local = threading.local()  # type: ignore[attr-defined]
    thread_local = fetch_moex_amortization_payload._thread_local  # type: ignore[attr-defined]
    if not hasattr(thread_local, "session"):
        thread_local.session = create_resilient_http_session(pool_size=max(16, config.MOEX_AMORTIZATION_MAX_WORKERS))

    url = config.MOEX_AMORTIZATION_URL_TEMPLATE.format(secid=secid)
    response = thread_local.session.get(url, timeout=config.MOEX_AMORTIZATION_REQUEST_TIMEOUT_SECONDS)
    response.raise_for_status()
    payload = response.json()
    section = payload.get("amortizations") or {}
    columns = section.get("columns") or []
    data_rows = section.get("data") or []

    rows: list[dict[str, str]] = []
    for item in data_rows:
        normalized = {
            str(col): str(item[idx]) if idx < len(item) and item[idx] is not None else ""
            for idx, col in enumerate(columns)
        }
        normalized["secid"] = normalized.get("secid") or secid
        normalized["source_url"] = url
        rows.append(normalized)
    return rows


def refresh_moex_amortizations_if_needed(
    conn: sqlite3.Connection, logger: logging.Logger, now_utc: datetime
) -> tuple[int, int, int]:
    ensure_moex_amortizations_table(conn)
    secids = collect_amortization_secids(conn)
    if not secids:
        return 0, 0, 0

    stale_secids = get_stale_moex_amortization_secids(conn, secids, now_utc)
    fetched_rows: list[dict[str, str]] = []
    errors = 0

    if stale_secids:
        with ThreadPoolExecutor(max_workers=config.MOEX_AMORTIZATION_MAX_WORKERS) as executor:
            futures = {executor.submit(fetch_moex_amortization_payload, secid): secid for secid in stale_secids}
            with progress(total=len(stale_secids), desc="MOEX amort fetch", unit="бумаг", position=1) as pbar:
                for future in as_completed(futures):
                    secid = futures[future]
                    try:
                        fetched_rows.extend(future.result())
                    except Exception as exc:
                        errors += 1
                        logger.warning("MOEX amortizations: ошибка SECID=%s: %s", secid, exc)
                    pbar.update(1)

    if stale_secids:
        conn.execute("BEGIN")
        placeholders = ", ".join(["?"] * len(stale_secids))
        conn.execute(
            f'DELETE FROM "{config.MOEX_AMORTIZATION_TABLE_NAME}" WHERE "secid" IN ({placeholders})',
            tuple(stale_secids),
        )
        if fetched_rows:
            now_iso = now_utc.isoformat()
            payload = [
                (
                    row.get("secid", ""),
                    row.get("isin", ""),
                    row.get("name", ""),
                    row.get("issuevalue", ""),
                    row.get("amortdate", ""),
                    row.get("facevalue", ""),
                    row.get("initialfacevalue", ""),
                    row.get("faceunit", ""),
                    row.get("valueprc", ""),
                    row.get("value", ""),
                    row.get("value_rub", ""),
                    row.get("data_source", ""),
                    row.get("primary_boardid", ""),
                    row.get("source_url", ""),
                    now_iso,
                )
                for row in fetched_rows
                if row.get("secid", "")
            ]
            conn.executemany(
                f'''
                INSERT OR REPLACE INTO "{config.MOEX_AMORTIZATION_TABLE_NAME}" (
                    "secid", "isin", "name", "issuevalue", "amortdate", "facevalue",
                    "initialfacevalue", "faceunit", "valueprc", "value", "value_rub",
                    "data_source", "primary_boardid", "source_url", "updated_at_utc"
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                payload,
            )
        conn.commit()

    logger.info(
        "MOEX amortizations: SECID в Merge*=%s, к запросу по TTL=%s, строк сохранено=%s, ошибок=%s, TTL=%s ч",
        len(secids),
        len(stale_secids),
        len(fetched_rows),
        errors,
        config.MOEX_AMORTIZATION_CACHE_TTL_HOURS,
    )
    return len(secids), len(stale_secids), len(fetched_rows)


def apply_amortization_start_date_to_merge_table(conn: sqlite3.Connection, table_name: str) -> int:
    ensure_merge_table(conn, table_name)
    rows = conn.execute(
        f'''
        SELECT m."ISIN", MIN(a."amortdate")
        FROM "{table_name}" m
        LEFT JOIN "{config.MOEX_AMORTIZATION_TABLE_NAME}" a
            ON TRIM(COALESCE(m."SECID", '')) = TRIM(COALESCE(a."secid", ''))
        GROUP BY m."ISIN"
        '''
    ).fetchall()

    conn.execute("BEGIN")
    conn.executemany(
        f'UPDATE "{table_name}" SET "{AMORTIZATION_START_COLUMN}" = ? WHERE "ISIN" = ?',
        [(str(amortdate or ""), str(isin)) for isin, amortdate in rows if str(isin or "").strip()],
    )
    conn.commit()
    return sum(1 for _, amortdate in rows if str(amortdate or "").strip())


def export_moex_amortization_snapshot(conn: sqlite3.Connection) -> int:
    sampled_secids = conn.execute(
        f'''
        SELECT DISTINCT "secid"
        FROM "{config.MOEX_AMORTIZATION_TABLE_NAME}"
        WHERE TRIM(COALESCE("secid", '')) <> ''
        ORDER BY RANDOM()
        LIMIT 5
        '''
    ).fetchall()
    secids = [str(row[0]).strip() for row in sampled_secids if row and str(row[0]).strip()]

    rows: list[tuple] = []
    headers = [
        "secid",
        "isin",
        "name",
        "issuevalue",
        "amortdate",
        "facevalue",
        "initialfacevalue",
        "faceunit",
        "valueprc",
        "value",
        "value_rub",
        "data_source",
        "primary_boardid",
        "source_url",
        "updated_at_utc",
    ]
    if secids:
        placeholders = ", ".join(["?"] * len(secids))
        cursor = conn.execute(
            f'''
            SELECT "secid", "isin", "name", "issuevalue", "amortdate", "facevalue", "initialfacevalue",
                   "faceunit", "valueprc", "value", "value_rub", "data_source", "primary_boardid", "source_url", "updated_at_utc"
            FROM "{config.MOEX_AMORTIZATION_TABLE_NAME}"
            WHERE "secid" IN ({placeholders})
            ORDER BY "secid", "amortdate"
            ''',
            tuple(secids),
        )
        rows = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "moex_amortizations_snapshot"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / config.MOEX_AMORTIZATION_SNAPSHOT_FILENAME
    wb.save(snapshot_path)
    return len(rows)


def _get_table_columns(conn: sqlite3.Connection, table_name: str) -> list[str]:
    return [str(row[1]) for row in conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()]


def apply_corpbonds_inner_join_to_merge_table(conn: sqlite3.Connection, table_name: str) -> int:
    ensure_merge_table(conn, table_name)
    merge_columns = _get_table_columns(conn, table_name)
    require_match = bool(getattr(config, "MERGE_REQUIRE_CORPBONDS_SECID_MATCH", True))
    select_columns: list[str] = []
    for column in merge_columns:
        if column in CORPBONDS_COLUMNS_MAP:
            source_column = CORPBONDS_COLUMNS_MAP[column]
            select_columns.append(f'c."{source_column}" AS "{column}"')
        else:
            select_columns.append(f'm."{column}"')

    join_type = "INNER" if require_match else "LEFT"
    secid_condition = "WHERE TRIM(COALESCE(m.\"SECID\", '')) <> ''" if require_match else ""

    conn.execute("BEGIN")
    conn.execute(f'DROP TABLE IF EXISTS "{table_name}__tmp_corpbonds"')
    conn.execute(
        f'''
        CREATE TABLE "{table_name}__tmp_corpbonds" AS
        SELECT {", ".join(select_columns)}
        FROM "{table_name}" m
        {join_type} JOIN "{config.CORPBONDS_TABLE_NAME}" c
            ON TRIM(COALESCE(m."SECID", '')) = TRIM(COALESCE(c."SECID", ''))
        {secid_condition}
        '''
    )
    conn.execute(f'DELETE FROM "{table_name}"')
    conn.execute(
        f'''
        INSERT OR REPLACE INTO "{table_name}" ({", ".join(f'"{column}"' for column in merge_columns)})
        SELECT {", ".join(f'"{column}"' for column in merge_columns)}
        FROM "{table_name}__tmp_corpbonds"
        '''
    )
    conn.execute(f'DROP TABLE "{table_name}__tmp_corpbonds"')
    rows_after = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
    conn.commit()
    return int(rows_after)


def apply_smartlab_inner_join_to_merge_table(conn: sqlite3.Connection, table_name: str) -> int:
    ensure_merge_table(conn, table_name)
    merge_columns = _get_table_columns(conn, table_name)
    require_match = bool(getattr(config, "MERGE_REQUIRE_SMARTLAB_SECID_MATCH", True))
    select_columns: list[str] = []
    for column in merge_columns:
        if column in SMARTLAB_COLUMNS_MAP:
            source_column = SMARTLAB_COLUMNS_MAP[column]
            select_columns.append(f's."{source_column}" AS "{column}"')
        else:
            select_columns.append(f'm."{column}"')

    join_type = "INNER" if require_match else "LEFT"
    secid_condition = "WHERE TRIM(COALESCE(m.\"SECID\", '')) <> ''" if require_match else ""

    conn.execute("BEGIN")
    conn.execute(f'DROP TABLE IF EXISTS "{table_name}__tmp_smartlab"')
    conn.execute(
        f'''
        CREATE TABLE "{table_name}__tmp_smartlab" AS
        SELECT {", ".join(select_columns)}
        FROM "{table_name}" m
        {join_type} JOIN "{config.SMARTLAB_TABLE_NAME}" s
            ON TRIM(COALESCE(m."SECID", '')) = TRIM(COALESCE(s."SECID", ''))
        {secid_condition}
        '''
    )
    conn.execute(f'DELETE FROM "{table_name}"')
    conn.execute(
        f'''
        INSERT OR REPLACE INTO "{table_name}" ({", ".join(f'"{column}"' for column in merge_columns)})
        SELECT {", ".join(f'"{column}"' for column in merge_columns)}
        FROM "{table_name}__tmp_smartlab"
        '''
    )
    conn.execute(f'DROP TABLE "{table_name}__tmp_smartlab"')
    rows_after = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
    conn.commit()
    return int(rows_after)


def export_corpbonds_snapshot(conn: sqlite3.Connection) -> int:
    cursor = conn.execute(
        f'''
        SELECT *
        FROM "{config.CORPBONDS_TABLE_NAME}"
        WHERE rowid IN (
            SELECT MIN(rowid)
            FROM "{config.CORPBONDS_TABLE_NAME}"
            GROUP BY "SECID"
            ORDER BY RANDOM()
            LIMIT 5
        )
        '''
    )
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = "corpbonds_snapshot"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / config.CORPBONDS_SNAPSHOT_FILENAME
    wb.save(snapshot_path)
    return len(rows)


def export_smartlab_snapshot(conn: sqlite3.Connection) -> int:
    cursor = conn.execute(
        f'''
        SELECT *
        FROM "{config.SMARTLAB_TABLE_NAME}"
        WHERE rowid IN (
            SELECT MIN(rowid)
            FROM "{config.SMARTLAB_TABLE_NAME}"
            GROUP BY "SECID"
            ORDER BY RANDOM()
            LIMIT 5
        )
        '''
    )
    rows = cursor.fetchall()
    headers = [description[0] for description in cursor.description]

    wb = Workbook()
    ws = wb.active
    ws.title = "smartlab_snapshot"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))

    snapshot_path = config.BASE_SNAPSHOTS_DIR / config.SMARTLAB_SNAPSHOT_FILENAME
    wb.save(snapshot_path)
    return len(rows)


def ensure_screener_table(conn: sqlite3.Connection) -> None:
    columns_sql = ['"ISIN" TEXT PRIMARY KEY']
    columns_sql.extend(f'"{column}" TEXT' for column in SCREENER_COLUMNS if column != "ISIN")
    conn.execute(f'CREATE TABLE IF NOT EXISTS "{SCREENER_TABLE_NAME}" ({", ".join(columns_sql)})')
    ensure_table_columns(conn, SCREENER_TABLE_NAME, [column for column in SCREENER_COLUMNS if column != "ISIN"])
    conn.commit()


def _normalize_isin(raw_value: object) -> str | None:
    cleaned = re.sub(r"[^A-Z0-9]", "", str(raw_value or "").strip().upper().replace(" ", ""))
    return cleaned or None


def _normalize_override_header(raw_value: object) -> str:
    return str(raw_value or "").strip().replace("﻿", "").casefold()


def _safe_save_workbook_atomic(wb: Workbook, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        backup_path = destination.with_name(f"{destination.stem}.backup{destination.suffix}")
        if not backup_path.exists():
            backup_path.write_bytes(destination.read_bytes())

    with tempfile.NamedTemporaryFile(
        prefix=f"{destination.stem}.",
        suffix=destination.suffix,
        dir=destination.parent,
        delete=False,
    ) as tmp_file:
        temp_path = Path(tmp_file.name)

    try:
        wb.save(temp_path)
        os.replace(temp_path, destination)
    finally:
        if temp_path.exists():
            temp_path.unlink(missing_ok=True)


def _extract_sheet_headers(ws: Worksheet) -> list[str]:
    headers: list[str] = []
    for cell in ws[1]:
        value = str(cell.value or "").strip()
        if not value and len(headers) > 0:
            break
        headers.append(value)
    return headers


def _extract_sheet_rows_by_headers(ws: Worksheet, headers: list[str]) -> list[dict[str, str]]:
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
    header_map = {
        _normalize_override_header(value): idx
        for idx, value in enumerate(header_row)
        if _normalize_override_header(value)
    }
    rows: list[dict[str, str]] = []
    max_col = max(header_map.values()) + 1 if header_map else len(headers)
    for raw_row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
        row_dict: dict[str, str] = {}
        for header in headers:
            idx = header_map.get(_normalize_override_header(header))
            value = raw_row[idx] if idx is not None and idx < len(raw_row) else ""
            row_dict[header] = str(value or "").strip()
        if any(str(v).strip() for v in row_dict.values()):
            rows.append(row_dict)
    return rows


def _apply_default_bond_overrides_sheet_ux(ws: Worksheet) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:G1"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 54
    ws.column_dimensions["G"].width = 22
    bool_validation = DataValidation(type="list", formula1='"✅,❌"', allow_blank=True)
    ws.add_data_validation(bool_validation)
    for col in ("B", "C", "D", "E"):
        bool_validation.add(f"{col}2:{col}5000")


def ensure_bond_overrides_excel(logger: logging.Logger | None = None) -> Path:
    overrides_path = config.BASE_DIR / getattr(config, "BOND_OVERRIDES_XLSX_FILENAME", "BondOverrides.xlsx")
    backup_path = overrides_path.with_name(f"{overrides_path.stem}.backup{overrides_path.suffix}")
    sheet_name = BOND_OVERRIDES_SHEET_NAME
    required_headers = ["ISIN", "Enabled", "Drop", "Квал", "Суборд", "CouponFormulaOverride", "Тип купона"]
    logger = logger or logging.getLogger("bonds_main")
    logger.info(
        "BondOverrides ensure: file_exists=%s backup_exists=%s path=%s backup_path=%s",
        overrides_path.exists(),
        backup_path.exists(),
        overrides_path,
        backup_path,
    )

    if not overrides_path.exists() and not backup_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(BOND_OVERRIDES_HEADERS)
        _apply_default_bond_overrides_sheet_ux(ws)

        _safe_save_workbook_atomic(wb, overrides_path)
        wb.close()

        verify_wb = load_workbook(overrides_path, read_only=True, data_only=True)
        try:
            verify_ws = verify_wb[sheet_name] if sheet_name in verify_wb.sheetnames else verify_wb.active
            headers_after_reopen = _extract_sheet_headers(verify_ws)
            if "Тип купона" not in headers_after_reopen:
                raise RuntimeError(
                    "BondOverrides verification failed after create: missing expected header 'Тип купона'. "
                    f"file={overrides_path}; sheet={sheet_name}; headers={headers_after_reopen}"
                )
            logger.info(
                "BondOverrides ensure(create): headers_after_reopen=%s rows=%s",
                headers_after_reopen,
                verify_ws.max_row,
            )
        finally:
            verify_wb.close()
        return overrides_path

    wb = load_workbook(overrides_path) if overrides_path.exists() else Workbook()
    backup_wb = load_workbook(backup_path, data_only=True) if backup_path.exists() else None
    try:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            ws.title = sheet_name

        backup_ws = None
        if backup_wb is not None:
            backup_ws = backup_wb[sheet_name] if sheet_name in backup_wb.sheetnames else backup_wb.active

        headers_backup = _extract_sheet_headers(backup_ws) if backup_ws is not None else []
        headers_before = _extract_sheet_headers(ws)
        rows_before = ws.max_row
        logger.info(
            "BondOverrides ensure: headers_backup=%s headers_current_before_repair=%s rows_before=%s",
            headers_backup,
            headers_before,
            rows_before,
        )

        current_rows = _extract_sheet_rows_by_headers(ws, required_headers)
        backup_rows = _extract_sheet_rows_by_headers(backup_ws, required_headers) if backup_ws is not None else []

        current_by_isin: dict[str, dict[str, str]] = {}
        current_order: list[str] = []
        for row in current_rows:
            isin = _normalize_isin(row.get("ISIN"))
            if not isin:
                continue
            current_by_isin[isin] = row
            if isin not in current_order:
                current_order.append(isin)

        appended_from_backup = 0
        for row in backup_rows:
            isin = _normalize_isin(row.get("ISIN"))
            if not isin or isin in current_by_isin:
                continue
            current_by_isin[isin] = row
            current_order.append(isin)
            appended_from_backup += 1

        was_coupon_type_present = "Тип купона" in headers_before
        removed_extra_columns = len([
            header for header in headers_before if header and header not in required_headers
        ])

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        if ws.max_column > len(required_headers):
            ws.delete_cols(len(required_headers) + 1, ws.max_column - len(required_headers))
        elif ws.max_column < len(required_headers):
            ws.insert_cols(ws.max_column + 1, len(required_headers) - ws.max_column)

        for col_idx, header in enumerate(required_headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        for isin in current_order:
            row_data = current_by_isin.get(isin, {})
            ws.append([row_data.get(header, "") for header in required_headers])

        _apply_default_bond_overrides_sheet_ux(ws)
        _safe_save_workbook_atomic(wb, overrides_path)

        logger.info(
            "BondOverrides ensure: coupon_type_added=%s removed_extra_columns=%s rows_transferred=%s appended_from_backup=%s",
            not was_coupon_type_present,
            removed_extra_columns,
            len(current_order),
            appended_from_backup,
        )
    finally:
        wb.close()
        if backup_wb is not None:
            backup_wb.close()

    verify_wb = load_workbook(overrides_path, read_only=True, data_only=True)
    try:
        verify_ws = verify_wb[sheet_name] if sheet_name in verify_wb.sheetnames else verify_wb.active
        headers_after_reopen = _extract_sheet_headers(verify_ws)
        rows_after = verify_ws.max_row
        logger.info("BondOverrides ensure: headers_after_reopen=%s rows_after=%s", headers_after_reopen, rows_after)
        if headers_after_reopen != required_headers:
            raise RuntimeError(
                "BondOverrides verification failed after save+reopen: unexpected headers. "
                f"file={overrides_path}; sheet={verify_ws.title}; headers={headers_after_reopen}"
            )
        if rows_after < rows_before:
            raise RuntimeError(
                "BondOverrides verification failed after save+reopen: row count decreased. "
                f"file={overrides_path}; sheet={verify_ws.title}; rows_before={rows_before}; rows_after={rows_after}"
            )

        verify_isins = set()
        for raw_row in verify_ws.iter_rows(min_row=2, values_only=True):
            isin = _normalize_isin(raw_row[0] if raw_row else None)
            if isin:
                verify_isins.add(isin)
        if len(verify_isins) < len(current_by_isin):
            raise RuntimeError(
                "BondOverrides verification failed after save+reopen: ISIN rows were lost. "
                f"file={overrides_path}; sheet={verify_ws.title}; expected_isin_count={len(current_by_isin)}; actual_isin_count={len(verify_isins)}"
            )
    finally:
        verify_wb.close()

    return overrides_path


def _parse_override_bool(raw_value: object) -> bool | None:
    if raw_value is None:
        return None
    value = str(raw_value).strip().casefold()
    if not value:
        return None
    if value in {"✅", "1", "true", "да", "yes", "y"}:
        return True
    if value in {"❌", "0", "false", "нет", "no", "n"}:
        return False
    return None


def load_bond_overrides(logger: logging.Logger) -> tuple[dict[str, dict[str, object]], int, int]:
    overrides_path = ensure_bond_overrides_excel(logger)
    wb = load_workbook(overrides_path, read_only=True, data_only=True)
    try:
        ws = wb[BOND_OVERRIDES_SHEET_NAME] if BOND_OVERRIDES_SHEET_NAME in wb.sheetnames else wb.active
        overrides: dict[str, dict[str, object]] = {}
        total_rows = 0
        enabled_rows = 0

        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        header_map = {
            _normalize_override_header(value): idx
            for idx, value in enumerate(header_row)
            if _normalize_override_header(value)
        }

        isin_idx = header_map.get(_normalize_override_header("ISIN"), 0)
        enabled_idx = header_map.get(_normalize_override_header("Enabled"), 1)
        drop_idx = header_map.get(_normalize_override_header("Drop"), 2)
        kval_idx = header_map.get(_normalize_override_header("Квал"), 3)
        subord_idx = header_map.get(_normalize_override_header("Суборд"), 4)
        coupon_formula_idx = header_map.get(_normalize_override_header("CouponFormulaOverride"), 5)
        coupon_type_idx = header_map.get(_normalize_override_header("Тип купона"))
        coupon_type_read_count = 0

        max_col = max(
            idx
            for idx in [isin_idx, enabled_idx, drop_idx, kval_idx, subord_idx, coupon_formula_idx, coupon_type_idx]
            if idx is not None
        ) + 1

        for row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
            isin_norm = _normalize_isin(row[isin_idx] if isin_idx < len(row) else None)
            if not isin_norm:
                continue
            total_rows += 1
            enabled = _parse_override_bool(row[enabled_idx] if enabled_idx < len(row) else None)
            if enabled is not True:
                continue
            enabled_rows += 1
            coupon_formula_override = str(row[coupon_formula_idx] or "").strip() if coupon_formula_idx < len(row) else ""
            coupon_type_override = (
                str(row[coupon_type_idx] or "").strip()
                if coupon_type_idx is not None and coupon_type_idx < len(row)
                else ""
            )
            if coupon_type_override:
                coupon_type_read_count += 1
            overrides[isin_norm] = {
                "enabled": True,
                "drop": _parse_override_bool(row[drop_idx] if drop_idx < len(row) else None),
                "kval": _parse_override_bool(row[kval_idx] if kval_idx < len(row) else None),
                "subord": _parse_override_bool(row[subord_idx] if subord_idx < len(row) else None),
                "coupon_formula_override": coupon_formula_override or None,
                "coupon_type_override": coupon_type_override or None,
            }
    finally:
        wb.close()

    logger.info(
        "BondOverrides load: path=%s total_rows=%s enabled_rows=%s coupon_type_values_read=%s",
        overrides_path,
        total_rows,
        enabled_rows,
        coupon_type_read_count,
    )

    return overrides, total_rows, enabled_rows


def rebuild_screener_table(conn: sqlite3.Connection) -> dict[str, int]:
    ensure_screener_table(conn)
    conn.execute(f'DELETE FROM "{SCREENER_TABLE_NAME}"')

    totals: dict[str, int] = {"Green": 0, "Yellow": 0}
    ytm_fixed_count = 0
    ytm_floater_count = 0
    ytm_linker_count = 0
    ytm_other_count = 0
    logger = logging.getLogger("bonds_main")
    overrides_by_isin, overrides_total_rows, overrides_enabled_rows = load_bond_overrides(logger)
    overrides_drop_applied_count = 0
    overrides_kval_applied_count = 0
    overrides_subord_applied_count = 0
    overrides_formula_applied_count = 0
    overrides_coupon_type_applied_count = 0

    cbr_data = get_cbr_reference_data(conn, logger, datetime.now(timezone.utc))
    for source_table, source_list, score in SCREENER_SOURCE_TABLES:
        rows = conn.execute(
            f'''
            SELECT
                src."SECID", src."ISIN", src."Название", src."BOND_TYPE", src."BOND_SUBTYPE", src."IS_QUALIFIED_INVESTORS", src."Smartlab_Только для квалов?",
                src."Субординированная (да/нет)", src."Corpbonds_Наличие амортизации", src."Corpbonds_Купон лесенкой",
                src."{AMORTIZATION_START_COLUMN}", src."MATDATE", src."Corpbonds_Дата ближайшей оферты", src."Smartlab_Дата оферты",
                src."Corpbonds_Дата ближайшего купона", src."Corpbonds_Тип купона", src."Smartlab_Длительность купона, дней",
                src."Corpbonds_НКД", src."НКД", COALESCE(NULLIF(src."Corpbonds_Ставка купона", ''), src."Текущий купон, %") AS "Текущий купон, %",
                src."Купон (раз/год)", src."Corpbonds_Формула купона",
                src."FACEVALUE", src."FACEUNIT", src."Коэф. Ликвидности (max=100)", src."Corpbonds_Цена последняя",
                src."Цена Доход", src."Smartlab_Котировка облигации, %", src."PRICE"
            FROM "{source_table}" src
            WHERE TRIM(COALESCE(src."ISIN", '')) <> ''
            '''
        ).fetchall()

        facevalues_by_secid: dict[str, float] = {}
        secids_with_amort: set[str] = set()
        filtered_rows: list[tuple[object, ...]] = []
        for row in rows:
            row_values = list(row)
            isin_norm = _normalize_isin(row_values[1])
            override = overrides_by_isin.get(isin_norm or "") if isin_norm else None
            if override and override.get("enabled") is True:
                if override.get("drop") is True:
                    overrides_drop_applied_count += 1
                    continue
                if override.get("kval") is not None:
                    row_values[5] = 1 if bool(override.get("kval")) else 0
                    row_values[6] = ""
                    overrides_kval_applied_count += 1
                if override.get("subord") is not None:
                    row_values[7] = "Да" if bool(override.get("subord")) else "Нет"
                    overrides_subord_applied_count += 1
                if override.get("coupon_formula_override"):
                    row_values[21] = str(override.get("coupon_formula_override") or "").strip()
                    overrides_formula_applied_count += 1
                if override.get("coupon_type_override"):
                    row_values[15] = str(override.get("coupon_type_override") or "").strip()
                    overrides_coupon_type_applied_count += 1
            filtered_rows.append(tuple(row_values))

        rows = filtered_rows
        totals[source_list] = len(rows)

        for row in rows:
            secid = str(row[0] or "").strip()
            facevalue_value = _parse_decimal_value(row[22])
            if secid and facevalue_value is not None:
                facevalues_by_secid[secid] = facevalue_value
            if secid and not _is_false_like(row[8]):
                secids_with_amort.add(secid)

        amortization_map = _load_amortization_schedule(conn, secids_with_amort, facevalues_by_secid)

        records: list[tuple[str, ...]] = []
        for row in rows:
            offer_date = _pick_offer_date(row[12], row[13])
            secid = str(row[0] or "").strip()
            coupon_type = row[15]
            if _is_floater_coupon_type(coupon_type):
                ytm_value = _calculate_floater_ytm(
                    coupon_formula=row[21],
                    subord_flag=row[7],
                    coupon_frequency=row[20],
                    coupon_period_days=row[16],
                    next_coupon_date=row[14],
                    faceunit=row[23],
                    corpbonds_nkd=row[17],
                    fallback_nkd=row[18],
                    facevalue=row[22],
                    matdate=row[11],
                    offerdate=offer_date,
                    corpbonds_price=row[25],
                    dohod_price=row[26],
                    smartlab_price=row[27],
                    moex_price=row[28],
                    amortization_schedule=amortization_map.get(secid),
                    cbr_data=cbr_data,
                    logger=logging.getLogger("bonds_main"),
                    secid=secid,
                )
                if ytm_value:
                    ytm_floater_count += 1
            elif is_linker(secid, row[3], row[4]):
                ytm_value = _calculate_linker_ytm(
                    secid=secid,
                    bond_type=row[3],
                    bond_subtype=row[4],
                    coupon_percent=row[19],
                    coupon_frequency=row[20],
                    coupon_period_days=row[16],
                    next_coupon_date=row[14],
                    faceunit=row[23],
                    corpbonds_nkd=row[17],
                    fallback_nkd=row[18],
                    facevalue=row[22],
                    matdate=row[11],
                    offerdate=offer_date,
                    corpbonds_price=row[25],
                    dohod_price=row[26],
                    smartlab_price=row[27],
                    moex_price=row[28],
                    logger=logging.getLogger("bonds_main"),
                )
                if ytm_value:
                    ytm_linker_count += 1
            elif _is_other_coupon_type(coupon_type):
                ytm_value = _calculate_other_coupon_ytm(
                    subord_flag=row[7],
                    coupon_percent=row[19],
                    coupon_frequency=row[20],
                    coupon_period_days=row[16],
                    next_coupon_date=row[14],
                    amort_flag=row[8],
                    faceunit=row[23],
                    corpbonds_nkd=row[17],
                    fallback_nkd=row[18],
                    facevalue=row[22],
                    matdate=row[11],
                    offerdate=offer_date,
                    corpbonds_price=row[25],
                    dohod_price=row[26],
                    smartlab_price=row[27],
                    moex_price=row[28],
                    amortization_schedule=amortization_map.get(secid),
                    logger=logging.getLogger("bonds_main"),
                    secid=secid,
                )
                if ytm_value:
                    ytm_other_count += 1
            else:
                ytm_value = _calculate_fixed_coupon_ytm(
                    subord_flag=row[7],
                    amort_flag=row[8],
                    coupon_type=coupon_type,
                    coupon_percent=row[19],
                    coupon_frequency=row[20],
                    coupon_period_days=row[16],
                    next_coupon_date=row[14],
                    faceunit=row[23],
                    corpbonds_nkd=row[17],
                    fallback_nkd=row[18],
                    facevalue=row[22],
                    matdate=row[11],
                    offerdate=offer_date,
                    corpbonds_price=row[25],
                    dohod_price=row[26],
                    smartlab_price=row[27],
                    moex_price=row[28],
                    amortization_schedule=amortization_map.get(secid),
                    logger=logging.getLogger("bonds_main"),
                    secid=secid,
                )
                if ytm_value:
                    ytm_fixed_count += 1

            records.append(
                (
                    str(row[1] or "").strip(),
                    str(row[2] or "").strip(),
                    str(_merge_qualified(row[5], row[6])),
                    str(row[7] or "").strip(),
                    str(row[8] or "").strip(),
                    str(row[9] or "").strip(),
                    _normalize_date_to_iso(row[10]),
                    _normalize_date_to_iso(row[11]),
                    offer_date,
                    _normalize_date_to_iso(row[14]),
                    str(row[15] or "").strip(),
                    _parse_decimal_value(row[16]),
                    _parse_decimal_value(row[17]) if _parse_decimal_value(row[17]) is not None else _parse_decimal_value(row[18]),
                    _parse_decimal_value(row[19]),
                    _parse_decimal_value(ytm_value),
                    str(row[21] or "").strip(),
                    _parse_decimal_value(row[22]),
                    str(row[23] or "").strip(),
                    _parse_decimal_value(row[24]),
                    _parse_decimal_value(row[25]),
                    _parse_decimal_value(row[26]),
                    _parse_decimal_value(row[27]),
                    _parse_decimal_value(row[28]),
                    int(score),
                    source_list,
                )
            )

            ytm_num = _parse_decimal_value(ytm_value)
            if (not ytm_value) or (ytm_num is not None and (ytm_num < -10 or ytm_num > 200)):
                price_used = _pick_price_for_ytm(row[25], row[26], row[27], row[28])
                nkd_used = _parse_decimal_value(row[17]) if _parse_decimal_value(row[17]) is not None else _parse_decimal_value(row[18])
                facevalue_value = _parse_decimal_value(row[22])
                dirty_price = None
                floater_terms = parse_floater_terms(row[21]) if _is_floater_coupon_type(coupon_type) else None
                if price_used is not None and facevalue_value is not None:
                    dirty_price = _normalize_purchase_price(price_used, facevalue_value, nkd_used)
                _write_ytm_debug_record(
                    {
                        "secid": secid,
                        "isin": str(row[1] or "").strip(),
                        "coupon_formula": str(row[21] or "").strip(),
                        "parsed_terms": {
                            "index_type": floater_terms[0] if floater_terms else None,
                            "tenor": floater_terms[1] if floater_terms else None,
                            "premium": floater_terms[2] if floater_terms else None,
                        },
                        "faceunit": str(row[23] or "").strip(),
                        "facevalue": facevalue_value,
                        "price_used": price_used,
                        "nkd_used": nkd_used,
                        "dirty_price": dirty_price,
                        "coupon_type": str(coupon_type or ""),
                        "coupon_rate": _parse_decimal_value(row[19]),
                        "freq": _parse_decimal_value(row[20]),
                        "target_date": offer_date or _normalize_date_to_iso(row[11]),
                        "cbr_snapshot": {
                            "key_rate": cbr_data.get("key_rate"),
                            "ruonia": cbr_data.get("ruonia"),
                            "zcyc_tenor_used": floater_terms[1] if floater_terms else None,
                            "zcyc_yield_used": pick_zcyc_point(cbr_data.get("zcyc") if isinstance(cbr_data.get("zcyc"), dict) else {}, floater_terms[1]) if floater_terms else None,
                        },
                        "result_yield": ytm_num,
                        "reason": "empty" if not ytm_value else "anomaly",
                    }
                )

        if records:
            column_names = ", ".join(f'"{column}"' for column in SCREENER_COLUMNS)
            placeholders = ", ".join(["?"] * len(SCREENER_COLUMNS))
            conn.executemany(
                f'INSERT OR REPLACE INTO "{SCREENER_TABLE_NAME}" ({column_names}) VALUES ({placeholders})',
                records,
            )

    conn.commit()
    logger.info(
        "BondOverrides: total_rows=%s, enabled_rows=%s, drop_applied=%s, kval_applied=%s, subord_applied=%s, formula_applied=%s, coupon_type_applied=%s",
        overrides_total_rows,
        overrides_enabled_rows,
        overrides_drop_applied_count,
        overrides_kval_applied_count,
        overrides_subord_applied_count,
        overrides_formula_applied_count,
        overrides_coupon_type_applied_count,
    )
    total = conn.execute(f'SELECT COUNT(*) FROM "{SCREENER_TABLE_NAME}"').fetchone()[0]
    return {
        "green": totals["Green"],
        "yellow": totals["Yellow"],
        "total": int(total),
        "ytm_fixed_count": ytm_fixed_count,
        "ytm_floater_count": ytm_floater_count,
        "ytm_linker_count": ytm_linker_count,
        "ytm_other_count": ytm_other_count,
    }


def _symbolize_boolean(raw_value: str | None) -> str:
    return "✅" if _to_binary_flag(raw_value) == 1 else "❌"


def _symbolize_yes_no(raw_value: str | None) -> str:
    value = str(raw_value or "").strip().casefold()
    if value == "да":
        return "✅"
    if value == "нет":
        return "❌"
    return ""


def export_screener_excel(conn: sqlite3.Connection) -> dict[str, int]:
    wb = Workbook()
    ws_green = wb.active
    ws_green.title = "Green"
    ws_yellow = wb.create_sheet("Yellow")

    headers = [target for _, target in SCREENER_EXPORT_COLUMNS]
    for ws in (ws_green, ws_yellow):
        ws.append(headers)

    counts: dict[str, int] = {"Green": 0, "Yellow": 0}
    for sheet_name in ("Green", "Yellow"):
        ws = ws_green if sheet_name == "Green" else ws_yellow
        rows = conn.execute(
            f'''
            SELECT {", ".join(f'"{name}"' for name, _ in SCREENER_EXPORT_COLUMNS)}
            FROM "{SCREENER_TABLE_NAME}"
            WHERE "SourceList" = ?
            ''',
            (sheet_name,),
        ).fetchall()
        if getattr(config, "SCREENER_SORT_BY_AMORT_START_DATE", True):
            rows = sorted(rows, key=_screener_sort_key)
        row_idx_debug = 0
        for row in rows:
            row_values = list(row)
            row_values[2] = _symbolize_boolean(row_values[2])
            row_values[3] = _symbolize_yes_no(row_values[3])
            row_values[4] = _symbolize_yes_no(row_values[4])
            row_values[5] = _symbolize_yes_no(row_values[5])
            row_values = _prepare_screener_export_row(headers, row_values)
            if row_idx_debug < 20:
                for debug_col in ("YTM", "Купон, %", "НКД"):
                    idx = headers.index(debug_col)
                    debug_value = row_values[idx]
                    logging.getLogger("bonds_main").debug(
                        "Excel type check %s row=%s type=%s value=%r",
                        debug_col,
                        row_idx_debug + 1,
                        type(debug_value).__name__,
                        debug_value,
                    )
            row_idx_debug += 1
            ws.append(row_values)
        counts[sheet_name] = len(rows)

        header_fill = PatternFill(fill_type="solid", fgColor=config.EMITENTS_HEADER_FILL_COLOR)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        for column_cells in ws.columns:
            max_len = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[column_letter].width = min(max_len + 2, 80)

        date_col_indices = [
            headers.index(column_name) + 1
            for column_name in ("AmortStarrtDate", "MATDATE", "Offerdate", "Ближайший купон")
        ]
        number_formats = {
            "YTM": "0.00",
            "Купон, %": "0.00",
            "НКД": "0.00",
            "FACEVALUE": "0.00",
            "Цена Corpbonds": "0.00",
            "Цена Доход": "0.00",
            "Цена Smartlab": "0.00",
            "Цена MOEX": "0.00",
            "КупонПериод": "0",
            "Ликвидность": "0.00",
        }
        number_format_indices = {
            headers.index(column_name) + 1: number_format
            for column_name, number_format in number_formats.items()
        }
        if ws.max_row >= 2:
            for row_idx in range(2, ws.max_row + 1):
                for col_idx in date_col_indices:
                    ws.cell(row=row_idx, column=col_idx).number_format = "yyyy-mm-dd"

        liquidity_col_idx = headers.index("Ликвидность") + 1
        liquidity_col_letter = ws.cell(row=1, column=liquidity_col_idx).column_letter
        if ws.max_row >= 2:
            for row_idx in range(2, ws.max_row + 1):
                for col_idx, number_format in number_format_indices.items():
                    ws.cell(row=row_idx, column=col_idx).number_format = number_format
            data_bar_rule = Rule(
                type="dataBar",
                dataBar=DataBar(
                    cfvo=[FormatObject(type="min"), FormatObject(type="max")],
                    color="63C384",
                    showValue=True,
                ),
            )
            ws.conditional_formatting.add(f"{liquidity_col_letter}2:{liquidity_col_letter}{ws.max_row}", data_bar_rule)

    screener_path = config.BASE_DIR / config.SCREENER_XLSX_FILENAME
    wb.save(screener_path)
    counts["total"] = counts["Green"] + counts["Yellow"]
    return counts


def main() -> None:
    logger = setup_logging()
    if getattr(config, "ENABLE_YTM_DEBUG_LOG", True):
        (config.LOGS_DIR / getattr(config, "YTM_DEBUG_FILENAME", "ytm_debug.jsonl")).write_text("", encoding="utf-8")
    stage_times: dict[str, float] = {}
    presorter_summary: dict[str, dict[str, int]] = {}
    started = perf_counter()

    db_path = config.DB_DIR / config.DB_FILENAME
    ratings_db_path = config.DB_DIR / config.RAITINGS_DB_FILENAME

    try:
        print("=====\nЭтап 1: Подготовка окружения")
        s = perf_counter()
        with progress(total=3, desc="Подготовка", unit="шаг") as pbar:
            ensure_directories()
            migrate_legacy_db_if_needed()
            ensure_bond_overrides_excel()
            pbar.update(1)
            pbar.set_description("Подготовка БД")
            with connect_db(db_path) as conn:
                init_meta_table(conn)
                migrate_legacy_rates_table_if_needed(conn)
                ensure_emitents_table(conn)
                ensure_dohod_table(conn)
            pbar.update(1)
            pbar.set_description("Инициализация витрин")
            pbar.update(1)
        stage_times["Этап 1: Подготовка окружения"] = perf_counter() - s

        print("Этап 2: Проверка TTL кэша и обновление данных")
        s = perf_counter()
        with progress(total=3, desc="Обновление данных", unit="шаг") as pbar:
            now_utc = datetime.now(timezone.utc)
            pbar.update(1)
            pbar.set_description("Работа с SQLite")
            with connect_db(db_path) as conn:
                init_meta_table(conn)
                refreshed, row_count = refresh_data_if_needed(conn, logger, now_utc)
            pbar.update(1)
            pbar.set_description("Финализация")
            logger.info("Режим: %s", "обновлено из сети" if refreshed else "использован локальный кэш")
            logger.info("Количество строк в таблице %s: %s", config.RATES_TABLE_NAME, row_count)
            pbar.update(1)
        stage_times["Этап 2: Проверка TTL кэша и обновление данных"] = perf_counter() - s

        print("Этап 3: Доходъ (Playwright Excel + SQLite + snapshot)")
        s = perf_counter()
        with progress(total=3, desc="Dohod", unit="шаг") as pbar:
            with connect_db(db_path) as conn:
                refreshed, dohod_rows = refresh_dohod_data_if_needed(conn, logger, datetime.now(timezone.utc))
                pbar.update(1)
                dohod_snapshot = export_dohod_snapshot(conn)
            pbar.update(1)
            logger.info(
                "Доходъ: режим=%s, строк в таблице=%s, строк в snapshot=%s",
                "обновлено из сети" if refreshed else "использован локальный кэш",
                dohod_rows,
                dohod_snapshot,
            )
            pbar.update(1)
        stage_times["Этап 3: Доходъ (Playwright Excel + SQLite + snapshot)"] = perf_counter() - s

        print("Этап 4: Формирование Excel-среза MOEX")
        s = perf_counter()
        with progress(total=2, desc="MOEX snapshot", unit="шаг") as pbar:
            with connect_db(db_path) as conn:
                count = export_random_snapshot(conn)
            pbar.update(1)
            logger.info("Сформирован Excel-срез MOEX: строк=%s", count)
            pbar.update(1)
        stage_times["Этап 4: Формирование Excel-среза MOEX"] = perf_counter() - s

        print("Этап 5: Рейтинги агентств (НРА + АКРА + НКР + RAEX, отдельная SQLite)")
        s = perf_counter()
        with progress(total=10, desc="NRA/ACRA/NKR/RAEX", unit="шаг", position=1) as pbar:
            now_utc = datetime.now(timezone.utc)
            with connect_db(ratings_db_path) as nra_conn:
                init_meta_table(nra_conn)
                ensure_nra_tables(nra_conn)
                ensure_acra_tables(nra_conn)
                ensure_nkr_tables(nra_conn)
                ensure_raex_tables(nra_conn)
                nra_refreshed, nra_rows = refresh_nra_data_if_needed(nra_conn, logger, now_utc)
                nra_snapshot_rows = export_nra_snapshot(nra_conn)
                pbar.update(1)
                acra_refreshed, acra_rows, acra_cards = refresh_acra_data_if_needed(nra_conn, logger, now_utc)
                acra_snapshot_rows = export_acra_snapshot(nra_conn)
                pbar.update(1)
                nkr_refreshed, nkr_rows = refresh_nkr_data_if_needed(nra_conn, logger, now_utc)
                nkr_snapshot_rows = export_nkr_snapshot(nra_conn)
                pbar.update(1)
                raex_refreshed, raex_rows, raex_inns, raex_errors = refresh_raex_data_if_needed(nra_conn, db_path, logger, now_utc)
                raex_snapshot_rows = export_raex_snapshot(nra_conn)
            pbar.update(1)
            pbar.set_description("Фиксация результата")
            logger.info(
                "НРА: режим=%s, строк в источнике=%s, строк в snapshot=%s",
                "обновлено из сети" if nra_refreshed else "использован локальный кэш",
                nra_rows,
                nra_snapshot_rows,
            )
            logger.info(
                "АКРА: режим=%s, строк в базе=%s, карточек запрошено=%s, строк в snapshot=%s",
                "обновлено из сети" if acra_refreshed else "использован локальный кэш",
                acra_rows,
                acra_cards,
                acra_snapshot_rows,
            )
            logger.info(
                "НКР: режим=%s, строк в источнике=%s, строк в snapshot=%s",
                "обновлено из сети" if nkr_refreshed else "использован локальный кэш",
                nkr_rows,
                nkr_snapshot_rows,
            )
            logger.info(
                "RAEX: режим=%s, обработано INN=%s, актуальных=%s, ошибок=%s, строк в snapshot=%s",
                "обновлено из сети" if raex_refreshed else "использован локальный кэш",
                raex_inns,
                raex_rows,
                raex_errors,
                raex_snapshot_rows,
            )
            pbar.update(1)
            pbar.update(1)
            pbar.update(1)
            pbar.update(1)
            pbar.update(1)
            pbar.update(1)
        stage_times["Этап 5: Рейтинги агентств (НРА + АКРА + НКР + RAEX, отдельная SQLite)"] = perf_counter() - s

        print("Этап 6: Витрина эмитентов (SQL + Excel)")
        s = perf_counter()
        with progress(total=8, desc="Emitents", unit="шаг") as pbar:
            today_str = datetime.now().strftime(config.DATE_SCORING_FORMAT)
            with connect_db(db_path) as conn:
                ensure_emitents_table(conn)
                pulled = pull_scoring_from_excel(conn, logger, today_str)
                pbar.update(1)
                synced = sync_emitents_from_rates(conn, logger)
                pbar.update(1)
                with connect_db(ratings_db_path) as nra_conn:
                    nra_synced = sync_nra_rate_to_emitents(conn, nra_conn, logger)
                    acra_synced = sync_acra_rate_to_emitents(conn, nra_conn, logger)
                    nkr_synced = sync_nkr_rate_to_emitents(conn, nra_conn, logger)
                    raex_synced = sync_raex_rate_to_emitents(conn, nra_conn, logger)
                pbar.update(1)
                dates_fixed = ensure_scoring_dates(conn, logger, today_str)
                pbar.update(1)
                emitents_count = export_emitents_excel(conn)
                pbar.update(1)
                emitents_snapshot = export_emitents_snapshot(conn)
                pbar.update(1)

            logger.info(
                "Витрина эмитентов: перенос из Excel=%s, upsert из %s=%s, NRA_Rate=%s, Acra_Rate=%s, NKR_Rate=%s, RAEX_Rate=%s, авто-дат=%s, строк в Emitents.xlsx=%s, строк в snapshot=%s",
                pulled,
                config.RATES_TABLE_NAME,
                synced,
                nra_synced,
                acra_synced,
                nkr_synced,
                raex_synced,
                dates_fixed,
                emitents_count,
                emitents_snapshot,
            )
            pbar.update(1)
        stage_times["Этап 6: Витрина эмитентов (SQL + Excel)"] = perf_counter() - s

        print("Этап 7: Merge Green/Yellow (SQL)")
        s = perf_counter()
        with progress(total=2, desc="Merge bonds", unit="шаг") as pbar:
            with connect_db(db_path) as conn:
                green_rows = rebuild_merge_table_by_scoring(conn, config.MERGE_GREEN_TABLE_NAME, "Green")
                pbar.update(1)
                yellow_rows = rebuild_merge_table_by_scoring(conn, config.MERGE_YELLOW_TABLE_NAME, "Yellow")
                pbar.update(1)

            logger.info(
                "Merge Green: строк=%s; Merge Yellow: строк=%s",
                green_rows,
                yellow_rows,
            )
        stage_times["Этап 7: Merge Green/Yellow (SQL)"] = perf_counter() - s

        print("Этап 8: Presorter для Merge-таблиц")
        s = perf_counter()
        with progress(total=2, desc="Presorter", unit="шаг") as pbar:
            with connect_db(db_path) as conn:
                green_presort = presort_merge_table(conn, config.MERGE_GREEN_TABLE_NAME)
                pbar.update(1)
                yellow_presort = presort_merge_table(conn, config.MERGE_YELLOW_TABLE_NAME)
                pbar.update(1)
            presorter_summary["MergeGreen"] = green_presort
            presorter_summary["MergeYellow"] = yellow_presort
            logger.info(
                "Presorter: MergeGreen rows=%s->%s excluded(matdate=%s, dohod_nearest=%s, bond_type=%s, offerdate=%s, amortstart=%s, total=%s); "
                "MergeYellow rows=%s->%s excluded(matdate=%s, dohod_nearest=%s, bond_type=%s, offerdate=%s, amortstart=%s, total=%s); "
                "dohod_nearest_rule_enabled=%s",
                green_presort["rows_before"],
                green_presort["rows_after"],
                green_presort["excluded_by_matdate_rule"],
                green_presort["excluded_by_dohod_nearest_rule"],
                green_presort["excluded_by_bond_type_rule"],
                green_presort["excluded_by_offerdate_rule"],
                green_presort["excluded_by_amortstartdate_rule"],
                green_presort["excluded_total"],
                yellow_presort["rows_before"],
                yellow_presort["rows_after"],
                yellow_presort["excluded_by_matdate_rule"],
                yellow_presort["excluded_by_dohod_nearest_rule"],
                yellow_presort["excluded_by_bond_type_rule"],
                yellow_presort["excluded_by_offerdate_rule"],
                yellow_presort["excluded_by_amortstartdate_rule"],
                yellow_presort["excluded_total"],
                config.PRESORTER_USE_DOHOD_NEAREST_DATE,
            )
        stage_times["Этап 8: Presorter для Merge-таблиц"] = perf_counter() - s

        print("Этап 9: Обогащение Merge* из Corpbonds")
        s = perf_counter()
        with progress(total=6, desc="Corpbonds enrich", unit="шаг") as pbar:
            with connect_db(db_path) as conn:
                secids_total, secids_requested, secids_saved = refresh_corpbonds_data_if_needed(
                    conn, logger, datetime.now(timezone.utc)
                )
                pbar.update(1)
                green_rows_after_corpbonds = apply_corpbonds_inner_join_to_merge_table(conn, config.MERGE_GREEN_TABLE_NAME)
                pbar.update(1)
                yellow_rows_after_corpbonds = apply_corpbonds_inner_join_to_merge_table(conn, config.MERGE_YELLOW_TABLE_NAME)
                pbar.update(1)
                corpbonds_snapshot = export_corpbonds_snapshot(conn)
                pbar.update(1)
                green_snapshot = export_merge_snapshot(
                    conn,
                    config.MERGE_GREEN_TABLE_NAME,
                    config.MERGE_GREEN_SNAPSHOT_FILENAME,
                    "merge_green_snapshot",
                )
                pbar.update(1)
                yellow_snapshot = export_merge_snapshot(
                    conn,
                    config.MERGE_YELLOW_TABLE_NAME,
                    config.MERGE_YELLOW_SNAPSHOT_FILENAME,
                    "merge_yellow_snapshot",
                )
                pbar.update(1)
            logger.info(
                "Corpbonds: SECID в Merge=%s, к запросу по TTL=%s, успешно сохранено=%s, "
                "INNER JOIN обновил MergeGreen=%s, MergeYellow=%s, corpbonds_snapshot=%s, "
                "merge_snapshots_after_corpbonds: green=%s, yellow=%s",
                secids_total,
                secids_requested,
                secids_saved,
                green_rows_after_corpbonds,
                yellow_rows_after_corpbonds,
                corpbonds_snapshot,
                green_snapshot,
                yellow_snapshot,
            )
        stage_times["Этап 9: Обогащение Merge* из Corpbonds"] = perf_counter() - s

        print("Этап 10: Обогащение Merge* из MOEX Amortizations")
        s = perf_counter()
        with progress(total=6, desc="MOEX amort enrich", unit="шаг") as pbar:
            with connect_db(db_path) as conn:
                secids_total, secids_requested, rows_saved = refresh_moex_amortizations_if_needed(
                    conn, logger, datetime.now(timezone.utc)
                )
                pbar.update(1)
                green_amort_applied = apply_amortization_start_date_to_merge_table(conn, config.MERGE_GREEN_TABLE_NAME)
                pbar.update(1)
                yellow_amort_applied = apply_amortization_start_date_to_merge_table(conn, config.MERGE_YELLOW_TABLE_NAME)
                pbar.update(1)
                amort_snapshot = export_moex_amortization_snapshot(conn)
                pbar.update(1)
                green_snapshot = export_merge_snapshot(
                    conn,
                    config.MERGE_GREEN_TABLE_NAME,
                    config.MERGE_GREEN_SNAPSHOT_FILENAME,
                    "merge_green_snapshot",
                )
                pbar.update(1)
                yellow_snapshot = export_merge_snapshot(
                    conn,
                    config.MERGE_YELLOW_TABLE_NAME,
                    config.MERGE_YELLOW_SNAPSHOT_FILENAME,
                    "merge_yellow_snapshot",
                )
                pbar.update(1)
            logger.info(
                "MOEX amortizations: SECID в Merge*=%s, к запросу по TTL=%s, строк сохранено=%s, "
                "проставлено AmortStarrtDate: MergeGreen=%s, MergeYellow=%s, moex_amort_snapshot=%s, "
                "merge_snapshots_after_moex_amort: green=%s, yellow=%s",
                secids_total,
                secids_requested,
                rows_saved,
                green_amort_applied,
                yellow_amort_applied,
                amort_snapshot,
                green_snapshot,
                yellow_snapshot,
            )
        stage_times["Этап 10: Обогащение Merge* из MOEX Amortizations"] = perf_counter() - s

        print("Этап 11: Обогащение Merge* из Smartlab")
        s = perf_counter()
        with progress(total=6, desc="Smartlab enrich", unit="шаг") as pbar:
            with connect_db(db_path) as conn:
                secids_total, secids_requested, secids_saved = refresh_smartlab_data_if_needed(
                    conn, logger, datetime.now(timezone.utc)
                )
                pbar.update(1)
                green_rows_after_smartlab = apply_smartlab_inner_join_to_merge_table(conn, config.MERGE_GREEN_TABLE_NAME)
                pbar.update(1)
                yellow_rows_after_smartlab = apply_smartlab_inner_join_to_merge_table(conn, config.MERGE_YELLOW_TABLE_NAME)
                pbar.update(1)
                smartlab_snapshot = export_smartlab_snapshot(conn)
                pbar.update(1)
                green_snapshot = export_merge_snapshot(
                    conn,
                    config.MERGE_GREEN_TABLE_NAME,
                    config.MERGE_GREEN_SNAPSHOT_FILENAME,
                    "merge_green_snapshot",
                )
                pbar.update(1)
                yellow_snapshot = export_merge_snapshot(
                    conn,
                    config.MERGE_YELLOW_TABLE_NAME,
                    config.MERGE_YELLOW_SNAPSHOT_FILENAME,
                    "merge_yellow_snapshot",
                )
                pbar.update(1)
            logger.info(
                "Smartlab: SECID в Merge=%s, к запросу по TTL=%s, успешно сохранено=%s, "
                "INNER JOIN обновил MergeGreen=%s, MergeYellow=%s, smartlab_snapshot=%s, "
                "merge_snapshots_after_smartlab: green=%s, yellow=%s",
                secids_total,
                secids_requested,
                secids_saved,
                green_rows_after_smartlab,
                yellow_rows_after_smartlab,
                smartlab_snapshot,
                green_snapshot,
                yellow_snapshot,
            )
        stage_times["Этап 11: Обогащение Merge* из Smartlab"] = perf_counter() - s

        print("Этап 11.5: Повторный Presorter после обогащений")
        s = perf_counter()
        with progress(total=2, desc="Presorter rerun", unit="шаг") as pbar:
            with connect_db(db_path) as conn:
                green_presort = presort_merge_table(conn, config.MERGE_GREEN_TABLE_NAME)
                pbar.update(1)
                yellow_presort = presort_merge_table(conn, config.MERGE_YELLOW_TABLE_NAME)
                pbar.update(1)
            presorter_summary["MergeGreen_after_enrichment"] = green_presort
            presorter_summary["MergeYellow_after_enrichment"] = yellow_presort
            logger.info(
                "Presorter rerun after enrichment: "
                "MergeGreen rows=%s->%s excluded(matdate=%s, dohod_nearest=%s, bond_type=%s, offerdate=%s, amortstart=%s, total=%s); "
                "MergeYellow rows=%s->%s excluded(matdate=%s, dohod_nearest=%s, bond_type=%s, offerdate=%s, amortstart=%s, total=%s); "
                "dohod_nearest_rule_enabled=%s",
                green_presort["rows_before"],
                green_presort["rows_after"],
                green_presort["excluded_by_matdate_rule"],
                green_presort["excluded_by_dohod_nearest_rule"],
                green_presort["excluded_by_bond_type_rule"],
                green_presort["excluded_by_offerdate_rule"],
                green_presort["excluded_by_amortstartdate_rule"],
                green_presort["excluded_total"],
                yellow_presort["rows_before"],
                yellow_presort["rows_after"],
                yellow_presort["excluded_by_matdate_rule"],
                yellow_presort["excluded_by_dohod_nearest_rule"],
                yellow_presort["excluded_by_bond_type_rule"],
                yellow_presort["excluded_by_offerdate_rule"],
                yellow_presort["excluded_by_amortstartdate_rule"],
                yellow_presort["excluded_total"],
                config.PRESORTER_USE_DOHOD_NEAREST_DATE,
            )
        stage_times["Этап 11.5: Повторный Presorter после обогащений"] = perf_counter() - s

        print("Этап 12: Screener (SQL + Excel)")
        s = perf_counter()
        self_check_enabled = bool(getattr(config, "YTM_SELFCHECK_ENABLED", True))
        screener_steps = 3 if self_check_enabled else 2
        with progress(total=screener_steps, desc="Screener", unit="шаг") as pbar:
            with connect_db(db_path) as conn:
                screener_stats = rebuild_screener_table(conn)
                pbar.update(1)
                screener_export = export_screener_excel(conn)
                pbar.update(1)
            if self_check_enabled:
                ytm_self_check_errors = _run_ytm_self_check()
                if ytm_self_check_errors:
                    logger.warning("YTM self-check: %s", " | ".join(ytm_self_check_errors))
                    if bool(getattr(config, "YTM_SELFCHECK_STRICT", False)):
                        raise RuntimeError("YTM self-check failed")
                else:
                    logger.info("YTM self-check: OK")
                pbar.update(1)
            logger.info(
                "Screener: SQL rows green=%s yellow=%s total=%s ytm_fixed=%s; Excel rows green=%s yellow=%s total=%s",
                screener_stats.get("green", 0),
                screener_stats.get("yellow", 0),
                screener_stats.get("total", 0),
                screener_stats.get("ytm_fixed_count", 0),
                screener_export.get("Green", 0),
                screener_export.get("Yellow", 0),
                screener_export.get("total", 0),
            )
        stage_times["Этап 12: Screener (SQL + Excel)"] = perf_counter() - s

        print("=====\nГотово")
    except Exception as exc:
        logger.exception("Ошибка выполнения: %s", exc)
        raise
    finally:
        total = perf_counter() - started
        print("=====\nSummary")
        for stage_name, duration in stage_times.items():
            print(f"{stage_name}: {duration:.2f} сек")
        if presorter_summary:
            print("Этап Presorter")
            for merge_name in ("MergeGreen", "MergeYellow"):
                item = presorter_summary.get(merge_name, {})
                print(merge_name)
                print(f"Строк до/после Presorter: {item.get('rows_before', 0)} -> {item.get('rows_after', 0)}")
                print(
                    f"Исключено бумаг по правилу MATDATE < {config.PRESORTER_MIN_DAYS_TO_EVENT} дней: {item.get('excluded_by_matdate_rule', 0)}"
                )
                print(
                    f"Исключено бумаг по правилу Доходъ (ближайшая дата) < {config.PRESORTER_MIN_DAYS_TO_EVENT} дней: {item.get('excluded_by_dohod_nearest_rule', 0)}"
                )
                print(
                    f"Исключено бумаг по правилу Bond_TYPE: {item.get('excluded_by_bond_type_rule', 0)}"
                )
                print(
                    f"Исключено бумаг по правилу Offerdate < {config.PRESORTER_MIN_DAYS_TO_EVENT} дней: {item.get('excluded_by_offerdate_rule', 0)}"
                )
                print(
                    f"Исключено бумаг по правилу AmortStarrtDate < {config.PRESORTER_MIN_DAYS_TO_EVENT} дней: {item.get('excluded_by_amortstartdate_rule', 0)}"
                )
        print(f"Всего: {total:.2f} сек")


if __name__ == "__main__":
    main()
