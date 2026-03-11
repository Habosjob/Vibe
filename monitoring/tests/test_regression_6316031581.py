from __future__ import annotations

import logging
import sqlite3
import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from monitoring import config
import monitoring.main as monitoring_main
from monitoring.main import EmitentRow, SCHEMA_SQL, build_latest_event_by_inn, export_portfolio, parse_reports_page, stage_reports_prepare


class TestEDisclosureRegression6316031581(unittest.TestCase):
    def test_parse_type4_row_extracts_fileload_and_date(self) -> None:
        html = """
        <html><body>
        <table class='zebra'>
          <tr>
            <td>Годовая консолидированная финансовая отчетность</td>
            <td>2025 год</td>
            <td>10.02.2026</td>
            <td>10.02.2026</td>
            <td><span>Документ</span> <a href="/portal/FileLoad.ashx?Fileid=123456">скачать</a></td>
          </tr>
        </table>
        </body></html>
        """
        rows, top_hash = parse_reports_page(
            html=html,
            company_id="225",
            type_id=4,
            type_name="Консолидированная",
            known_state=None,
        )
        self.assertTrue(top_hash)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["placement_date"], "2026-02-10")
        self.assertIn("FileLoad.ashx", rows[0]["file_url"])
        self.assertTrue(rows[0]["file_url"].startswith("https://www.e-disclosure.ru/portal/FileLoad.ashx"))

    def test_parse_reports_page_fallback_to_page_url_without_fileload(self) -> None:
        html = """
        <html><body>
        <table class='zebra'>
          <tr>
            <td>Бухгалтерская (финансовая) отчетность</td>
            <td>2025 год</td>
            <td>15.02.2026</td>
            <td>16.02.2026</td>
            <td><a href='/portal/files.aspx?id=225&type=3'>карточка</a></td>
          </tr>
        </table>
        </body></html>
        """
        rows, _ = parse_reports_page(
            html=html,
            company_id="225",
            type_id=3,
            type_name="Финансовая",
            known_state=None,
        )
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["file_url"], "")
        self.assertTrue(rows[0]["page_url"].endswith("files.aspx?id=225&type=3"))

    def test_missing_history_forces_recheck_even_if_schedule_not_due(self) -> None:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        conn.executescript(SCHEMA_SQL)

        inn = "6316031581"
        future_check = (datetime.now() + timedelta(days=10)).isoformat(timespec="seconds")
        conn.execute(
            """
            INSERT INTO company_map (inn, company_id, company_name, company_url, verified_inn, validation_status, last_success_at, full_scan_at, fast_scan_at, last_checked_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (inn, "225", "TEST", "", inn, "verified", "", "", "", datetime.now().isoformat(timespec="seconds")),
        )
        conn.execute(
            """
            INSERT INTO emitent_schedule (inn, company_id, last_checked_at, next_check_at, last_change_at, stable_run_count, last_mode, last_event_gate_at, last_files_scan_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (inn, "225", datetime.now().isoformat(timespec="seconds"), future_check, "", 5, "event_gate_only", "", ""),
        )
        conn.commit()

        emitents = [
            EmitentRow(
                inn=inn,
                company_name="TEST",
                scoring="",
                scoring_date="",
                nra_rate="",
                acra_rate="",
                nkr_rate="",
                raex_rate="",
            )
        ]

        tasks, skipped, _, _, _, prep = stage_reports_prepare(conn, emitents)
        self.assertEqual(prep["processed_emitents"], 1)
        self.assertEqual(len(tasks), 1)
        self.assertEqual(len(skipped), 0)
        self.assertTrue(tasks[0]["force_missing_recheck"])


    def test_latest_event_by_inn_uses_newest_event_date(self) -> None:
        report_events = [
            {
                "event_hash": "old",
                "inn": "6316031581",
                "event_date": "2024-01-10",
                "event_type": "Опубликована новая отчетность",
                "event_url": "https://www.e-disclosure.ru/portal/files.aspx?id=225&type=4",
                "source": "e-disclosure",
            },
            {
                "event_hash": "new",
                "inn": "6316031581",
                "event_date": "2025-02-15",
                "event_type": "Опубликована новая отчетность",
                "event_url": "https://www.e-disclosure.ru/portal/FileLoad.ashx?Fileid=123456",
                "source": "e-disclosure",
            },
        ]
        latest = build_latest_event_by_inn(report_events)
        self.assertIn("6316031581", latest)
        self.assertEqual(latest["6316031581"]["event_hash"], "new")


    def test_portfolio_export_and_news_include_edisclosure_for_stock_inn(self) -> None:
        portfolio_items = [
            {
                "instrument_type": "Stock",
                "instrument_code": "NVTK",
                "inn": "6316031581",
                "company_name": "Новатэк",
            }
        ]
        report_events = [
            {
                "event_hash": "evt6316",
                "inn": "6316031581",
                "company_name": "Новатэк",
                "scoring_date": "2026-02-10",
                "event_date": "2026-02-10",
                "event_type": "Опубликована новая отчетность",
                "event_url": "https://www.e-disclosure.ru/portal/files.aspx?id=225&type=4",
                "source": "e-disclosure",
                "is_new": True,
            }
        ]
        latest = build_latest_event_by_inn(report_events)
        latest_news_by_key = {
            ("Stock", "NVTK"): {
                "title": "Smartlab headline",
                "news_date": "2026-02-11",
                "url": "https://smart-lab.ru/blog/1.php",
                "is_new": True,
            }
        }
        news_rows = [
            {
                "event_hash": "news6316",
                "instrument_type": "Stock",
                "instrument_code": "NVTK",
                "inn": "6316031581",
                "company_name": "Новатэк",
                "news_date": "2026-02-11",
                "title": "Smartlab headline",
                "url": "https://smart-lab.ru/blog/1.php",
                "source": "Smartlab",
                "is_new": True,
            }
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            old_path = config.PORTFOLIO_XLSX
            old_main_path = monitoring_main.config.PORTFOLIO_XLSX
            config.PORTFOLIO_XLSX = Path(tmp_dir) / "Portfolio.xlsx"
            monitoring_main.config.PORTFOLIO_XLSX = config.PORTFOLIO_XLSX
            try:
                export_portfolio(
                    portfolio_items=portfolio_items,
                    latest_event_by_inn=latest,
                    latest_news_by_key=latest_news_by_key,
                    news_rows=news_rows,
                    report_rows=report_events,
                    logger=logging.getLogger("test_export"),
                )

                wb = load_workbook(config.PORTFOLIO_XLSX, data_only=True)
                ws_all = wb["Portfolio_All"]
                ws_unique = wb["Portfolio_UniqueEmitents"]
                ws_news = wb["News"]

                all_row = next(r for r in ws_all.iter_rows(min_row=2, values_only=True) if str(r[2]) == "6316031581")
                self.assertTrue(all_row[5])
                self.assertTrue(all_row[6])
                self.assertEqual(all_row[7], "e-disclosure")
                self.assertTrue(all_row[8])

                unique_row = next(r for r in ws_unique.iter_rows(min_row=2, values_only=True) if str(r[0]) == "6316031581")
                self.assertTrue(unique_row[5])
                self.assertTrue(unique_row[6])
                self.assertEqual(unique_row[7], "e-disclosure")
                self.assertTrue(unique_row[8])

                has_report = any(
                    str(r[0]) == "Report" and str(r[3]) == "6316031581" and str(r[7]) == "e-disclosure"
                    for r in ws_news.iter_rows(min_row=2, values_only=True)
                )
                self.assertTrue(has_report)
            finally:
                config.PORTFOLIO_XLSX = old_path
                monitoring_main.config.PORTFOLIO_XLSX = old_main_path


if __name__ == "__main__":
    unittest.main()
