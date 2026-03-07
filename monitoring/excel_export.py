from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from . import config


URL_COLUMNS = {"Ссылка", "Ссылка на последнее событие", "Ссылка на последнюю новость"}


def _autowidth(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max_len + 2, config.MAX_EXCEL_COL_WIDTH)


def _common_format(ws, url_headers: set[str]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill(start_color=config.NEW_ITEM_FILL_COLOR, end_color=config.NEW_ITEM_FILL_COLOR, fill_type="solid")
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        for header in url_headers:
            cell = row_dict.get(header)
            if cell and cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
        if row_dict.get("_is_new") and row_dict["_is_new"].value:
            for cell in row:
                cell.fill = fill
    # hide technical column
    for i, head in enumerate(headers, start=1):
        if head == "_is_new":
            ws.column_dimensions[get_column_letter(i)].hidden = True
    _autowidth(ws)


def export_reports_xlsx(events: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    headers = ["ИНН", "Наименование", "Дата скоринга", "Дата события", "Событие", "Ссылка", "_is_new"]
    ws.append(headers)
    rows = sorted(events, key=lambda x: x.get("event_date", ""), reverse=True)
    for row in rows:
        ws.append(
            [
                row.get("inn", ""),
                row.get("company_name", ""),
                row.get("scoring_date", ""),
                row.get("event_date", ""),
                row.get("event_type", ""),
                row.get("event_url", ""),
                "1" if row.get("is_new") else "",
            ]
        )
    _common_format(ws, {"Ссылка"})
    wb.save(config.REPORTS_XLSX)


def export_simple_snapshot(path, sheet_name: str, headers: list[str], rows: list[list[str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _autowidth(ws)
    wb.save(path)
