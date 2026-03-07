from __future__ import annotations

from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from . import config


def _autowidth(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max_len + 2, config.MAX_EXCEL_COL_WIDTH)


def _format(ws, url_cols: set[str], highlight_col: str = "_is_new") -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    fill = PatternFill(start_color=config.NEW_ITEM_FILL_COLOR, end_color=config.NEW_ITEM_FILL_COLOR, fill_type="solid")
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        row_map = {headers[i]: row[i] for i in range(len(headers))}
        for url_col in url_cols:
            cell = row_map.get(url_col)
            if cell and cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
        marker = row_map.get(highlight_col)
        if marker and marker.value:
            for cell in row:
                cell.fill = fill
    for i, header in enumerate(headers, start=1):
        if header == highlight_col:
            ws.column_dimensions[get_column_letter(i)].hidden = True
    _autowidth(ws)


def export_portfolio_xlsx(
    portfolio_items: list[dict[str, str]],
    latest_event_by_inn: dict[str, dict[str, str]],
    latest_news_by_key: dict[tuple[str, str], dict[str, str]],
    news_rows: list[dict[str, str]],
) -> None:
    wb = Workbook()

    ws_all = wb.active
    ws_all.title = "Portfolio_All"
    ws_all.append([
        "Тип", "ISIN / Тикер", "ИНН", "Наименование", "Дата скоринга", "Последнее событие",
        "Дата последнего события", "Ссылка на последнее событие", "Последняя новость",
        "Дата последней новости", "Ссылка на последнюю новость", "_is_new",
    ])

    for item in portfolio_items:
        evt = latest_event_by_inn.get(item.get("inn", ""), {})
        key = (item.get("instrument_type", ""), item.get("instrument_code", ""))
        news = latest_news_by_key.get(key, {})
        ws_all.append([
            item.get("instrument_type", ""),
            item.get("instrument_code", ""),
            item.get("inn", ""),
            item.get("company_name", ""),
            evt.get("scoring_date", ""),
            evt.get("event_type", ""),
            evt.get("event_date", ""),
            evt.get("event_url", ""),
            news.get("title", ""),
            news.get("news_date", ""),
            news.get("url", ""),
            "1" if news.get("is_new") else "",
        ])
    _format(ws_all, {"Ссылка на последнее событие", "Ссылка на последнюю новость"})

    ws_unique = wb.create_sheet("Portfolio_UniqueEmitents")
    ws_unique.append([
        "ИНН", "Наименование", "Кол-во инструментов в портфеле", "Инструменты", "Дата скоринга",
        "Последнее событие", "Дата последнего события", "Ссылка на последнее событие", "Последняя новость",
        "Дата последней новости", "Ссылка на последнюю новость", "_is_new",
    ])
    grouped = defaultdict(list)
    for item in portfolio_items:
        grouped[item.get("inn", "")].append(item)

    for inn, items in grouped.items():
        first = items[0] if items else {}
        evt = latest_event_by_inn.get(inn, {})
        merged_news = sorted(
            [latest_news_by_key.get((it.get("instrument_type", ""), it.get("instrument_code", "")), {}) for it in items],
            key=lambda x: x.get("news_date", ""),
            reverse=True,
        )
        news = merged_news[0] if merged_news else {}
        ws_unique.append([
            inn,
            first.get("company_name", ""),
            len(items),
            ", ".join(it.get("instrument_code", "") for it in items),
            evt.get("scoring_date", ""),
            evt.get("event_type", ""),
            evt.get("event_date", ""),
            evt.get("event_url", ""),
            news.get("title", ""),
            news.get("news_date", ""),
            news.get("url", ""),
            "1" if news.get("is_new") else "",
        ])
    _format(ws_unique, {"Ссылка на последнее событие", "Ссылка на последнюю новость"})

    ws_news = wb.create_sheet("News")
    ws_news.append(["Тип", "ISIN / Тикер", "ИНН", "Наименование", "Дата новости", "Заголовок", "Ссылка", "Источник", "Новое", "_is_new"])
    rows = sorted(news_rows, key=lambda x: x.get("news_date", ""), reverse=True)
    for row in rows:
        ws_news.append([
            row.get("instrument_type", ""),
            row.get("instrument_code", ""),
            row.get("inn", ""),
            row.get("company_name", ""),
            row.get("news_date", ""),
            row.get("title", ""),
            row.get("url", ""),
            row.get("source", "Smartlab"),
            "✓ НОВОЕ" if row.get("is_new") else "",
            "1" if row.get("is_new") else "",
        ])
    _format(ws_news, {"Ссылка"})

    wb.save(config.PORTFOLIO_XLSX)
