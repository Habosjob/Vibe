from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from . import config
from .helpers import md5_short, sanitize_str


def _normalize(name: str) -> str:
    return sanitize_str(name).replace(" ", "").lower()


def find_portfolio_file() -> Path | None:
    if config.PORTFOLIO_SOURCE_FILE:
        explicit = Path(config.PORTFOLIO_SOURCE_FILE)
        if explicit.exists():
            return explicit
    candidates: list[Path] = []
    for pattern in config.PORTFOLIO_GLOBS:
        candidates.extend(Path.cwd().glob(pattern))
    filtered = []
    for path in candidates:
        abs_path = path.resolve()
        if str(config.BASE_DIR.resolve()).lower() in str(abs_path).lower():
            continue
        if abs_path.name in {config.PORTFOLIO_XLSX.name, config.REPORTS_XLSX.name}:
            continue
        filtered.append(abs_path)
    if not filtered:
        return None
    return sorted(filtered, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def _extract_sheet(ws, instrument_type: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    headers = [_normalize(c.value or "") for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    def get(values, *keys: str) -> str:
        for key in keys:
            pos = idx.get(_normalize(key))
            if pos is not None and pos < len(values):
                value = sanitize_str(values[pos])
                if value:
                    return value
        return ""

    for values in ws.iter_rows(min_row=2, values_only=True):
        values = list(values)
        inn = get(values, "ИНН")
        company = get(values, "Наименование эмитента")
        ticker = get(values, "Тикер")
        isin = get(values, "ISIN")
        issuer_ticker = get(values, "Тикер эмитента")

        if instrument_type == "Stock":
            code = ticker
        else:
            code = isin or issuer_ticker or ticker
            if not code:
                code = md5_short(f"{inn}_{company}", 12)

        if not code and not inn and not company:
            continue
        rows.append(
            {
                "instrument_type": instrument_type,
                "instrument_code": code,
                "inn": inn,
                "company_name": company,
            }
        )
    return rows


def load_portfolio_items(source_file: Path | None, logger) -> list[dict[str, str]]:
    if source_file is None or not source_file.exists():
        logger.info("Portfolio source not found, using empty portfolio")
        return []

    items: list[dict[str, str]] = []
    try:
        wb = load_workbook(source_file, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.error("Cannot open portfolio file %s: %s", source_file, exc)
        return []

    for sheet_name, instrument_type in [("Акции", "Stock"), ("Облигации", "Bond")]:
        if sheet_name not in wb.sheetnames:
            logger.warning("Sheet %s not found in %s", sheet_name, source_file)
            continue
        try:
            ws = wb[sheet_name]
            items.extend(_extract_sheet(ws, instrument_type))
        except Exception as exc:  # noqa: BLE001
            logger.error("Failed processing sheet %s: %s", sheet_name, exc)
            continue
    dedup = {(row["instrument_type"], row["instrument_code"]): row for row in items if row.get("instrument_code")}
    return list(dedup.values())
