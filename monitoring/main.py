from __future__ import annotations

import csv
import hashlib
import json
import logging
import random
import re
import sqlite3
import statistics
import sys
import threading
import time
from collections import deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from requests.adapters import HTTPAdapter

# Важно для запуска на Windows как: python monitoring/main.py
CURRENT_DIR = Path(__file__).resolve().parent
if str(CURRENT_DIR) not in sys.path:
    sys.path.insert(0, str(CURRENT_DIR))

import config  # noqa: E402


# -----------------------------
# Базовые утилиты и лог
# -----------------------------
def ensure_dirs() -> None:
    for path in [
        config.CACHE_DIR,
        config.RAW_DIR,
        config.DB_DIR,
        config.LOGS_DIR,
        config.BASE_SNAPSHOTS_DIR,
        config.CACHE_DIR / "edisclosure",
        config.CACHE_DIR / "news",
    ]:
        path.mkdir(parents=True, exist_ok=True)


def setup_logger() -> logging.Logger:
    ensure_dirs()
    logger = logging.getLogger("monitoring")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    handler = logging.FileHandler(config.LOG_FILE, mode="w", encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
    logger.addHandler(handler)
    logger.propagate = False
    return logger


def sanitize_str(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def md5_short(value: str, size: int = 16) -> str:
    return hashlib.md5(value.encode("utf-8", errors="ignore")).hexdigest()[:size]


def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def today_iso() -> str:
    return date.today().isoformat()


def parse_date(value: Any) -> datetime | None:
    text = sanitize_str(value)
    if not text:
        return None
    for fmt in [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d.%m.%Y",
        "%d.%m.%Y %H:%M:%S",
        "%d/%m/%Y",
        "%d/%m/%y",
    ]:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def to_iso_date_str(value: Any) -> str:
    dt = parse_date(value)
    return dt.date().isoformat() if dt else ""


def _extract_status_code_from_error(exc: Exception) -> int | None:
    response = getattr(exc, "response", None)
    if response is not None and getattr(response, "status_code", None):
        return int(response.status_code)
    text = sanitize_str(exc)
    m = re.search(r"HTTP\s+(\d{3})", text)
    return int(m.group(1)) if m else None


def _extract_retry_after_seconds(response: requests.Response | None, exc: Exception | None) -> float | None:
    header = ""
    if response is not None:
        header = sanitize_str(response.headers.get("Retry-After"))
    if not header and exc is not None:
        err_response = getattr(exc, "response", None)
        if err_response is not None:
            header = sanitize_str(err_response.headers.get("Retry-After"))
    if not header:
        return None
    if header.isdigit():
        return float(header)
    dt = parse_date(header)
    if dt is None:
        return None
    return max(0.0, (dt - datetime.now()).total_seconds())


def _detect_endpoint_kind(url: str, request_kind: str | None = None) -> str:
    if request_kind:
        return request_kind
    lower = url.lower()
    if "files.aspx" in lower:
        return "files"
    if "fileload.ashx" in lower:
        return "fileload"
    if "/api/events/page" in lower:
        return "events"
    if "/api/search/companies" in lower:
        return "search"
    if "poisk-po-kompaniyam" in lower or "company.aspx" in lower:
        return "company"
    return "generic"


def request_with_retries(
    session: requests.Session,
    method: str,
    url: str,
    logger: logging.Logger,
    timeout: tuple[float, float] | None = None,
    request_kind: str | None = None,
    **kwargs: Any,
) -> requests.Response:
    timeout = timeout or (config.CONNECT_TIMEOUT_SECONDS, config.READ_TIMEOUT_SECONDS)
    endpoint_kind = _detect_endpoint_kind(url, request_kind)
    is_files_request = endpoint_kind == "files"
    is_search_request = endpoint_kind == "search"
    last_error: Exception | None = None

    for attempt in range(config.HTTP_RETRIES + 1):
        response: requests.Response | None = None
        semaphore_acquired = False
        started = time.perf_counter()
        try:
            if is_search_request:
                search_burst_controller.before_search_request()
            if is_files_request:
                _files_semaphore.acquire()
                semaphore_acquired = True

            response = session.request(method=method, url=url, timeout=timeout, **kwargs)
            if response.status_code == 429 or response.status_code >= 500:
                err = requests.HTTPError(f"HTTP {response.status_code}: {url}")
                err.response = response
                raise err
            latency = time.perf_counter() - started
            runtime_state.register_request_event(RequestTelemetryEvent(endpoint_kind, int(response.status_code), latency, False, False))
            return response
        except Exception as exc:  # noqa: BLE001
            latency = time.perf_counter() - started
            status = _extract_status_code_from_error(exc)
            is_timeout = isinstance(exc, requests.Timeout) or "timed out" in sanitize_str(exc).lower()
            is_retryable_http = status == 429 or (status is not None and status >= 500)
            is_retryable_exception = isinstance(exc, (requests.Timeout, requests.ConnectionError))
            runtime_state.register_request_event(RequestTelemetryEvent(endpoint_kind, int(status or 0), latency, is_timeout, status == 429))
            last_error = exc

            retry_after = _extract_retry_after_seconds(response, exc)
            if is_search_request and (status == 429 or is_timeout):
                search_burst_controller.register_search_error(is_429=status == 429, is_timeout=is_timeout, retry_after=retry_after if status == 429 else None)

            if (not is_retryable_http and not is_retryable_exception) or attempt >= config.HTTP_RETRIES:
                break

            retry_jitter_ms = random.randint(config.EDISCLOSURE_RETRY_JITTER_MIN_MS, config.EDISCLOSURE_RETRY_JITTER_MAX_MS)
            sleep_for = config.BACKOFF_BASE_SECONDS * (2 ** attempt) + retry_jitter_ms / 1000.0
            if status == 429 and retry_after is not None:
                sleep_for = max(sleep_for, min(retry_after, config.HTTP_RETRY_AFTER_MAX_SECONDS))
            elif status == 429:
                sleep_for = max(sleep_for, config.BACKOFF_BASE_SECONDS * 2 * (attempt + 1))
            sleep_for = min(sleep_for, config.HTTP_MAX_BACKOFF_SECONDS)
            logger.warning("Retry %s for %s %s due to %s (sleep %.2fs)", attempt + 1, method, url, exc, sleep_for)
            time.sleep(sleep_for)
        finally:
            if semaphore_acquired:
                _files_semaphore.release()

    raise RuntimeError(f"Request failed: {method} {url}: {last_error}")


def timed(func):
    started = time.perf_counter()
    result = func()
    return result, time.perf_counter() - started


# -----------------------------
# SQLite
# -----------------------------
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS company_map (
    inn TEXT PRIMARY KEY,
    company_id TEXT,
    company_name TEXT,
    company_url TEXT,
    verified_inn TEXT,
    validation_status TEXT,
    last_success_at TEXT,
    full_scan_at TEXT,
    fast_scan_at TEXT,
    last_checked_at TEXT
);
CREATE TABLE IF NOT EXISTS report_events (
    event_hash TEXT PRIMARY KEY,
    inn TEXT,
    company_name TEXT,
    scoring_date TEXT,
    event_date TEXT,
    event_type TEXT,
    event_url TEXT,
    source TEXT,
    payload_json TEXT,
    first_seen_at TEXT,
    last_seen_at TEXT
);
CREATE TABLE IF NOT EXISTS report_state (
    inn TEXT,
    company_id TEXT,
    report_type_id TEXT,
    latest_hash TEXT,
    latest_placement_date TEXT,
    latest_foundation_date TEXT,
    top_row_hash TEXT,
    page_checked_at TEXT,
    last_checked_at TEXT,
    PRIMARY KEY (inn, report_type_id)
);
CREATE TABLE IF NOT EXISTS emitent_schedule (
    inn TEXT PRIMARY KEY,
    company_id TEXT,
    last_checked_at TEXT,
    next_check_at TEXT,
    last_change_at TEXT,
    stable_run_count INTEGER,
    last_mode TEXT,
    last_event_gate_at TEXT,
    last_files_scan_at TEXT
);
CREATE TABLE IF NOT EXISTS emitents_snapshot (
    inn TEXT PRIMARY KEY,
    company_name TEXT,
    scoring TEXT,
    scoring_date TEXT,
    nra_rate TEXT,
    acra_rate TEXT,
    nkr_rate TEXT,
    raex_rate TEXT,
    snapshot_at TEXT
);
CREATE TABLE IF NOT EXISTS ratings_monitoring_snapshot (
    inn TEXT,
    source TEXT,
    rating TEXT,
    assigned_date TEXT,
    loaded_at TEXT,
    PRIMARY KEY (inn, source)
);
CREATE TABLE IF NOT EXISTS news_events (
    event_hash TEXT PRIMARY KEY,
    instrument_type TEXT,
    instrument_code TEXT,
    inn TEXT,
    company_name TEXT,
    news_date TEXT,
    title TEXT,
    url TEXT,
    source TEXT,
    first_seen_at TEXT
);
CREATE TABLE IF NOT EXISTS portfolio_items (
    instrument_type TEXT,
    instrument_code TEXT,
    inn TEXT,
    company_name TEXT,
    source_file TEXT,
    loaded_at TEXT,
    PRIMARY KEY (instrument_type, instrument_code)
);
CREATE TABLE IF NOT EXISTS meta (
    key TEXT PRIMARY KEY,
    value TEXT
);
"""


def db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(config.DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA temp_store=MEMORY")
    conn.execute("PRAGMA foreign_keys=ON")
    conn.executescript(SCHEMA_SQL)
    company_cols = {r[1] for r in conn.execute("PRAGMA table_info(company_map)").fetchall()}
    for col in ["full_scan_at", "fast_scan_at", "verified_inn", "validation_status", "last_success_at"]:
        if col not in company_cols:
            conn.execute(f"ALTER TABLE company_map ADD COLUMN {col} TEXT")
    report_state_cols = {r[1] for r in conn.execute("PRAGMA table_info(report_state)").fetchall()}
    for col in ["top_row_hash", "page_checked_at"]:
        if col not in report_state_cols:
            conn.execute(f"ALTER TABLE report_state ADD COLUMN {col} TEXT")
    schedule_cols = {r[1] for r in conn.execute("PRAGMA table_info(emitent_schedule)").fetchall()}
    for col in ["company_id", "last_checked_at", "next_check_at", "last_change_at", "stable_run_count", "last_mode", "last_event_gate_at", "last_files_scan_at"]:
        if col not in schedule_cols:
            col_type = "INTEGER" if col == "stable_run_count" else "TEXT"
            conn.execute(f"ALTER TABLE emitent_schedule ADD COLUMN {col} {col_type}")
    conn.commit()
    return conn


# -----------------------------
# E-disclosure
# -----------------------------


_thread_local = threading.local()
_files_semaphore = threading.BoundedSemaphore(max(1, int(config.EDISCLOSURE_FILES_SEMAPHORE_DEFAULT)))
_current_workers = max(1, int(config.EDISCLOSURE_FETCH_WORKERS_DEFAULT))
_current_files_semaphore = max(1, int(config.EDISCLOSURE_FILES_SEMAPHORE_DEFAULT))


def configure_runtime_concurrency(workers: int, files_semaphore: int) -> None:
    global _files_semaphore, _current_workers, _current_files_semaphore
    _current_workers = max(1, int(workers))
    _current_files_semaphore = max(1, int(files_semaphore))
    _files_semaphore = threading.BoundedSemaphore(_current_files_semaphore)


@dataclass
class RequestTelemetryEvent:
    endpoint: str
    status_code: int
    latency_sec: float
    is_timeout: bool
    is_429: bool


class MonitoringRuntimeState:
    def __init__(self) -> None:
        self.lock = threading.Lock()
        self.total_requests = 0
        self.files_requests = 0
        self.search_requests = 0
        self.events_requests = 0
        self.fileload_requests = 0
        self.status_429 = 0
        self.timeout_count = 0
        self.request_latencies: list[float] = []
        self.max_consecutive_429 = 0
        self.max_consecutive_timeouts = 0
        self._consecutive_429 = 0
        self._consecutive_timeouts = 0

    def register_request_event(self, event: RequestTelemetryEvent) -> None:
        with self.lock:
            self.total_requests += 1
            self.request_latencies.append(event.latency_sec)
            if event.endpoint == "files":
                self.files_requests += 1
            elif event.endpoint == "search":
                self.search_requests += 1
            elif event.endpoint == "events":
                self.events_requests += 1
            elif event.endpoint == "fileload":
                self.fileload_requests += 1
            if event.is_429:
                self.status_429 += 1
                self._consecutive_429 += 1
                self.max_consecutive_429 = max(self.max_consecutive_429, self._consecutive_429)
            else:
                self._consecutive_429 = 0
            if event.is_timeout:
                self.timeout_count += 1
                self._consecutive_timeouts += 1
                self.max_consecutive_timeouts = max(self.max_consecutive_timeouts, self._consecutive_timeouts)
            else:
                self._consecutive_timeouts = 0


runtime_state = MonitoringRuntimeState()


class SearchBurstController:
    def __init__(self) -> None:
        self.lock = threading.Lock()
        self.cooldown_until = 0.0
        self.error_window: deque[tuple[float, str]] = deque()
        self.search_429_count = 0
        self.search_timeout_count = 0
        self.search_cooldown_events = 0
        self.total_search_cooldown_seconds = 0.0

    def _cleanup(self, now_ts: float) -> None:
        window = max(1.0, float(config.SEARCH_BURST_WINDOW_SECONDS))
        while self.error_window and now_ts - self.error_window[0][0] > window:
            self.error_window.popleft()

    def before_search_request(self) -> None:
        with self.lock:
            now_ts = time.time()
            sleep_for = max(0.0, self.cooldown_until - now_ts)
        if sleep_for > 0:
            time.sleep(sleep_for)

    def register_search_error(self, *, is_429: bool, is_timeout: bool, retry_after: float | None = None) -> None:
        if not is_429 and not is_timeout:
            return
        with self.lock:
            now_ts = time.time()
            self._cleanup(now_ts)
            if is_429:
                self.search_429_count += 1
                self.error_window.append((now_ts, "429"))
            if is_timeout:
                self.search_timeout_count += 1
                self.error_window.append((now_ts, "timeout"))

            window_429 = sum(1 for _, kind in self.error_window if kind == "429")
            window_timeout = sum(1 for _, kind in self.error_window if kind == "timeout")
            need_cooldown = (
                window_429 >= max(1, int(config.SEARCH_BURST_429_THRESHOLD))
                or window_timeout >= max(1, int(config.SEARCH_BURST_TIMEOUT_THRESHOLD))
            )
            if not need_cooldown:
                return

            cooldown = float(config.SEARCH_COOLDOWN_SECONDS)
            if retry_after is not None:
                cooldown = max(cooldown, float(retry_after))
            cooldown = max(0.1, min(float(config.SEARCH_COOLDOWN_MAX_SECONDS), cooldown))
            new_until = now_ts + cooldown
            if new_until > self.cooldown_until:
                self.cooldown_until = new_until
            self.search_cooldown_events += 1
            self.total_search_cooldown_seconds += cooldown

    def snapshot(self) -> dict[str, Any]:
        with self.lock:
            return {
                "search_429_count": self.search_429_count,
                "search_timeout_count": self.search_timeout_count,
                "search_cooldown_events": self.search_cooldown_events,
                "total_search_cooldown_seconds": round(self.total_search_cooldown_seconds, 3),
            }


search_burst_controller = SearchBurstController()


def reset_runtime_state() -> None:
    global runtime_state, search_burst_controller
    runtime_state = MonitoringRuntimeState()
    search_burst_controller = SearchBurstController()


def get_thread_local_edisclosure_client(logger: logging.Logger) -> "EDisclosureClient":
    client = getattr(_thread_local, "edisclosure_client", None)
    if client is None:
        client = EDisclosureClient(logger)
        _thread_local.edisclosure_client = client
    return client

class EDisclosureClient:
    def __init__(self, logger: logging.Logger):
        self.logger = logger
        self.session = requests.Session()
        adapter = HTTPAdapter(
            pool_connections=max(_current_workers, 8),
            pool_maxsize=max(_current_workers * 2, 16),
            max_retries=0,
        )
        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)
        self.session.headers.update(
            {
                "User-Agent": config.BROWSER_USER_AGENT,
                "Accept": "application/json, text/javascript, */*; q=0.01",
                "Accept-Language": "ru,en;q=0.9",
                "Accept-Encoding": "gzip, deflate, br",
                "Connection": "keep-alive",
                "Origin": "https://www.e-disclosure.ru",
                "Referer": "https://www.e-disclosure.ru/poisk-po-kompaniyam",
                "X-Requested-With": "XMLHttpRequest",
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            }
        )

    def _cache_file(self, company_id: str, data_type: str) -> Path:
        return config.CACHE_DIR / "edisclosure" / f"{md5_short(f'{company_id}_{data_type}', 10)}.json"

    def _load_cache(self, company_id: str, data_type: str, ttl_hours: int) -> dict[str, Any] | None:
        path = self._cache_file(company_id, data_type)
        if not path.exists():
            return None
        age = time.time() - path.stat().st_mtime
        if age > ttl_hours * 3600:
            return None
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return None

    def _save_cache(self, company_id: str, data_type: str, payload: dict[str, Any]) -> None:
        path = self._cache_file(company_id, data_type)
        path.parent.mkdir(parents=True, exist_ok=True)
        content = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        if path.exists():
            try:
                if path.read_text(encoding="utf-8") == content:
                    return
            except Exception:  # noqa: BLE001
                pass
        tmp = path.with_suffix(path.suffix + ".tmp")
        tmp.write_text(content, encoding="utf-8")
        tmp.replace(path)

    def search_company_by_inn(self, inn: str) -> list[dict[str, str]]:
        payload = {
            "textfield": inn,
            "radReg": "FederalDistricts",
            "districtsCheckboxGroup": "-1",
            "regionsCheckboxGroup": "-1",
            "branchesCheckboxGroup": "-1",
            "lastPageSize": "10",
            "lastPageNumber": "1",
            "query": inn,
            "mode": "companies",
        }
        response = request_with_retries(
            self.session,
            "POST",
            "https://www.e-disclosure.ru/api/search/companies",
            self.logger,
            data=payload,
        )
        data = response.json() if response.text else {}
        items = data.get("foundCompaniesList") or []
        out = []
        for row in items:
            company_id = sanitize_str(row.get("id"))
            if not company_id:
                continue
            out.append(
                {
                    "id": company_id,
                    "name": sanitize_str(row.get("name")),
                    "district": sanitize_str(row.get("district")),
                    "region": sanitize_str(row.get("region")),
                    "branch": sanitize_str(row.get("branch")),
                    "lastActivity": sanitize_str(row.get("lastActivity")),
                    "docCount": sanitize_str(row.get("docCount")),
                    "url": f"https://www.e-disclosure.ru/portal/company.aspx?id={company_id}",
                }
            )
        return out

    def get_company_card(self, company_id: str) -> dict[str, str]:
        cached = self._load_cache(company_id, "card", config.EDISCLOSURE_CARD_TTL_HOURS)
        if cached:
            return cached
        url = f"https://www.e-disclosure.ru/portal/company.aspx?id={company_id}"
        html = request_with_retries(self.session, "GET", url, self.logger).text
        inn_match = re.search(r"ИНН\s*:?\s*(\d{10,12})", html, flags=re.IGNORECASE)
        ogrn_match = re.search(r"ОГРН\s*:?\s*(\d{13,15})", html, flags=re.IGNORECASE)
        reg_match = re.search(r"Дата\s+регистрац(?:ии|ии:)\s*:?\s*(\d{2}[./]\d{2}[./]\d{4})", html, flags=re.IGNORECASE)
        soup = BeautifulSoup(html, "lxml")
        address = ""
        for tr in soup.select("tr"):
            cols = tr.find_all("td")
            if len(cols) < 2:
                continue
            key = sanitize_str(cols[0].get_text(" ", strip=True)).lower()
            if "адрес" in key:
                address = sanitize_str(cols[1].get_text(" ", strip=True))
                break
        card = {
            "inn": sanitize_str(inn_match.group(1)) if inn_match else "",
            "ogrn": sanitize_str(ogrn_match.group(1)) if ogrn_match else "",
            "registration_date": sanitize_str(reg_match.group(1)) if reg_match else "",
            "address": address,
            "url": url,
        }
        self._save_cache(company_id, "card", card)
        return card

    def choose_best_candidate(self, inn: str, candidates: list[dict[str, str]], company_name: str) -> dict[str, str] | None:
        if not candidates:
            return None
        for candidate in candidates:
            card = self.get_company_card(candidate["id"])
            if sanitize_str(card.get("inn")) == sanitize_str(inn):
                return candidate
        low_name = sanitize_str(company_name).lower()
        ranked = sorted(
            candidates,
            key=lambda x: (
                1 if low_name and low_name in x.get("name", "").lower() else 0,
                int(sanitize_str(x.get("docCount", "0")) or "0"),
                sanitize_str(x.get("lastActivity", "")),
            ),
            reverse=True,
        )
        return ranked[0]

    def get_reports_page_cached(self, company_id: str, type_id: int, force_refresh: bool = False) -> str:
        cache_key = f"reports_page_{type_id}"
        if not force_refresh:
            cached = self._load_cache(company_id, cache_key, config.EDISCLOSURE_REPORTS_TTL_HOURS)
            if cached and isinstance(cached.get("html"), str):
                return cached["html"]
        page_url = f"https://www.e-disclosure.ru/portal/files.aspx?id={company_id}&type={type_id}"
        html = request_with_retries(self.session, "GET", page_url, self.logger, request_kind="files").text
        self._save_cache(company_id, cache_key, {"html": html})
        return html

    def get_company_events(self, company_id: str, days_back: int = 60) -> list[dict[str, str]]:
        current_year = datetime.now().year
        min_date = datetime.now() - timedelta(days=max(1, days_back))
        rows: list[dict[str, str]] = []
        for year in [current_year, current_year - 1]:
            cache_key = f"events_{year}"
            cached = self._load_cache(company_id, cache_key, config.EDISCLOSURE_EVENTS_TTL_HOURS)
            payload = None
            if cached and isinstance(cached.get("payload"), dict):
                payload = cached["payload"]
            if payload is None:
                url = f"https://www.e-disclosure.ru/api/events/page?companyId={company_id}&year={year}"
                resp = request_with_retries(self.session, "GET", url, self.logger, request_kind="events")
                payload = resp.json() if resp.text else {}
                self._save_cache(company_id, cache_key, {"payload": payload})
            items = payload.get("events") or payload.get("items") or payload.get("data") or []
            if not isinstance(items, list):
                continue
            for item in items:
                pseudo_guid = sanitize_str(item.get("pseudoGUID") or item.get("eventId") or item.get("id"))
                event_name = sanitize_str(item.get("eventName") or item.get("name"))
                event_date = to_iso_date_str(item.get("eventDate"))
                pub_date = to_iso_date_str(item.get("pubDate") or item.get("publishDate"))
                if not pseudo_guid:
                    continue
                dt = parse_date(pub_date or event_date)
                if dt and dt < min_date:
                    continue
                rows.append({
                    "pseudoGUID": pseudo_guid,
                    "eventName": event_name,
                    "eventDate": event_date,
                    "pubDate": pub_date,
                    "isCorrectedByAnotherEvent": sanitize_str(item.get("isCorrectedByAnotherEvent")),
                    "event_url": f"https://www.e-disclosure.ru/portal/event.aspx?EventId={pseudo_guid}",
                })
        return rows

    def get_financial_reports(self, company_id: str) -> list[dict[str, str]]:
        cached = self._load_cache(company_id, "reports", config.EDISCLOSURE_REPORTS_TTL_HOURS)
        if cached and isinstance(cached.get("items"), list):
            return cached["items"]

        report_types = {2: "Годовая", 3: "Финансовая", 4: "Консолидированная", 5: "Отчет эмитента"}
        keywords = ("отчет", "бухгалтер", "финанс", "баланс", "прибыль", "убыток", "аудитор", "годовой", "промежуточный")
        rows: list[dict[str, str]] = []

        for type_id, type_name in report_types.items():
            page_url = f"https://www.e-disclosure.ru/portal/files.aspx?id={company_id}&type={type_id}"
            try:
                html = request_with_retries(self.session, "GET", page_url, self.logger).text
            except Exception as exc:  # noqa: BLE001
                self.logger.warning("files load failed company=%s type=%s: %s", company_id, type_id, exc)
                continue
            soup = BeautifulSoup(html, "lxml")
            table = soup.find("table", class_="zebra")
            if not table:
                continue
            for tr in table.select("tr"):
                tds = tr.find_all("td")
                if len(tds) < 4:
                    continue
                doc_type = sanitize_str(tds[0].get_text(" ", strip=True))
                period = sanitize_str(tds[1].get_text(" ", strip=True)) if len(tds) > 1 else ""
                foundation_date = to_iso_date_str(tds[2].get_text(" ", strip=True)) if len(tds) > 2 else ""
                placement_date = to_iso_date_str(tds[3].get_text(" ", strip=True)) if len(tds) > 3 else ""
                if not (any(k in doc_type.lower() for k in keywords) or period):
                    continue
                anchor = (tds[4] if len(tds) > 4 else tds[-1]).find("a", href=True)
                file_url = ""
                if anchor:
                    href = sanitize_str(anchor.get("href"))
                    if href.startswith("/"):
                        href = f"https://www.e-disclosure.ru{href}"
                    if "FileLoad.ashx" in href:
                        file_url = href
                rows.append(
                    {
                        "hash": md5_short(f"{company_id}_{type_id}_{doc_type}_{period}_{placement_date}", 16),
                        "company_id": company_id,
                        "type_id": str(type_id),
                        "report_type": type_name,
                        "doc_type": doc_type,
                        "period": period,
                        "foundation_date": foundation_date,
                        "placement_date": placement_date,
                        "file_url": file_url,
                        "page_url": page_url,
                    }
                )
        dedup = list({x["hash"]: x for x in rows}.values())
        dedup.sort(key=lambda x: x.get("placement_date") or x.get("foundation_date") or "", reverse=True)
        self._save_cache(company_id, "reports", {"items": dedup})
        return dedup


# -----------------------------
# Emitents snapshot + rating changes
# -----------------------------
OUTLOOK_MARKERS = {
    "позитив",
    "положитель",
    "stable",
    "стабиль",
    "negative",
    "негатив",
    "developing",
    "развива",
    "positive",
    "watch",
    "revision",
}


@dataclass
class EmitentRow:
    inn: str
    company_name: str
    scoring: str
    scoring_date: str
    nra_rate: str
    acra_rate: str
    nkr_rate: str
    raex_rate: str


def load_emitents_rows(path: Path) -> list[EmitentRow]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [sanitize_str(c.value).replace(" ", "").lower() for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    def get(values: list[Any], key: str) -> str:
        pos = idx.get(key)
        if pos is None or pos >= len(values):
            return ""
        value = values[pos]
        if hasattr(value, "isoformat"):
            try:
                return value.date().isoformat()
            except Exception:  # noqa: BLE001
                return value.isoformat()
        return sanitize_str(value)

    result: list[EmitentRow] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        values = list(raw)
        inn = get(values, "inn")
        if not inn:
            continue
        result.append(
            EmitentRow(
                inn=inn,
                company_name=get(values, "emitentname"),
                scoring=get(values, "scoring"),
                scoring_date=get(values, "datescoring"),
                nra_rate=get(values, "nra_rate"),
                acra_rate=get(values, "acra_rate"),
                nkr_rate=get(values, "nkr_rate"),
                raex_rate=get(values, "raex_rate"),
            )
        )
    return result


def save_emitents_snapshot_excel(rows: list[EmitentRow]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "emitents_snapshot"
    ws.append(["INN", "EMITENTNAME", "Scoring", "DateScoring", "NRA_Rate", "Acra_Rate", "NKR_Rate", "RAEX_Rate", "SnapshotAt"])
    for row in rows:
        ws.append([row.inn, row.company_name, row.scoring, row.scoring_date, row.nra_rate, row.acra_rate, row.nkr_rate, row.raex_rate, today_iso()])
    wb.save(config.EMITENTS_SNAPSHOT_XLSX)


def split_rating_and_outlook(text: str) -> tuple[str, str]:
    low = sanitize_str(text).lower()
    if not low:
        return "", ""
    outlook = [m for m in OUTLOOK_MARKERS if m in low]
    rate_part = low
    for m in outlook:
        rate_part = rate_part.replace(m, "")
    return " ".join(rate_part.split()), "|".join(sorted(outlook))


def classify_rating_change(old: str, new: str) -> str | None:
    old_clean, new_clean = sanitize_str(old), sanitize_str(new)
    if old_clean == new_clean:
        return None
    if old_clean and not new_clean:
        return "Рейтинг отозван / снят"
    old_rate, old_outlook = split_rating_and_outlook(old_clean)
    new_rate, new_outlook = split_rating_and_outlook(new_clean)
    if old_rate == new_rate and old_outlook != new_outlook:
        return "Изменен прогноз"
    return "Изменен рейтинг"


def _normalize_rating_row_inn(row: sqlite3.Row, source: str) -> str:
    if source == "NKR":
        return sanitize_str(row["tin"]) if "tin" in row.keys() else ""
    return sanitize_str(row["inn"]) if "inn" in row.keys() else ""


def _pick_rating_value(row: sqlite3.Row) -> str:
    keys = ("rating", "value", "rate", "rating_value", "assigned_rating")
    for key in keys:
        if key in row.keys():
            value = sanitize_str(row[key])
            if value:
                return value
    return ""


def _pick_rating_date(row: sqlite3.Row) -> str:
    keys = (
        "rating_date",
        "assigned_date",
        "date_assigned",
        "date",
        "published_at",
        "pub_date",
        "created_at",
    )
    for key in keys:
        if key in row.keys():
            iso = to_iso_date_str(row[key])
            if iso:
                return iso
    return ""


def load_ratings_snapshot_from_db(logger: logging.Logger) -> dict[tuple[str, str], dict[str, str]]:
    db_path = Path(config.RATINGS_DB_FILE)
    if not db_path.exists():
        logger.warning("Ratings DB not found: %s", db_path)
        return {}

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    result: dict[tuple[str, str], dict[str, str]] = {}
    try:
        for source, table in config.RATINGS_SOURCE_TABLES.items():
            exists = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                (table,),
            ).fetchone()
            if not exists:
                logger.warning("Ratings table not found in DB: %s", table)
                continue

            rows = conn.execute(f'SELECT * FROM "{table}"').fetchall()
            for row in rows:
                inn = _normalize_rating_row_inn(row, source)
                if not inn:
                    continue
                rating = _pick_rating_value(row)
                if not rating:
                    continue
                assigned_date = _pick_rating_date(row)
                key = (inn, source)
                prev = result.get(key)
                if not prev or assigned_date >= prev.get("assigned_date", ""):
                    result[key] = {
                        "inn": inn,
                        "source": source,
                        "rating": rating,
                        "assigned_date": assigned_date,
                    }
    finally:
        conn.close()
    return result


# -----------------------------
# Portfolio loader
# -----------------------------
def ensure_portfolio_workbook(path: Path, logger: logging.Logger) -> None:
    manual_headers = {
        "Акции": ["Тикер", "Наименование эмитента", "ИНН"],
        "Облигации": ["ISIN", "Наименование эмитента", "ИНН"],
    }
    if path.exists():
        wb = load_workbook(path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Акции"
        ws.append(manual_headers["Акции"])
    for sheet_name, headers in manual_headers.items():
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        elif wb[sheet_name].max_row < 1:
            wb[sheet_name].append(headers)
    wb.save(path)
    logger.info("Portfolio workbook ready: %s", path)


def load_portfolio_items(path: Path, logger: logging.Logger) -> list[dict[str, str]]:
    ensure_portfolio_workbook(path, logger)
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.error("Cannot open portfolio file %s: %s", path, exc)
        return []

    def norm(name: str) -> str:
        return sanitize_str(name).replace(" ", "").lower()

    def parse_sheet(ws, instrument_type: str) -> list[dict[str, str]]:
        items = []
        headers = [norm(c.value or "") for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}

        def get(values, *keys: str) -> str:
            for key in keys:
                pos = idx.get(norm(key))
                if pos is not None and pos < len(values):
                    value = sanitize_str(values[pos])
                    if value:
                        return value
            return ""

        for values in ws.iter_rows(min_row=2, values_only=True):
            values = list(values)
            inn = get(values, "ИНН")
            company_name = get(values, "Наименование эмитента")
            code = get(values, "Тикер") if instrument_type == "Stock" else get(values, "ISIN")
            if not code and not inn and not company_name:
                continue
            if not code:
                continue
            items.append(
                {
                    "instrument_type": instrument_type,
                    "instrument_code": code,
                    "inn": inn,
                    "company_name": company_name,
                }
            )
        return items

    result: list[dict[str, str]] = []
    for sheet_name, instrument_type in [("Акции", "Stock"), ("Облигации", "Bond")]:
        try:
            result.extend(parse_sheet(wb[sheet_name], instrument_type))
        except Exception as exc:  # noqa: BLE001
            logger.error("Failed to process sheet %s: %s", sheet_name, exc)
            continue
    return list({(x["instrument_type"], x["instrument_code"]): x for x in result if x.get("instrument_code")}.values())


# -----------------------------
# News
# -----------------------------
class NewsCacheManager:
    def __init__(self, cache_path: Path):
        self.cache_path = cache_path
        self.cache_path.parent.mkdir(parents=True, exist_ok=True)
        self.rows = self._load()
        self.known_hashes = {row["hash"] for row in self.rows if row.get("hash")}

    def _load(self) -> list[dict[str, str]]:
        if not self.cache_path.exists():
            return []
        with self.cache_path.open("r", encoding="utf-8", newline="") as f:
            return list(csv.DictReader(f))

    def is_new(self, hash_value: str) -> bool:
        return hash_value not in self.known_hashes

    def add(self, row: dict[str, str]) -> None:
        if row["hash"] in self.known_hashes:
            return
        self.rows.append(row)
        self.known_hashes.add(row["hash"])

    def save(self) -> None:
        with self.cache_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=["hash", "company_name", "company_inn", "date", "title", "source", "url", "added_date"],
            )
            writer.writeheader()
            for row in self.rows:
                writer.writerow(row)


class SmartlabNewsCollector:
    def __init__(self, logger: logging.Logger):
        self.logger = logger
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": config.BROWSER_USER_AGENT,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "ru,en;q=0.9",
                "Connection": "keep-alive",
            }
        )

    def _normalize_date(self, text: str) -> datetime:
        now = datetime.now()
        t = sanitize_str(text)
        try:
            if "/" in t and len(t) <= 5:
                d, m = t.split("/")
                dt = datetime(year=now.year, month=int(m), day=int(d))
                if dt.date() > now.date():
                    dt = dt.replace(year=now.year - 1)
                return dt
            if ":" in t and len(t) <= 5:
                hh, mm = t.split(":")
                return now.replace(hour=int(hh), minute=int(mm), second=0, microsecond=0)
            for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(t, fmt)
                except ValueError:
                    pass
        except Exception:  # noqa: BLE001
            pass
        return now

    def _parse_news_lines(self, html: str) -> list[dict[str, str]]:
        rows = []
        soup = BeautifulSoup(html, "lxml")
        for block in soup.select("div.news__line")[:50]:
            date_node = block.select_one("div.news__date")
            link_node = block.select_one("div.news__link > a")
            if not link_node:
                continue
            title = sanitize_str(link_node.get_text(" ", strip=True))
            href = sanitize_str(link_node.get("href"))
            if href.startswith("/"):
                href = f"https://smartlab.news{href}"
            dt = self._normalize_date(date_node.get_text(" ", strip=True) if date_node else "")
            if dt < datetime.now() - timedelta(days=config.NEWS_DAYS_BACK):
                continue
            rows.append({"title": title, "url": href, "news_date": dt.date().isoformat()})
        return rows

    def _tag_name(self, company_name: str) -> str:
        text = sanitize_str(company_name).lower().replace('"', "")
        for token in ["пао", "ао", "ооо", "зао", "публичное акционерное общество", "акционерное общество"]:
            text = text.replace(token, "")
        return quote(text.strip())

    def _relevant_title(self, title: str, company_name: str) -> bool:
        stop_words = {"пао", "ао", "ооо", "зао", "публичное", "акционерное", "общество"}
        words = [w for w in sanitize_str(company_name).lower().replace('"', "").split() if len(w) > 2 and w not in stop_words]
        low_title = sanitize_str(title).lower()
        return any(w in low_title for w in words) if words else True

    def collect(self, item: dict[str, str]) -> list[dict[str, str]]:
        ticker = item.get("instrument_code", "")
        company_name = item.get("company_name", "")
        result: list[dict[str, str]] = []

        if ticker:
            try:
                html = request_with_retries(self.session, "GET", f"https://smartlab.news/company/{ticker}", self.logger).text
                result = self._parse_news_lines(html)
            except Exception as exc:  # noqa: BLE001
                self.logger.warning("Smartlab ticker strategy failed %s: %s", ticker, exc)
            time.sleep(config.NEWS_REQUEST_PAUSE_SECONDS)

        if not result:
            tag = self._tag_name(company_name)
            try:
                html = request_with_retries(self.session, "GET", f"https://smartlab.news/tag/{tag}", self.logger).text
                result = [x for x in self._parse_news_lines(html) if self._relevant_title(x["title"], company_name)]
            except Exception as exc:  # noqa: BLE001
                self.logger.warning("Smartlab fallback failed %s: %s", company_name, exc)
            time.sleep(config.NEWS_REQUEST_PAUSE_SECONDS)

        return result


# -----------------------------
# Excel exporters
# -----------------------------
def apply_ws_style(ws, url_headers: set[str]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    headers = [c.value for c in ws[1]]
    fill = PatternFill(start_color=config.NEW_ITEM_FILL_COLOR, end_color=config.NEW_ITEM_FILL_COLOR, fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        row_map = {headers[i]: row[i] for i in range(len(headers))}
        for h in url_headers:
            c = row_map.get(h)
            if c and c.value:
                c.hyperlink = str(c.value)
                c.style = "Hyperlink"
        if row_map.get("_is_new") and row_map["_is_new"].value:
            for c in row:
                c.fill = fill

    for i, header in enumerate(headers, start=1):
        if header == "_is_new":
            ws.column_dimensions[get_column_letter(i)].hidden = True

    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max_len + 2, config.MAX_EXCEL_COL_WIDTH)


def export_reports(events: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    ws.append(["ИНН", "Наименование", "Дата скоринга", "Дата события", "Событие", "Ссылка", "_is_new"])
    for row in sorted(events, key=lambda x: x.get("event_date", ""), reverse=True):
        ws.append([
            row.get("inn", ""),
            row.get("company_name", ""),
            row.get("scoring_date", ""),
            row.get("event_date", ""),
            row.get("event_type", ""),
            row.get("event_url", ""),
            "1" if row.get("is_new") else "",
        ])
    apply_ws_style(ws, {"Ссылка"})
    wb.save(config.REPORTS_XLSX)




def build_latest_event_by_inn(report_events: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    ordered = sorted(
        report_events,
        key=lambda x: (parse_date(x.get("event_date")) or datetime.min, sanitize_str(x.get("event_hash"))),
        reverse=True,
    )
    latest_event_by_inn: dict[str, dict[str, str]] = {}
    for row in ordered:
        inn = sanitize_str(row.get("inn", ""))
        if not inn:
            continue
        latest_event_by_inn.setdefault(inn, row)
    return latest_event_by_inn
def export_portfolio(
    portfolio_items: list[dict[str, str]],
    latest_event_by_inn: dict[str, dict[str, str]],
    latest_news_by_key: dict[tuple[str, str], dict[str, str]],
    news_rows: list[dict[str, str]],
    report_rows: list[dict[str, str]],
) -> None:
    ensure_portfolio_workbook(config.PORTFOLIO_XLSX, logging.getLogger("monitoring"))
    wb = load_workbook(config.PORTFOLIO_XLSX)

    for sheet_name in ["Portfolio_All", "Portfolio_UniqueEmitents", "News"]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]

    ws_all = wb.create_sheet("Portfolio_All")
    ws_all.append([
        "Тип", "ISIN / Тикер", "ИНН", "Наименование", "Дата скоринга", "Последнее событие",
        "Дата последнего события", "Источник события", "Ссылка на последнее событие", "Последняя новость", "Дата последней новости",
        "Ссылка на последнюю новость", "_is_new",
    ])
    for item in portfolio_items:
        evt = latest_event_by_inn.get(item.get("inn", ""), {})
        news = latest_news_by_key.get((item.get("instrument_type", ""), item.get("instrument_code", "")), {})
        ws_all.append([
            item.get("instrument_type", ""),
            item.get("instrument_code", ""),
            item.get("inn", ""),
            item.get("company_name", ""),
            evt.get("scoring_date", ""),
            evt.get("event_type", ""),
            evt.get("event_date", ""),
            evt.get("source", ""),
            evt.get("event_url", ""),
            news.get("title", ""),
            news.get("news_date", ""),
            news.get("url", ""),
            "1" if news.get("is_new") else "",
        ])
    apply_ws_style(ws_all, {"Ссылка на последнее событие", "Ссылка на последнюю новость"})

    ws_unique = wb.create_sheet("Portfolio_UniqueEmitents")
    ws_unique.append([
        "ИНН", "Наименование", "Кол-во инструментов в портфеле", "Инструменты", "Дата скоринга",
        "Последнее событие", "Дата последнего события", "Источник события", "Ссылка на последнее событие", "Последняя новость",
        "Дата последней новости", "Ссылка на последнюю новость", "_is_new",
    ])
    grouped: dict[str, list[dict[str, str]]] = {}
    for item in portfolio_items:
        grouped.setdefault(item.get("inn", ""), []).append(item)
    for inn, items in grouped.items():
        first = items[0] if items else {}
        evt = latest_event_by_inn.get(inn, {})
        news_candidates = [latest_news_by_key.get((x.get("instrument_type", ""), x.get("instrument_code", "")), {}) for x in items]
        news_candidates.sort(key=lambda x: x.get("news_date", ""), reverse=True)
        news = news_candidates[0] if news_candidates else {}
        ws_unique.append([
            inn,
            first.get("company_name", ""),
            len(items),
            ", ".join(x.get("instrument_code", "") for x in items),
            evt.get("scoring_date", ""),
            evt.get("event_type", ""),
            evt.get("event_date", ""),
            evt.get("source", ""),
            evt.get("event_url", ""),
            news.get("title", ""),
            news.get("news_date", ""),
            news.get("url", ""),
            "1" if news.get("is_new") else "",
        ])
    apply_ws_style(ws_unique, {"Ссылка на последнее событие", "Ссылка на последнюю новость"})

    ws_news = wb.create_sheet("News")
    ws_news.append(["Тип", "ISIN / Тикер", "ИНН", "Наименование", "Дата новости", "Заголовок", "Ссылка", "Источник", "Новое", "_is_new"])

    instruments_by_inn: dict[str, list[dict[str, str]]] = {}
    for item in portfolio_items:
        inn = sanitize_str(item.get("inn", ""))
        if not inn:
            continue
        instruments_by_inn.setdefault(inn, []).append(item)

    merged_rows: list[dict[str, str]] = []
    for row in news_rows:
        merged_rows.append(
            {
                "instrument_type": sanitize_str(row.get("instrument_type", "")),
                "instrument_code": sanitize_str(row.get("instrument_code", "")),
                "inn": sanitize_str(row.get("inn", "")),
                "company_name": sanitize_str(row.get("company_name", "")),
                "news_date": sanitize_str(row.get("news_date", "")),
                "title": sanitize_str(row.get("title", "")),
                "url": sanitize_str(row.get("url", "")),
                "source": sanitize_str(row.get("source", "Smartlab")) or "Smartlab",
                "is_new": "1" if row.get("is_new") else "",
            }
        )

    for row in report_rows:
        inn = sanitize_str(row.get("inn", ""))
        linked_items = instruments_by_inn.get(inn)
        if not linked_items:
            continue
        for linked_item in linked_items:
            merged_rows.append(
                {
                    "instrument_type": sanitize_str(linked_item.get("instrument_type", "")),
                    "instrument_code": sanitize_str(linked_item.get("instrument_code", "")),
                    "inn": inn,
                    "company_name": sanitize_str(row.get("company_name", "")),
                    "news_date": sanitize_str(row.get("event_date", "")),
                    "title": sanitize_str(row.get("event_type", "")),
                    "url": sanitize_str(row.get("event_url", "")),
                    "source": sanitize_str(row.get("source", "")),
                    "is_new": "1" if row.get("is_new") else "",
                }
            )

    merged_rows.sort(key=lambda x: x.get("news_date", ""), reverse=True)
    for row in merged_rows:
        ws_news.append([
            row.get("instrument_type", ""),
            row.get("instrument_code", ""),
            row.get("inn", ""),
            row.get("company_name", ""),
            row.get("news_date", ""),
            row.get("title", ""),
            row.get("url", ""),
            row.get("source", ""),
            "✓ НОВОЕ" if row.get("is_new") else "",
            "1" if row.get("is_new") else "",
        ])
    apply_ws_style(ws_news, {"Ссылка"})

    wb.save(config.PORTFOLIO_XLSX)


REPORT_TYPE_PRIORITY = [(3, "Финансовая"), (4, "Консолидированная"), (5, "Отчет эмитента"), (2, "Годовая")]
REPORT_KEYWORDS = ("отчет", "финансов", "бухгалтер", "консолид", "эмитент", "годовой", "промежуточный", "аудитор")


@dataclass
class ReportFetchResult:
    inn: str
    company_name: str
    scoring_date: str
    company_map_row: dict[str, str] | None
    report_events: list[dict[str, Any]]
    report_state_rows: list[dict[str, str]]
    schedule_row: dict[str, Any] | None
    latest_report_date: str
    skipped_unchanged: bool
    telemetry: dict[str, Any]
    elapsed_sec: float
    error: str = ""


def _is_due(schedule_row: dict[str, Any] | None, now_dt: datetime, force_full: bool) -> bool:
    if force_full or not schedule_row:
        return True
    next_check = parse_date(schedule_row.get("next_check_at"))
    if not next_check:
        return True
    return next_check <= now_dt


def _calc_next_check(last_change_at: str, stable_run_count: int) -> str:
    last_change_dt = parse_date(last_change_at)
    if last_change_dt and last_change_dt >= datetime.now() - timedelta(days=30):
        return (datetime.now() + timedelta(hours=config.EDISCLOSURE_RECENT_CHANGE_RECHECK_HOURS)).isoformat(timespec="seconds")
    if last_change_dt and last_change_dt >= datetime.now() - timedelta(days=90):
        return (datetime.now() + timedelta(hours=config.EDISCLOSURE_ACTIVE_RECHECK_HOURS)).isoformat(timespec="seconds")
    hours = config.EDISCLOSURE_STABLE_RECHECK_HOURS if stable_run_count < 5 else config.EDISCLOSURE_STABLE_RECHECK_HOURS * 2
    return (datetime.now() + timedelta(hours=hours)).isoformat(timespec="seconds")


def stage_reports_prepare(conn: sqlite3.Connection, emitents: list[EmitentRow]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], set[str], dict[tuple[str, str], dict[str, str]], int, dict[str, int]]:
    emitent_map = {sanitize_str(row.inn): row for row in emitents if sanitize_str(row.inn)}
    company_rows = conn.execute("SELECT * FROM company_map").fetchall()
    mappings = {sanitize_str(r["inn"]): dict(r) for r in company_rows if sanitize_str(r["inn"]) in emitent_map}
    schedule_rows = conn.execute("SELECT * FROM emitent_schedule").fetchall()
    schedule_map = {sanitize_str(r["inn"]): dict(r) for r in schedule_rows}

    states: dict[tuple[str, str], dict[str, str]] = {}
    for r in conn.execute("SELECT * FROM report_state").fetchall():
        states[(sanitize_str(r["inn"]), sanitize_str(r["report_type_id"]))] = dict(r)

    has_state_by_inn: dict[str, bool] = {}
    for inn_key, _type in states:
        if inn_key:
            has_state_by_inn[inn_key] = True

    rows = conn.execute(
        "SELECT inn, COUNT(1) AS c FROM report_events WHERE source='e-disclosure' GROUP BY inn"
    ).fetchall()
    has_edisclosure_event_by_inn = {sanitize_str(r["inn"]): int(r["c"] or 0) > 0 for r in rows}

    tasks: list[dict[str, Any]] = []
    skipped_tasks: list[dict[str, Any]] = []
    force_full = bool(config.EDISCLOSURE_FORCE_FULL_SCAN)
    now_dt = datetime.now()
    for inn, row in emitent_map.items():
        missing_reports_history = not has_edisclosure_event_by_inn.get(inn, False)
        missing_state = not has_state_by_inn.get(inn, False)
        force_missing_recheck = bool(config.EDISCLOSURE_FORCE_RECHECK_MISSING_HISTORY) and (missing_reports_history or missing_state)

        task = {
            "inn": inn,
            "company_name": sanitize_str(row.company_name),
            "scoring_date": sanitize_str(row.scoring_date),
            "mapping": mappings.get(inn),
            "schedule": schedule_map.get(inn),
            "force_full": force_full,
            "force_missing_recheck": force_missing_recheck,
        }
        if force_missing_recheck or _is_due(task["schedule"], now_dt, force_full):
            tasks.append(task)
        else:
            skipped_tasks.append(task)
    run_no = int((conn.execute("SELECT value FROM meta WHERE key = 'edisclosure_run_no'").fetchone() or ["0"])[0] or "0") + 1
    conn.execute("INSERT INTO meta (key, value) VALUES ('edisclosure_run_no', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (str(run_no),))
    conn.commit()
    prep_stats = {
        "total_emitents": len(emitent_map),
        "processed_emitents": len(tasks),
        "skipped_by_cache": len(skipped_tasks),
    }
    return tasks, skipped_tasks, set(), states, run_no, prep_stats


def fetch_existing_event_hashes(conn: sqlite3.Connection, hashes: set[str], sources: tuple[str, ...]) -> set[str]:
    if not hashes or not sources:
        return set()
    chunk_size = 900
    existing: set[str] = set()
    source_placeholders = ",".join("?" for _ in sources)
    hash_list = list(hashes)
    for i in range(0, len(hash_list), chunk_size):
        chunk = hash_list[i : i + chunk_size]
        hash_placeholders = ",".join("?" for _ in chunk)
        query = (
            f"SELECT event_hash FROM report_events "
            f"WHERE source IN ({source_placeholders}) AND event_hash IN ({hash_placeholders})"
        )
        params = [*sources, *chunk]
        existing.update(r[0] for r in conn.execute(query, params).fetchall())
    return existing


def parse_reports_page(
    html: str,
    company_id: str,
    type_id: int,
    type_name: str,
    known_state: dict[str, str] | None,
    preview_limit: int | None = None,
    max_new_rows: int | None = None,
) -> tuple[list[dict[str, str]], str]:
    soup = BeautifulSoup(html, "lxml")
    table = soup.find("table", class_="zebra") or soup.find("table")
    if not table:
        return [], ""
    rows: list[dict[str, str]] = []
    known_hash = sanitize_str((known_state or {}).get("latest_hash"))
    top_row_hash = ""

    for idx, tr in enumerate(table.find_all("tr")):
        if preview_limit is not None and idx >= preview_limit:
            break
        tds = tr.find_all("td")
        if not tds:
            continue

        row_text = sanitize_str(tr.get_text(" ", strip=True))
        row_dates = [to_iso_date_str(d) for d in re.findall(r"\b\d{2}[./]\d{2}[./]\d{4}\b", row_text)]
        row_dates = [d for d in row_dates if d]

        doc_type = sanitize_str(tds[0].get_text(" ", strip=True)) if tds else ""
        period = sanitize_str(tds[1].get_text(" ", strip=True)) if len(tds) > 1 else ""
        foundation_date = to_iso_date_str(tds[2].get_text(" ", strip=True)) if len(tds) > 2 else ""
        placement_date = to_iso_date_str(tds[3].get_text(" ", strip=True)) if len(tds) > 3 else ""

        if not foundation_date and row_dates:
            foundation_date = row_dates[0]
        if not placement_date:
            placement_date = row_dates[-1] if row_dates else ""

        if not doc_type:
            doc_type = sanitize_str(row_text)

        anchors = tr.find_all("a", href=True)
        file_url = ""
        for anchor in anchors:
            href = sanitize_str(anchor.get("href"))
            if "fileload.ashx" not in href.lower():
                continue
            file_url = urljoin("https://www.e-disclosure.ru", href)
            break

        is_relevant = any(k in row_text.lower() for k in REPORT_KEYWORDS) or bool(period) or bool(file_url)
        if not is_relevant:
            continue

        row_hash = md5_short(f"{company_id}_{type_id}_{doc_type}_{period}_{placement_date}", 16)
        if not top_row_hash:
            top_row_hash = row_hash
        if known_hash and row_hash == known_hash:
            break

        rows.append({
            "hash": row_hash,
            "row_hash": row_hash,
            "company_id": company_id,
            "type_id": str(type_id),
            "report_type": type_name,
            "doc_type": doc_type,
            "period": period,
            "foundation_date": foundation_date,
            "placement_date": placement_date,
            "file_url": file_url,
            "row_text": row_text,
            "page_url": f"https://www.e-disclosure.ru/portal/files.aspx?id={company_id}&type={type_id}",
        })
        if max_new_rows and len(rows) >= max_new_rows:
            break
    return rows, top_row_hash


def _event_is_relevant(event_name: str) -> bool:
    low = sanitize_str(event_name).lower()
    return any(k in low for k in REPORT_KEYWORDS)


def choose_company_fast(client: EDisclosureClient, inn: str, company_name: str, mapping: dict[str, Any] | None, telemetry: dict[str, Any]) -> dict[str, str] | None:
    if mapping and sanitize_str(mapping.get("company_id")):
        company_id = sanitize_str(mapping.get("company_id"))
        verified_inn = sanitize_str(mapping.get("verified_inn"))
        validation_status = sanitize_str(mapping.get("validation_status")).lower()
        checked = parse_date(mapping.get("last_checked_at"))

        if company_id and (validation_status == "verified" or (verified_inn and verified_inn == inn)):
            telemetry["company_map_hits"] += 1
            return {"id": company_id, "name": sanitize_str(mapping.get("company_name")), "url": sanitize_str(mapping.get("company_url"))}

        if checked and checked >= datetime.now() - timedelta(days=config.COMPANY_MAP_HARD_TTL_DAYS) and (not verified_inn or verified_inn == inn):
            telemetry["company_map_hits"] += 1
            return {"id": company_id, "name": sanitize_str(mapping.get("company_name")), "url": sanitize_str(mapping.get("company_url"))}

    telemetry["company_search_requests"] += 1
    candidates = client.search_company_by_inn(inn)
    if not candidates:
        return None
    return client.choose_best_candidate(inn, candidates, company_name)


def _append_report_rows(
    rows: list[dict[str, str]],
    inn: str,
    company_name: str,
    scoring_date: str,
    company_display_name: str,
    report_events: list[dict[str, Any]],
    latest_report_date: str,
) -> str:
    for rep in rows:
        event_date = rep.get("placement_date") or rep.get("foundation_date")
        if event_date and (not latest_report_date or event_date > latest_report_date):
            latest_report_date = event_date
        report_events.append({
            "event_hash": rep["hash"],
            "inn": inn,
            "company_name": company_name or company_display_name,
            "scoring_date": scoring_date,
            "event_date": event_date,
            "event_type": "Опубликована новая отчетность",
            "event_url": rep.get("file_url") or rep.get("page_url", ""),
            "source": "e-disclosure",
            "payload": rep,
        })
    return latest_report_date


def _scan_reports_types(
    client: EDisclosureClient,
    *,
    inn: str,
    company_name: str,
    scoring_date: str,
    company_id: str,
    company_display_name: str,
    state_by_type: dict[tuple[str, str], dict[str, str]],
    report_events: list[dict[str, Any]],
    report_states: list[dict[str, str]],
    telemetry: dict[str, Any],
    now_ts: str,
    scan_types: list[tuple[int, str]],
    force_refresh: bool,
    max_new_rows: int | None,
    latest_report_date: str,
    logger: logging.Logger,
    scan_diag: dict[str, Any] | None = None,
) -> str:
    if scan_diag is None:
        scan_diag = {}
    scan_diag.setdefault("scanned_types", [])
    scan_diag.setdefault("found_rows_count", 0)
    scan_diag.setdefault("found_file_urls_count", 0)

    for type_id, type_name in scan_types:
        scan_diag["scanned_types"].append(str(type_id))
        known_state = state_by_type.get((inn, str(type_id)))
        html = client.get_reports_page_cached(company_id, type_id, force_refresh=force_refresh)
        telemetry["files_page_requests"] += 1
        rows, top_row_hash = parse_reports_page(
            html,
            company_id,
            type_id,
            type_name,
            known_state,
            max_new_rows=max_new_rows,
        )
        scan_diag["found_rows_count"] += len(rows)
        scan_diag["found_file_urls_count"] += sum(1 for row in rows if sanitize_str(row.get("file_url")))

        if rows and not any(sanitize_str(row.get("file_url")) for row in rows):
            logger.warning("files page parsed but FileLoad link not found | inn=%s company_id=%s type_id=%s", inn, company_id, type_id)

        if rows:
            top = rows[0]
            report_states.append({
                "inn": inn,
                "company_id": company_id,
                "report_type_id": str(type_id),
                "latest_hash": top.get("hash", ""),
                "latest_placement_date": top.get("placement_date", ""),
                "latest_foundation_date": top.get("foundation_date", ""),
                "top_row_hash": top_row_hash,
                "page_checked_at": now_ts,
                "last_checked_at": now_ts,
            })
        elif known_state:
            if top_row_hash and top_row_hash == sanitize_str(known_state.get("top_row_hash")):
                telemetry["preview_skips"] += 1
            report_states.append({
                "inn": inn,
                "company_id": company_id,
                "report_type_id": str(type_id),
                "latest_hash": sanitize_str(known_state.get("latest_hash")),
                "latest_placement_date": sanitize_str(known_state.get("latest_placement_date")),
                "latest_foundation_date": sanitize_str(known_state.get("latest_foundation_date")),
                "top_row_hash": top_row_hash or sanitize_str(known_state.get("top_row_hash")),
                "page_checked_at": now_ts,
                "last_checked_at": now_ts,
            })
        latest_report_date = _append_report_rows(rows, inn, company_name, scoring_date, company_display_name, report_events, latest_report_date)
    return latest_report_date


def fetch_one_emitent_reports(task: dict[str, Any], state_by_type: dict[tuple[str, str], dict[str, str]], logger: logging.Logger, run_no: int) -> ReportFetchResult:
    started = time.perf_counter()
    telemetry = {
        "company_search_requests": 0,
        "events_requests": 0,
        "files_page_requests": 0,
        "company_map_hits": 0,
        "event_gate_only": 0,
        "deep_scanned": 0,
        "preview_skips": 0,
        "full_scans": 0,
        "direct_fallback_scans": 0,
        "event_gate_result": "not_used",
    }
    inn = task["inn"]
    company_name = task["company_name"]
    scoring_date = task["scoring_date"]
    mapping = task.get("mapping") or {}
    schedule = task.get("schedule") or {}
    force_full = bool(task.get("force_full"))
    force_missing_recheck = bool(task.get("force_missing_recheck"))
    client = get_thread_local_edisclosure_client(logger)
    try:
        company = choose_company_fast(client, inn, company_name, mapping, telemetry)
        if not company or not company.get("id"):
            return ReportFetchResult(inn, company_name, scoring_date, None, [], [], None, "", False, telemetry, time.perf_counter() - started)
        company_id = sanitize_str(company.get("id"))

        report_events: list[dict[str, Any]] = []
        report_states: list[dict[str, str]] = []
        last_known_report_date = ""
        for type_id, _ in REPORT_TYPE_PRIORITY:
            known = state_by_type.get((inn, str(type_id)))
            if known and sanitize_str(known.get("latest_placement_date")) > last_known_report_date:
                last_known_report_date = sanitize_str(known.get("latest_placement_date"))

        known_state_missing = not any(state_by_type.get((inn, str(type_id))) for type_id, _ in REPORT_TYPE_PRIORITY)
        missing_type4_state = state_by_type.get((inn, "4")) is None
        suspicious_state = known_state_missing or not last_known_report_date or missing_type4_state
        needs_deep = force_full or force_missing_recheck or not mapping or not schedule or suspicious_state
        scan_diag: dict[str, Any] = {"scanned_types": [], "found_rows_count": 0, "found_file_urls_count": 0}
        event_gate_result = "not_used"
        if not needs_deep:
            telemetry["events_requests"] += 1
            try:
                events = client.get_company_events(company_id, days_back=60)
                rel = [e for e in events if _event_is_relevant(e.get("eventName", ""))]
                newest_rel = max([sanitize_str(x.get("pubDate") or x.get("eventDate")) for x in rel] or [""])
                if newest_rel and (not last_known_report_date or newest_rel > last_known_report_date):
                    needs_deep = True
                    event_gate_result = "newer_event"
                else:
                    telemetry["event_gate_only"] += 1
                    event_gate_result = "empty" if not rel else "no_newer"
            except Exception as exc:  # noqa: BLE001
                logger.warning("event gate failed, forcing direct scan | inn=%s company_id=%s err=%s", inn, company_id, exc)
                needs_deep = True
                event_gate_result = "error"

        telemetry["event_gate_result"] = event_gate_result

        latest_report_date = last_known_report_date
        scan_types = REPORT_TYPE_PRIORITY[:2]
        now_ts = now_iso()
        if needs_deep:
            telemetry["deep_scanned"] += 1
            if force_full or not mapping or not schedule:
                telemetry["full_scans"] += 1
                scan_types = REPORT_TYPE_PRIORITY
            latest_report_date = _scan_reports_types(
                client,
                inn=inn,
                company_name=company_name,
                scoring_date=scoring_date,
                company_id=company_id,
                company_display_name=sanitize_str(company.get("name")),
                state_by_type=state_by_type,
                report_events=report_events,
                report_states=report_states,
                telemetry=telemetry,
                now_ts=now_ts,
                scan_types=scan_types,
                force_refresh=True,
                max_new_rows=config.EDISCLOSURE_PARSE_MAX_NEW_ROWS_PER_TYPE,
                latest_report_date=latest_report_date,
                logger=logger,
                scan_diag=scan_diag,
            )

        known_company_id = bool(company_id) and bool(sanitize_str(mapping.get("company_id")) or company_id)
        fallback_required = bool(company_id) and known_company_id and (not report_events) and (
            suspicious_state
            or force_full
            or force_missing_recheck
            or event_gate_result in {"empty", "error", "not_used"}
            or not last_known_report_date
        )
        if fallback_required:
            telemetry["direct_fallback_scans"] += 1
            latest_report_date = _scan_reports_types(
                client,
                inn=inn,
                company_name=company_name,
                scoring_date=scoring_date,
                company_id=company_id,
                company_display_name=sanitize_str(company.get("name")),
                state_by_type=state_by_type,
                report_events=report_events,
                report_states=report_states,
                telemetry=telemetry,
                now_ts=now_ts,
                scan_types=[(4, "Консолидированная"), (3, "Финансовая"), (5, "Отчет эмитента"), (2, "Годовая")],
                force_refresh=True,
                max_new_rows=max(config.EDISCLOSURE_PARSE_MAX_NEW_ROWS_PER_TYPE, 10),
                latest_report_date=latest_report_date,
                logger=logger,
                scan_diag=scan_diag,
            )

        if company_id and not report_events:
            logger.warning(
                "known company_id but no reports parsed | inn=%s company_id=%s scanned_types=%s found_rows_count=%s found_file_urls_count=%s event_gate_result=%s report_state=%s",
                inn,
                company_id,
                ",".join(scan_diag.get("scanned_types", [])),
                scan_diag.get("found_rows_count", 0),
                scan_diag.get("found_file_urls_count", 0),
                event_gate_result,
                json.dumps({k[1]: v for k, v in state_by_type.items() if k[0] == inn}, ensure_ascii=False),
            )

        stable_count = int(schedule.get("stable_run_count") or 0)
        if report_events:
            stable_count = 0
        else:
            stable_count += 1
        schedule_row = {
            "inn": inn,
            "company_id": company_id,
            "last_checked_at": now_ts,
            "next_check_at": _calc_next_check(latest_report_date, stable_count),
            "last_change_at": latest_report_date or sanitize_str(schedule.get("last_change_at")),
            "stable_run_count": stable_count,
            "last_mode": "deep_missing_recheck" if force_missing_recheck else ("deep" if needs_deep else "event_gate_only"),
            "last_event_gate_at": now_ts,
            "last_files_scan_at": now_ts if (needs_deep or fallback_required) else sanitize_str(schedule.get("last_files_scan_at")),
        }
        company_row = {
            "inn": inn,
            "company_id": company_id,
            "company_name": sanitize_str(company.get("name")),
            "company_url": sanitize_str(company.get("url")),
            "verified_inn": inn,
            "validation_status": "verified",
            "last_success_at": now_ts,
            "full_scan_at": now_ts if force_full else sanitize_str(mapping.get("full_scan_at")),
            "fast_scan_at": now_ts,
            "last_checked_at": now_ts,
        }
        dedup_events = list({ev["event_hash"]: ev for ev in report_events}.values())
        return ReportFetchResult(inn, company_name, scoring_date, company_row, dedup_events, report_states, schedule_row, latest_report_date, not needs_deep, telemetry, time.perf_counter() - started)
    except Exception as exc:  # noqa: BLE001
        return ReportFetchResult(inn, company_name, scoring_date, None, [], [], None, "", False, telemetry, time.perf_counter() - started, error=str(exc))


def _read_autotune_meta(conn: sqlite3.Connection) -> tuple[int, int]:
    workers_raw = conn.execute("SELECT value FROM meta WHERE key='edisclosure_autotune_workers'").fetchone()
    files_raw = conn.execute("SELECT value FROM meta WHERE key='edisclosure_autotune_files_semaphore'").fetchone()

    has_meta = bool(workers_raw and files_raw)
    if has_meta:
        workers = int((workers_raw or [str(config.EDISCLOSURE_FETCH_WORKERS_DEFAULT)])[0])
        files = int((files_raw or [str(config.EDISCLOSURE_FILES_SEMAPHORE_DEFAULT)])[0])
    elif config.EDISCLOSURE_AUTOTUNE_COLD_START_MAX:
        workers = int(config.EDISCLOSURE_FETCH_WORKERS_MAX)
        files = int(config.EDISCLOSURE_FILES_SEMAPHORE_MAX)
    else:
        workers = int(config.EDISCLOSURE_FETCH_WORKERS_DEFAULT)
        files = int(config.EDISCLOSURE_FILES_SEMAPHORE_DEFAULT)

    workers = max(config.EDISCLOSURE_FETCH_WORKERS_MIN, min(config.EDISCLOSURE_FETCH_WORKERS_MAX, workers))
    files = max(config.EDISCLOSURE_FILES_SEMAPHORE_MIN, min(config.EDISCLOSURE_FILES_SEMAPHORE_MAX, files))

    if not has_meta:
        _save_autotune_meta(conn, workers, files)

    return workers, files


def _save_autotune_meta(conn: sqlite3.Connection, workers: int, files: int) -> None:
    conn.execute("INSERT INTO meta (key, value) VALUES ('edisclosure_autotune_workers', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (str(workers),))
    conn.execute("INSERT INTO meta (key, value) VALUES ('edisclosure_autotune_files_semaphore', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (str(files),))
    conn.commit()


def _autotune_concurrency(conn: sqlite3.Connection, logger: logging.Logger, workers: int, files: int) -> tuple[int, int, bool]:
    if not config.EDISCLOSURE_AUTOTUNE_ENABLED:
        return workers, files, False

    with runtime_state.lock:
        total_requests = max(1, runtime_state.total_requests)
        err_rate = (runtime_state.status_429 + runtime_state.timeout_count) / total_requests

    burst_snapshot = search_burst_controller.snapshot()
    had_search_burst = burst_snapshot.get("search_cooldown_events", 0) > 0

    bad_streak_raw = conn.execute("SELECT value FROM meta WHERE key='edisclosure_autotune_bad_run_streak'").fetchone()
    bad_streak = int((bad_streak_raw or ["0"])[0] or "0")

    overloaded = err_rate > config.EDISCLOSURE_AUTOTUNE_ERROR_RATE_THRESHOLD
    if overloaded:
        bad_streak += 1
    else:
        bad_streak = 0

    next_workers = workers
    next_files = files
    changed = False

    if overloaded and bad_streak >= max(2, int(config.EDISCLOSURE_AUTOTUNE_SCALE_DOWN_STREAK)):
        # Scale down only after several consecutive overloaded runs.
        next_workers = max(config.EDISCLOSURE_FETCH_WORKERS_MIN, workers - max(1, int(config.EDISCLOSURE_AUTOTUNE_SCALE_DOWN_STEP)))
        next_files = max(config.EDISCLOSURE_FILES_SEMAPHORE_MIN, files - max(1, int(config.EDISCLOSURE_AUTOTUNE_SCALE_DOWN_STEP // 2 or 1)))
        changed = (next_workers != workers) or (next_files != files)
    elif not overloaded:
        if err_rate <= config.EDISCLOSURE_AUTOTUNE_FAST_GROW_ERROR_RATE_THRESHOLD and not had_search_burst:
            next_workers = min(config.EDISCLOSURE_FETCH_WORKERS_MAX, workers + max(1, int(config.EDISCLOSURE_AUTOTUNE_FAST_GROW_STEP)))
            next_files = min(config.EDISCLOSURE_FILES_SEMAPHORE_MAX, files + max(1, int(config.EDISCLOSURE_AUTOTUNE_FAST_GROW_STEP // 2 or 1)))
        else:
            next_workers = min(config.EDISCLOSURE_FETCH_WORKERS_MAX, workers + max(1, int(config.EDISCLOSURE_AUTOTUNE_GROW_STEP)))
            next_files = min(config.EDISCLOSURE_FILES_SEMAPHORE_MAX, files + max(1, int(config.EDISCLOSURE_AUTOTUNE_GROW_STEP // 2 or 1)))
        changed = (next_workers != workers) or (next_files != files)

    logger.info(
        "autotune next workers=%s files_semaphore=%s changed=%s err_rate=%.4f bad_streak=%s thresholds=(fast<=%.4f, down>%.4f) search_cooldown_events=%s",
        next_workers,
        next_files,
        changed,
        err_rate,
        bad_streak,
        config.EDISCLOSURE_AUTOTUNE_FAST_GROW_ERROR_RATE_THRESHOLD,
        config.EDISCLOSURE_AUTOTUNE_ERROR_RATE_THRESHOLD,
        burst_snapshot.get("search_cooldown_events", 0),
    )
    conn.execute("INSERT INTO meta (key, value) VALUES ('edisclosure_autotune_bad_run_streak', ?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (str(bad_streak),))
    _save_autotune_meta(conn, next_workers, next_files)
    return next_workers, next_files, changed


def stage_reports_fetch_parallel(tasks: list[dict[str, Any]], report_state: dict[tuple[str, str], dict[str, str]], logger: logging.Logger, run_no: int) -> list[ReportFetchResult]:
    if not tasks:
        return []
    results: list[ReportFetchResult] = []
    workers = max(1, int(_current_workers))
    pbar = tqdm(total=len(tasks), desc="Сбор отчетности", position=0, leave=True)
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [pool.submit(fetch_one_emitent_reports, task, report_state, logger, run_no) for task in tasks]
        for future in as_completed(futures):
            results.append(future.result())
            pbar.update(1)
    pbar.close()
    return results


def update_emitent_schedule_batch(conn: sqlite3.Connection, rows: list[tuple[Any, ...]]) -> None:
    if not rows:
        return
    conn.executemany(
        """
        INSERT INTO emitent_schedule (inn, company_id, last_checked_at, next_check_at, last_change_at, stable_run_count, last_mode, last_event_gate_at, last_files_scan_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(inn) DO UPDATE SET
            company_id=excluded.company_id,
            last_checked_at=excluded.last_checked_at,
            next_check_at=excluded.next_check_at,
            last_change_at=excluded.last_change_at,
            stable_run_count=excluded.stable_run_count,
            last_mode=excluded.last_mode,
            last_event_gate_at=excluded.last_event_gate_at,
            last_files_scan_at=excluded.last_files_scan_at
        """,
        rows,
    )


def update_report_state_batch(conn: sqlite3.Connection, rows: list[tuple[str, str, str, str, str, str, str, str, str]]) -> None:
    if not rows:
        return
    conn.executemany(
        """
        INSERT INTO report_state (inn, company_id, report_type_id, latest_hash, latest_placement_date, latest_foundation_date, top_row_hash, page_checked_at, last_checked_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(inn, report_type_id) DO UPDATE SET
            company_id=excluded.company_id,
            latest_hash=excluded.latest_hash,
            latest_placement_date=excluded.latest_placement_date,
            latest_foundation_date=excluded.latest_foundation_date,
            top_row_hash=excluded.top_row_hash,
            page_checked_at=excluded.page_checked_at,
            last_checked_at=excluded.last_checked_at
        """,
        rows,
    )


def stage_reports_flush_db(conn: sqlite3.Connection, results: list[ReportFetchResult], existing_hashes: set[str], all_new_event_hashes: set[str], logger: logging.Logger, prep_stats: dict[str, int]) -> dict[str, Any]:
    now = now_iso()
    company_rows = []
    report_rows = []
    state_rows = []
    schedule_rows = []
    current_event_hashes: set[str] = set()
    for res in results:
        for ev in res.report_events:
            current_event_hashes.add(ev["event_hash"])
    existing_hashes = fetch_existing_event_hashes(conn, current_event_hashes, ("e-disclosure",))
    stats = {
        "reports_found": 0,
        "new_events": 0,
        "errors": 0,
        "company_search_requests": 0,
        "events_requests": 0,
        "files_page_requests": 0,
        "fileload_requests": 0,
        "company_map_hits": 0,
        "event_gate_only": 0,
        "deep_scanned": 0,
        "preview_skips": 0,
        "full_scans": 0,
        "direct_fallback_scans": 0,
        **prep_stats,
    }
    durations = []
    slowest = []
    for res in results:
        durations.append(res.elapsed_sec)
        slowest.append((res.elapsed_sec, res.inn, res.error, 0.0, 0.0, 0.0, 0.0, ""))
        if res.error:
            stats["errors"] += 1
            logger.warning("Failed reports INN=%s: %s", res.inn, res.error)
        for key in ["company_search_requests", "events_requests", "files_page_requests", "company_map_hits", "event_gate_only", "deep_scanned", "preview_skips", "full_scans", "direct_fallback_scans"]:
            stats[key] += int(res.telemetry.get(key, 0) or 0)
        if res.company_map_row:
            cm = res.company_map_row
            company_rows.append((cm["inn"], cm["company_id"], cm["company_name"], cm["company_url"], cm["verified_inn"], cm["validation_status"], cm["last_success_at"], cm.get("full_scan_at", ""), cm.get("fast_scan_at", ""), cm["last_checked_at"]))
        if res.schedule_row:
            sc = res.schedule_row
            schedule_rows.append((sc["inn"], sc["company_id"], sc["last_checked_at"], sc["next_check_at"], sc["last_change_at"], sc["stable_run_count"], sc["last_mode"], sc["last_event_gate_at"], sc["last_files_scan_at"]))
        for st in res.report_state_rows:
            state_rows.append((st["inn"], st["company_id"], st["report_type_id"], st["latest_hash"], st["latest_placement_date"], st["latest_foundation_date"], st.get("top_row_hash", ""), st.get("page_checked_at", ""), st["last_checked_at"]))
        stats["reports_found"] += len(res.report_events)
        for ev in res.report_events:
            payload_json = json.dumps(ev["payload"], ensure_ascii=False, separators=(",", ":"))
            report_rows.append((ev["event_hash"], ev["inn"], ev["company_name"], ev["scoring_date"], ev["event_date"], ev["event_type"], ev["event_url"], ev["source"], payload_json, now, now))
            if ev["event_hash"] not in existing_hashes:
                stats["new_events"] += 1
                all_new_event_hashes.add(ev["event_hash"])
                existing_hashes.add(ev["event_hash"])

    if company_rows:
        conn.executemany(
            """
        INSERT INTO company_map (inn, company_id, company_name, company_url, verified_inn, validation_status, last_success_at, full_scan_at, fast_scan_at, last_checked_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(inn) DO UPDATE SET
            company_id=excluded.company_id,
            company_name=excluded.company_name,
            company_url=excluded.company_url,
            verified_inn=excluded.verified_inn,
            validation_status=excluded.validation_status,
            last_success_at=excluded.last_success_at,
            full_scan_at=excluded.full_scan_at,
            fast_scan_at=excluded.fast_scan_at,
            last_checked_at=excluded.last_checked_at
        """,
            company_rows,
        )
    flush_report_events_batch(conn, report_rows)
    update_report_state_batch(conn, state_rows)
    update_emitent_schedule_batch(conn, schedule_rows)
    conn.commit()
    stats["durations"] = durations
    stats["slowest"] = sorted(slowest, key=lambda x: x[0], reverse=True)[:30]
    return stats


def flush_report_events_batch(conn: sqlite3.Connection, rows: list[tuple[Any, ...]]) -> None:
    if not rows:
        return
    conn.executemany(
        """
        INSERT INTO report_events (event_hash, inn, company_name, scoring_date, event_date, event_type, event_url, source, payload_json, first_seen_at, last_seen_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(event_hash) DO UPDATE SET
            company_name=excluded.company_name,
            scoring_date=excluded.scoring_date,
            event_date=excluded.event_date,
            event_type=excluded.event_type,
            event_url=excluded.event_url,
            source=excluded.source,
            payload_json=excluded.payload_json,
            last_seen_at=excluded.last_seen_at
        """,
        rows,
    )


def print_perf_summary(stats: dict[str, Any], total_emitents: int, total_seconds: float, logger: logging.Logger) -> None:
    durations = stats.get("durations", [])
    avg = sum(durations) / len(durations) if durations else 0
    median = statistics.median(durations) if durations else 0
    p95 = statistics.quantiles(durations, n=100)[94] if len(durations) >= 20 else (max(durations) if durations else 0)
    emitents_per_min = (total_emitents / total_seconds * 60.0) if total_seconds > 0 else 0.0

    with runtime_state.lock:
        request_latencies = list(runtime_state.request_latencies)
        total_requests = runtime_state.total_requests
        total_429 = runtime_state.status_429
        total_timeout = runtime_state.timeout_count
        files_requests = runtime_state.files_requests
        company_search_requests = runtime_state.search_requests
        events_requests = runtime_state.events_requests
        fileload_requests = runtime_state.fileload_requests

    req_median = statistics.median(request_latencies) if request_latencies else 0.0
    req_p95 = statistics.quantiles(request_latencies, n=100)[94] if len(request_latencies) >= 20 else (max(request_latencies) if request_latencies else 0.0)
    search_burst = search_burst_controller.snapshot()

    saved_company_map = max(0, stats.get("processed_emitents", 0) - stats.get("company_search_requests", 0))
    saved_event_gate = max(0, stats.get("event_gate_only", 0) * 2)
    saved_preview = max(0, stats.get("preview_skips", 0))

    print(
        f"Summary reports: total_emitents={stats.get('total_emitents', total_emitents)}, processed_emitents={stats.get('processed_emitents', total_emitents)}, "
        f"skipped_by_cache={stats.get('skipped_by_cache', 0)}, event_gated={stats.get('event_gate_only', 0)}, deep_scanned={stats.get('deep_scanned', 0)}, "
        f"reports_found={stats.get('reports_found', 0)}, new_events_found={stats.get('new_events', 0)}"
    )
    print(
        f"HTTP saved: company_map={saved_company_map}, event_gate={saved_event_gate}, preview_skip={saved_preview}, "
        f"company_search_requests={company_search_requests}, events_requests={events_requests}, files_requests={files_requests}, fileload_requests={fileload_requests}"
    )
    print(
        f"Search burst: search_429_count={search_burst.get('search_429_count', 0)}, search_timeout_count={search_burst.get('search_timeout_count', 0)}, "
        f"search_cooldown_events={search_burst.get('search_cooldown_events', 0)}, total_search_cooldown_seconds={search_burst.get('total_search_cooldown_seconds', 0.0):.2f}, "
        f"workers_used={stats.get('workers_used', _current_workers)}, files_semaphore_used={stats.get('files_semaphore_used', _current_files_semaphore)}, "
        f"autotune_changed_next_run={stats.get('autotune_changed_next_run', 'no')}"
    )
    print(
        f"Timing: total_stage_seconds={total_seconds:.2f}s avg={avg:.2f}s median={median:.2f}s p95={p95:.2f}s "
        f"avg_emitents_per_minute={emitents_per_min:.2f}"
    )

    logger.info(
        "stage_reports telemetry total_emitents=%s processed_emitents=%s skipped_by_cache=%s event_gated=%s deep_scanned=%s company_search_requests=%s events_requests=%s files_requests=%s fileload_requests=%s preview_skips=%s reports_found=%s new_events_found=%s total_stage_seconds=%.3f avg_emitents_per_minute=%.3f median_request_latency=%.3f p95_request_latency=%.3f 429_count=%s timeout_count=%s workers_used=%s files_semaphore_used=%s next_workers=%s next_files_semaphore=%s autotune_changed_next_run=%s search_429_count=%s search_timeout_count=%s search_cooldown_events=%s total_search_cooldown_seconds=%.3f http_saved_company_map=%s http_saved_event_gate=%s http_saved_preview_skip=%s",
        stats.get("total_emitents", total_emitents),
        stats.get("processed_emitents", total_emitents),
        stats.get("skipped_by_cache", 0),
        stats.get("event_gate_only", 0),
        stats.get("deep_scanned", 0),
        company_search_requests,
        events_requests,
        files_requests,
        fileload_requests,
        stats.get("preview_skips", 0),
        stats.get("reports_found", 0),
        stats.get("new_events", 0),
        total_seconds,
        emitents_per_min,
        req_median,
        req_p95,
        total_429,
        total_timeout,
        stats.get("workers_used", _current_workers),
        stats.get("files_semaphore_used", _current_files_semaphore),
        stats.get("chosen_workers", _current_workers),
        stats.get("chosen_files_semaphore", _current_files_semaphore),
        stats.get("autotune_changed_next_run", "no"),
        search_burst.get("search_429_count", 0),
        search_burst.get("search_timeout_count", 0),
        search_burst.get("search_cooldown_events", 0),
        search_burst.get("total_search_cooldown_seconds", 0.0),
        saved_company_map,
        saved_event_gate,
        saved_preview,
    )

    for row in stats.get("slowest", [])[:20]:
        logger.info(
            "slow_emitent inn=%s elapsed=%.3f err=%s breakdown(search=%.3f card=%.3f files=%.3f parse=%.3f bottleneck=%s)",
            row[1],
            row[0],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
            row[7],
        )



# -----------------------------
# Pipeline
# -----------------------------
def run_monitoring() -> None:
    logger = setup_logger()
    conn = db_connect()
    stage_times: dict[str, float] = {}
    all_new_event_hashes: set[str] = set()

    print("=====\nЭтап 1: Загрузка эмитентов")
    emitents, elapsed = timed(lambda: load_emitents_rows(config.EMITENTS_SOURCE_FILE))
    stage_times["Этап 1: Загрузка эмитентов"] = elapsed

    print("Этап 2: Сбор отчетности")

    def stage_reports() -> None:
        reset_runtime_state()
        workers, files_semaphore = _read_autotune_meta(conn)
        configure_runtime_concurrency(workers, files_semaphore)
        tasks, skipped_tasks, existing_hashes, report_state, run_no, prep_stats = stage_reports_prepare(conn, emitents)
        stage_started = time.perf_counter()
        results = stage_reports_fetch_parallel(tasks, report_state, logger, run_no)
        stats = stage_reports_flush_db(conn, results, existing_hashes, all_new_event_hashes, logger, prep_stats)
        next_workers, next_files_semaphore, autotune_changed = _autotune_concurrency(conn, logger, workers, files_semaphore)
        stats["workers_used"] = workers
        stats["files_semaphore_used"] = files_semaphore
        stats["chosen_workers"] = next_workers
        stats["chosen_files_semaphore"] = next_files_semaphore
        stats["autotune_changed_next_run"] = "yes" if autotune_changed else "no"
        print_perf_summary(stats, prep_stats.get("processed_emitents", len(tasks)), time.perf_counter() - stage_started, logger)

    _, elapsed = timed(stage_reports)
    stage_times["Этап 2: Сбор отчетности"] = elapsed

    print("Этап 3: События по рейтингам")

    def stage_ratings() -> None:
        now = now_iso()

        conn.execute("DELETE FROM emitents_snapshot")
        conn.executemany(
            """
            INSERT INTO emitents_snapshot (inn, company_name, scoring, scoring_date, nra_rate, acra_rate, nkr_rate, raex_rate, snapshot_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (row.inn, row.company_name, row.scoring, row.scoring_date, row.nra_rate, row.acra_rate, row.nkr_rate, row.raex_rate, now)
                for row in emitents
            ],
        )
        save_emitents_snapshot_excel(emitents)

        if not bool(config.RATINGS_MONITORING_ENABLED):
            conn.commit()
            return

        ratings_snapshot = load_ratings_snapshot_from_db(logger)
        prev_rows = conn.execute("SELECT inn, source, rating, assigned_date FROM ratings_monitoring_snapshot").fetchall()
        prev = {(sanitize_str(r["inn"]), sanitize_str(r["source"])): dict(r) for r in prev_rows if sanitize_str(r["inn"]) and sanitize_str(r["source"])}

        existing_hashes = {
            r[0]
            for r in conn.execute(
                "SELECT event_hash FROM report_events WHERE source IN ('NRA', 'ACRA', 'NKR', 'RAEX')"
            ).fetchall()
        }

        update_rows: list[tuple[str, str]] = []
        insert_rows: list[tuple[str, str, str, str, str, str, str, str, str, str, str]] = []
        emitent_by_inn = {sanitize_str(e.inn): e for e in emitents if sanitize_str(e.inn)}

        for (inn, source), current in tqdm(ratings_snapshot.items(), desc="Рейтинговые изменения", position=0):
            old = prev.get((inn, source), {})
            old_rating = sanitize_str(old.get("rating", ""))
            new_rating = sanitize_str(current.get("rating", ""))
            event_type = classify_rating_change(old_rating, new_rating)
            if not event_type:
                continue

            emitent = emitent_by_inn.get(inn)
            company_name = emitent.company_name if emitent else ""
            scoring_date = emitent.scoring_date if emitent else ""
            event_date = current.get("assigned_date") or scoring_date or today_iso()
            event_hash = md5_short(f"rate_db_{inn}_{source}_{old_rating}_{new_rating}_{event_date}", 16)

            if event_hash in existing_hashes:
                update_rows.append((now, event_hash))
                continue

            insert_rows.append(
                (
                    event_hash,
                    inn,
                    company_name,
                    scoring_date,
                    to_iso_date_str(event_date) or today_iso(),
                    event_type,
                    "",
                    source,
                    json.dumps({"old": old_rating, "new": new_rating, "assigned_date": current.get("assigned_date", "")}, ensure_ascii=False),
                    now,
                    now,
                )
            )
            existing_hashes.add(event_hash)
            all_new_event_hashes.add(event_hash)

        if update_rows:
            conn.executemany("UPDATE report_events SET last_seen_at = ? WHERE event_hash = ?", update_rows)
        if insert_rows:
            conn.executemany(
                """
                INSERT INTO report_events (event_hash, inn, company_name, scoring_date, event_date, event_type, event_url,
                source, payload_json, first_seen_at, last_seen_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                insert_rows,
            )

        conn.execute("DELETE FROM ratings_monitoring_snapshot")
        if ratings_snapshot:
            conn.executemany(
                """
                INSERT INTO ratings_monitoring_snapshot (inn, source, rating, assigned_date, loaded_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                [
                    (inn, source, data.get("rating", ""), data.get("assigned_date", ""), now)
                    for (inn, source), data in ratings_snapshot.items()
                ],
            )
        conn.commit()

    _, elapsed = timed(stage_ratings)
    stage_times["Этап 3: События по рейтингам"] = elapsed

    print("Этап 4: Загрузка портфеля")

    def stage_portfolio() -> list[dict[str, str]]:
        src = config.PORTFOLIO_XLSX
        items = load_portfolio_items(src, logger)
        conn.execute("DELETE FROM portfolio_items")
        for item in items:
            conn.execute(
                """
                INSERT INTO portfolio_items (instrument_type, instrument_code, inn, company_name, source_file, loaded_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(instrument_type, instrument_code) DO UPDATE SET
                inn=excluded.inn, company_name=excluded.company_name, source_file=excluded.source_file, loaded_at=excluded.loaded_at
                """,
                (
                    item.get("instrument_type", ""),
                    item.get("instrument_code", ""),
                    item.get("inn", ""),
                    item.get("company_name", ""),
                    str(src),
                    now_iso(),
                ),
            )
        conn.commit()

        return items

    portfolio_items, elapsed = timed(stage_portfolio)
    stage_times["Этап 4: Загрузка портфеля"] = elapsed

    print("Этап 5: Новости портфеля")

    def stage_news() -> list[dict[str, str]]:
        collector = SmartlabNewsCollector(logger)
        cache = NewsCacheManager(config.CACHE_DIR / "news" / "news_cache.csv")
        rows: list[dict[str, str]] = []
        news_insert_rows: list[tuple[str, str, str, str, str, str, str, str, str, str]] = []
        now = now_iso()
        for item in tqdm(portfolio_items, desc="Сбор новостей", position=0):
            try:
                for news in collector.collect(item):
                    h = md5_short(f"{news['url']}_{sanitize_str(news['title'])[:50]}_{news['news_date']}", 16)
                    is_new = cache.is_new(h)
                    if is_new:
                        cache.add(
                            {
                                "hash": h,
                                "company_name": item.get("company_name", ""),
                                "company_inn": item.get("inn", ""),
                                "date": news["news_date"],
                                "title": news["title"],
                                "source": "Smartlab",
                                "url": news["url"],
                                "added_date": today_iso(),
                            }
                        )
                    source = news.get("source", "Smartlab")
                    news_insert_rows.append(
                        (
                            h,
                            item.get("instrument_type", ""),
                            item.get("instrument_code", ""),
                            item.get("inn", ""),
                            item.get("company_name", ""),
                            news["news_date"],
                            news["title"],
                            news["url"],
                            source,
                            now,
                        )
                    )
                    rows.append(
                        {
                            "event_hash": h,
                            "instrument_type": item.get("instrument_type", ""),
                            "instrument_code": item.get("instrument_code", ""),
                            "inn": item.get("inn", ""),
                            "company_name": item.get("company_name", ""),
                            "news_date": news["news_date"],
                            "title": news["title"],
                            "url": news["url"],
                            "source": source,
                            "is_new": is_new,
                        }
                    )
            except Exception as exc:  # noqa: BLE001
                logger.exception("News failed for %s: %s", item, exc)
        if news_insert_rows:
            conn.executemany(
                """
                INSERT OR IGNORE INTO news_events (
                    event_hash, instrument_type, instrument_code, inn, company_name, news_date, title, url, source, first_seen_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                news_insert_rows,
            )
            conn.commit()
        cache.save()
        return rows

    news_rows, elapsed = timed(stage_news)
    stage_times["Этап 5: Новости портфеля"] = elapsed

    print("Этап 6: Экспорт витрин")

    def stage_export() -> None:
        report_events = [
            dict(r)
            for r in conn.execute(
                "SELECT event_hash, inn, company_name, scoring_date, event_date, event_type, event_url, source FROM report_events"
            ).fetchall()
        ]
        for row in report_events:
            row["is_new"] = row.get("event_hash") in all_new_event_hashes
        export_reports(report_events)

        latest_event_by_inn = build_latest_event_by_inn(report_events)

        logger.info(
            "export diagnostics | report_events=%s latest_event_by_inn=%s portfolio_items=%s has_6316031581_report=%s has_6316031581_latest=%s has_6316031581_portfolio=%s",
            len(report_events),
            len(latest_event_by_inn),
            len(portfolio_items),
            any(sanitize_str(r.get("inn")) == "6316031581" for r in report_events),
            "6316031581" in latest_event_by_inn,
            any(sanitize_str(item.get("inn")) == "6316031581" for item in portfolio_items),
        )

        latest_news_by_key: dict[tuple[str, str], dict[str, str]] = {}
        for row in sorted(news_rows, key=lambda x: x.get("news_date", ""), reverse=True):
            latest_news_by_key.setdefault((row.get("instrument_type", ""), row.get("instrument_code", "")), row)

        export_portfolio(portfolio_items, latest_event_by_inn, latest_news_by_key, news_rows, report_events)

    _, elapsed = timed(stage_export)
    stage_times["Этап 6: Экспорт витрин"] = elapsed

    conn.close()

    print("=====\nSummary")
    total = sum(stage_times.values())
    for stage, sec in stage_times.items():
        print(f"- {stage}: {sec:.2f} сек")
    print(f"- Итого: {total:.2f} сек")


if __name__ == "__main__":
    run_monitoring()
