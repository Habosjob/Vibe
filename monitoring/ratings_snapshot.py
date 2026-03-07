from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from . import config
from .helpers import md5_short, parse_date, sanitize_str, today_iso

OUTLOOK_MARKERS = {
    "позитив", "положитель", "stable", "стабиль", "negative", "негатив",
    "developing", "развива", "positive", "watch", "revision",
}


@dataclass
class EmitentRow:
    inn: str
    company_name: str
    scoring: str
    scoring_date: str
    nra_rate: str
    acra_rate: str
    nkr_rate: str
    raex_rate: str


def _normalize_header(name: str) -> str:
    return sanitize_str(name).replace(" ", "").lower()


def load_emitents_xlsx(path: Path) -> list[EmitentRow]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [_normalize_header(c.value or "") for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    def get(values: list[Any], key: str) -> str:
        pos = idx.get(key)
        if pos is None or pos >= len(values):
            return ""
        value = values[pos]
        if hasattr(value, "isoformat"):
            try:
                return value.date().isoformat()  # datetime
            except Exception:  # noqa: BLE001
                return value.isoformat()
        return sanitize_str(value)

    rows: list[EmitentRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = list(row)
        inn = get(values, "inn")
        if not inn:
            continue
        rows.append(
            EmitentRow(
                inn=inn,
                company_name=get(values, "emitentname"),
                scoring=get(values, "scoring"),
                scoring_date=get(values, "datescoring"),
                nra_rate=get(values, "nra_rate"),
                acra_rate=get(values, "acra_rate"),
                nkr_rate=get(values, "nkr_rate"),
                raex_rate=get(values, "raex_rate"),
            )
        )
    return rows


def save_emitents_snapshot_xlsx(rows: list[EmitentRow]) -> None:
    config.BASE_SNAPSHOTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "emitents_snapshot"
    ws.append([
        "INN", "EMITENTNAME", "Scoring", "DateScoring", "NRA_Rate", "Acra_Rate", "NKR_Rate", "RAEX_Rate", "SnapshotAt"
    ])
    snap_at = today_iso()
    for row in rows:
        ws.append([
            row.inn, row.company_name, row.scoring, row.scoring_date, row.nra_rate,
            row.acra_rate, row.nkr_rate, row.raex_rate, snap_at,
        ])
    wb.save(config.EMITENTS_SNAPSHOT_XLSX)


def _split_rating_and_outlook(value: str) -> tuple[str, str]:
    text = sanitize_str(value).lower()
    if not text:
        return "", ""
    rating_part = text
    outlook_hits = [m for m in OUTLOOK_MARKERS if m in text]
    for marker in outlook_hits:
        rating_part = rating_part.replace(marker, "")
    rating_part = " ".join(rating_part.split())
    return rating_part, "|".join(sorted(outlook_hits))


def _classify_change(old: str, new: str) -> str | None:
    old_clean, new_clean = sanitize_str(old), sanitize_str(new)
    if old_clean == new_clean:
        return None
    if old_clean and not new_clean:
        return "Рейтинг отозван / снят"
    old_rate, old_outlook = _split_rating_and_outlook(old_clean)
    new_rate, new_outlook = _split_rating_and_outlook(new_clean)
    if old_rate == new_rate and old_outlook != new_outlook:
        return "Изменен прогноз"
    return "Изменен рейтинг"


def build_rating_events(
    current_rows: list[EmitentRow],
    previous_by_inn: dict[str, dict[str, Any]],
) -> list[dict[str, str]]:
    events: list[dict[str, str]] = []
    for row in current_rows:
        prev = previous_by_inn.get(row.inn)
        if not prev:
            continue
        for field, label in [
            ("nra_rate", "NRA"),
            ("acra_rate", "ACRA"),
            ("nkr_rate", "NKR"),
            ("raex_rate", "RAEX"),
        ]:
            old = sanitize_str(prev.get(field, ""))
            new = sanitize_str(getattr(row, field))
            event_type = _classify_change(old, new)
            if not event_type:
                continue
            event_date = row.scoring_date or today_iso()
            event_hash = md5_short(f"rate_{row.inn}_{field}_{old}_{new}_{event_date}", 16)
            events.append(
                {
                    "event_hash": event_hash,
                    "inn": row.inn,
                    "company_name": row.company_name,
                    "scoring_date": row.scoring_date,
                    "event_date": parse_date(event_date).date().isoformat() if parse_date(event_date) else today_iso(),
                    "event_type": event_type,
                    "event_url": "",
                    "source": label,
                    "payload": {"old": old, "new": new, "field": field},
                }
            )
    return events
