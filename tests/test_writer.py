from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

from moex_bond_screener.writer import save_bonds_file, save_emitents_excel


@pytest.fixture
def bonds_sample() -> list[dict[str, object]]:
    return [
        {
            "SECID": "SU26218RMFS6",
            "SHORTNAME": "ОФЗ 26218",
            "ISIN": "RU000A0JVW48",
            "FACEUNIT": "SUR",
            "LISTLEVEL": 1,
            "PREVLEGALCLOSEPRICE": 99.13,
            "MATDATE": "2029-05-16",
        }
    ]


def test_save_bonds_file_as_excel(tmp_path: Path, bonds_sample: list[dict[str, object]]) -> None:
    target = tmp_path / "output" / "bonds.xlsx"

    save_bonds_file(str(target), bonds_sample)

    assert target.exists()
    workbook = load_workbook(target)
    sheet = workbook["MOEX_BONDS"]

    assert "SUMMARY" in workbook.sheetnames
    assert workbook.sheetnames[0] == "SUMMARY"
    summary_sheet = workbook["SUMMARY"]
    assert summary_sheet["A2"].value == "Дата и время формирования"
    assert summary_sheet["A3"].value == "Количество бумаг"
    assert summary_sheet["B3"].value == 1

    assert sheet["A1"].value == "Служебная информация"
    headers = [sheet.cell(row=2, column=idx).value for idx in range(1, sheet.max_column + 1)]

    assert "SECID" in headers
    assert "SHORTNAME" in headers
    assert "CURRENCYID" in headers
    assert "PREVLEGALCLOSEPRICE" in headers
    assert "MATDATE" in headers
    assert "|" not in headers

    shortname_column = headers.index("SHORTNAME") + 1
    matdate_column = headers.index("MATDATE") + 1
    assert sheet.cell(row=3, column=shortname_column).value == "ОФЗ 26218"
    assert sheet.cell(row=3, column=matdate_column).value == datetime(2029, 5, 16)
    assert sheet.cell(row=3, column=matdate_column).number_format == "DD.MM.YYYY"


def test_save_bonds_file_as_csv_with_bom(tmp_path: Path, bonds_sample: list[dict[str, object]]) -> None:
    target = tmp_path / "output" / "bonds.csv"

    save_bonds_file(str(target), bonds_sample)

    assert target.exists()
    content = target.read_text(encoding="utf-8-sig")
    assert "ОФЗ 26218" in content
    assert "MATDATE" in content
    assert "16.05.2029" in content


def test_save_bonds_excel_applies_readable_formatting(tmp_path: Path, bonds_sample: list[dict[str, object]]) -> None:
    target = tmp_path / "output" / "bonds.xlsx"

    save_bonds_file(str(target), bonds_sample)

    workbook = load_workbook(target)
    sheet = workbook["MOEX_BONDS"]

    assert sheet.freeze_panes == "A3"
    assert sheet.auto_filter.ref == f"A2:{sheet.cell(row=2, column=sheet.max_column).column_letter}3"
    assert sheet["A1"].font.bold is True
    assert sheet.column_dimensions["B"].width >= 10
    assert sheet["B3"].fill.fgColor.rgb == "00F2F7FF"
    assert sheet["A1"].alignment.wrap_text is True
    assert sheet.row_dimensions[1].height == 42


def test_save_bonds_file_removes_unwanted_and_merged_columns(tmp_path: Path) -> None:
    target = tmp_path / "output" / "bonds.xlsx"
    bonds = [
        {
            "SECID": "SU26218RMFS6",
            "SHORTNAME": "ОФЗ 26218",
            "FACEUNIT": "SUR",
            "CURRENCYID": "SUR",
            "BOARDID": "TQOB",
            "SECTORID": "GOV",
            "ISIN": "RU000A0JVW48",
            "MATDATE": "2029-05-16",
        }
    ]

    save_bonds_file(str(target), bonds)

    workbook = load_workbook(target)
    sheet = workbook["MOEX_BONDS"]
    headers = [sheet.cell(row=2, column=idx).value for idx in range(1, sheet.max_column + 1)]

    assert "BOARDID" not in headers
    assert "SECTORID" not in headers
    assert "FACEUNIT" not in headers
    assert "CURRENCYID" in headers


def test_save_bonds_excel_keeps_separator_columns_and_group_outline(tmp_path: Path) -> None:
    target = tmp_path / "output" / "bonds.xlsx"
    bonds = [
        {
            "SECID": "SU26218RMFS6",
            "SHORTNAME": "ОФЗ 26218",
            "ISIN": "RU000A0JVW48",
            "CURRENCYID": "SUR",
            "COUPONVALUE": 62.33,
        }
    ]

    save_bonds_file(str(target), bonds)

    workbook = load_workbook(target)
    sheet = workbook["MOEX_BONDS"]
    first_row = [sheet.cell(row=1, column=idx).value for idx in range(1, sheet.max_column + 1)]
    headers = [sheet.cell(row=2, column=idx).value for idx in range(1, sheet.max_column + 1)]

    assert not sheet.merged_cells.ranges

    separator_columns = [idx for idx, header in enumerate(headers, start=1) if header in ("", None)]
    assert len(separator_columns) >= 3

    first_separator = separator_columns[0]
    assert first_row[first_separator - 1] == "Служебная информация"
    assert sheet.column_dimensions[sheet.cell(row=1, column=first_separator).column_letter].width >= 18
    assert sheet.cell(row=1, column=first_separator).alignment.wrap_text is True

    for idx, value in enumerate(first_row, start=1):
        if idx in separator_columns:
            assert value in {
                "Служебная информация",
                "Торги и доходность",
                "Купоны и номинал",
                "Даты",
                "Прочее",
            }
        else:
            assert value in ("", None)

    data_columns_outline = [
        idx
        for idx, header in enumerate(headers, start=1)
        if header not in ("", None)
        and sheet.column_dimensions[sheet.cell(row=2, column=idx).column_letter].outlineLevel == 1
    ]
    assert data_columns_outline
    assert headers[first_separator] == "SHORTNAME"

    separator_letter = sheet.cell(row=2, column=first_separator).column_letter
    separator_fill_row_2 = sheet.cell(row=2, column=first_separator).fill.fgColor.rgb
    separator_fill_last_row = sheet.cell(row=sheet.max_row, column=first_separator).fill.fgColor.rgb
    assert separator_letter
    assert separator_fill_row_2 == separator_fill_last_row


def test_save_bonds_excel_formats_numeric_columns_and_empty_zero_dates(tmp_path: Path) -> None:
    target = tmp_path / "output" / "bonds.xlsx"
    bonds = [
        {
            "SHORTNAME": "ОФЗ 26218",
            "ISSUESIZE": 1000000000,
            "ISSUESIZEPLACED": 950000000,
            "COUPONVALUE": 62.33,
            "BIDVALUE": 100500.556,
            "ACCRUEDINT": 100500,
            "MATDATE": "0000-00-00",
            "NEXTCOUPON": "2027-01-01",
        }
    ]

    save_bonds_file(str(target), bonds)

    workbook = load_workbook(target)
    sheet = workbook["MOEX_BONDS"]
    headers = [sheet.cell(row=2, column=idx).value for idx in range(1, sheet.max_column + 1)]

    issuesize_column = headers.index("ISSUESIZE") + 1
    issuesizeplaced_column = headers.index("ISSUESIZEPLACED") + 1
    coupon_column = headers.index("COUPONVALUE") + 1
    bidvalue_column = headers.index("BIDVALUE") + 1
    accruedint_column = headers.index("ACCRUEDINT") + 1
    matdate_column = headers.index("MATDATE") + 1
    nextcoupon_column = headers.index("NEXTCOUPON") + 1

    assert sheet.cell(row=3, column=matdate_column).value in ("", None)
    assert sheet.cell(row=3, column=issuesize_column).number_format == "#,##0"
    assert sheet.cell(row=3, column=issuesizeplaced_column).number_format == "#,##0"
    assert sheet.cell(row=3, column=coupon_column).number_format == "#,##0.00"
    assert sheet.cell(row=3, column=bidvalue_column).number_format == "#,##0.00"
    assert sheet.cell(row=3, column=accruedint_column).number_format == "#,##0"
    assert sheet.cell(row=3, column=nextcoupon_column).value == datetime(2027, 1, 1)
    assert sheet.cell(row=3, column=nextcoupon_column).number_format == "DD.MM.YYYY"


def test_save_emitents_excel(tmp_path: Path) -> None:
    target = tmp_path / "output" / "emitents.xlsx"
    rows = [
        {
            "Полное наименование": "ПАО Тест",
            "ИНН": "7701000000",
            "Тикеры акций": "TST",
            "ISIN облигаций": "RU000000001",
            "missing_full_name": "0",
            "missing_inn": "0",
            "Флаг качества": "ok",
        }
    ]

    save_emitents_excel(str(target), rows)

    workbook = load_workbook(target)
    sheet = workbook["EMITENTS"]
    assert sheet["A1"].value == "Полное наименование"
    assert sheet["B2"].value == "7701000000"
    assert sheet["G1"].value == "Флаг качества"
    assert sheet["G2"].value == "ok"
    assert sheet.freeze_panes == "A2"


def test_save_bonds_excel_converts_large_numeric_strings_with_any_thousands_groups(tmp_path: Path) -> None:
    target = tmp_path / "output" / "bonds.xlsx"
    bonds = [
        {
            "SHORTNAME": "Тест большой эмиссии",
            "ISSUESIZE": "1000000 000",
            "ISSUESIZEPLACED": "1 250 000 000 000",
            "COUPONVALUE": "62,33",
        }
    ]

    save_bonds_file(str(target), bonds)

    workbook = load_workbook(target)
    sheet = workbook["MOEX_BONDS"]
    headers = [sheet.cell(row=2, column=idx).value for idx in range(1, sheet.max_column + 1)]

    issuesize_column = headers.index("ISSUESIZE") + 1
    issuesizeplaced_column = headers.index("ISSUESIZEPLACED") + 1
    coupon_column = headers.index("COUPONVALUE") + 1

    assert sheet.cell(row=3, column=issuesize_column).value == 1_000_000_000
    assert sheet.cell(row=3, column=issuesizeplaced_column).value == 1_250_000_000_000
    assert sheet.cell(row=3, column=coupon_column).value == 62.33
    assert sheet.cell(row=3, column=issuesize_column).number_format == "#,##0"
    assert sheet.cell(row=3, column=issuesizeplaced_column).number_format == "#,##0"
    assert sheet.cell(row=3, column=coupon_column).number_format == "#,##0.00"




def test_save_bonds_excel_converts_numeric_strings_with_narrow_nbsp(tmp_path: Path) -> None:
    target = tmp_path / "output" / "bonds.xlsx"
    bonds = [
        {
            "SHORTNAME": "Тест узкого неразрывного пробела",
            "ISSUESIZE": "1 000 000 000",
            "ISSUESIZEPLACED": "750 000 000",
        }
    ]

    save_bonds_file(str(target), bonds)

    workbook = load_workbook(target)
    sheet = workbook["MOEX_BONDS"]
    headers = [sheet.cell(row=2, column=idx).value for idx in range(1, sheet.max_column + 1)]

    issuesize_column = headers.index("ISSUESIZE") + 1
    issuesizeplaced_column = headers.index("ISSUESIZEPLACED") + 1

    assert sheet.cell(row=3, column=issuesize_column).value == 1_000_000_000
    assert sheet.cell(row=3, column=issuesizeplaced_column).value == 750_000_000
    assert sheet.cell(row=3, column=issuesize_column).number_format == "#,##0"
    assert sheet.cell(row=3, column=issuesizeplaced_column).number_format == "#,##0"


def test_save_bonds_excel_writes_summary_values_from_run_metadata(tmp_path: Path) -> None:
    target = tmp_path / "output" / "bonds.xlsx"
    bonds = [{"SHORTNAME": "ОФЗ 26218", "ISSUESIZE": "1000000 000"}]

    save_bonds_file(
        str(target),
        bonds,
        summary={"bonds_count": 77, "errors_count": 3, "elapsed_seconds": 12.34},
    )

    workbook = load_workbook(target)
    summary_sheet = workbook["SUMMARY"]

    assert summary_sheet["B3"].value == 77
    assert summary_sheet["B4"].value == 3
    assert summary_sheet["B5"].value == 12.34
    assert summary_sheet["B5"].number_format == "0.00"



def test_save_bonds_excel_writes_extra_summary_metrics(tmp_path: Path) -> None:
    target = tmp_path / "output" / "bonds.xlsx"
    bonds = [{"SECID": "A", "SHORTNAME": "A"}]

    save_bonds_file(
        str(target),
        bonds,
        summary={
            "bonds_count": 1,
            "errors_count": 0,
            "elapsed_seconds": 1.25,
            "excluded_offer_lt_1y": 3,
        },
    )

    workbook = load_workbook(target)
    summary_sheet = workbook["SUMMARY"]

    labels = [summary_sheet.cell(row=row, column=1).value for row in range(1, summary_sheet.max_row + 1)]
    values = [summary_sheet.cell(row=row, column=2).value for row in range(1, summary_sheet.max_row + 1)]

    assert "Excluded offer lt 1y" in labels
    row_index = labels.index("Excluded offer lt 1y")
    assert values[row_index] == 3

def test_save_bonds_file_with_unsupported_extension(tmp_path: Path, bonds_sample: list[dict[str, object]]) -> None:
    target = tmp_path / "output" / "bonds.txt"

    with pytest.raises(ValueError, match=".xlsx и .csv"):
        save_bonds_file(str(target), bonds_sample)


def test_save_bonds_excel_places_amortization_start_date_into_dates_group(tmp_path: Path) -> None:
    target = tmp_path / "output" / "bonds.xlsx"
    bonds = [{"SECID": "A", "SHORTNAME": "A", "Amortization_start_date": "2025-06-01"}]

    save_bonds_file(str(target), bonds)

    workbook = load_workbook(target)
    sheet = workbook["MOEX_BONDS"]

    headers = [sheet.cell(row=2, column=idx).value for idx in range(1, sheet.max_column + 1)]
    amort_column = headers.index("Amortization_start_date") + 1

    assert sheet.cell(row=1, column=amort_column - 1).value == "Даты"
    assert sheet.cell(row=3, column=amort_column).value == datetime(2025, 6, 1)


def test_save_bonds_excel_marks_approx_coupon_in_yellow(tmp_path: Path) -> None:
    target = tmp_path / "output" / "bonds.xlsx"
    bonds = [
        {
            "SECID": "RU1",
            "SHORTNAME": "Тест",
            "COUPONPERCENT": 13.25,
            "_COUPONPERCENT_APPROX": True,
        }
    ]

    save_bonds_file(str(target), bonds)

    workbook = load_workbook(target)
    sheet = workbook["MOEX_BONDS"]
    headers = [sheet.cell(row=2, column=idx).value for idx in range(1, sheet.max_column + 1)]
    coupon_col = headers.index("COUPONPERCENT") + 1

    assert sheet.cell(row=3, column=coupon_col).fill.fgColor.rgb == "00FFF59D"
